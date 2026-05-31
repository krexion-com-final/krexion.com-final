"""
Real User Traffic — unified Real-Traffic + Form-Filler runner.

Each visit = one browser session through a residential proxy with a
UA-parsed device fingerprint. Filters happen BEFORE the click is sent:
    · allowed_os        — parse each UA, skip if OS not allowed
    · allowed_countries — probe proxy exit-IP, skip if country not allowed
    · skip_vpn          — skip if exit-IP is flagged proxy/hosting
    · skip_duplicate_ip — skip if exit-IP already exists in user's clicks
    · no_repeated_proxy — each proxy line used at most once per run

If form_fill_enabled is on, after the tracker click we multi-step-fill the
landing form with a row from the uploaded Excel / Google Sheet, take a
final-page screenshot, capture TrustedForm/LeadID proof, and zip everything.

Output:  results.zip  containing  screenshots/*.png  +  report.xlsx
"""
from __future__ import annotations
import asyncio
import io
import os
import random
import time
import uuid
import zipfile
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Callable, Awaitable

if not os.environ.get("PLAYWRIGHT_BROWSERS_PATH") and os.path.isdir("/pw-browsers"):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "/pw-browsers"


# Serialize concurrent "ensure chromium installed" attempts. The preview pod's
# ephemeral filesystem can wipe the chromium binary across restarts, and the
# non-blocking startup hook in server.py may not finish before the first
# RUT job fires — so before each job launch we synchronously verify the
# browser is present, installing it with a lock if missing. This guarantees
# the very first visit NEVER fails with "Executable doesn't exist".
_CHROMIUM_INSTALL_LOCK = asyncio.Lock()
# Tracks whether an install is currently in progress so the engine-status
# API can report "installing" instead of just "missing" while the binary
# is being downloaded.
_CHROMIUM_INSTALL_IN_PROGRESS = False


# ──────────────────────────────────────────────────────────────────────
# 2026-05 — Offer-side duplicate / VPN block retry signal
# ──────────────────────────────────────────────────────────────────────
# When the offer's landing page rejects an exit-IP as "Duplicate IP" or
# "VPN/proxy detected", we want the visit slot to TRANSPARENTLY retry
# with a fresh ProxyJet IP — without recording a wasted entry in the
# report. This internal exception is raised by `process_one` and caught
# by the `worker` wrapper which loops up to `proxyjet_unique_retry_cap`
# times. The IP that triggered the block is already burned & persisted
# to `rut_burnt_ips` before the exception, so subsequent retries within
# the SAME job AND every future job will skip it at the on-demand
# probe stage.
class _OfferBlockRetryNeeded(Exception):
    """Internal signal: offer rejected this IP, retry with fresh IP without recording."""
    def __init__(self, reason: str = "", burnt_ip: str = ""):
        super().__init__(f"{reason}:{burnt_ip}")
        self.reason = reason
        self.burnt_ip = burnt_ip


def get_engine_status() -> Dict[str, Any]:
    """Return the current state of the Playwright chromium-headless-shell
    binary so the frontend can show a coloured "Engine Status" badge:
        ready      → binary present at the EXACT revision Playwright wants
        installing → install_in_progress flag is set
        missing    → binary absent and no install in progress
        error      → couldn't read browsers.json (unexpected)
    """
    browsers_root = os.environ.get("PLAYWRIGHT_BROWSERS_PATH", "/pw-browsers")
    expected: Optional[str] = None
    try:
        import json as _json
        import playwright as _pw
        bj = Path(_pw.__file__).parent / "driver" / "package" / "browsers.json"
        if bj.exists():
            with open(bj, "r") as fh:
                data = _json.load(fh)
            for entry in data.get("browsers", []):
                if entry.get("name") == "chromium-headless-shell":
                    expected = str(entry.get("revision") or "").strip() or None
                    break
    except Exception:
        expected = None

    if not expected:
        return {
            "status": "error",
            "message": "Cannot read Playwright revision metadata",
            "expected_revision": None,
            "browser_path": None,
        }

    binary_path = Path(browsers_root) / f"chromium_headless_shell-{expected}" / "chrome-linux" / "headless_shell"
    # 2026-01: Detect which engine is actually being used at runtime so
    # the admin engine-status badge can show "Full Chromium (--headless=new)"
    # vs "Headless Shell (legacy)". Helps confirm the anti-detect upgrade
    # took effect on a deployed instance.
    engine_mode = "full-chromium-headless-new" if _use_full_chromium() else "chromium-headless-shell"
    full_chromium_path = _full_chromium_binary_path()
    extra = {
        "engine_mode": engine_mode,
        "full_chromium_installed": full_chromium_path is not None,
        "full_chromium_path": str(full_chromium_path) if full_chromium_path else None,
    }
    # If either binary is present we're ready. Prefer full chromium for
    # anti-detect, but the legacy headless_shell remains a valid fallback.
    if binary_path.exists() or full_chromium_path is not None:
        ready_via = "full chromium (--headless=new)" if full_chromium_path is not None else f"headless_shell rev {expected}"
        return {
            "status": "ready",
            "message": f"Chromium ready · using {ready_via}",
            "expected_revision": expected,
            "browser_path": str(full_chromium_path or binary_path),
            **extra,
        }
    if _CHROMIUM_INSTALL_IN_PROGRESS:
        return {
            "status": "installing",
            "message": f"Downloading Chromium rev {expected}…",
            "expected_revision": expected,
            "browser_path": str(binary_path),
            **extra,
        }
    return {
        "status": "missing",
        "message": f"Chromium rev {expected} not installed yet",
        "expected_revision": expected,
        "browser_path": str(binary_path),
        **extra,
    }


async def _ensure_chromium_available() -> bool:
    """Returns True when the EXACT chromium-headless-shell revision required
    by the installed Playwright Python package is present (installing it
    first if missing). Safe to call before every job — no-op when binary
    is already present.

    NOTE: Earlier versions of this helper used a glob pattern
    `chromium_headless_shell-*` which matched ANY revision present on disk
    (e.g. an old 1208 left over from a previous Playwright upgrade) and
    falsely returned True even though Playwright 1.49.x specifically wanted
    revision 1148 → BrowserType.launch() blew up with "Executable doesn't
    exist at /pw-browsers/chromium_headless_shell-1148/...". We now read
    the EXACT revision from Playwright's bundled browsers.json and verify
    that specific path exists.
    """
    browsers_root = os.environ.get("PLAYWRIGHT_BROWSERS_PATH", "/pw-browsers")

    def _expected_revision() -> Optional[str]:
        """Read the chromium-headless-shell revision Playwright expects.
        Falls back to None if the JSON layout changes."""
        try:
            import json as _json
            import playwright as _pw
            pw_root = Path(_pw.__file__).parent
            bj = pw_root / "driver" / "package" / "browsers.json"
            if not bj.exists():
                return None
            with open(bj, "r") as fh:
                data = _json.load(fh)
            for entry in data.get("browsers", []):
                if entry.get("name") == "chromium-headless-shell":
                    rev = str(entry.get("revision") or "").strip()
                    return rev or None
        except Exception as e:
            logger.debug(f"_expected_revision: {e}")
        return None

    def _binary_for(rev: Optional[str]) -> Optional[Path]:
        if not rev:
            return None
        return Path(browsers_root) / f"chromium_headless_shell-{rev}" / "chrome-linux" / "headless_shell"

    expected = _expected_revision()

    def _exists() -> bool:
        # Strict check: the EXACT revision Playwright wants must be present.
        if expected:
            bp = _binary_for(expected)
            if bp and bp.exists():
                return True
            return False
        # Fallback (only when we can't read browsers.json): glob check.
        try:
            for p in Path(browsers_root).glob("chromium_headless_shell-*"):
                if (p / "chrome-linux" / "headless_shell").exists():
                    return True
        except Exception:
            pass
        return False

    if _exists():
        # ── 2026-01 Anti-detect: kick off background install of FULL
        # chromium (no-op if already present). Lets existing deploys
        # automatically upgrade to --headless=new mode on next job.
        try:
            if _full_chromium_binary_path() is None:
                asyncio.create_task(_install_full_chromium_background())
        except Exception:
            pass
        return True

    # ── 2026-01 Anti-detect: if the FULL chromium binary is already
    # installed, we don't need chromium-headless-shell at all — the
    # launcher will use full chromium with --headless=new. Treat the
    # engine as ready and skip the (sometimes-failing) shell install.
    if _full_chromium_binary_path() is not None:
        return True

    # Missing — install with a lock to prevent duplicate installs when
    # multiple jobs start in parallel on a fresh pod.
    async with _CHROMIUM_INSTALL_LOCK:
        # Re-check after acquiring lock (another coroutine may have just
        # finished the install while we waited).
        if _exists():
            return True
        global _CHROMIUM_INSTALL_IN_PROGRESS
        _CHROMIUM_INSTALL_IN_PROGRESS = True
        try:
            logger.warning(
                f"Playwright chromium-headless-shell rev {expected or '?'} missing — "
                f"installing now (this may take ~60s)…"
            )
            try:
                proc = await asyncio.create_subprocess_exec(
                    "playwright", "install", "chromium-headless-shell",
                    env={**os.environ, "PLAYWRIGHT_BROWSERS_PATH": browsers_root},
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                )
                try:
                    _out, err = await asyncio.wait_for(proc.communicate(), timeout=300)
                except asyncio.TimeoutError:
                    try: proc.kill()
                    except Exception: pass
                    logger.error("Playwright install timed out after 5 min")
                    return False
                if proc.returncode != 0:
                    logger.error(
                        f"Playwright install returned {proc.returncode}: "
                        f"{(err or b'').decode(errors='ignore')[:300]}"
                    )
                    return False
                logger.info(
                    f"Playwright chromium-headless-shell install: OK (rev {expected or '?'})"
                )
            except Exception as e:
                logger.error(f"Playwright install failed: {e}")
                return False
            # Final strict check — must satisfy the EXACT revision Playwright wants
            ok = _exists()
        finally:
            _CHROMIUM_INSTALL_IN_PROGRESS = False

    # ── 2026-01 Anti-detect: also ensure FULL chromium is installed ──
    # Fires in the background — never blocks job execution. On the first
    # run after upgrade, the next job (or two) still uses chromium-
    # headless-shell, then once full chromium finishes downloading
    # (~165MB), `_use_full_chromium()` flips to True and subsequent jobs
    # automatically launch with `--headless=new` for maximum stealth.
    try:
        if ok and _full_chromium_binary_path() is None:
            asyncio.create_task(_install_full_chromium_background())
    except Exception:
        pass
    return ok


import httpx
import pandas as pd
from user_agents import parse as ua_parse
from playwright.async_api import async_playwright, Page, BrowserContext, Browser


# ─── Full Chromium detection (2026-01: anti-detect upgrade) ────────
# When the FULL chromium binary is installed alongside the lightweight
# chromium-headless-shell, we prefer it because:
#   1. Identical code paths to headed Chrome → bypasses headless-shell
#      detection heuristics (Anura Premium, IPQS Deep, PerimeterX).
#   2. Full font set, real GPU pipeline (SwiftShader), full audio
#      subsystem — canvas / WebGL / audio fingerprints match real Chrome.
#   3. With `--headless=new` flag, runs the SAME binary as headed Chrome
#      just without a visible window. Detectors can't distinguish.
#
# Falls back to chromium-headless-shell automatically if full chromium
# isn't installed (e.g. on customer VPS that hasn't run the upgrade).
# This keeps the codebase backwards-compatible.

def _full_chromium_binary_path() -> Optional[Path]:
    """Return the path to the full chromium binary if installed, else None.
    Reads the expected revision from Playwright's browsers.json (same one
    used for headless-shell) so the rev always stays in sync."""
    browsers_root = os.environ.get("PLAYWRIGHT_BROWSERS_PATH", "/pw-browsers")
    try:
        import json as _json
        import playwright as _pw
        bj = Path(_pw.__file__).parent / "driver" / "package" / "browsers.json"
        if not bj.exists():
            return None
        with open(bj, "r") as fh:
            data = _json.load(fh)
        rev = None
        for entry in data.get("browsers", []):
            if entry.get("name") == "chromium":
                rev = str(entry.get("revision") or "").strip() or None
                break
        if not rev:
            return None
        bp = Path(browsers_root) / f"chromium-{rev}" / "chrome-linux" / "chrome"
        return bp if bp.exists() else None
    except Exception:
        return None


def _use_full_chromium() -> bool:
    """Should we launch the full chromium (with --headless=new) instead of
    the lightweight chromium-headless-shell? Yes when the binary is
    installed AND the env override `KREXION_FORCE_HEADLESS_SHELL=1` is NOT
    set (operators can flip this to revert if a new bug appears)."""
    if os.environ.get("KREXION_FORCE_HEADLESS_SHELL", "").strip() in ("1", "true", "yes"):
        return False
    return _full_chromium_binary_path() is not None


async def _install_full_chromium_background() -> None:
    """Best-effort background install of the FULL chromium binary so that
    later jobs can use --headless=new mode. Logs success/failure but
    never raises — chromium-headless-shell remains the safety net."""
    if _full_chromium_binary_path() is not None:
        return
    browsers_root = os.environ.get("PLAYWRIGHT_BROWSERS_PATH", "/pw-browsers")
    try:
        logger.info("Installing full chromium binary for --headless=new mode…")
        proc = await asyncio.create_subprocess_exec(
            "playwright", "install", "chromium", "--no-shell",
            env={**os.environ, "PLAYWRIGHT_BROWSERS_PATH": browsers_root},
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        try:
            _out, err = await asyncio.wait_for(proc.communicate(), timeout=600)
        except asyncio.TimeoutError:
            try: proc.kill()
            except Exception: pass
            logger.warning("Full chromium install timed out — staying on headless-shell")
            return
        if proc.returncode == 0:
            logger.info("Full chromium installed — --headless=new mode enabled for subsequent jobs")
        else:
            logger.warning(
                f"Full chromium install returned {proc.returncode}: "
                f"{(err or b'').decode(errors='ignore')[:200]} — staying on headless-shell"
            )
    except Exception as e:
        logger.warning(f"Full chromium install failed: {e} — staying on headless-shell")


# Shared launch-args list — kept here so both the primary launch and the
# in-flight crash-recovery relaunch use the EXACT same flags. Order matters
# for some flags (e.g. --headless=new must precede --disable-features when
# they share keys), so we keep this as a single source of truth.
#
# 2026-05 — single consolidated --disable-features=... entry: Chromium
# honours ONLY the LAST --disable-features arg, so all keys MUST be in
# one comma-separated list:
#   • WebRtcHideLocalIpsWithMdns,AutomationControlled — original anti-detect
#   • UseDnsHttpsSvcb — block HTTPS DNS records that bypass our AAAA-skip
#     and reintroduce IPv6 on Chrome 100+
_BROWSER_LAUNCH_ARGS_BASE = [
    "--no-sandbox",
    "--disable-dev-shm-usage",
    "--disable-features=WebRtcHideLocalIpsWithMdns,AutomationControlled,UseDnsHttpsSvcb",
    "--disable-blink-features=AutomationControlled",
    "--force-webrtc-ip-handling-policy=disable_non_proxied_udp",
    "--disable-extensions",
    "--disable-background-networking",
    "--disable-background-timer-throttling",
    "--disable-backgrounding-occluded-windows",
    "--disable-renderer-backgrounding",
    "--disable-sync",
    "--disable-translate",
    "--disable-default-apps",
    "--disable-component-update",
    "--no-first-run",
    "--no-default-browser-check",
    "--metrics-recording-only",
    # 2026-01 Anti-detect: removed `--mute-audio` — real Chrome doesn't
    # launch with audio muted by default, and detectors comparing
    # AudioContext.state can flag the discrepancy. Audio output stays
    # silent anyway because no <audio>/<video> auto-plays on lead pages.
]


async def _launch_anti_detect_browser(pw) -> Browser:
    """Launch Chromium for RUT jobs with the strongest anti-detect setup
    available on this host:

      1. If the FULL chromium binary is installed → launch it with
         `--headless=new` flag. This is the SAME binary as headed Chrome,
         just without a visible window. Detectors that compare
         headless-specific signals (font rendering, audio subsystem,
         GPU pipeline) cannot distinguish it from a real user's browser.

      2. If only chromium-headless-shell is installed → launch it the
         legacy way. Fully backwards-compatible — customer VPS instances
         that haven't run the upgrade keep working exactly as before.

    Returns the launched Browser. Caller is responsible for closing it.
    """
    if _use_full_chromium():
        # Full chromium: pass --headless=new explicitly via args. We
        # ALSO set headless=False so Playwright doesn't pass --headless
        # (legacy mode) on top of our --headless=new. The two together
        # would force old headless and defeat the whole point.
        try:
            return await pw.chromium.launch(
                channel="chromium",
                headless=False,
                args=["--headless=new", *_BROWSER_LAUNCH_ARGS_BASE],
            )
        except Exception as e:
            # Full chromium failed (missing system lib, GPU-init crash,
            # etc.) — fall back transparently to headless-shell so the
            # job still runs. Logged so operators can investigate.
            logger.warning(
                f"Full chromium launch failed ({e}) — falling back to "
                f"chromium-headless-shell"
            )
    return await pw.chromium.launch(
        headless=True,
        args=list(_BROWSER_LAUNCH_ARGS_BASE),
    )


from form_filler import (
    load_rows_from_excel,
    load_rows_from_google_sheet,
    _page_has_captcha,
    _dismiss_popups,
    _ensure_form_visible,
    _fill_form,
    _click_submit,
    _dismiss_review_modal,
    _tick_consent_checkboxes,
)

# 2026-01: New additive automation extensions (iframe-aware lookup,
# cookie-banner auto-dismiss, bot-block detection, extract variables,
# wait_for_text / wait_for_url, formatter pipeline, retry, friendly errors,
# lint). Pure helpers — when a step does not opt-in to these, behaviour
# is unchanged.
try:
    from automation_extensions import (
        auto_dismiss_cookie_banners as _ext_dismiss_cookies,
        detect_bot_block as _ext_detect_bot,
        find_frame_with_selector as _ext_find_frame,
        selector_exists_in_shadow as _ext_in_shadow,
        extract_to_row as _ext_extract,
        wait_for_text as _ext_wait_text,
        wait_for_url as _ext_wait_url,
        apply_formatters as _ext_apply_formatters,
        split_placeholder_pipeline as _ext_split_pipeline,
        friendly_error as _ext_friendly_error,
        run_with_retry as _ext_run_with_retry,
        latest_popup_page as _ext_latest_popup,
    )
    _EXT_LOADED = True
except Exception as _ext_err:  # pragma: no cover
    _EXT_LOADED = False
    logging.getLogger(__name__).warning(f"automation_extensions failed to load: {_ext_err}")

logger = logging.getLogger(__name__)

RESULTS_ROOT = Path("/app/backend/real_user_traffic_results")
RESULTS_ROOT.mkdir(parents=True, exist_ok=True)

RUT_JOBS: Dict[str, Dict[str, Any]] = {}


# ──────────────────────────────────────────────────────────────────────
# UA → Referer auto-detection
# ──────────────────────────────────────────────────────────────────────
# When a residential visit's user-agent string identifies an in-app
# browser (TikTok, Facebook, Instagram, etc.), the destination tracker
# expects the Referer header to match the platform — otherwise the
# visit looks like a bare-browser hit and lifetime-value attribution
# breaks. This helper inspects the UA and returns the matching Referer
# URL. The engine injects it into `extra_http_headers` when creating
# the Playwright context (see `_get_referer_from_ua` call sites).
#
# Match order: most specific first (long substrings before generic).
# All match strings are case-INsensitive against the original UA.
# Returns "" for plain browser UAs — engine will either leave Referer
# unset or fall back to a Google search referer per its existing logic.
# ──────────────────────────────────────────────────────────────────────
_UA_REFERER_MAP: List[Tuple[Tuple[str, ...], str]] = [
    # In-app browsers — these strings appear in mobile webview UAs
    (("musical_ly", "musically", "tiktok", "ttwebview"), "https://www.tiktok.com/"),
    (("fban/", "fbav/", "fbios", "fb_iab", "fb4a"), "https://www.facebook.com/"),
    (("messenger", "fbms"),                        "https://www.facebook.com/"),
    (("instagram",),                               "https://www.instagram.com/"),
    (("snapchat",),                                "https://www.snapchat.com/"),
    (("pinterest",),                               "https://www.pinterest.com/"),
    (("reddit",),                                  "https://www.reddit.com/"),
    (("twitter", "twitterandroid", "twitterios"),  "https://twitter.com/"),
    (("linkedinapp",),                             "https://www.linkedin.com/"),
    (("whatsapp", "wa-android", "wa-ios"),         "https://www.whatsapp.com/"),
    (("telegram",),                                "https://t.me/"),
    (("youtube",),                                 "https://www.youtube.com/"),
    (("tumblr",),                                  "https://www.tumblr.com/"),
    (("kakaotalk",),                               "https://www.kakaocorp.com/"),
    (("line/",),                                   "https://line.me/"),
    (("wechat", "micromessenger"),                 "https://www.wechat.com/"),
    (("discord",),                                 "https://discord.com/"),
    (("slack",),                                   "https://slack.com/"),
    # Search-engine bots / crawlers — keep Referer empty (don't fake)
    (("googlebot", "bingbot", "yandexbot", "duckduckbot"), ""),
]


def _get_referer_from_ua(ua: str) -> str:
    """Return the matching Referer URL for an in-app-browser UA, or "" if
    the UA is a plain mobile/desktop browser.

    Examples:
        "Mozilla/5.0 ... TikTok/26.5.0 ..."     → "https://www.tiktok.com/"
        "Mozilla/5.0 ... [FBAN/FBIOS;FBAV/...]" → "https://www.facebook.com/"
        "Mozilla/5.0 ... Instagram 280..."      → "https://www.instagram.com/"
        "Mozilla/5.0 (iPhone; CPU iPhone OS...) Safari/..."  → ""  (plain Safari)
    """
    if not ua:
        return ""
    ua_lower = ua.lower()
    for needles, referer in _UA_REFERER_MAP:
        for needle in needles:
            if needle in ua_lower:
                return referer
    return ""


# ──────────────────────────────────────────────────────────────────────
# 2026-01: ENHANCED ANTI-DETECT HELPERS
# ──────────────────────────────────────────────────────────────────────
# These helpers feed the per-visit fingerprint with extra fields that
# the stealth init-script uses to mask the headless Chrome signature
# against advanced fraud-detection tools (MaxMind minFraud, IPQuality
# Score, Anura, ArkoseLabs). All helpers are pure-Python, dependency-
# free, and never raise — failures fall back to safe defaults so the
# RUT engine's existing flow is NEVER blocked by these enhancements.
# ──────────────────────────────────────────────────────────────────────
import re as _re

_CHROME_VERSION_RE = _re.compile(r"(?:Chrome|CriOS|Edg|EdgA|EdgiOS|Chromium)/(\d+)(?:\.(\d+))?", _re.IGNORECASE)
_SAFARI_VERSION_RE = _re.compile(r"Version/(\d+)(?:\.(\d+))?", _re.IGNORECASE)


def _extract_chrome_version(ua: str) -> Tuple[int, int]:
    """Return (major, minor) of the Chrome/Chromium version in the UA.

    Handles all Chromium-family browsers including:
      · Chrome desktop:  ".../Chrome/147.0.7727.102 ..."
      · Chrome iOS:      ".../CriOS/147.0.7727.102 ..."
      · Edge desktop:    ".../Edg/147.0.7727.102 ..."
      · Edge Android:    ".../EdgA/147.0.7727.102 ..."
      · Edge iOS:        ".../EdgiOS/147.0.7727.102 ..."
      · Chromium:        ".../Chromium/147.0.7727.102 ..."
      · Any in-app webview that embeds "Chrome/X" (Instagram, Facebook,
        TikTok Android, etc. — extracts the underlying Chrome version)

    Returns (142, 0) as a safe fallback for non-Chromium UAs (pure
    Safari / Firefox / WhatsApp Darwin) so Sec-CH-UA generation never
    crashes; callers also gate Sec-CH-UA emission on is_chromium so
    a Safari UA never gets bogus Chrome hints.
    """
    if ua:
        m = _CHROME_VERSION_RE.search(ua)
        if m:
            return int(m.group(1)), int(m.group(2) or 0)
    return 142, 0


def _extract_safari_version(ua: str) -> Tuple[int, int]:
    """Return (major, minor) of the Safari version in the UA, or (17, 0)."""
    if ua:
        m = _SAFARI_VERSION_RE.search(ua)
        if m:
            return int(m.group(1)), int(m.group(2) or 0)
    return 17, 0


# Per-OS realistic system font lists. Fingerprint libraries enumerate a
# known font set and bucket the result; matching one of these distributions
# makes the fingerprint indistinguishable from a real user's machine.
_OS_FONTS = {
    "windows": [
        "Arial", "Arial Black", "Arial Narrow", "Bahnschrift", "Calibri",
        "Cambria", "Cambria Math", "Candara", "Comic Sans MS", "Consolas",
        "Constantia", "Corbel", "Courier", "Courier New", "Ebrima",
        "Franklin Gothic Medium", "Gabriola", "Gadugi", "Georgia",
        "Impact", "Ink Free", "Javanese Text", "Leelawadee UI",
        "Lucida Console", "Lucida Sans Unicode", "Malgun Gothic", "Marlett",
        "Microsoft Himalaya", "Microsoft JhengHei", "Microsoft New Tai Lue",
        "Microsoft PhagsPa", "Microsoft Sans Serif", "Microsoft Tai Le",
        "Microsoft YaHei", "Microsoft Yi Baiti", "MingLiU-ExtB",
        "Mongolian Baiti", "MS Gothic", "MV Boli", "Myanmar Text",
        "Nirmala UI", "Palatino Linotype", "Segoe MDL2 Assets",
        "Segoe Print", "Segoe Script", "Segoe UI", "Segoe UI Emoji",
        "Segoe UI Historic", "Segoe UI Symbol", "SimSun", "Sitka",
        "Sylfaen", "Symbol", "Tahoma", "Times New Roman", "Trebuchet MS",
        "Verdana", "Webdings", "Wingdings", "Yu Gothic",
    ],
    "macos": [
        "American Typewriter", "Andale Mono", "Arial", "Arial Black",
        "Arial Narrow", "Arial Rounded MT Bold", "Avenir", "Avenir Next",
        "Avenir Next Condensed", "Baskerville", "Big Caslon", "Bodoni 72",
        "Bradley Hand", "Brush Script MT", "Chalkboard", "Chalkboard SE",
        "Chalkduster", "Charter", "Cochin", "Comic Sans MS", "Copperplate",
        "Courier", "Courier New", "Didot", "DIN Alternate", "DIN Condensed",
        "Futura", "Geneva", "Georgia", "Gill Sans", "Helvetica",
        "Helvetica Neue", "Herculanum", "Hoefler Text", "Impact",
        "Lucida Grande", "Luminari", "Marker Felt", "Menlo", "Monaco",
        "Noteworthy", "Optima", "Palatino", "Papyrus", "Phosphate",
        "Rockwell", "Savoye LET", "SignPainter", "Skia", "Snell Roundhand",
        "Tahoma", "Times", "Times New Roman", "Trattatello", "Trebuchet MS",
        "Verdana", "Zapfino",
    ],
    "android": [
        "Roboto", "Noto Sans", "Noto Serif", "Noto Color Emoji",
        "Droid Sans", "Droid Sans Mono", "Droid Serif", "Cutive Mono",
        "Coming Soon", "Dancing Script", "Carrois Gothic", "Carrois Gothic SC",
        "sans-serif", "sans-serif-condensed", "sans-serif-light",
        "sans-serif-medium", "sans-serif-thin", "serif", "monospace",
    ],
    "ios": [
        ".SF NS Text", "Academy Engraved LET", "Al Nile", "American Typewriter",
        "Apple Color Emoji", "Apple SD Gothic Neo", "Arial", "Arial Hebrew",
        "Arial Rounded MT Bold", "Avenir", "Avenir Next", "Avenir Next Condensed",
        "Baskerville", "Bodoni 72", "Bradley Hand", "Chalkboard SE", "Chalkduster",
        "Charter", "Cochin", "Copperplate", "Courier", "Courier New",
        "Damascus", "Devanagari Sangam MN", "Didot", "Euphemia UCAS", "Farah",
        "Futura", "Galvji", "Geeza Pro", "Georgia", "Gill Sans",
        "Grantha Sangam MN", "Helvetica", "Helvetica Neue", "Impact",
        "Iowan Old Style", "Kefa", "Khmer Sangam MN", "Kohinoor Bangla",
        "Kohinoor Devanagari", "Marker Felt", "Menlo", "Monaco", "Myanmar Sangam MN",
        "Noteworthy", "Optima", "Palatino", "Papyrus", "Party LET",
        "PingFang HK", "PingFang SC", "PingFang TC", "Rockwell", "Savoye LET",
        "Snell Roundhand", "Symbol", "Thonburi", "Times New Roman", "Trebuchet MS",
        "Verdana", "Zapfino",
    ],
    "linux": [
        "DejaVu Sans", "DejaVu Sans Mono", "DejaVu Serif", "Liberation Mono",
        "Liberation Sans", "Liberation Serif", "Nimbus Mono", "Nimbus Sans",
        "Nimbus Roman", "Ubuntu", "Ubuntu Condensed", "Ubuntu Mono",
        "Cantarell", "Noto Sans", "Noto Serif", "Noto Mono", "FreeMono",
        "FreeSans", "FreeSerif", "URW Bookman", "URW Gothic", "URW Palladio",
    ],
}


def _build_client_hint_headers(fp: Dict[str, Any], ua: str) -> Dict[str, str]:
    """Build Sec-CH-UA / Sec-CH-UA-Mobile / Sec-CH-UA-Platform headers
    that match the user-agent's Chrome version and OS. Modern fraud
    detectors cross-check these against the UA string — a mismatch is
    a HARD bot signal. Returns {} for non-Chromium UAs so the existing
    Playwright defaults remain untouched.
    """
    headers: Dict[str, str] = {}
    os_key = fp.get("os", "")
    ua_l = (ua or "").lower()

    platform_label = {
        "windows": "Windows", "macos": "macOS", "ios": "iOS",
        "android": "Android", "linux": "Linux",
    }.get(os_key, "")
    if platform_label:
        headers["Sec-CH-UA-Platform"] = f'"{platform_label}"'

    headers["Sec-CH-UA-Mobile"] = "?1" if fp.get("is_mobile") else "?0"

    is_chromium = any(tok in ua_l for tok in (
        "chrome/", "crios/", "edg/", "edga/", "edgios/", "chromium/"
    ))
    if is_chromium:
        chrome_major, _ = _extract_chrome_version(ua)
        not_brand_variants = [
            '"Not_A Brand";v="24"',
            '"Not(A:Brand";v="24"',
            '"Not.A/Brand";v="24"',
            '" Not;A=Brand";v="99"',
        ]
        not_brand = random.choice(not_brand_variants)
        headers["Sec-CH-UA"] = (
            f'"Chromium";v="{chrome_major}", '
            f'"Google Chrome";v="{chrome_major}", '
            f'{not_brand}'
        )
        platform_version = {
            "windows": "15.0.0", "macos": "14.4.0", "ios": "17.4.0",
            "android": "14.0.0", "linux": "6.5.0",
        }.get(os_key, "")
        if platform_version:
            headers["Sec-CH-UA-Platform-Version"] = f'"{platform_version}"'

    return headers


# ──────────────────────────────────────────────────────────────────────
# 2026-01: HUMAN-LIKE BEHAVIOUR HELPERS
# ──────────────────────────────────────────────────────────────────────
# Anura / IPQS / ArkoseLabs measure mouse-movement entropy, scroll-
# velocity variance, and dwell time. A bot that jumps cursor straight
# to a button + fills a form in 800ms = instant fraud flag even if every
# JS API is masked. These helpers add realistic warm-up behaviour at
# the right moments without touching the existing form_filler flow.
# ──────────────────────────────────────────────────────────────────────
async def _human_warmup(page: Any, fp: Dict[str, Any]) -> None:
    """Simulate a real user landing on a page: read pause + mouse jitter
    + a few realistic scrolls. NEVER raises — every failure is swallowed
    so an unrelated runtime quirk in the page can't abort the visit."""
    try:
        viewport = fp.get("viewport") or {"width": 1280, "height": 800}
        vw = max(320, int(viewport.get("width", 1280)))
        vh = max(480, int(viewport.get("height", 800)))

        # Initial read pause — real users don't move for ~0.8-2.5s on land
        await asyncio.sleep(random.uniform(0.8, 2.5))

        # Mouse jitter — 3-6 small bezier-style moves
        if not fp.get("is_mobile"):
            try:
                x, y = random.randint(50, vw - 50), random.randint(50, vh - 50)
                await page.mouse.move(x, y, steps=random.randint(8, 20))
                await asyncio.sleep(random.uniform(0.15, 0.45))
                for _ in range(random.randint(3, 6)):
                    x = max(10, min(vw - 10, x + random.randint(-180, 180)))
                    y = max(10, min(vh - 10, y + random.randint(-120, 120)))
                    await page.mouse.move(x, y, steps=random.randint(6, 14))
                    await asyncio.sleep(random.uniform(0.08, 0.32))
            except Exception:
                pass

        # Scroll down 1-3 times then occasional scroll back up
        try:
            for _ in range(random.randint(1, 3)):
                await page.mouse.wheel(0, random.randint(180, 620))
                await asyncio.sleep(random.uniform(0.35, 1.4))
            if random.random() < 0.4:
                await page.mouse.wheel(0, -random.randint(80, 240))
                await asyncio.sleep(random.uniform(0.25, 0.8))
        except Exception:
            pass

        # Final settle pause
        await asyncio.sleep(random.uniform(0.3, 1.1))
    except Exception:
        pass


async def _human_pre_submit_dwell(page: Any, fp: Dict[str, Any]) -> None:
    """Pause + small jitter right before submit so the form-fill-to-submit
    interval matches a real human's review cadence (3-9s)."""
    try:
        await asyncio.sleep(random.uniform(2.2, 6.5))
        if not fp.get("is_mobile"):
            try:
                viewport = fp.get("viewport") or {"width": 1280, "height": 800}
                vw = max(320, int(viewport.get("width", 1280)))
                vh = max(480, int(viewport.get("height", 800)))
                x = random.randint(50, vw - 50)
                y = random.randint(50, vh - 50)
                await page.mouse.move(x, y, steps=random.randint(6, 16))
            except Exception:
                pass
        await asyncio.sleep(random.uniform(0.4, 1.5))
    except Exception:
        pass




def _device_name_from_ua(ua_str: str) -> str:
    """Extract a human-readable device name from a user-agent string.

    Strategy:
      1. Let the `user_agents` library try first (handles most Android device
         codes like "SM-S918U" → "Samsung SM-S918U" and "Pixel 8 Pro").
      2. For iOS UAs (which never contain the exact model), fall back to a
         regex over CPU / hardware hints, else return "iPhone" / "iPad".
      3. Windows/macOS/Linux: return OS name + short hardware hint.
    """
    import re
    s = ua_str or ""
    try:
        ua = ua_parse(s)
    except Exception:
        ua = None

    # 1. Library-parsed brand + model (works great for Android)
    if ua and ua.device:
        brand = (ua.device.brand or "").strip()
        model = (ua.device.model or "").strip()
        family = (ua.device.family or "").strip()
        if brand and model and brand.lower() not in model.lower():
            return f"{brand} {model}".strip()
        if model and model.lower() not in ("generic", "smartphone", "other"):
            return model
        if family and family.lower() not in ("generic", "smartphone", "other"):
            return family

    sl = s.lower()

    # 2. iOS — pull iPhone/iPad + model hint if we can, otherwise fall back
    if "iphone" in sl or ("cpu iphone os" in sl):
        # Try to find an explicit model (rare but possible): "iPhone 15 Pro"
        m = re.search(r"iPhone\s*(?:OS\s*)?(\d{1,2}(?:[._]\d{1,2})?)?\s*(?:Max|Pro|Plus|mini)?",
                      s, flags=re.I)
        # Prefer an iOS version tag ("iOS 17.1") to differentiate
        iosv = ""
        m2 = re.search(r"iphone os (\d+[._]\d+)", sl)
        if m2:
            iosv = m2.group(1).replace("_", ".")
        return f"iPhone (iOS {iosv})" if iosv else "iPhone"
    if "ipad" in sl:
        return "iPad"

    # 3. Android fallback when the library couldn't detect a model
    if "android" in sl:
        # Pull the "build" code inside parentheses, e.g. "(Linux; Android 14; SM-S918U)"
        m = re.search(r"android\s+[\d.]+;\s*([^;)\s][^;)]*?)(?:\)|;|\s+build)", s, flags=re.I)
        if m:
            model = m.group(1).strip()
            if model and not model.lower().startswith("linux"):
                return f"Android — {model}"
        return "Android"

    # 4. Desktop
    if "windows" in sl:
        m = re.search(r"windows nt (\d+\.\d+)", sl)
        return f"Windows {m.group(1)}" if m else "Windows PC"
    if "macintosh" in sl or "mac os x" in sl:
        return "Mac"
    if "linux" in sl:
        return "Linux"
    return "Unknown"


# ─── UA → device fingerprint ─────────────────────────────────────
def _os_key_from_ua(ua_str: str) -> str:
    """Return a lowercase OS key matching the frontend's allowed_os chips."""
    try:
        ua = ua_parse(ua_str or "")
        fam = (ua.os.family or "").lower()
    except Exception:
        fam = ""
    if "android" in fam:
        return "android"
    if "ios" in fam or "iphone" in fam or "ipad" in fam:
        return "ios"
    if "windows" in fam:
        return "windows"
    if "mac" in fam:
        return "macos"
    if any(k in fam for k in ("linux", "ubuntu", "fedora", "debian", "chromium os", "chrome os")):
        return "linux"
    return "other"


def _fingerprint_from_ua(ua_str: str) -> Dict[str, Any]:
    """Derive viewport / DPR / platform / mobile flags from a user-agent."""
    try:
        ua = ua_parse(ua_str or "")
    except Exception:
        ua = None

    os_key = _os_key_from_ua(ua_str)
    is_mobile = bool(ua and ua.is_mobile)
    is_tablet = bool(ua and ua.is_tablet)

    # Plausible ranges per OS — picked FRESH each visit for true uniqueness.
    if os_key == "ios":
        platform = "iPhone" if not is_tablet else "iPad"
        vendor = "Apple Computer, Inc."
        viewport = {"width": 390, "height": 844} if not is_tablet else {"width": 820, "height": 1180}
        dpr = 3 if not is_tablet else 2
        hc = random.choice([4, 6])
        dm = random.choice([4, 6, 8])
        webgl_vendor = "Apple Inc."
        webgl_renderer = random.choice([
            "Apple GPU", "Apple A15 GPU", "Apple A16 GPU", "Apple A17 Pro GPU",
        ])
    elif os_key == "android":
        platform = "Linux armv8l"
        vendor = "Google Inc."
        viewport = {"width": 412, "height": 915} if not is_tablet else {"width": 800, "height": 1280}
        dpr = random.choice([2.0, 2.625, 3.0])
        hc = random.choice([6, 8])
        dm = random.choice([4, 6, 8])
        webgl_vendor = "Google Inc. (Qualcomm)"
        webgl_renderer = random.choice([
            "ANGLE (Qualcomm, Adreno (TM) 740, OpenGL ES 3.2)",
            "ANGLE (Qualcomm, Adreno (TM) 730, OpenGL ES 3.2)",
            "ANGLE (ARM, Mali-G78 MP24, OpenGL ES 3.2)",
            "ANGLE (Qualcomm, Adreno (TM) 650, OpenGL ES 3.2)",
        ])
    elif os_key == "windows":
        platform = "Win32"
        vendor = "Google Inc."
        viewport = random.choice([
            {"width": 1920, "height": 1080},
            {"width": 1536, "height": 864},
            {"width": 1366, "height": 768},
            {"width": 1680, "height": 1050},
        ])
        dpr = random.choice([1.0, 1.25, 1.5])
        hc = random.choice([4, 8, 12, 16])
        dm = random.choice([4, 8, 16, 32])
        webgl_vendor = "Google Inc. (NVIDIA)"
        webgl_renderer = random.choice([
            "ANGLE (NVIDIA, NVIDIA GeForce GTX 1650 Direct3D11 vs_5_0 ps_5_0)",
            "ANGLE (NVIDIA, NVIDIA GeForce RTX 3060 Direct3D11 vs_5_0 ps_5_0)",
            "ANGLE (Intel, Intel(R) UHD Graphics 630 Direct3D11 vs_5_0 ps_5_0)",
            "ANGLE (AMD, AMD Radeon RX 6600 Direct3D11 vs_5_0 ps_5_0)",
        ])
    elif os_key == "macos":
        platform = "MacIntel"
        vendor = "Google Inc."
        viewport = random.choice([
            {"width": 1440, "height": 900},
            {"width": 1512, "height": 982},
            {"width": 1680, "height": 1050},
        ])
        dpr = 2
        hc = random.choice([8, 10, 12])
        dm = random.choice([8, 16])
        webgl_vendor = "Google Inc. (Apple)"
        webgl_renderer = random.choice([
            "ANGLE (Apple, Apple M1 Pro, OpenGL 4.1)",
            "ANGLE (Apple, Apple M2, OpenGL 4.1)",
            "ANGLE (Apple, Apple M3, OpenGL 4.1)",
        ])
    elif os_key == "linux":
        platform = "Linux x86_64"
        vendor = "Google Inc."
        viewport = {"width": 1920, "height": 1080}
        dpr = 1
        hc = random.choice([4, 8, 12])
        dm = random.choice([4, 8, 16])
        webgl_vendor = "Google Inc. (Intel)"
        webgl_renderer = "ANGLE (Intel, Mesa Intel(R) UHD Graphics 620)"
    else:
        platform = "Linux x86_64"
        vendor = "Google Inc."
        viewport = {"width": 1366, "height": 768}
        dpr = 1
        hc, dm = 4, 8
        webgl_vendor = "Google Inc."
        webgl_renderer = "ANGLE (Intel, Mesa Intel(R) HD Graphics)"

    # Small jitter on top so even two visits from the same preset look distinct
    viewport = {
        "width": max(320, viewport["width"] + random.randint(-4, 4)),
        "height": max(568, viewport["height"] + random.randint(-8, 8)),
    }

    # ── 2026-01: extra fingerprint fields for deep anti-detect ──
    # All ADDITIVE — existing callers that only read the old keys are
    # unaffected. Values are realistic per-OS so MaxMind / IPQS / Anura
    # cannot fingerprint the headless-Chrome signature.
    chrome_major, _chrome_minor = _extract_chrome_version(ua_str)
    safari_major, _safari_minor = _extract_safari_version(ua_str)

    # Screen dimensions match viewport + realistic toolbar/taskbar.
    # Desktop OSs have a taskbar (~40-100px); mobile = viewport.
    if os_key == "windows":
        screen_w = max(1024, viewport["width"])
        screen_h = max(768, viewport["height"] + random.choice([40, 60, 80]))
        avail_w = screen_w
        avail_h = max(768, screen_h - random.choice([40, 48, 60]))  # taskbar
        outer_w_delta = 0  # outerWidth == innerWidth (window borders thin in modern Win)
        outer_h_delta = random.choice([74, 87, 117, 138])  # tab bar + URL bar + bookmarks
        color_depth = 24
        max_touch = 0
    elif os_key == "macos":
        screen_w = max(1280, viewport["width"])
        screen_h = max(800, viewport["height"] + random.choice([22, 25, 28]))
        avail_w = screen_w
        avail_h = max(800, screen_h - random.choice([22, 25, 28]))  # menu bar
        outer_w_delta = 0
        outer_h_delta = random.choice([74, 87, 105, 130])
        color_depth = 30 if dpr >= 2 else 24
        max_touch = 0
    elif os_key == "linux":
        screen_w = max(1024, viewport["width"])
        screen_h = max(768, viewport["height"] + random.choice([24, 40, 60]))
        avail_w = screen_w
        avail_h = max(768, screen_h - random.choice([24, 40]))
        outer_w_delta = 0
        outer_h_delta = random.choice([74, 100, 130])
        color_depth = 24
        max_touch = 0
    elif os_key == "android":
        screen_w = viewport["width"]
        screen_h = viewport["height"]
        avail_w = viewport["width"]
        avail_h = viewport["height"]
        outer_w_delta = 0
        outer_h_delta = 0
        color_depth = 24
        max_touch = random.choice([5, 10])
    elif os_key == "ios":
        screen_w = viewport["width"]
        screen_h = viewport["height"]
        avail_w = viewport["width"]
        avail_h = viewport["height"]
        outer_w_delta = 0
        outer_h_delta = 0
        color_depth = 24
        max_touch = 5
    else:
        screen_w = max(1024, viewport["width"])
        screen_h = max(768, viewport["height"] + 40)
        avail_w = screen_w
        avail_h = screen_h - 40
        outer_w_delta = 0
        outer_h_delta = 80
        color_depth = 24
        max_touch = 0

    # Network Information API — realistic per-device-type effective type
    if os_key in ("android", "ios"):
        effective_type = random.choice(["4g", "4g", "4g", "3g"])  # mostly 4g
        downlink = round(random.uniform(2.5, 10.0), 1)
        rtt = random.choice([50, 100, 150, 200, 300])
        connection_type = "cellular"
    else:
        effective_type = "4g"
        downlink = round(random.uniform(5.0, 20.0), 1)
        rtt = random.choice([25, 50, 75, 100])
        connection_type = "wifi"

    # Battery API — realistic snapshot. Mobile usually mid-charge,
    # desktop often plugged in.
    if os_key in ("android", "ios"):
        battery_level = round(random.uniform(0.18, 0.92), 2)
        battery_charging = random.random() < 0.35  # 35% plugged-in chance
    else:
        battery_level = round(random.uniform(0.45, 0.98), 2)
        battery_charging = random.random() < 0.7  # 70% plugged in for desktop

    return {
        "os": os_key,
        "platform": platform,
        "vendor": vendor,
        "viewport": viewport,
        "device_scale_factor": dpr,
        "is_mobile": is_mobile or is_tablet or os_key in ("android", "ios"),
        "has_touch": is_mobile or is_tablet or os_key in ("android", "ios"),
        "hardware_concurrency": hc,
        "device_memory": dm,
        "webgl_vendor": webgl_vendor,
        "webgl_renderer": webgl_renderer,
        # Canvas noise seed — unique per visit so canvas fingerprint differs too
        "canvas_seed": random.randint(1, 2**30),
        "label": f"{(ua.os.family + ' ' + ua.os.version_string) if ua else os_key}".strip() or ua_str[:40],
        # ── 2026-01 additions ──
        "chrome_version": chrome_major,
        "safari_version": safari_major,
        "audio_seed": random.randint(1, 2**30),
        "font_seed": random.randint(1, 2**30),
        "screen_width": screen_w,
        "screen_height": screen_h,
        "avail_width": avail_w,
        "avail_height": avail_h,
        "outer_width_delta": outer_w_delta,
        "outer_height_delta": outer_h_delta,
        "color_depth": color_depth,
        "max_touch_points": max_touch,
        "effective_type": effective_type,
        "downlink": downlink,
        "rtt": rtt,
        "connection_type": connection_type,
        "battery_level": battery_level,
        "battery_charging": battery_charging,
        "fonts": _OS_FONTS.get(os_key, _OS_FONTS["windows"]),
    }


_UNFILLED_MACRO_URL_RX = __import__("re").compile(r"\{\{[^}]+\}\}|%7[bB]%7[bB]")


# Telemetry queue (job_id → list of macro-leak records) consumed at
# job-finalise time and flushed into MongoDB `rut_diagnostics`. We keep
# it in-memory during the run so the route handler stays non-blocking
# and we don't need to pass the Mongo handle through Playwright closures.
_MACRO_LEAK_BUFFER: Dict[str, List[Dict[str, Any]]] = {}
# Stuck-visit buffer (job_id → list of stuck-event records). Same flush
# semantics as the macro-leak buffer.
_STUCK_EVENT_BUFFER: Dict[str, List[Dict[str, Any]]] = {}


def _record_macro_leak(job_id: str, visit_index: int, url: str, resource_type: str) -> None:
    """Append a macro-leak event to the in-memory buffer. Called from the
    route handler — must stay sync + non-throwing so a Playwright handler
    error never crashes the visit."""
    try:
        buf = _MACRO_LEAK_BUFFER.setdefault(job_id, [])
        if len(buf) >= 200:  # cap per-job to avoid runaway memory
            return
        buf.append({
            "job_id": job_id,
            "visit_index": int(visit_index or 0),
            "blocked_url": (url or "")[:1024],
            "resource_type": (resource_type or "")[:32],
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        pass


def _record_stuck_event(job_id: str, visit_index: int, url: str, seconds_stuck: float,
                       last_step: int = -1, snapshot_name: str = "",
                       body_snippet: str = "") -> None:
    """Append a stuck-visit event. Called by the per-visit watchdog when
    the page URL hasn't changed for the configured threshold."""
    try:
        buf = _STUCK_EVENT_BUFFER.setdefault(job_id, [])
        if len(buf) >= 200:
            return
        buf.append({
            "job_id": job_id,
            "visit_index": int(visit_index or 0),
            "stuck_url": (url or "")[:1024],
            "seconds_stuck": round(float(seconds_stuck or 0), 1),
            "last_step_index": int(last_step) if last_step is not None else -1,
            "snapshot_name": (snapshot_name or "")[:128],
            "body_snippet": (body_snippet or "")[:2000],
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        pass


async def _stuck_watchdog(page, job_id: str, visit_index: int,
                          threshold_s: float = 240.0, poll_s: float = 5.0,
                          shots_dir=None, on_stuck=None,
                          state: Optional[Dict[str, Any]] = None) -> None:
    """Watch the page while a visit's automation steps are running. If
    the page hasn't progressed for `threshold_s` seconds we consider
    the visit stuck and:
        1. Capture a screenshot + first 2 KB of body text (debug aid).
        2. Emit `_record_stuck_event` so the operator can see WHICH page
           the job dies on.
        3. Invoke the `on_stuck` callback (if provided) — typically the
           visit loop's `steps_task.cancel()` so the run doesn't waste
           the rest of its automation budget on a dead page.
    We only record ONCE per stuck-period so a 5-minute hang doesn't
    generate 12 identical rows.

    2026-01: default threshold raised from 25s → 60s so that legitimate
    long form-submit sequences (form fill + Continue + cascading
    wait_for_load steps that can total 25-30s of intentional waiting)
    don't get killed prematurely.

    2026-05: default raised again from 60s → 240s (4 min) for slow
    survey-style offer pages (e.g. anyunclaimedassets.com/indexform.php)
    where multi-step SPA transitions + form-fill + submit + thank-you
    redirect can legitimately take 90-180s end-to-end. The
    chrome-error:// fast-path below still fires INSTANTLY (no wait) so
    dead proxies / DNS failures / SSL errors are still caught
    immediately — only the "page is alive but slow" case gets the
    extended budget.

    2026-01 (revised): "Progress" is now defined as EITHER the main-frame
    URL changing OR the DOM-progress fingerprint changing (body text
    length, body child count, document title, location hash). This is
    essential for offers running as single-page apps (SPAs) — e.g.
    anyunclaimedassets.com/indexform.php, which transitions through
    Profile → Cash & Assets Survey → Benefits → Thank-you ALL on the
    same URL via JavaScript. The old URL-only watchdog killed every
    such visit at 60s even though the DOM was progressing through
    multiple survey steps. Now we only abort if BOTH the URL and the
    DOM fingerprint stay frozen for the full threshold."""
    # Cheap "is the page progressing?" signal — combines URL + a handful
    # of DOM characteristics that change on any meaningful UI transition
    # (route, modal open, new survey question rendered). Computing this
    # is one Playwright evaluate call (single round-trip) instead of a
    # full innerHTML snapshot, keeping per-poll cost ~10 ms.
    #
    # ── 2026-05 (improved sensitivity) ──
    # Issue: survey-style SPAs (e.g. anyunclaimedassets indexform.php)
    # swap question text WITHOUT changing length / node count appreciably,
    # so the bot's automation was progressing but the watchdog couldn't
    # see it and killed the visit at 60s. Fixes:
    #   1. Add a fast 32-bit FNV-1a hash of body.innerText — detects ANY
    #      content change even if length stays constant.
    #   2. Add `location.search` (query string) — many SPAs encode step
    #      in `?step=3` etc.
    #   3. Add visible button/input count — survey transitions usually
    #      add/remove form controls.
    #   4. Add the FIRST 80 chars of trimmed innerText so a quick eyeball
    #      shows the actual screen content the watchdog sees.
    # Any change in any of these signals resets the stuck timer.
    _PROGRESS_JS = (
        "()=>{try{"
        "var b=document.body;"
        "var raw=(b&&b.innerText)?b.innerText:'';"
        "var t=raw.length;"
        "var c=(b&&b.children)?b.children.length:0;"
        "var d=(b&&b.querySelectorAll)?b.querySelectorAll('*').length:0;"
        # FNV-1a 32-bit hash — fast, deterministic, no deps
        "var h=2166136261;"
        "for(var i=0;i<raw.length;i++){h^=raw.charCodeAt(i);h=(h*16777619)>>>0;}"
        # Visible interactive control count
        "var ctrls=0;"
        "try{"
        " var sel=b?b.querySelectorAll('button,a[href],input,select,textarea,[role=button]'):[];"
        " for(var k=0;k<sel.length;k++){"
        "  var el=sel[k];"
        "  var s=window.getComputedStyle(el);"
        "  if(s&&s.display!=='none'&&s.visibility!=='hidden'&&el.offsetWidth>0&&el.offsetHeight>0)ctrls++;"
        " }"
        "}catch(_e){}"
        "var snip=(raw||'').replace(/\\s+/g,' ').trim().slice(0,80);"
        "return [document.title||'',(location.hash||''),(location.search||''),t,c,d,h>>>0,ctrls,snip];"
        "}catch(e){return ['','','',0,0,0,0,0,''];}}"
    )

    async def _probe():
        try:
            return await asyncio.wait_for(page.evaluate(_PROGRESS_JS), timeout=2.0)
        except Exception:
            return None

    last_url = ""
    try:
        last_url = page.url or ""
    except Exception:
        pass
    last_progress = await _probe()
    last_changed_at = __import__("time").monotonic()
    recorded_this_period = False
    # 2026-05: track progression so the caller can distinguish between
    # "page never moved" (truly dead) vs "page progressed then went idle"
    # (submit succeeded, just no more user-driven steps). state is a
    # shared dict the caller can read AFTER the watchdog fires.
    progression_count = 0
    if state is not None:
        state.setdefault("progressed", False)
        state.setdefault("progression_count", 0)
    while True:
        try:
            await asyncio.sleep(poll_s)
        except asyncio.CancelledError:
            return
        try:
            cur = page.url or ""
        except Exception:
            return
        now = __import__("time").monotonic()
        # ── Fast path: chrome-error:// means the tab died mid-flight
        #    (proxy tunnel dropped, DNS lookup failed, SSL error, etc.).
        #    Don't wait for the full 25s threshold — fire the stuck
        #    record + abort callback immediately so the run can move on.
        if cur.startswith("chrome-error://") or cur.startswith("chrome://network-error"):
            if not recorded_this_period:
                _record_stuck_event(
                    job_id, visit_index, cur,
                    max(0.0, now - last_changed_at),
                    snapshot_name="",
                    body_snippet="(chrome-error fast-path — no DOM to capture)",
                )
                recorded_this_period = True
                try:
                    push_live_step(
                        job_id, visit_index, "stuck", "warn",
                        f"Chrome error page detected ({cur[:60]}) — aborting visit",
                    )
                except Exception:
                    pass
                if on_stuck is not None:
                    try:
                        on_stuck()
                    except Exception:
                        pass
                return
        # Did the URL change?
        _url_changed = cur != last_url
        if _url_changed:
            last_url = cur
            last_changed_at = now
            recorded_this_period = False
            last_progress = await _probe()
            progression_count += 1
            if state is not None:
                state["progressed"] = True
                state["progression_count"] = progression_count
            continue
        # URL is unchanged — check DOM-progress fingerprint. If it
        # moved (different title / hash / search / text-content-hash /
        # body-text length / child count / total node count / visible
        # controls count), the page IS progressing (SPA transition,
        # survey question rendered, etc.) and we reset the stuck timer.
        # ── 2026-05 ──
        # New tuple layout: [title, hash, search, text_len, child, node,
        # text_hash, ctrls, snip]
        cur_progress = await _probe()
        if cur_progress is not None and last_progress is not None:
            try:
                _title_chg = cur_progress[0] != last_progress[0]
                _hash_chg = cur_progress[1] != last_progress[1]
                _search_chg = (
                    cur_progress[2] != last_progress[2]
                    if len(cur_progress) > 2 and len(last_progress) > 2
                    else False
                )
                _text_delta = abs(int(cur_progress[3]) - int(last_progress[3])) if len(cur_progress) > 3 else 0
                _child_delta = abs(int(cur_progress[4]) - int(last_progress[4])) if len(cur_progress) > 4 else 0
                _node_delta = abs(int(cur_progress[5]) - int(last_progress[5])) if len(cur_progress) > 5 else 0
                _text_hash_chg = (
                    int(cur_progress[6]) != int(last_progress[6])
                    if len(cur_progress) > 6 and len(last_progress) > 6
                    else False
                )
                _ctrls_chg = (
                    int(cur_progress[7]) != int(last_progress[7])
                    if len(cur_progress) > 7 and len(last_progress) > 7
                    else False
                )
                _dom_changed = (
                    _title_chg or _hash_chg or _search_chg
                    or _text_hash_chg                  # ANY content change (even same length)
                    or _ctrls_chg                      # form controls added/removed
                    or _text_delta >= 1                # 2026-05: lowered from 4 — even single-char delta resets
                    or _child_delta >= 1
                    or _node_delta >= 1                # 2026-05: lowered from 3 — any DOM node add/remove resets
                )
            except Exception:
                _dom_changed = False
            if _dom_changed:
                last_progress = cur_progress
                last_changed_at = now
                recorded_this_period = False
                progression_count += 1
                if state is not None:
                    state["progressed"] = True
                    state["progression_count"] = progression_count
                continue
        elif cur_progress is not None:
            # First successful probe — seed the baseline so the next
            # iteration can detect deltas.
            last_progress = cur_progress
        elapsed = now - last_changed_at
        if elapsed >= threshold_s and not recorded_this_period:
            # ── Capture debug artefacts BEFORE aborting ────────────────
            snapshot_name = ""
            body_snippet = ""
            try:
                if shots_dir is not None:
                    sp = Path(str(shots_dir)) / f"visit_{int(visit_index):05d}_stuck.png"
                    await asyncio.wait_for(
                        page.screenshot(path=str(sp), full_page=True),
                        timeout=5.0,
                    )
                    snapshot_name = sp.name
            except Exception:
                pass
            try:
                body_snippet = await asyncio.wait_for(
                    page.evaluate(
                        "() => (document.body && document.body.innerText) "
                        "? document.body.innerText.slice(0, 2000) : ''"
                    ),
                    timeout=4.0,
                ) or ""
            except Exception:
                body_snippet = ""

            _record_stuck_event(
                job_id, visit_index, cur, elapsed,
                snapshot_name=snapshot_name,
                body_snippet=body_snippet,
            )
            recorded_this_period = True
            try:
                push_live_step(
                    job_id, visit_index, "stuck", "warn",
                    f"Visit stuck on {cur[:120]} for ~{int(elapsed)}s (URL+DOM frozen) — aborting + see diagnostics",
                )
            except Exception:
                pass
            # ── Fire the abort callback so the visit loop can cancel
            #    the automation steps task and move on to the next visit.
            #    Then exit the watchdog (parent will await us).
            if on_stuck is not None:
                try:
                    on_stuck()
                except Exception:
                    pass
            return




def _make_macro_guard(job_id: str, visit_index: int):
    """Return a Playwright route handler closure bound to this visit's
    `(job_id, visit_index)` so the macro-leak telemetry record contains
    the right context. Returned handler is `async def`."""
    async def _handler(route, request):
        try:
            url = request.url or ""
            if _UNFILLED_MACRO_URL_RX.search(url):
                try:
                    rtype = (request.resource_type or "").lower()
                except Exception:
                    rtype = ""
                _record_macro_leak(job_id, visit_index, url, rtype)
                if rtype == "document":
                    try:
                        await route.fulfill(
                            status=200,
                            content_type="text/html",
                            body=(
                                "<!doctype html><html><head><meta charset='utf-8'>"
                                "<title>blocked</title></head><body>"
                                "<script>try{history.back();}catch(e){}</script>"
                                "</body></html>"
                            ),
                        )
                    except Exception:
                        try:
                            await route.abort()
                        except Exception:
                            pass
                else:
                    try:
                        await route.abort()
                    except Exception:
                        pass
                return
            await route.continue_()
        except Exception:
            try:
                await route.continue_()
            except Exception:
                pass
    return _handler


async def _block_unfilled_macro_request(route, request) -> None:
    """Abort any request whose URL still contains an unfilled tracker
    macro like `{{ccpa}}`, `{{sub_id}}`, etc.

    Affiliate landing pages frequently embed CCPA / opt-out / privacy
    anchors with macro placeholders the tracker is expected to replace
    server-side. When the inbound tracker URL is missing the macro, the
    literal `{{...}}` leaks into the rendered HTML and any aggressive
    "force-navigation" JS step in the user's automation script can end
    up clicking it, dead-ending the visit on a 404. Killing those
    navigations at the network layer keeps the page on the legitimate
    flow URL so subsequent form-fill / answer-click steps still run.

    For top-level document navigations we cannot just `route.abort()`
    because that leaves Chrome on `chrome-error://chromewebdata/`. We
    instead fulfill the request with a tiny HTML that calls
    `history.back()` so the browser returns to the legitimate offer
    page and the automation can resume from there.

    NOTE: This generic handler does NOT record telemetry — used by demo
    scripts only. Production visits use `_make_macro_guard(job_id, idx)`
    so each blocked request is logged against the originating job.
    """
    try:
        url = request.url or ""
        if _UNFILLED_MACRO_URL_RX.search(url):
            try:
                rtype = (request.resource_type or "").lower()
            except Exception:
                rtype = ""
            if rtype == "document":
                # Fulfill with a back-nav stub so we don't blank the tab.
                try:
                    await route.fulfill(
                        status=200,
                        content_type="text/html",
                        body=(
                            "<!doctype html><html><head><meta charset='utf-8'>"
                            "<title>blocked</title></head><body>"
                            "<script>try{history.back();}catch(e){}</script>"
                            "</body></html>"
                        ),
                    )
                except Exception:
                    try:
                        await route.abort()
                    except Exception:
                        pass
            else:
                try:
                    await route.abort()
                except Exception:
                    pass
            return
        await route.continue_()
    except Exception:
        try:
            await route.continue_()
        except Exception:
            pass



# ─── Proxy helpers ───────────────────────────────────────────────
def _parse_proxy_line(line: str) -> Optional[Dict[str, Any]]:
    s = (line or "").strip()
    if not s:
        return None
    scheme = "http"
    if s.startswith("http://"):
        s = s[7:]
    elif s.startswith("https://"):
        s = s[8:]
        scheme = "https"
    user, pwd = None, None
    if "@" in s:
        auth, s = s.rsplit("@", 1)
        if ":" in auth:
            user, pwd = auth.split(":", 1)
        else:
            user = auth
    parts = s.split(":")
    if len(parts) == 2:
        host, port = parts
    elif len(parts) == 4:
        host, port, user, pwd = parts
    else:
        return None
    try:
        int(port)
    except ValueError:
        return None
    out: Dict[str, Any] = {"server": f"{scheme}://{host}:{port}", "raw": line.strip()}
    if user:
        out["username"] = user
    if pwd:
        out["password"] = pwd
    return out


async def _probe_proxy_geo(proxy: Dict[str, Any], ua: str) -> Dict[str, Any]:
    """Probe proxy through ip-api — returns exit IP + country + city + timezone +
    locale + accept_language + is_vpn flag."""
    result = {
        "exit_ip": None, "country": "US", "country_name": "United States",
        "city": "New York", "region": "NY", "region_name": "New York",
        "lat": 40.7128, "lon": -74.0060,
        "timezone": "America/New_York", "accept_language": "en-US,en;q=0.9",
        "locale": "en-US", "is_vpn": False, "ok": False,
    }
    server = proxy["server"]
    if proxy.get("username"):
        prefix, rest = server.split("://", 1)
        server = f"{prefix}://{proxy['username']}:{proxy.get('password','')}@{rest}"

    # Some commercial residential proxies (proxy-jet, brightdata, etc.) ONLY accept
    # HTTPS CONNECT tunnels and reject plain `GET http://…` forward-proxy requests,
    # so we try an HTTPS geolocation endpoint first. If that fails we fall back to
    # the original HTTP ip-api.com endpoint (which works on proxies that do allow
    # plain HTTP forwarding).
    async def _try_https_ipwhois(cli: httpx.AsyncClient) -> bool:
        try:
            r = await cli.get("https://ipwho.is/")
            if r.status_code == 200:
                data = r.json()
                if data.get("success") is True:
                    result["exit_ip"] = data.get("ip")
                    result["country_name"] = data.get("country") or result["country_name"]
                    result["country"] = data.get("country_code") or result["country"]
                    result["region_name"] = data.get("region") or result["region_name"]
                    result["region"] = data.get("region_code") or result["region"]
                    result["city"] = data.get("city") or result["city"]
                    try:
                        result["lat"] = float(data.get("latitude") or result["lat"])
                        result["lon"] = float(data.get("longitude") or result["lon"])
                    except (TypeError, ValueError):
                        pass
                    tz = data.get("timezone") or {}
                    if isinstance(tz, dict):
                        result["timezone"] = tz.get("id") or result["timezone"]
                    elif isinstance(tz, str):
                        result["timezone"] = tz or result["timezone"]
                    conn = data.get("connection") or {}
                    result["is_vpn"] = bool(
                        conn.get("type") in ("hosting", "datacenter")
                        or (str(conn.get("org") or "").lower().find("hosting") >= 0)
                    )
                    return True
        except Exception as e:
            logger.debug(f"ipwho.is probe failed: {e}")
        return False

    async def _try_http_ipapi(cli: httpx.AsyncClient) -> bool:
        try:
            r = await cli.get(
                "http://ip-api.com/json/?fields=status,country,countryCode,region,regionName,city,"
                "timezone,lat,lon,query,proxy,hosting"
            )
            if r.status_code == 200:
                data = r.json()
                if data.get("status") == "success":
                    result["exit_ip"] = data.get("query")
                    result["country_name"] = data.get("country") or result["country_name"]
                    result["country"] = data.get("countryCode") or result["country"]
                    result["region"] = data.get("region") or result["region"]
                    result["region_name"] = data.get("regionName") or result["region_name"]
                    result["city"] = data.get("city") or result["city"]
                    result["lat"] = float(data.get("lat") or result["lat"])
                    result["lon"] = float(data.get("lon") or result["lon"])
                    result["timezone"] = data.get("timezone") or result["timezone"]
                    result["is_vpn"] = bool(data.get("proxy") or data.get("hosting"))
                    return True
        except Exception as e:
            logger.debug(f"ip-api.com probe failed: {e}")
        return False

    try:
        # Longer timeout because residential proxies can take 10-15s to route.
        # Retry up to 3 times — residential proxies (proxy-jet, brightdata,
        # etc.) have ~10-20% per-request failure rate due to rotating exit
        # nodes; retrying the same proxy usually succeeds with a different
        # exit IP on the next attempt.
        timeout_cfg = httpx.Timeout(30.0, connect=20.0)
        ok = False
        for attempt in range(3):
            try:
                async with httpx.AsyncClient(proxy=server, timeout=timeout_cfg, headers={"User-Agent": ua}, verify=False, http2=False) as cli:
                    ok = await _try_https_ipwhois(cli)
                    if not ok:
                        ok = await _try_http_ipapi(cli)
                if ok:
                    break
            except Exception as e:
                logger.debug(f"Proxy probe attempt {attempt+1} failed: {e}")
            # Brief backoff before next attempt
            if attempt < 2:
                await asyncio.sleep(1.5 * (attempt + 1))
        if ok:
            cc = (result["country"] or "").lower()
            lang_map = {
                "us": "en-US,en;q=0.9", "gb": "en-GB,en;q=0.9", "ca": "en-CA,en;q=0.9",
                "au": "en-AU,en;q=0.9", "nz": "en-NZ,en;q=0.9",
                "de": "de-DE,de;q=0.9,en;q=0.7", "fr": "fr-FR,fr;q=0.9,en;q=0.7",
                "es": "es-ES,es;q=0.9,en;q=0.7", "it": "it-IT,it;q=0.9,en;q=0.7",
                "nl": "nl-NL,nl;q=0.9,en;q=0.7", "pt": "pt-PT,pt;q=0.9,en;q=0.7",
                "br": "pt-BR,pt;q=0.9,en;q=0.7", "mx": "es-MX,es;q=0.9,en;q=0.7",
                "jp": "ja-JP,ja;q=0.9,en;q=0.7", "kr": "ko-KR,ko;q=0.9,en;q=0.7",
                "in": "en-IN,en;q=0.9,hi;q=0.8", "pk": "en-PK,en;q=0.9,ur;q=0.8",
                "ae": "ar-AE,ar;q=0.9,en;q=0.8", "sa": "ar-SA,ar;q=0.9,en;q=0.8",
            }
            locale_map = {
                "us": "en-US", "gb": "en-GB", "ca": "en-CA", "au": "en-AU", "nz": "en-NZ",
                "de": "de-DE", "fr": "fr-FR", "es": "es-ES", "it": "it-IT", "nl": "nl-NL",
                "pt": "pt-PT", "br": "pt-BR", "mx": "es-MX", "jp": "ja-JP", "kr": "ko-KR",
                "in": "en-IN", "pk": "en-PK", "ae": "ar-AE", "sa": "ar-SA",
            }
            result["accept_language"] = lang_map.get(cc, "en-US,en;q=0.9")
            result["locale"] = locale_map.get(cc, "en-US")
            result["ok"] = True
    except Exception as e:
        logger.debug(f"Proxy geo probe failed: {e}")
    return result


async def _probe_proxy_target_reachable(
    proxy: Dict[str, Any], target_url: str, ua: str,
    timeout_s: float = 12.0,
) -> Tuple[bool, str]:
    """2026-01 — Pre-flight reachability check: verify the proxy can
    actually reach the *target* URL (not just the internet) BEFORE we
    spawn a browser context for the visit.

    Why: residential pools (proxy-jet, brightdata, etc.) regularly serve
    exit nodes that pass ipwho.is / ip-api.com geo probes (because those
    endpoints are extremely permissive) but get refused / firewalled /
    blackholed by the actual offer landing page. Without this pre-check
    those visits used to spawn a full Chromium context, attempt a
    `page.goto()`, time out after 30-60s with "Proxy tunnel failed",
    and waste a UA/lead from the user's batches.

    Returns (ok, diagnostic). On `ok=False` the caller should skip this
    visit (no browser launch, no UA/lead consumption) and move on to
    the next proxy in the pool.

    Uses a cheap HEAD request first; falls back to a GET with tiny
    Range header if HEAD is rejected by the target (some Cloudflare
    setups do this). Anything that returns HTTP 2xx/3xx/4xx counts as
    "reachable" — we only fail on transport errors (connect refused,
    SSL handshake fail, timeout, tunnel error). 4xx responses (403,
    404, etc.) mean the proxy DID reach the target — the offer's own
    rules will then handle access; that's the existing engine's job.
    """
    server = proxy.get("server", "")
    if proxy.get("username"):
        try:
            prefix, rest = server.split("://", 1)
            server = f"{prefix}://{proxy['username']}:{proxy.get('password','')}@{rest}"
        except ValueError:
            return False, "Malformed proxy server string"

    headers = {
        "User-Agent": ua or "Mozilla/5.0",
        # Tiny range so even servers that don't support HEAD return cheaply
        "Range": "bytes=0-0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    timeout_cfg = httpx.Timeout(timeout_s, connect=min(8.0, timeout_s))
    try:
        async with httpx.AsyncClient(
            proxy=server, timeout=timeout_cfg, headers=headers,
            verify=False, http2=False, follow_redirects=False,
        ) as cli:
            # HEAD first (cheapest)
            try:
                r = await cli.head(target_url)
                # Any HTTP response — even 4xx/5xx — proves the proxy
                # reached the target. Only transport failures count.
                return True, f"HEAD {r.status_code}"
            except httpx.HTTPError:
                pass
            # Fallback: tiny GET (HEAD blocked at edge / not implemented)
            r = await cli.get(target_url)
            return True, f"GET {r.status_code}"
    except httpx.ProxyError as e:
        return False, f"Proxy tunnel: {str(e)[:80]}"
    except httpx.ConnectError as e:
        return False, f"Connect: {str(e)[:80]}"
    except httpx.ReadTimeout:
        return False, "Read timeout"
    except httpx.ConnectTimeout:
        return False, "Connect timeout"
    except Exception as e:
        # Any other transport-level error → consider unreachable so the
        # engine moves on. (Browser would have failed similarly later
        # but at a much higher cost.)
        return False, f"{type(e).__name__}: {str(e)[:80]}"


# ─── Stealth init script ────────────────────────────────────────
# 2026-01: comprehensive anti-detect coverage. Spoofs every JS API that
# MaxMind minFraud / IPQualityScore / Anura / ArkoseLabs / FingerprintJS
# / CreepJS / BotD probe. Per-visit unique noise so two visits from the
# same OS+UA still get different audio/canvas/font fingerprints.
def _build_stealth_script(fp: Dict[str, Any], geo: Dict[str, Any]) -> str:
    import json as _json
    langs = [s.split(";")[0].strip() for s in geo["accept_language"].split(",") if s.strip()]
    langs = [lg for lg in langs if lg]
    langs = langs[:4] or ["en-US", "en"]

    # Inject all per-visit constants into a single __KX namespace at the top
    # of the JS so the rest of the script is a normal raw JS string (no
    # f-string brace escaping nightmare).
    kx = {
        "platform": fp["platform"],
        "vendor": fp["vendor"],
        "languages": langs,
        "primaryLang": langs[0],
        "hardwareConcurrency": int(fp["hardware_concurrency"]),
        "deviceMemory": int(fp["device_memory"]),
        "webglVendor": fp["webgl_vendor"],
        "webglRenderer": fp["webgl_renderer"],
        "canvasSeed": int(fp["canvas_seed"]),
        "audioSeed": int(fp.get("audio_seed", random.randint(1, 2**30))),
        "fontSeed": int(fp.get("font_seed", random.randint(1, 2**30))),
        "screenW": int(fp.get("screen_width", fp["viewport"]["width"])),
        "screenH": int(fp.get("screen_height", fp["viewport"]["height"])),
        "availW": int(fp.get("avail_width", fp["viewport"]["width"])),
        "availH": int(fp.get("avail_height", fp["viewport"]["height"])),
        "outerHDelta": int(fp.get("outer_height_delta", 0)),
        "outerWDelta": int(fp.get("outer_width_delta", 0)),
        "colorDepth": int(fp.get("color_depth", 24)),
        "maxTouchPoints": int(fp.get("max_touch_points", 0)),
        "effectiveType": fp.get("effective_type", "4g"),
        "downlink": float(fp.get("downlink", 10.0)),
        "rtt": int(fp.get("rtt", 50)),
        "saveData": False,
        "batteryLevel": float(fp.get("battery_level", 0.85)),
        "batteryCharging": bool(fp.get("battery_charging", True)),
        "fonts": fp.get("fonts", []),
        "os": fp.get("os", "windows"),
        "isMobile": bool(fp.get("is_mobile", False)),
        "chromeVersion": int(fp.get("chrome_version", 142)),
        "dpr": float(fp.get("device_scale_factor", 1.0)),
        "tz": geo.get("timezone", "America/New_York"),
    }
    config_js = "const __KX = " + _json.dumps(kx) + ";"

    # Big raw JS body — uses __KX.* values, no Python interpolation here.
    js_body = r"""
const safe = (fn) => { try { fn(); } catch (e) {} };
const safeDefine = (obj, prop, getter) => { try { Object.defineProperty(obj, prop, { get: getter, configurable: true }); } catch (e) {} };
// Tiny seeded PRNG — same for canvas, audio, font jitter so each visit is
// deterministic per-seed but unique vs other visits.
const makeRng = (seed) => { let s = seed >>> 0 || 1; return () => { s = (s * 1664525 + 1013904223) >>> 0; return s / 4294967296; }; };

// ── navigator.* core props ──────────────────────────────────────
safe(() => safeDefine(navigator, 'webdriver', () => false));
safe(() => safeDefine(navigator, 'platform', () => __KX.platform));
safe(() => safeDefine(navigator, 'vendor', () => __KX.vendor));
safe(() => safeDefine(navigator, 'hardwareConcurrency', () => __KX.hardwareConcurrency));
safe(() => safeDefine(navigator, 'deviceMemory', () => __KX.deviceMemory));
safe(() => safeDefine(navigator, 'languages', () => __KX.languages));
safe(() => safeDefine(navigator, 'language', () => __KX.primaryLang));
safe(() => safeDefine(navigator, 'maxTouchPoints', () => __KX.maxTouchPoints));
safe(() => safeDefine(navigator, 'doNotTrack', () => null));
safe(() => safeDefine(navigator, 'cookieEnabled', () => true));
safe(() => safeDefine(navigator, 'pdfViewerEnabled', () => true));

// ── window.chrome stub ──────────────────────────────────────────
safe(() => {
  if (!window.chrome) window.chrome = {};
  if (!window.chrome.runtime) {
    window.chrome.runtime = {
      OnInstalledReason: { CHROME_UPDATE: 'chrome_update', INSTALL: 'install', SHARED_MODULE_UPDATE: 'shared_module_update', UPDATE: 'update' },
      OnRestartRequiredReason: { APP_UPDATE: 'app_update', OS_UPDATE: 'os_update', PERIODIC: 'periodic' },
      PlatformArch: { ARM: 'arm', ARM64: 'arm64', MIPS: 'mips', MIPS64: 'mips64', X86_32: 'x86-32', X86_64: 'x86-64' },
      PlatformNaclArch: { ARM: 'arm', MIPS: 'mips', MIPS64: 'mips64', X86_32: 'x86-32', X86_64: 'x86-64' },
      PlatformOs: { ANDROID: 'android', CROS: 'cros', LINUX: 'linux', MAC: 'mac', OPENBSD: 'openbsd', WIN: 'win' },
      RequestUpdateCheckStatus: { NO_UPDATE: 'no_update', THROTTLED: 'throttled', UPDATE_AVAILABLE: 'update_available' },
    };
  }
  if (!window.chrome.app) {
    window.chrome.app = {
      isInstalled: false,
      InstallState: { DISABLED: 'disabled', INSTALLED: 'installed', NOT_INSTALLED: 'not_installed' },
      RunningState: { CANNOT_RUN: 'cannot_run', READY_TO_RUN: 'ready_to_run', RUNNING: 'running' },
      getDetails: function () { return null; },
      getIsInstalled: function () { return false; },
      runningState: function () { return 'cannot_run'; },
    };
  }
  if (!window.chrome.csi) window.chrome.csi = function () { return { startE: Date.now(), onloadT: Date.now(), pageT: 0, tran: 15 }; };
  if (!window.chrome.loadTimes) window.chrome.loadTimes = function () { return { commitLoadTime: Date.now() / 1000, finishDocumentLoadTime: Date.now() / 1000, finishLoadTime: Date.now() / 1000, firstPaintAfterLoadTime: 0, firstPaintTime: Date.now() / 1000, navigationType: 'Other', npnNegotiatedProtocol: 'h2', requestTime: Date.now() / 1000, startLoadTime: Date.now() / 1000, wasAlternateProtocolAvailable: false, wasFetchedViaSpdy: true, wasNpnNegotiated: true }; };
});

// ── permissions.query: realistic states ─────────────────────────
safe(() => {
  const orig = navigator.permissions.query.bind(navigator.permissions);
  navigator.permissions.query = (p) => {
    if (p && p.name === 'notifications') return Promise.resolve({ state: Notification.permission || 'default', onchange: null });
    if (p && p.name === 'geolocation') return Promise.resolve({ state: 'prompt', onchange: null });
    if (p && p.name === 'midi') return Promise.resolve({ state: 'prompt', onchange: null });
    if (p && p.name === 'camera') return Promise.resolve({ state: 'prompt', onchange: null });
    if (p && p.name === 'microphone') return Promise.resolve({ state: 'prompt', onchange: null });
    return orig(p);
  };
});

// ── plugins / mimeTypes (real Chrome has PDF viewer plugin) ────
safe(() => {
  if (navigator.plugins.length === 0) {
    const fakePlugins = [
      { name: 'PDF Viewer', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
      { name: 'Chrome PDF Viewer', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
      { name: 'Chromium PDF Viewer', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
      { name: 'Microsoft Edge PDF Viewer', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
      { name: 'WebKit built-in PDF', filename: 'internal-pdf-viewer', description: 'Portable Document Format' },
    ];
    safeDefine(navigator, 'plugins', () => fakePlugins);
    safeDefine(navigator, 'mimeTypes', () => [
      { type: 'application/pdf', suffixes: 'pdf', description: 'Portable Document Format' },
      { type: 'text/pdf', suffixes: 'pdf', description: 'Portable Document Format' },
    ]);
  }
});

// ── WebGL UNMASKED_VENDOR / UNMASKED_RENDERER ──────────────────
safe(() => {
  const patch = (proto) => {
    if (!proto) return;
    const orig = proto.getParameter;
    proto.getParameter = function (p) {
      if (p === 37445) return __KX.webglVendor;
      if (p === 37446) return __KX.webglRenderer;
      // VENDOR / RENDERER (non-unmasked) — also return realistic strings
      if (p === 7936) return 'WebKit';
      if (p === 7937) return 'WebKit WebGL';
      return orig.call(this, p);
    };
  };
  patch(WebGLRenderingContext && WebGLRenderingContext.prototype);
  if (window.WebGL2RenderingContext) patch(WebGL2RenderingContext.prototype);
});

// ── Canvas fingerprint noise (per-visit seed) ──────────────────
safe(() => {
  const rng = makeRng(__KX.canvasSeed);
  const origToDataURL = HTMLCanvasElement.prototype.toDataURL;
  HTMLCanvasElement.prototype.toDataURL = function () {
    safe(() => {
      const ctx = this.getContext('2d');
      const w = this.width, h = this.height;
      if (ctx && w > 0 && h > 0 && w * h < 2000000) {
        const data = ctx.getImageData(0, 0, w, h);
        for (let i = 0; i < data.data.length; i += 4) {
          const r = (rng() * 4) | 0;
          data.data[i] ^= (r & 1);
          data.data[i + 1] ^= (r >> 1) & 1;
        }
        ctx.putImageData(data, 0, 0);
      }
    });
    return origToDataURL.apply(this, arguments);
  };
  const origGetImageData = CanvasRenderingContext2D.prototype.getImageData;
  CanvasRenderingContext2D.prototype.getImageData = function () {
    const d = origGetImageData.apply(this, arguments);
    if (d && d.data) {
      for (let i = 0; i < d.data.length; i += 4) {
        const r = (rng() * 2) | 0;
        d.data[i + 3] ^= r;
      }
    }
    return d;
  };
  // Also nudge measureText so font fingerprinting via text width differs
  const origMeasure = CanvasRenderingContext2D.prototype.measureText;
  CanvasRenderingContext2D.prototype.measureText = function (txt) {
    const m = origMeasure.apply(this, arguments);
    try {
      const jitter = (rng() - 0.5) * 0.04;
      const proxy = Object.create(Object.getPrototypeOf(m));
      Object.getOwnPropertyNames(m).forEach((k) => {
        try { proxy[k] = m[k]; } catch (e) {}
      });
      ['width', 'actualBoundingBoxLeft', 'actualBoundingBoxRight'].forEach((k) => {
        if (typeof m[k] === 'number') {
          try { Object.defineProperty(proxy, k, { value: m[k] + jitter, writable: false, configurable: true }); } catch (e) {}
        }
      });
      return proxy;
    } catch (e) { return m; }
  };
});

// ── AudioContext fingerprint noise ─────────────────────────────
// IPQS specifically uses OfflineAudioContext fingerprinting. We add
// tiny noise to every float-frequency / time-domain readback so the
// hash changes per visit but the audio itself remains audible.
safe(() => {
  const arng = makeRng(__KX.audioSeed);
  const patchFloat32Output = (proto, methodName) => {
    if (!proto || !proto[methodName]) return;
    const orig = proto[methodName];
    proto[methodName] = function (arr) {
      const r = orig.apply(this, arguments);
      try {
        if (arr && arr.length) {
          for (let i = 0; i < arr.length; i += 100) {
            arr[i] = arr[i] + (arng() - 0.5) * 1e-7;
          }
        }
      } catch (e) {}
      return r;
    };
  };
  if (window.AnalyserNode) {
    patchFloat32Output(AnalyserNode.prototype, 'getFloatFrequencyData');
    patchFloat32Output(AnalyserNode.prototype, 'getByteFrequencyData');
    patchFloat32Output(AnalyserNode.prototype, 'getFloatTimeDomainData');
    patchFloat32Output(AnalyserNode.prototype, 'getByteTimeDomainData');
  }
  if (window.AudioBuffer) {
    const origGCD = AudioBuffer.prototype.getChannelData;
    AudioBuffer.prototype.getChannelData = function () {
      const arr = origGCD.apply(this, arguments);
      try {
        for (let i = 0; i < arr.length; i += 100) {
          arr[i] = arr[i] + (arng() - 0.5) * 1e-7;
        }
      } catch (e) {}
      return arr;
    };
  }
  // OfflineAudioContext rendering — return value identity changes per seed
  if (window.OfflineAudioContext) {
    const origStart = OfflineAudioContext.prototype.startRendering;
    OfflineAudioContext.prototype.startRendering = function () {
      const p = origStart.apply(this, arguments);
      return p.then((buf) => {
        try {
          for (let ch = 0; ch < buf.numberOfChannels; ch++) {
            const data = buf.getChannelData(ch);
            for (let i = 0; i < data.length; i += 500) {
              data[i] = data[i] + (arng() - 0.5) * 1e-7;
            }
          }
        } catch (e) {}
        return buf;
      });
    };
  }
});

// ── screen.* and window outer dimensions ───────────────────────
safe(() => safeDefine(screen, 'width', () => __KX.screenW));
safe(() => safeDefine(screen, 'height', () => __KX.screenH));
safe(() => safeDefine(screen, 'availWidth', () => __KX.availW));
safe(() => safeDefine(screen, 'availHeight', () => __KX.availH));
safe(() => safeDefine(screen, 'colorDepth', () => __KX.colorDepth));
safe(() => safeDefine(screen, 'pixelDepth', () => __KX.colorDepth));
safe(() => {
  if (window.outerWidth === 0 || window.outerWidth === window.innerWidth) {
    safeDefine(window, 'outerWidth', () => window.innerWidth + __KX.outerWDelta);
  }
  if (window.outerHeight === 0 || window.outerHeight === window.innerHeight) {
    safeDefine(window, 'outerHeight', () => window.innerHeight + __KX.outerHDelta);
  }
});
safe(() => safeDefine(window, 'devicePixelRatio', () => __KX.dpr));

// ── userAgentData (Sec-CH-UA equivalent in JS) ─────────────────
safe(() => {
  if (navigator.userAgent && /Chrome\//.test(navigator.userAgent)) {
    const cv = String(__KX.chromeVersion);
    const platformName = ({ windows: 'Windows', macos: 'macOS', ios: 'iOS', android: 'Android', linux: 'Linux' })[__KX.os] || 'Windows';
    const brands = [
      { brand: 'Chromium', version: cv },
      { brand: 'Google Chrome', version: cv },
      { brand: 'Not_A Brand', version: '24' },
    ];
    const uaData = {
      brands: brands,
      mobile: __KX.isMobile,
      platform: platformName,
      getHighEntropyValues: function (hints) {
        return Promise.resolve({
          architecture: 'x86', bitness: '64', brands: brands,
          fullVersionList: brands.map((b) => ({ brand: b.brand, version: cv + '.0.0.0' })),
          mobile: __KX.isMobile, model: '', platform: platformName, platformVersion: '15.0.0',
          uaFullVersion: cv + '.0.0.0', wow64: false,
        });
      },
      toJSON: function () { return { brands: brands, mobile: __KX.isMobile, platform: platformName }; },
    };
    safeDefine(navigator, 'userAgentData', () => uaData);
  }
});

// ── Battery API — realistic per-device snapshot ────────────────
safe(() => {
  const battery = {
    charging: __KX.batteryCharging,
    chargingTime: __KX.batteryCharging ? 1800 : Infinity,
    dischargingTime: __KX.batteryCharging ? Infinity : 7200,
    level: __KX.batteryLevel,
    onchargingchange: null, onchargingtimechange: null,
    ondischargingtimechange: null, onlevelchange: null,
    addEventListener: function () {}, removeEventListener: function () {},
    dispatchEvent: function () { return true; },
  };
  navigator.getBattery = function () { return Promise.resolve(battery); };
});

// ── Network Information API ────────────────────────────────────
safe(() => {
  const conn = {
    effectiveType: __KX.effectiveType,
    downlink: __KX.downlink, rtt: __KX.rtt, saveData: __KX.saveData,
    onchange: null, type: __KX.effectiveType === '4g' ? 'cellular' : 'wifi',
    addEventListener: function () {}, removeEventListener: function () {},
  };
  safeDefine(navigator, 'connection', () => conn);
  // Also expose deprecated mozConnection / webkitConnection
  safeDefine(navigator, 'mozConnection', () => conn);
  safeDefine(navigator, 'webkitConnection', () => conn);
});

// ── speechSynthesis voices (Chrome has 21+ by default) ─────────
safe(() => {
  if (!window.speechSynthesis || (window.speechSynthesis.getVoices && window.speechSynthesis.getVoices().length === 0)) {
    const realisticVoices = [
      { name: 'Google US English', lang: 'en-US', localService: false, voiceURI: 'Google US English', default: true },
      { name: 'Google UK English Female', lang: 'en-GB', localService: false, voiceURI: 'Google UK English Female', default: false },
      { name: 'Google UK English Male', lang: 'en-GB', localService: false, voiceURI: 'Google UK English Male', default: false },
      { name: 'Google español', lang: 'es-ES', localService: false, voiceURI: 'Google español', default: false },
      { name: 'Google français', lang: 'fr-FR', localService: false, voiceURI: 'Google français', default: false },
      { name: 'Google Deutsch', lang: 'de-DE', localService: false, voiceURI: 'Google Deutsch', default: false },
      { name: 'Google italiano', lang: 'it-IT', localService: false, voiceURI: 'Google italiano', default: false },
      { name: 'Google português do Brasil', lang: 'pt-BR', localService: false, voiceURI: 'Google português do Brasil', default: false },
      { name: 'Google русский', lang: 'ru-RU', localService: false, voiceURI: 'Google русский', default: false },
      { name: 'Google 日本語', lang: 'ja-JP', localService: false, voiceURI: 'Google 日本語', default: false },
      { name: 'Google 한국의', lang: 'ko-KR', localService: false, voiceURI: 'Google 한국의', default: false },
      { name: 'Google हिन्दी', lang: 'hi-IN', localService: false, voiceURI: 'Google हिन्दी', default: false },
      { name: 'Google 中文（普通话）', lang: 'zh-CN', localService: false, voiceURI: 'Google 中文（普通话）', default: false },
    ];
    if (window.speechSynthesis) {
      try { window.speechSynthesis.getVoices = function () { return realisticVoices; }; } catch (e) {}
    }
  }
});

// ── document.fonts — realistic OS font list ────────────────────
// Fingerprinters enumerate a known set; we return true for our OS list
// and let the browser handle the rest.
safe(() => {
  if (document.fonts && document.fonts.check) {
    const origCheck = document.fonts.check.bind(document.fonts);
    const fontSet = new Set(__KX.fonts.map((f) => f.toLowerCase()));
    document.fonts.check = function (font, text) {
      try {
        const m = String(font || '').match(/['"]([^'"]+)['"]/);
        if (m && fontSet.has(m[1].toLowerCase())) return true;
      } catch (e) {}
      return origCheck(font, text);
    };
  }
});

// ── WebRTC IP leak — block local IPs in ICE candidates ─────────
safe(() => {
  if (window.RTCPeerConnection) {
    const OrigRTC = window.RTCPeerConnection;
    const PatchedRTC = function (config) {
      const pc = new OrigRTC(config);
      const origAdd = pc.addIceCandidate.bind(pc);
      pc.addIceCandidate = function (cand) {
        try {
          if (cand && cand.candidate) {
            // Filter host/local IPs — only allow srflx (server reflexive, =proxy IP)
            if (/typ host/.test(cand.candidate) || /\.local /.test(cand.candidate)) {
              return Promise.resolve();
            }
          }
        } catch (e) {}
        return origAdd(cand);
      };
      // Also strip local candidates from generated ICE
      const origCreateOffer = pc.createOffer.bind(pc);
      pc.createOffer = function () {
        return origCreateOffer.apply(this, arguments).then((offer) => {
          try {
            offer.sdp = offer.sdp.replace(/^a=candidate.+typ host.+\r?\n/gm, '');
            offer.sdp = offer.sdp.replace(/^a=candidate.+\.local .+\r?\n/gm, '');
          } catch (e) {}
          return offer;
        });
      };
      return pc;
    };
    PatchedRTC.prototype = OrigRTC.prototype;
    window.RTCPeerConnection = PatchedRTC;
    if (window.webkitRTCPeerConnection) window.webkitRTCPeerConnection = PatchedRTC;
    if (window.mozRTCPeerConnection) window.mozRTCPeerConnection = PatchedRTC;
  }
  // Block enumerateDevices from exposing too many media devices
  if (navigator.mediaDevices && navigator.mediaDevices.enumerateDevices) {
    const origEnum = navigator.mediaDevices.enumerateDevices.bind(navigator.mediaDevices);
    navigator.mediaDevices.enumerateDevices = function () {
      return origEnum().then((devices) => {
        // Return realistic 2-4 devices for desktop, 2-3 for mobile
        return devices.length > 0 ? devices : [
          { deviceId: 'default', kind: 'audioinput', label: '', groupId: 'g1' },
          { deviceId: 'default', kind: 'audiooutput', label: '', groupId: 'g1' },
          ...(__KX.isMobile ? [] : [{ deviceId: 'default', kind: 'videoinput', label: '', groupId: 'g2' }]),
        ];
      });
    };
  }
});

// ── Hide Function.prototype.toString tampering (CDP detection) ─
safe(() => {
  const origToString = Function.prototype.toString;
  const proxyMap = new WeakMap();
  Function.prototype.toString = function () {
    if (proxyMap.has(this)) return proxyMap.get(this);
    return origToString.apply(this, arguments);
  };
  // Mark common patched fns to return "native code" string
  const nativeShim = (name) => `function ${name}() { [native code] }`;
  try { proxyMap.set(navigator.permissions.query, nativeShim('query')); } catch (e) {}
  try { proxyMap.set(WebGLRenderingContext.prototype.getParameter, nativeShim('getParameter')); } catch (e) {}
  try { proxyMap.set(HTMLCanvasElement.prototype.toDataURL, nativeShim('toDataURL')); } catch (e) {}
  try { proxyMap.set(CanvasRenderingContext2D.prototype.getImageData, nativeShim('getImageData')); } catch (e) {}
  try { proxyMap.set(navigator.getBattery, nativeShim('getBattery')); } catch (e) {}
});

// ── Notification API ───────────────────────────────────────────
safe(() => {
  if (window.Notification) {
    safeDefine(window.Notification, 'permission', () => 'default');
  }
});

// ── Date timezone sanity — match the proxy region ──────────────
// Playwright already sets timezone via context option; we just ensure
// the Intl object reports the same in case some detector compares.
safe(() => {
  try {
    const origResolved = Intl.DateTimeFormat.prototype.resolvedOptions;
    Intl.DateTimeFormat.prototype.resolvedOptions = function () {
      const r = origResolved.apply(this, arguments);
      r.timeZone = __KX.tz;
      return r;
    };
  } catch (e) {}
});

// ── window.history.length — realistic browsing depth ───────────
safe(() => {
  if (window.history && window.history.length <= 1) {
    safeDefine(window.history, 'length', () => 2);
  }
});

// ── Block automation-only debugger statements ──────────────────
// Some bot-test pages call `debugger;` in a tight loop to detect CDP.
// We can't stop them but they don't hurt our runtime.

// ── Final touch: hide ourselves from common detector libs ──────
safe(() => {
  // BotD / CreepJS look at Notification.requestPermission's
  // result-after-call behaviour; make it pass.
  if (window.Notification && Notification.requestPermission) {
    const orig = Notification.requestPermission;
    Notification.requestPermission = function () {
      return orig.apply(this, arguments).catch(() => 'default');
    };
  }
});
"""
    return "(()=>{" + config_js + js_body + "})();"


# ─── Validation-error detection (invalid data on landing) ──────────
# After a form submit, we scan the page for classic server-side / inline
# validation errors. DISABLED BY DEFAULT — too many landing pages show
# consent / marketing banners that use `.alert-danger` / `.text-danger`
# classes and give false positives on the form page itself. User opts in
# via `invalid_detection_enabled=true` on the create-job call.
# Selectors are the TIGHT set — only field-level validation.
_VALIDATION_ERROR_SELECTORS = [
    ".invalid-feedback",                       # Bootstrap per-field
    ".field-error",                            # common custom class
    ".form-field-error",
    ".ng-invalid-message",                     # Angular Material
    ".Mui-error + .MuiFormHelperText-root",    # MUI
    "input.is-invalid + .invalid-feedback",
    "[aria-invalid='true'] + .error",
    "[aria-invalid='true'] + .form-error",
]

_VALIDATION_ERROR_PHRASES = [
    # Must combine validation verb AND field noun to avoid matching
    # promotional / consent text on the form page itself.
    "invalid email", "invalid e-mail", "invalid zip", "invalid zipcode",
    "invalid postal", "invalid postcode", "invalid phone", "invalid number",
    "invalid address", "invalid date",
    "please enter a valid email", "please enter a valid phone",
    "please enter a valid zip", "please enter a valid address",
    "not a valid email", "not a valid phone", "not a valid zip",
    "enter a valid email address", "enter a valid phone number",
    "enter a valid zip code", "enter a valid postal code",
    # duplicate / already — be specific
    "already registered", "already submitted", "duplicate submission",
    "duplicate email", "already exists in our system", "already in our system",
    # explicit submission failure
    "submission failed", "could not be submitted", "validation failed",
]


# ─── 2026-05: Offer-site "Duplicate IP" landing-page detection ─────
# Some offers (e.g. getmyoffer.app) have their OWN anti-fraud layer
# that shows a hard-block "Duplicate IP" page BEFORE the form even
# renders, whenever the visiting exit-IP has been seen by the offer
# before. Our `duplicate_ip_set` only covers OUR clicks DB — it can't
# know what the offer site already saw. So we must scan the landing
# page body for these phrases and, on hit, burn the IP from THIS
# job's in-memory dup-set so the on-demand ProxyJet probe will reject
# it on the next visit (and any subsequent visit in the same job).
#
# Phrases are kept VERY specific — generic "access denied" alone is
# avoided because it false-positives on geo-blocks and Cloudflare
# challenges that are NOT IP-duplicate situations.
_DUPLICATE_IP_PAGE_PHRASES = [
    "duplicate ip",
    "this ip address has already been used",
    "this ip has already been used",
    "ip address has already been used",
    "ip already used",
    "ip has been used before",
    "your ip is blocked",
    "this ip is blocked",
    "ip address is blocked",
    "ip blocked",
    "duplicate visit",
    "duplicate entry detected",
    "duplicate click",
]


# ─── 2026-05: Offer-site "VPN detected" landing-page detection ─────
# Some offers (e.g. getmyoffer.app on certain Amazon survey campaigns)
# fingerprint the visiting exit-IP against their own VPN/datacenter
# block-list and show a "Please turn off your VPN" page BEFORE the
# form even renders. We treat this identically to the Duplicate-IP
# page: burn the exit-IP from the in-job dup-set AND persist it to
# the rut_burnt_ips collection so every FUTURE job (including those
# running on the same VPS days/weeks later) will skip this IP from
# the ProxyJet pool automatically.
#
# Phrases kept narrow — bare "vpn" alone is avoided because some
# legitimate landing pages mention VPN in unrelated marketing copy.
_VPN_BLOCK_PAGE_PHRASES = [
    "please turn off your vpn",
    "turn off your vpn",
    "turn off the vpn",
    "disable your vpn",
    "disable the vpn",
    "kindly disable your vpn",
    "vpn detected",
    "vpn is not allowed",
    "vpn or proxy detected",
    "vpn/proxy detected",
    "we detected a vpn",
    "you appear to be using a vpn",
    "vpn usage detected",
    "vpn blocked",
    "no vpn allowed",
]


async def _detect_offer_duplicate_ip_block(page: "Page") -> Tuple[bool, str]:
    """Detect offer-site IP-duplicate hard-block landing page.

    Returns (is_duplicate_ip_block, snippet).
    Safe — any exception → (False, '').
    """
    try:
        body_text = await page.evaluate(
            "() => (document.body ? document.body.innerText : '').toLowerCase().slice(0, 8000)"
        )
    except Exception:
        return False, ""
    if not body_text:
        return False, ""
    for phrase in _DUPLICATE_IP_PAGE_PHRASES:
        if phrase in body_text:
            idx = body_text.find(phrase)
            start = max(0, idx - 30)
            end = min(len(body_text), idx + 140)
            snippet = body_text[start:end].strip().replace("\n", " ")
            return True, snippet[:240]
    return False, ""


async def _detect_offer_vpn_block(page: "Page") -> Tuple[bool, str]:
    """Detect offer-site VPN/proxy-rejection landing page.

    Returns (is_vpn_block, snippet).
    Safe — any exception → (False, '').
    """
    try:
        body_text = await page.evaluate(
            "() => (document.body ? document.body.innerText : '').toLowerCase().slice(0, 8000)"
        )
    except Exception:
        return False, ""
    if not body_text:
        return False, ""
    for phrase in _VPN_BLOCK_PAGE_PHRASES:
        if phrase in body_text:
            idx = body_text.find(phrase)
            start = max(0, idx - 30)
            end = min(len(body_text), idx + 140)
            snippet = body_text[start:end].strip().replace("\n", " ")
            return True, snippet[:240]
    return False, ""


# ─── 2026-05 — PRE-BROWSER offer-side duplicate/VPN probe ──────────
# Issue: even when our `duplicate_ip_set` (clicks DB + rut_burnt_ips)
# says the exit-IP is "unique", the OFFER itself may still reject it
# because the offer's own dedup list is invisible to us. Previously
# we only discovered this AFTER spawning Chromium and navigating —
# which meant the Live Activity feed showed real "Duplicate IP"
# screenshots and the visit slot wasted ~10-30 seconds of browser
# overhead per rejected IP.
#
# This probe does a cheap GET via httpx (no browser, no JS) against
# the resolved offer URL through the SAME proxy. If the response
# body contains any of the duplicate / VPN block phrases, we bail
# out instantly and the worker retries with a fresh ProxyJet IP —
# never opening a browser, never producing a screenshot. The
# operator sees the visit slot transparently scroll through fresh
# IPs in the Live Activity feed until a clean one is found, with
# no "Duplicate IP" thumbnails in between.
#
# Falls back gracefully on transport errors — those are handled by
# the existing reachability probe.
async def _probe_offer_duplicate_via_proxy(
    proxy: Dict[str, Any], target_url: str, ua: str,
    timeout_s: float = 12.0,
) -> Tuple[bool, str, str]:
    """Pre-browser GET probe to detect offer-side duplicate-IP / VPN block pages.

    Returns ``(is_blocked, reason, snippet)`` where:
      • is_blocked  — True if the response body matches a known
                       duplicate-IP or VPN-block phrase
      • reason      — "duplicate_ip" | "vpn" | "" (never blocked)
      • snippet     — first ~240 chars of matched context (for logs)

    On transport errors → (False, "", "") so the caller falls through
    to the normal browser path; the existing reachability probe is
    responsible for catching dead proxies separately.
    """
    server = proxy.get("server", "")
    if proxy.get("username"):
        try:
            prefix, rest = server.split("://", 1)
            server = f"{prefix}://{proxy['username']}:{proxy.get('password','')}@{rest}"
        except ValueError:
            return False, "", ""

    headers = {
        "User-Agent": ua or "Mozilla/5.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        # Hint we want full HTML so the offer renders its real page
        # (not a 206 partial-content stub from a Range header).
        "Cache-Control": "no-cache",
    }
    timeout_cfg = httpx.Timeout(timeout_s, connect=min(8.0, timeout_s))
    try:
        async with httpx.AsyncClient(
            proxy=server, timeout=timeout_cfg, headers=headers,
            verify=False, http2=False, follow_redirects=True,
        ) as cli:
            r = await cli.get(target_url)
            # Only read first ~32 KB — duplicate / VPN markers are
            # ALWAYS in the visible top-of-page text; deep scans waste
            # bandwidth on multi-MB landing pages.
            body = (r.text or "")[:32000].lower()
    except Exception:
        # Any transport error → treat as "not blocked here", let the
        # existing engine handle it via browser. We don't want a
        # network blip to mark a clean IP as duplicate.
        return False, "", ""

    if not body:
        return False, "", ""

    # Check duplicate FIRST (more specific reason for the user)
    for phrase in _DUPLICATE_IP_PAGE_PHRASES:
        if phrase in body:
            idx = body.find(phrase)
            start = max(0, idx - 30)
            end = min(len(body), idx + 140)
            snippet = body[start:end].strip().replace("\n", " ")
            return True, "duplicate_ip", snippet[:240]
    for phrase in _VPN_BLOCK_PAGE_PHRASES:
        if phrase in body:
            idx = body.find(phrase)
            start = max(0, idx - 30)
            end = min(len(body), idx + 140)
            snippet = body[start:end].strip().replace("\n", " ")
            return True, "vpn", snippet[:240]
    return False, "", ""


# ─── 2026-05: Persistent burnt-IP block-list (cross-job memory) ────
# When an exit-IP gets flagged by the offer site (duplicate or VPN)
# we persist it to `rut_burnt_ips` so EVERY future job loads it into
# its `duplicate_ip_set` and the on-demand ProxyJet probe skips it
# during the unique-IP retry loop. Without this, the same dirty IPs
# would keep coming back from the ProxyJet pool job after job
# (because they never made it into the regular clicks collection —
# the visit failed before the click could be recorded server-side).
async def _persist_burnt_ip(
    db,
    ip: str,
    reason: str,
    user_id: str = "",
    offer_url: str = "",
    state: str = "",
    job_id: str = "",
) -> None:
    """Upsert a burnt exit-IP into the rut_burnt_ips collection.

    Never raises — any failure is logged and swallowed. Uses $addToSet
    so the same IP getting flagged by different jobs/offers merges
    cleanly into one document with a history.
    """
    if not ip or not isinstance(ip, str):
        return
    try:
        await db.rut_burnt_ips.update_one(
            {"ip": ip.strip()},
            {
                "$set": {
                    "ip": ip.strip(),
                    "last_reason": reason or "unknown",
                    "last_detected_at": datetime.now(timezone.utc).isoformat(),
                },
                "$addToSet": {
                    "reasons": reason or "unknown",
                    **({"user_ids": user_id} if user_id else {}),
                    **({"offer_urls": offer_url} if offer_url else {}),
                    **({"states": state.upper()} if state else {}),
                    **({"job_ids": job_id} if job_id else {}),
                },
                "$inc": {"hit_count": 1},
                "$setOnInsert": {
                    "first_detected_at": datetime.now(timezone.utc).isoformat(),
                },
            },
            upsert=True,
        )
    except Exception as e:  # noqa: BLE001
        try:
            logger.warning(f"[rut_burnt_ips] persist failed for {ip}: {e}")
        except Exception:
            pass


# ─── US state matching — map + normaliser ──────────────────────────
# Used by the "match lead state to proxy IP state" feature so a row from
# California only gets submitted via a CA-exit proxy, etc. We accept both
# 2-letter codes ("CA") and full names ("California") on both sides.
_US_STATES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming", "DC": "District of Columbia",
    "PR": "Puerto Rico", "GU": "Guam", "VI": "U.S. Virgin Islands",
}
_US_NAME_TO_CODE = {v.lower(): k for k, v in _US_STATES.items()}
_US_CODES = set(_US_STATES.keys())


def _normalize_state(s: Any) -> str:
    """Normalise a US state value to its 2-letter uppercase code.
    Accepts codes ("ca", "CA"), full names ("California"), mixed case, with
    trailing whitespace. Returns '' if not recognised."""
    if s is None:
        return ""
    txt = str(s).strip()
    if not txt:
        return ""
    up = txt.upper()
    if up in _US_CODES:
        return up
    low = txt.lower()
    if low in _US_NAME_TO_CODE:
        return _US_NAME_TO_CODE[low]
    # Handle "California, USA" / "California (CA)" / "NJ (New Jersey)" style —
    # try each part separately against codes and names
    import re
    # Split on any separator: comma / paren / slash / dash / pipe
    parts = re.split(r"[,()\/|\-]", txt)
    for raw in parts:
        part = raw.strip()
        if not part:
            continue
        if part.upper() in _US_CODES:
            return part.upper()
        if part.lower() in _US_NAME_TO_CODE:
            return _US_NAME_TO_CODE[part.lower()]
    # 2-letter-only fallback: strip non-alpha → if exactly 2 uppercase letters matching a code
    stripped = re.sub(r"[^A-Za-z]", "", txt).upper()
    if len(stripped) == 2 and stripped in _US_CODES:
        return stripped
    return ""


def _find_state_column(rows: List[Dict[str, Any]]) -> Optional[str]:
    """Return the key in the row dicts that holds the US-state value, or None.
    Looks for common name variations (state, State, region, st, state_code, etc.)."""
    if not rows:
        return None
    # Gather all unique keys across first few rows
    seen = []
    for r in rows[:10]:
        for k in r.keys():
            if k not in seen:
                seen.append(k)
    # Priority order
    priority = [
        "state", "State", "STATE",
        "state_code", "stateCode", "StateCode",
        "region", "Region", "REGION",
        "st", "ST",
        "province", "Province",
    ]
    for p in priority:
        if p in seen:
            return p
    # Fallback: case-insensitive match on any key ending in 'state'
    for k in seen:
        if k.strip().lower() in ("state", "st", "region", "state_code", "statecode"):
            return k
    return None


async def _detect_validation_errors(page: Page) -> Tuple[bool, str]:
    """Scan a page for inline / server-side validation errors.

    Returns (is_invalid, error_message).  Safe — any exception → (False, '').
    """
    # 1. Visible elements with error classes
    try:
        for sel in _VALIDATION_ERROR_SELECTORS:
            try:
                els = await page.query_selector_all(sel)
            except Exception:
                continue
            for el in els:
                try:
                    if not await el.is_visible():
                        continue
                    txt = ((await el.inner_text()) or "").strip()
                except Exception:
                    continue
                if txt and 2 < len(txt) < 400:
                    # Skip benign "required" labels that show up when field is empty
                    # (we only care about validation that FAILED after submit).
                    low = txt.lower()
                    if low in ("required", "*required", "required field"):
                        continue
                    return True, txt[:200]
    except Exception:
        pass

    # 2. Body-text phrase scan (catch server-rendered error banners)
    try:
        body_text = await page.evaluate(
            "() => (document.body ? document.body.innerText : '').toLowerCase().slice(0, 8000)"
        )
    except Exception:
        body_text = ""
    if body_text:
        for phrase in _VALIDATION_ERROR_PHRASES:
            if phrase in body_text:
                # grab a nearby snippet for context
                idx = body_text.find(phrase)
                start = max(0, idx - 30)
                end = min(len(body_text), idx + 120)
                snippet = body_text[start:end].strip().replace("\n", " ")
                return True, snippet[:200]

    return False, ""


# ── 2026-01: Server-side tracker bypass ────────────────────────────
# When the residential-proxy provider refuses to connect to OUR OWN
# backend domain (e.g. ProxyJet returning 502 Bad Gateway on
# `api.krexion.com` because of an internal block-list), there's no
# need to go through the proxy at all for that one HTTP hop — the
# backend can hit its own tracker endpoint via 127.0.0.1 and have it
# record the click as if it came from the proxy's exit IP by
# injecting the IP into the `X-Forwarded-For` header. After that the
# browser-via-proxy can navigate directly to the offer URL (which
# the proxy IS allowed to reach), and the rest of the visit
# (form-fill, conversion, screenshots) proceeds normally.
#
# Behaviour is OPT-IN by hostname: only triggers if the target URL's
# host is listed in the env var RUT_LOCALHOST_BYPASS_HOSTS (default
# "krexion.com,api.krexion.com,localhost,127.0.0.1") AND the proxy's
# tunnel failed with a 502 / tunnel error after all MAX_TUNNEL_RETRIES.
# Set RUT_LOCALHOST_BYPASS_HOSTS="" to fully disable the bypass.
#
# ── 2026-05: STRICT PROXY MODE ──────────────────────────────────────────
# The bypass mechanism uses the SERVER's network connection to register
# the click (with X-Forwarded-For set to the proxy IP). The customer's
# real network/server IP IS the TCP source for that one hop, which the
# user wants to avoid at all costs. We therefore default to STRICT mode:
# the direct bypass is DISABLED unless the operator explicitly sets
#     RUT_ALLOW_DIRECT_BYPASS=true
# in backend/.env. In strict mode, if no proxy can reach the tracker,
# the visit fails cleanly — guaranteeing that NO traffic ever leaves
# the customer's box bypassing the proxy.

def _bypass_hosts() -> set:
    raw = os.environ.get(
        "RUT_LOCALHOST_BYPASS_HOSTS",
        "krexion.com,api.krexion.com,localhost,127.0.0.1",
    )
    return {h.strip().lower() for h in raw.split(",") if h.strip()}


def _url_host_matches_bypass(url: str) -> bool:
    """True if the URL's hostname (or any parent domain) is in the
    bypass list — i.e. we know proxies can't reach it but localhost
    can."""
    if not url:
        return False
    try:
        from urllib.parse import urlparse
        host = (urlparse(url).hostname or "").lower()
    except Exception:
        return False
    if not host:
        return False
    bypass = _bypass_hosts()
    if not bypass:
        return False
    if host in bypass:
        return True
    # Also match parent domains (api.krexion.com matches krexion.com)
    for b in bypass:
        if host.endswith("." + b):
            return True
    return False


async def _get_exit_ip_via_proxy(
    proxy: Dict[str, Any], timeout: float = 10.0
) -> Optional[str]:
    """Fetch the residential proxy's exit IP by querying a neutral
    IP-echo endpoint THROUGH the proxy. Used only when the regular
    page-based detection can't run because the browser failed to
    reach the target. Returns None on any failure (caller must
    treat that as 'bypass not possible'). Never raises."""
    server = (proxy.get("server") or "").strip()
    username = (proxy.get("username") or "").strip()
    password = (proxy.get("password") or "").strip()
    if not server:
        return None
    # Build proxy URL for httpx
    if "://" in server:
        proto, hostport = server.split("://", 1)
    else:
        proto, hostport = "http", server
    if username:
        from urllib.parse import quote
        proxy_url = (
            f"{proto}://{quote(username, safe='')}:"
            f"{quote(password, safe='')}@{hostport}"
        )
    else:
        proxy_url = f"{proto}://{hostport}"
    # Try a couple of neutral IP-echo endpoints. These are known to
    # work through proxy-jet & most residential providers because
    # they are NOT on any tracker blocklist.
    candidates = (
        "https://api.ipify.org?format=json",
        "https://ipinfo.io/ip",
        "https://ifconfig.me/ip",
    )
    for url in candidates:
        try:
            async with httpx.AsyncClient(
                proxy=proxy_url, timeout=timeout, follow_redirects=True
            ) as c:
                r = await c.get(url)
                if r.status_code != 200:
                    continue
                txt = (r.text or "").strip()
                if not txt:
                    continue
                if txt.startswith("{"):
                    try:
                        import json as _json
                        data = _json.loads(txt)
                        ip = data.get("ip") or data.get("origin")
                        if ip:
                            return ip.split(",")[0].strip()
                    except Exception:
                        continue
                # Plain text IP response
                first = txt.splitlines()[0].strip()
                # Sanity: looks like an IPv4
                if first.count(".") == 3 and all(p.isdigit() for p in first.split(".")):
                    return first
        except Exception:
            continue
    return None


def _normalize_unresolvable_tracker_host(url: str) -> str:
    """2026-01 — DNS / SSL auto-fallback for misconfigured tracker subdomains.

    If `url`'s host is a `<sub>.<domain>` form (typically `api.krexion.com`)
    AND
       (a) it isn't publicly resolvable, OR
       (b) the path starts with `/api/t/` (Krexion's tracker route, which
           the same backend serves on the bare apex regardless of which
           host header is used)
    THEN we rewrite the host to the bare parent domain
    (e.g. `krexion.com`). Otherwise the URL is returned as-is.

    Why this exists: many Krexion deployments use `api.<domain>` as the
    tracker host but:
      • only configure a public DNS A record (or SSL certificate) for
        the bare apex/www domain, not the `api.` subdomain
      • OR configure DNS but never extend the apex SSL certificate to
        cover the subdomain — so connecting to `api.<domain>:443` via
        any residential proxy fails with `TLSV1_ALERT_INTERNAL_ERROR`
      • OR the `api.` subdomain is set up internally via nginx
        `server_name` or a local hosts file and is invisible to public
        DNS / proxies

    In every one of these failure modes, the proxy → 502 / SSL alert is
    fatal for the visit. Rewriting to the bare domain side-steps all of
    it, because the Krexion tracker route (`/api/t/<short_code>`) is
    served by the same FastAPI backend regardless of which host the
    request comes in on — and the apex domain ALWAYS has a working
    SSL certificate (it's where the rest of the user-facing site lives).

    Implementation note: we explicitly resolve via PUBLIC DNS
    (Cloudflare 1.1.1.1, with Google 8.8.8.8 + Quad9 fallback) rather
    than the local resolver. This is critical on the user's production
    VPS where `api.<domain>` IS resolvable locally (via /etc/hosts or
    nginx server_name) but may still have no PUBLIC A record / SSL —
    so a system-default `socket.gethostbyname()` would falsely report
    it as "resolvable" and skip the rewrite.

    Path-based rewrite kicks in even when DNS DOES resolve, because
    the more common failure mode in production is SSL cert mismatch
    (cheaper to detect by path pattern than by attempting a real TLS
    handshake on every job start).

    Safe to call repeatedly: per-host results are LRU-cached in
    process memory after the first DNS round-trip (~50-150ms).
    """
    try:
        from urllib.parse import urlparse, urlunparse
        parsed = urlparse(url)
        host = (parsed.hostname or "").lower()
        path = parsed.path or ""
        if not host:
            return url
        # Must look like a subdomain (at least 3 dot-separated labels)
        parts = host.split(".")
        if len(parts) < 3:
            return url

        # ── Path-based rewrite (always applied for Krexion tracker URLs) ─
        # Krexion's tracker route is `/api/t/<short_code>` and the
        # FastAPI backend serves it on every Host header the apex/www
        # accepts. If we spot this path on a subdomain'd host, prefer
        # the apex straight away — saves a DNS round-trip AND avoids
        # the apex-only-SSL-cert failure mode entirely.
        if path.startswith("/api/t/"):
            parent = ".".join(parts[1:])
            if parent and parent.count(".") >= 1:
                port = parsed.port
                netloc = f"{parent}:{port}" if port else parent
                return urlunparse((
                    parsed.scheme, netloc, parsed.path,
                    parsed.params, parsed.query, parsed.fragment,
                ))

        cache_key = host
        try:
            cached = _NORMALIZE_TRACKER_CACHE.get(cache_key)
        except NameError:
            cached = None

        def _resolves_public(h: str) -> bool:
            """Query a public DNS server directly so we ignore any
            local /etc/hosts entry or nginx server_name shortcut that
            only exists on the deployment server."""
            import socket
            try:
                import struct, random
                def _udp_query(server_ip: str, hostname: str, timeout: float = 3.0) -> bool:
                    try:
                        tid = random.randint(0, 65535)
                        header = struct.pack(">HHHHHH", tid, 0x0100, 1, 0, 0, 0)
                        qname = b""
                        for label in hostname.split("."):
                            qname += bytes([len(label)]) + label.encode("ascii", errors="ignore")
                        qname += b"\x00"
                        qbody = qname + struct.pack(">HH", 1, 1)  # A IN
                        pkt = header + qbody
                        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                        s.settimeout(timeout)
                        try:
                            s.sendto(pkt, (server_ip, 53))
                            data, _ = s.recvfrom(2048)
                        finally:
                            try:
                                s.close()
                            except Exception:
                                pass
                        if len(data) < 12:
                            return False
                        ancount = struct.unpack(">H", data[6:8])[0]
                        flags = struct.unpack(">H", data[2:4])[0]
                        rcode = flags & 0x0F
                        if rcode == 3:  # NXDOMAIN
                            return False
                        if rcode != 0:
                            return False
                        return ancount > 0
                    except Exception:
                        return False
                for resolver_ip in ("1.1.1.1", "8.8.8.8", "9.9.9.9"):
                    try:
                        if _udp_query(resolver_ip, h):
                            return True
                    except Exception:
                        continue
                return False
            except Exception:
                try:
                    socket.gethostbyname(h)
                    return True
                except Exception:
                    return False

        if cached is not None:
            new_host = cached
        else:
            if _resolves_public(host):
                new_host = host
            else:
                parent = ".".join(parts[1:])
                if _resolves_public(parent):
                    new_host = parent
                else:
                    new_host = host
            try:
                _NORMALIZE_TRACKER_CACHE[cache_key] = new_host
            except NameError:
                pass

        if new_host == host:
            return url
        port = parsed.port
        netloc = f"{new_host}:{port}" if port else new_host
        return urlunparse((
            parsed.scheme, netloc, parsed.path,
            parsed.params, parsed.query, parsed.fragment,
        ))
    except Exception:
        return url


# Process-wide cache to avoid repeating DNS lookups for the same host
# across thousands of RUT visits. Populated lazily by
# _normalize_unresolvable_tracker_host.
_NORMALIZE_TRACKER_CACHE: Dict[str, str] = {}


async def _resolve_tracker_via_localhost(
    target_url: str,
    exit_ip: str,
    user_agent: str,
    timeout: float = 15.0,
) -> Optional[str]:
    """Resolve the tracker server-side WITHOUT going through the
    residential proxy, while still recording the click as the
    proxy's exit IP via the X-Forwarded-For header.

    Architecture-aware: this fits two deployment styles —
      1. **Distributed**: the RUT browser runs on a customer PC but
         the tracker (`api.krexion.com/api/t/...`) lives on a remote
         VPS. The PC's backend (where RUT is running) can reach the
         VPS over its regular internet connection — no proxy needed.
      2. **All-local**: the RUT browser AND the tracker run on the
         same machine (e.g. a self-contained krexion install) and
         127.0.0.1 reaches the tracker too.

    Strategy: try the ORIGINAL target URL first (direct internet
    call from the RUT host, no proxy, with X-Forwarded-For injected).
    If that fails for any reason (DNS, host firewall, 404 because the
    link only exists on the local backend, etc.) we fall back to
    127.0.0.1:${LOCAL_BACKEND_PORT|PORT|8001}. Whichever returns a
    3xx redirect wins.

    Returns the Location URL from the 3xx response, or None on
    failure. Never raises.
    """
    if not (target_url and exit_ip):
        return None

    headers = {
        # Standard reverse-proxy headers — most FastAPI / Caddy /
        # Nginx setups read one of these to determine the real
        # client IP. We set all the common ones so whichever the
        # backend trusts will pick up the residential exit IP.
        "X-Forwarded-For": exit_ip,
        "X-Real-IP": exit_ip,
        "True-Client-IP": exit_ip,
        "X-Client-IP": exit_ip,
        # NOTE: We intentionally do NOT set CF-Connecting-IP here.
        # If the request actually transits Cloudflare, CF replaces
        # that header with the request's real source IP (so our
        # value would be ignored anyway). For non-CF setups, the
        # other four headers above cover the common cases.
        "User-Agent": user_agent or "Mozilla/5.0",
        "Accept": (
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "image/avif,image/webp,*/*;q=0.8"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    }

    candidates: List[str] = []
    try:
        from urllib.parse import urlparse, urlunparse
        parsed = urlparse(target_url)

        # Candidate #1 — call the ORIGINAL target URL directly (no
        # proxy). This is the right path for the typical owner setup
        # where the tracker is on a remote VPS but RUT runs on a PC.
        if parsed.scheme and parsed.netloc:
            candidates.append(target_url)

        # Candidate #2 — fall back to 127.0.0.1 on the configured
        # backend port. Right path for self-contained installs where
        # the tracker lives in the same box.
        local_port = (
            os.environ.get("LOCAL_BACKEND_PORT")
            or os.environ.get("PORT")
            or "8001"
        )
        local_url = urlunparse((
            "http",
            f"127.0.0.1:{local_port}",
            parsed.path or "/",
            parsed.params,
            parsed.query,
            parsed.fragment,
        ))
        candidates.append(local_url)
    except Exception:
        return None

    # Try each candidate; the first one to return a 3xx wins.
    for url in candidates:
        try:
            # The Host header for the 127.0.0.1 call should still be
            # the public hostname so the tracker's link lookup picks
            # up the right tenant (matters for multi-tenant setups).
            req_headers = dict(headers)
            if url.startswith("http://127.0.0.1"):
                try:
                    public_host = urlparse(target_url).hostname or ""
                    if public_host:
                        req_headers["Host"] = public_host
                        req_headers["X-Forwarded-Host"] = public_host
                        req_headers["X-Forwarded-Proto"] = "https"
                except Exception:
                    pass

            async with httpx.AsyncClient(
                follow_redirects=False,
                timeout=timeout,
                trust_env=False,
                # Verify SSL for public URLs; for 127.0.0.1 we use
                # http:// so verify doesn't matter.
            ) as c:
                r = await c.get(url, headers=req_headers)
            if r.status_code in (301, 302, 303, 307, 308):
                loc = r.headers.get("Location") or r.headers.get("location")
                if loc:
                    loc_stripped = loc.strip()
                    # SAFETY: never propagate a customer-visible
                    # 127.0.0.1 / localhost URL back to the browser
                    # — the browser address bar must show only the
                    # real offer URL for a professional appearance.
                    # Reject any Location that resolves to a private
                    # / loopback host; the next candidate (or the
                    # original failure path) will be tried instead.
                    _loc_low = loc_stripped.lower()
                    if (
                        "127.0.0.1" in _loc_low
                        or "://localhost" in _loc_low
                        or _loc_low.startswith("localhost")
                    ):
                        continue
                    return loc_stripped
        except Exception:
            # Move on to the next candidate.
            continue
    return None


# ─── Job runner ──────────────────────────────────────────────────
async def run_real_user_traffic_job(
    job_id: str,
    target_url: str,
    proxies_raw: List[str],
    user_agents: List[str],
    total_clicks: int,
    concurrency: int,
    duration_minutes: float,
    allowed_os: List[str],
    allowed_countries_lc: List[str],
    skip_duplicate_ip: bool,
    skip_vpn: bool,
    follow_redirect: bool,
    no_repeated_proxy: bool,
    form_fill_enabled: bool,
    rows: Optional[List[Dict[str, Any]]],
    skip_captcha: bool,
    duplicate_ip_set: Optional[set],
    post_submit_wait: int = 3,  # Reduced from 6s to 3s for speed optimization
    automation_steps: Optional[List[Dict[str, Any]]] = None,
    self_heal: bool = True,
    state_match_enabled: bool = False,
    target_mode: str = "clicks",                # "clicks" | "conversions"
    target_conversions: int = 0,
    max_attempts: int = 0,
    invalid_detection_enabled: bool = False,    # OFF by default — consent-text
                                                # banners were causing false positives
    db=None,
    link_id: Optional[str] = None,
    link_owner_id: Optional[str] = None,
    link_short_code: Optional[str] = None,
    # Per-use immediate removal of consumed items from the saved
    # "Uploaded Things" batches. As soon as a proxy / UA is picked for a
    # visit (or a row index is successfully submitted), it is pulled
    # from the saved batch in MongoDB / overwritten in the on-disk XLSX.
    # User explicitly asked for this real-time behaviour rather than a
    # batched end-of-job consume.
    engine_user_id: Optional[str] = None,
    upload_proxy_id: Optional[str] = None,
    upload_ua_id: Optional[str] = None,                 # legacy single-batch (backward compat)
    upload_ua_ids: Optional[List[str]] = None,          # multi-batch — preferred
    ua_to_batch_map: Optional[Dict[str, str]] = None,   # ua_string → batch_id (for per-batch $pull)
    upload_data_file_id: Optional[str] = None,
    # Target Screenshot Verification — perceptual hash of the user-uploaded
    # reference image of the expected final/thank-you page. After every
    # visit we compare the live final-page screenshot against this hash;
    # distance ≤ threshold counts as a VERIFIED conversion. Heuristic
    # host-change "thank_you_reached" still fires for stats but is no
    # longer treated as conversion when target_screenshot_phash is set.
    target_screenshot_phash: str = "",
    target_screenshot_threshold: int = 12,
    # ── 2026-01: ProxyJet ROW-FIRST on-demand mode ─────────────────
    # When True, the engine ignores the pre-loaded proxies list and
    # instead fetches a fresh state-matched ProxyJet IP per visit
    # (pick row → get state → fetch IP for that state → retry until
    # exit-IP is unique → use). Pre-flight check: every row MUST have
    # a non-empty state value; otherwise the whole job fails with a
    # clear error per the user's spec.
    proxyjet_on_demand: bool = False,
    proxyjet_country: str = "US",
    proxyjet_default_state: Optional[str] = None,
    proxyjet_unique_retry_cap: int = 50,
    # 2026-01: per-job override for the stuck-watchdog inactivity
    # threshold (seconds the page's main-frame URL is allowed to stay
    # unchanged before the visit is force-aborted). Default raised from
    # 60 → 240 (2026-05) so slow survey-style offer pages get enough
    # time to complete multi-step SPA + form-submit + thank-you flows.
    # chrome-error:// fast-path still fires INSTANTLY so dead proxies
    # are still aborted immediately. Operators can lower it (faster
    # fail-fast) or raise it via UI for extreme cases.
    stuck_watchdog_seconds: float = 240.0,
    # ── 2026-05: Pure JSON Mode ─────────────────────────────────────
    # When True, the engine STRICTLY follows the recorded automation
    # JSON without any AI involvement:
    #   • self_heal is forced OFF (no Gemini vision rescue on failed
    #     steps — if a step fails, the visit fails)
    #   • AI answer-learning (Thompson sampling for survey picks) is
    #     bypassed — survey/random-pick steps stay purely random and
    #     no outcomes are recorded to bias future picks
    # When False (default), behaves exactly as before (AI features ON
    # subject to the existing self_heal flag).
    pure_json_mode: bool = False,
):
    """
    Main orchestrator. Emits progress into RUT_JOBS[job_id].
    """
    # Guarantee chromium is installed BEFORE launching any visits.
    # This is the single robust guard that recovers from pod restarts that
    # wipe ad-hoc browser installs. First job on a fresh pod will pause
    # here for ~30-60s while the install runs; subsequent jobs are no-ops
    # (binary already present).
    try:
        push_live_step(job_id, 0, "preflight", "info", "Verifying browser engine…")
    except Exception:
        pass
    ok = await _ensure_chromium_available()
    if not ok:
        await _finalise_and_persist(db, job_id, "failed",
                  "Playwright chromium-headless-shell could not be installed. "
                  "Please contact support or retry — the install will be attempted again on the next job.")
        return

    # ── 2026-01: Auto-fix unresolvable tracker subdomains ───────────
    # If the target URL's host is e.g. `api.krexion.com` and that
    # subdomain isn't publicly resolvable (residential proxies can't
    # reach it → 502 Bad Gateway for every visit), automatically fall
    # back to the bare parent domain (`krexion.com`). The Krexion
    # tracker route lives on the same backend regardless of which
    # host header the request comes in on, so this fix is transparent.
    _orig_target_url = target_url
    try:
        target_url = await asyncio.to_thread(
            _normalize_unresolvable_tracker_host, target_url,
        )
    except Exception:
        target_url = _orig_target_url
    if target_url != _orig_target_url:
        try:
            from urllib.parse import urlparse as _np
            _old_h = (_np(_orig_target_url).hostname or "")
            _new_h = (_np(target_url).hostname or "")
            push_live_step(
                job_id, 0, "preflight", "info",
                f"Tracker host '{_old_h}' isn't publicly resolvable — auto-falling back to '{_new_h}' so residential proxies can reach it.",
            )
        except Exception:
            pass

    parsed_proxies: List[Dict[str, Any]] = []
    for ln in proxies_raw:
        p = _parse_proxy_line(ln)
        if p:
            parsed_proxies.append(p)
    # ── 2026-01 ROW-FIRST mode ─────────────────────────────────────
    # In ProxyJet on-demand mode the engine fetches a fresh IP per
    # visit, so the pre-supplied proxies list is allowed to be empty
    # (server.py passes [] in this mode). For every other mode we
    # still require at least one valid proxy line up-front.
    if not parsed_proxies and not proxyjet_on_demand:
        await _finalise_and_persist(db, job_id, "failed", "No valid proxies after parsing")
        return

    uas = [u.strip() for u in user_agents if u and u.strip()]
    if not uas:
        await _finalise_and_persist(db, job_id, "failed", "No user agents provided")
        return

    # Pre-filter UAs by allowed_os
    allowed_os_set = set((allowed_os or []))
    if allowed_os_set:
        uas_ok = [u for u in uas if _os_key_from_ua(u) in allowed_os_set]
        if not uas_ok:
            sample_detect = [(u[:60], _os_key_from_ua(u)) for u in uas[:3]]
            await _finalise_and_persist(job_id=job_id, db=db, status="failed",
                      error=(
                          f"All UAs filtered by allowed_os={sorted(allowed_os_set)}. "
                          f"Detected: {sample_detect}"
                      ))
            return
    else:
        uas_ok = uas

    job_dir = RESULTS_ROOT / job_id
    shots_dir = job_dir / "screenshots"
    shots_dir.mkdir(parents=True, exist_ok=True)

    total = max(1, min(int(total_clicks), 100000))
    delay_between = (duration_minutes * 60.0 / total) if duration_minutes and duration_minutes > 0 else 0.0

    RUT_JOBS[job_id].update({
        "status": "running",
        "started_at": datetime.now(timezone.utc).isoformat(),
        "total": total,
        "processed": 0,
        "succeeded": 0,
        "conversions": 0,
        "skipped_captcha": 0,
        "skipped_country": 0,
        "skipped_os": 0,
        "skipped_duplicate_ip": 0,
        "skipped_vpn": 0,
        "skipped_state_mismatch": 0,
        "skipped_no_unique_ip": 0,
        "skipped_dead_proxy": 0,
        "invalid_data": 0,
        "failed": 0,
        "started_at": datetime.now(timezone.utc).isoformat(),
        "events": [],
        "form_fill_enabled": form_fill_enabled,
        "state_match_enabled": state_match_enabled,
        "invalid_detection_enabled": invalid_detection_enabled,
        "target_mode": target_mode if target_mode in ("clicks", "conversions") else "clicks",
        "target_conversions": int(target_conversions or 0) if target_mode == "conversions" else 0,
        "max_attempts": int(max_attempts or 0) if target_mode == "conversions" else 0,
        # Link context — enables RUT visits to be logged as clicks in the
        # link's user DB (mirrors the /api/t/ tracker behaviour so the
        # dashboard Clicks page shows these visits too).
        "link_id": link_id,
        "link_owner_id": link_owner_id,
        "link_short_code": link_short_code,
        # 2026-05: When ProxyJet Auto Mode is on, silently absorb
        # offer/tracker-side Duplicate-IP and VPN block pages —
        # they don't count toward `processed`, don't show in
        # Recent Visits, and their early-logged click_doc is
        # deleted on detection. Diagnostics counter
        # `silent_skip_count` tracks them. The dispatcher uses a
        # HARD_CAP-bounded while-loop so the user gets exactly
        # `total_clicks` VISIBLE visits even if 90 % of ProxyJet
        # IPs hit the tracker's duplicate filter.
        "silent_skip_burnt_ip": bool(proxyjet_on_demand),
        "silent_skip_count": 0,
        "silent_skip_breakdown": {},
    })
    if db is not None:
        await _persist(db, job_id)

    # State-matching config — only honour if rows actually have a state column
    state_col: Optional[str] = None
    state_index: Dict[str, List[int]] = {}  # state_code -> list of row indices
    if state_match_enabled and rows:
        state_col = _find_state_column(rows)
        if state_col:
            for idx, r in enumerate(rows):
                code = _normalize_state(r.get(state_col))
                if code:
                    state_index.setdefault(code, []).append(idx)
        if not state_col or not state_index:
            # Turn off the feature quietly if the file doesn't have a state column
            RUT_JOBS[job_id]["state_match_enabled"] = False
            state_match_enabled = False
    RUT_JOBS[job_id]["state_match_column"] = state_col or ""

    # ── 2026-01 ROW-FIRST pre-flight state validation ──────────────
    # In ProxyJet on-demand mode every visit picks a row FIRST and
    # then asks ProxyJet for an IP in that row's state. Per the user's
    # spec the data file MUST have a state column AND every row must
    # have a non-empty state; otherwise the whole job fails with a
    # clear error so the operator can fix the data before any visits
    # consume credits.
    if proxyjet_on_demand:
        if not rows:
            await _finalise_and_persist(
                db, job_id, "failed",
                "ProxyJet on-demand mode requires a data file with rows. "
                "Upload a data file (Excel/CSV/Google Sheet) and try again."
            )
            return
        # Re-detect state column even if state_match_enabled wasn't ticked —
        # this mode mandates state-based IPs.
        if not state_col:
            state_col = _find_state_column(rows)
        if not state_col:
            await _finalise_and_persist(
                db, job_id, "failed",
                "ProxyJet on-demand mode requires a STATE column in your data "
                "file (column named 'state', 'State', 'STATE', 'region', 'st', "
                "or 'state_code'). None was detected. Add a state column and try again."
            )
            return
        # Validate every single row has a non-empty, valid US state.
        empty_state_rows: List[int] = []
        invalid_state_rows: List[Tuple[int, Any]] = []
        for idx, r in enumerate(rows):
            raw_val = r.get(state_col)
            if raw_val is None or str(raw_val).strip() == "":
                empty_state_rows.append(idx + 1)
                continue
            code = _normalize_state(raw_val)
            if not code:
                invalid_state_rows.append((idx + 1, raw_val))
        if empty_state_rows or invalid_state_rows:
            bits: List[str] = []
            if empty_state_rows:
                sample = ", ".join(str(x) for x in empty_state_rows[:10])
                more = f" (+{len(empty_state_rows) - 10} more)" if len(empty_state_rows) > 10 else ""
                bits.append(f"{len(empty_state_rows)} row(s) have an EMPTY state — rows: {sample}{more}")
            if invalid_state_rows:
                sample = ", ".join(f"row {n}=\"{v}\"" for n, v in invalid_state_rows[:5])
                more = f" (+{len(invalid_state_rows) - 5} more)" if len(invalid_state_rows) > 5 else ""
                bits.append(f"{len(invalid_state_rows)} row(s) have an UNRECOGNISED state value — {sample}{more}")
            await _finalise_and_persist(
                db, job_id, "failed",
                "ProxyJet on-demand mode: every row must have a valid US state. "
                + " · ".join(bits)
                + ". Fix the data file (state column: '" + str(state_col) + "') and resubmit."
            )
            return
        # Re-index by state so the per-state row picker works.
        state_index = {}
        for idx, r in enumerate(rows):
            code = _normalize_state(r.get(state_col))
            if code:
                state_index.setdefault(code, []).append(idx)
        RUT_JOBS[job_id]["state_match_column"] = state_col
        RUT_JOBS[job_id]["proxyjet_on_demand"] = True
        try:
            push_live_step(
                job_id, 0, "preflight", "ok",
                f"ProxyJet ROW-FIRST mode: {len(rows)} rows across {len(state_index)} states verified. "
                f"IPs will be fetched on-demand (≤{int(proxyjet_unique_retry_cap)} retries for uniqueness)."
            )
        except Exception:
            pass

    # State-match round-robin pointer per state code
    state_rr: Dict[str, int] = {code: 0 for code in state_index}

    # State shared across tasks
    used_proxy_set: set = set()
    used_ua_set: set = set()  # distinct UA strings actually picked for visits
    consumed_row_indices: set = set()   # rows OK-submitted — NOT reused, removed from pending_leads
    invalid_row_indices: set = set()    # rows that triggered a validation error — ALSO removed from pending_leads
    state = {"proxy_idx": 0, "ua_idx": 0, "row_idx": 0, "start_time": time.time()}
    report: List[Dict[str, Any]] = []
    report_lock = asyncio.Lock()

    # ── 2026-05: Pure JSON Mode — disable all AI features ────────────
    # When the user has explicitly turned ON Pure JSON Mode, we force
    # self_heal OFF here so the engine NEVER falls back to Gemini
    # vision rescue when a recorded step fails. The visit either
    # succeeds purely on the JSON or it fails — predictable lifecycle.
    if pure_json_mode:
        if self_heal:
            try:
                push_live_step(
                    job_id, 0, "preflight", "info",
                    "🎯 Pure JSON Mode ON — self-heal forced OFF (no Gemini rescue).",
                )
            except Exception:
                pass
        self_heal = False

    # ── AI answer-learning ─────────────────────────────────────────
    # Load historical (q, a) → conversion stats for this offer host once,
    # at job start, and build a Thompson-sampling picker. Each visit
    # passes this picker into _multi_step_fill so survey clicks bias toward
    # historically high-converting answers. After thank_you detection we
    # call record_outcomes to update stats — so the picker improves with
    # every job run.
    #
    # When pure_json_mode is ON we SKIP building the picker entirely so
    # survey/random-pick steps stay purely random and no outcomes are
    # recorded back into the learning collection.
    ai_picker = None
    if pure_json_mode:
        try:
            push_live_step(
                job_id, 0, "ai_learning", "info",
                "🎯 Pure JSON Mode ON — AI answer-learning disabled (random picks only, no outcome recording).",
            )
        except Exception:
            pass
    else:
        try:
            from rut_answer_learning import load_stats, make_picker  # noqa: WPS433
            if db is not None:
                ai_stats = await load_stats(db, target_url)
                ai_picker = make_picker(ai_stats, explore_prob=0.15)
                try:
                    from rut_answer_learning import summarize_stats
                    summary = summarize_stats(ai_stats, min_clicks=3)
                    if summary:
                        push_live_step(
                            job_id, 0, "ai_learning", "info",
                            f"Loaded {len(summary)} learned questions for "
                            f"{target_url[:60]} — biasing answers toward best.",
                        )
                except Exception:  # noqa: BLE001
                    pass
        except Exception as e:  # noqa: BLE001
            logger.debug(f"AI learning init failed: {e}")
            ai_picker = None

    # ── Cancellation / stop support ─────────────────────────────────
    # Any code path (worker loop, stop endpoint) can set this flag;
    # new visits will short-circuit; in-flight visits finish their current
    # step and exit. Partial results are still zipped.
    cancel_event = asyncio.Event()
    RUT_JOBS[job_id]["_cancel_event"] = cancel_event

    # ── Graceful drain on target-reached (separate from hard-cancel) ─
    # When the job hits its `target_clicks` / `target_conversions` goal,
    # we set `target_drain_event` instead of `cancel_event`. The
    # difference matters:
    #   • cancel_event  = HARD STOP (user pressed Stop). In-flight visits
    #                     can short-circuit; abort ASAP to free the user.
    #   • target_drain_event = SOFT DRAIN (target met). DON'T start new
    #                     visits, but let every in-flight visit RUN TO
    #                     COMPLETION so the proxies/UAs/leads they
    #                     already picked up don't get wasted.
    # User ask (Roman Urdu): *"jab attempt pore ho jayein, naye attempt
    # na hun. Bas jo file chal rahi hai us ke pore hone ka wait karein
    # ta-ke wo file zaya na ho."*
    target_drain_event = asyncio.Event()
    RUT_JOBS[job_id]["_target_drain_event"] = target_drain_event

    # ── 2026-05 — Per-visit manual cancel registry ──────────────────
    # User ask: agar koi single visit stuck/error pe ruk gaya ho (e.g.
    # "User ineligible" page baar baar load ho rahi ho), to UI se us
    # tile par "kill" button daba ke us ek visit ko abort kar sakein
    # bina pura job rokne ke — taki concurrency slot free ho aur next
    # visit start ho jaye.
    #
    # Mechanism: every spawn of a `worker()` task registers itself
    # here under its 1-based visit_index. The new endpoint
    # POST /api/real-user-traffic/jobs/{job_id}/visits/{vid}/cancel
    # looks up the task and calls .cancel() on it. asyncio cleanly
    # raises CancelledError at the next await inside the visit,
    # Playwright contexts get closed via existing try/finally /
    # `async with` blocks, and the slot in `in_flight` is freed so
    # the dispatcher's spawn loop replenishes it.
    visit_tasks: Dict[str, asyncio.Task] = {}
    RUT_JOBS[job_id]["_visit_tasks"] = visit_tasks

    def _register_visit_task(i_zero_based: int, t: asyncio.Task) -> None:
        vid = str(i_zero_based + 1)
        visit_tasks[vid] = t
        def _cleanup(_task: asyncio.Task) -> None:
            visit_tasks.pop(vid, None)
        t.add_done_callback(_cleanup)

    # ── Per-use immediate deletion (real-time pruning) ──────────────
    # User asked: "ek line use hoe wo sath he delete ho jay" — so as soon
    # as a proxy / UA / row gets consumed in a visit we $pull it from the
    # saved upload batch (or rewrite the on-disk XLSX). Fire-and-forget
    # tasks so the visit isn't blocked by Mongo round-trips. We track every
    # task in `_live_pending_tasks` so the job can await all of them
    # before _finalise_and_persist — without this guard, the LAST visit's
    # $pull was reliably lost when the orchestrator finished too quickly
    # (testing agent caught this: consumed_count = N-1 instead of N).
    _live_proxy_pulled: set = set()  # avoid duplicate $pulls
    _live_ua_pulled: set = set()
    _live_pending_tasks: List[asyncio.Task] = []
    _data_file_lock = asyncio.Lock()  # serialise XLSX rewrites
    user_db_truncated = (engine_user_id or "").replace("-", "_")[:20]

    def _spawn_live(coro) -> None:
        """Schedule a live-remove coroutine and remember it so the job can
        await completion at the end. Replaces bare `asyncio.create_task`."""
        try:
            t = asyncio.create_task(coro)
            _live_pending_tasks.append(t)
        except Exception:
            pass

    async def _live_remove_proxy(raw: str):
        if not (engine_user_id and upload_proxy_id and db is not None and raw):
            return
        if raw in _live_proxy_pulled:
            return
        _live_proxy_pulled.add(raw)
        try:
            client = db.client
            user_db = client[f"krexion_user_{user_db_truncated}"]
            res = await user_db["uploaded_resources"].update_one(
                {"id": upload_proxy_id, "user_id": engine_user_id, "type": "proxies"},
                {
                    "$pull": {"items": raw},
                    "$set": {"updated_at": datetime.now(timezone.utc).isoformat()},
                    "$inc": {"consumed_count": 1, "item_count": -1},
                },
            )
            # If the batch is now empty, mark it depleted but PRESERVE
            # the DB entry so the user keeps their upload history.
            if res.modified_count:
                doc = await user_db["uploaded_resources"].find_one(
                    {"id": upload_proxy_id, "user_id": engine_user_id},
                    {"_id": 0, "items": 1, "depleted": 1, "gsheet_url": 1},
                )
                # ── Live Google Sheet row delete (when SA configured) ──
                # If the upload is gsheet-backed, mirror the consumption
                # to the source sheet so the user sees the row vanish in
                # real time. Best-effort: any failure logs a warning but
                # NEVER blocks the consume path (Mongo-side already done).
                gsheet_url = (doc or {}).get("gsheet_url") or ""
                if gsheet_url:
                    try:
                        import gsheet_writer  # lazy
                        loop = asyncio.get_running_loop()
                        deleted = await loop.run_in_executor(
                            None,
                            gsheet_writer.delete_rows_by_first_column,
                            gsheet_url, [raw],
                        )
                        if deleted:
                            logger.info(f"gsheet live delete proxy: removed {deleted} row(s) from sheet")
                        else:
                            logger.debug(f"gsheet live delete proxy: no row matched / SA write disabled")
                    except Exception as e:
                        logger.warning(f"gsheet live delete proxy failed: {type(e).__name__}: {e}")
                if doc and isinstance(doc.get("items"), list) and len(doc["items"]) == 0 and not doc.get("depleted"):
                    await user_db["uploaded_resources"].update_one(
                        {"id": upload_proxy_id, "user_id": engine_user_id},
                        {"$set": {
                            "depleted": True,
                            "depleted_at": datetime.now(timezone.utc).isoformat(),
                            "item_count": 0,
                        }},
                    )
        except Exception as e:
            logger.warning(f"_live_remove_proxy update_one failed: {type(e).__name__}: {e}")

    # Normalise UA-batch input: single id OR list OR map. Engine stores a
    # combined list of batch ids + a ua→batch map so live-remove knows
    # which batch to $pull each consumed UA from (each device's UAs can
    # live in a different batch — iPhone, Android, iPad etc).
    ua_batch_ids: List[str] = []
    if upload_ua_ids:
        ua_batch_ids = [b for b in upload_ua_ids if b]
    elif upload_ua_id:
        ua_batch_ids = [upload_ua_id]
    ua_batch_map: Dict[str, str] = dict(ua_to_batch_map or {})

    async def _live_remove_ua(ua: str):
        if not (engine_user_id and db is not None and ua and ua_batch_ids):
            return
        if ua in _live_ua_pulled:
            return
        _live_ua_pulled.add(ua)
        # Figure out which batch this UA belongs to. If we have a precise
        # map (built at load time), use it; otherwise $pull from every
        # selected batch — $pull is a no-op on batches that don't contain
        # the UA, so this is safe + correct.
        target_batches: List[str] = []
        mapped = ua_batch_map.get(ua)
        if mapped and mapped in ua_batch_ids:
            target_batches = [mapped]
        else:
            target_batches = list(ua_batch_ids)
        try:
            client = db.client
            user_db = client[f"krexion_user_{user_db_truncated}"]
            for batch_id in target_batches:
                res = await user_db["uploaded_resources"].update_one(
                    {"id": batch_id, "user_id": engine_user_id, "type": "user_agents"},
                    {
                        "$pull": {"items": ua},
                        "$set": {"updated_at": datetime.now(timezone.utc).isoformat()},
                        "$inc": {"consumed_count": 1, "item_count": -1},
                    },
                )
                if res.modified_count:
                    doc = await user_db["uploaded_resources"].find_one(
                        {"id": batch_id, "user_id": engine_user_id},
                        {"_id": 0, "items": 1, "depleted": 1, "gsheet_url": 1},
                    )
                    # ── Live Google Sheet row delete (when SA configured) ──
                    gsheet_url = (doc or {}).get("gsheet_url") or ""
                    if gsheet_url:
                        try:
                            import gsheet_writer  # lazy
                            loop = asyncio.get_running_loop()
                            deleted = await loop.run_in_executor(
                                None,
                                gsheet_writer.delete_rows_by_first_column,
                                gsheet_url, [ua],
                            )
                            if deleted:
                                logger.info(f"gsheet live delete UA: removed {deleted} row(s) from sheet")
                            else:
                                logger.debug(f"gsheet live delete UA: no row matched / SA write disabled")
                        except Exception as e:
                            logger.warning(f"gsheet live delete UA failed: {type(e).__name__}: {e}")
                    if doc and isinstance(doc.get("items"), list) and len(doc["items"]) == 0 and not doc.get("depleted"):
                        await user_db["uploaded_resources"].update_one(
                            {"id": batch_id, "user_id": engine_user_id},
                            {"$set": {
                                "depleted": True,
                                "depleted_at": datetime.now(timezone.utc).isoformat(),
                                "item_count": 0,
                            }},
                        )
                    break  # only $pull once — we found the right batch
        except Exception as e:
            logger.debug(f"_live_remove_ua failed: {e}")

    async def _live_remove_data_row(row_idx: int):
        """Rewrite the saved data-file XLSX with the consumed/invalid row
        removed. Lock-serialised so concurrent writes don't corrupt the
        file. The on-disk path is read fresh from the upload doc each
        time so a previous flush is always reflected.

        For Google-Sheet-backed uploads (no on-disk file) this also makes
        a live API call to delete the row from the source sheet using the
        row's email column as the matching key. Falls back silently when
        SA credentials are not configured.
        """
        if not (engine_user_id and upload_data_file_id and db is not None):
            return
        async with _data_file_lock:
            try:
                client = db.client
                user_db = client[f"krexion_user_{user_db_truncated}"]
                doc = await user_db["uploaded_resources"].find_one(
                    {"id": upload_data_file_id, "user_id": engine_user_id, "type": "data_file"},
                    {"_id": 0, "file_path": 1, "items": 1, "gsheet_url": 1, "consumed_keys": 1},
                )
                if not doc:
                    return

                # ── Path A: live Google Sheet (no on-disk XLSX) ───────────
                gsheet_url = (doc.get("gsheet_url") or "").strip()
                if gsheet_url:
                    if not (rows and 0 <= row_idx < len(rows)):
                        return
                    target_row = rows[row_idx]
                    target_email = ""
                    for k in ("email", "email_address", "emailaddress", "e_mail", "mail"):
                        v = (target_row or {}).get(k)
                        if v and str(v).strip():
                            target_email = str(v).strip()
                            break
                    deleted_from_sheet = False
                    if target_email:
                        try:
                            import gsheet_writer  # lazy
                            loop = asyncio.get_running_loop()
                            deleted_from_sheet = await loop.run_in_executor(
                                None,
                                gsheet_writer.delete_row_by_email,
                                gsheet_url, target_email,
                            )
                            logger.info(
                                f"gsheet live delete row #{row_idx + 1} email={target_email}: {'OK' if deleted_from_sheet else 'no match / write disabled'}"
                            )
                        except Exception as e:
                            logger.warning(f"gsheet live delete failed (row {row_idx}, {target_email}): {e}")
                    # Always track as consumed (audit trail, also keeps
                    # depleted-detection working when SA write fails)
                    key_for_audit = (target_email or f"row_{row_idx}").lower()
                    await user_db["uploaded_resources"].update_one(
                        {"id": upload_data_file_id, "user_id": engine_user_id},
                        {
                            "$addToSet": {"consumed_keys": key_for_audit},
                            "$inc": {"consumed_count": 1},
                            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()},
                        },
                    )

                    # Low-stock email alert: fire-and-forget, never blocks
                    # the consume path. The notifications module is
                    # idempotent (won't re-spam) and gracefully degrades
                    # when no email provider is configured.
                    try:
                        user_doc = await client["krexion"]["users"].find_one(
                            {"id": engine_user_id},
                            {"_id": 0, "email": 1, "notification_email": 1, "low_stock_alerts_enabled": 1},
                        )
                        if user_doc:
                            from notifications import maybe_send_low_stock_alert  # lazy
                            asyncio.create_task(
                                maybe_send_low_stock_alert(
                                    user_db=user_db,
                                    user_id=engine_user_id,
                                    upload_doc={"id": upload_data_file_id, "gsheet_url": gsheet_url},
                                    notification_email=user_doc.get("notification_email"),
                                    primary_email=user_doc.get("email"),
                                    alerts_enabled=bool(user_doc.get("low_stock_alerts_enabled", True)),
                                )
                            )
                    except Exception as e:
                        logger.debug(f"low-stock alert dispatch failed: {e}")
                    return

                # ── Path B: on-disk XLSX (legacy / Excel uploads) ─────────
                fp = doc.get("file_path") or ""
                if not fp or not Path(fp).exists():
                    return
                # Load, drop the row, save back. We use openpyxl directly
                # to keep things fast (no pandas roundtrip for 1 row).
                import openpyxl
                wb = openpyxl.load_workbook(fp)
                ws = wb.active
                # row_idx is 0-based against the original-data rows; the
                # XLSX has a header row at row 1, so the actual sheet row
                # is row_idx + 2. After previous deletions the sheet has
                # fewer rows than the original — we therefore work off
                # row VALUES, not indices: scan all data rows and find
                # the one whose original_row_index column matches.
                # Simpler approach: maintain a hidden "_orig_idx" column
                # added on first write so subsequent deletions work
                # against a stable identifier.
                header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
                if "_orig_idx" not in header:
                    # Add the column once, populate with current sheet
                    # row positions (they correspond 1:1 to the source
                    # data file order on first write).
                    col_idx = len(header) + 1
                    ws.cell(row=1, column=col_idx, value="_orig_idx")
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=col_idx, value=r - 2)
                    header.append("_orig_idx")
                orig_col = header.index("_orig_idx") + 1
                target_sheet_row = None
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=orig_col).value
                    try:
                        if int(val) == int(row_idx):
                            target_sheet_row = r
                            break
                    except (TypeError, ValueError):
                        continue
                if target_sheet_row:
                    ws.delete_rows(target_sheet_row, 1)
                wb.save(fp)
                wb.close()
                # Update count + bump consumed_count for analytics
                remaining = max(0, (ws.max_row or 1) - 1)
                await user_db["uploaded_resources"].update_one(
                    {"id": upload_data_file_id, "user_id": engine_user_id},
                    {
                        "$set": {
                            "row_count": remaining,
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                        },
                        "$inc": {"consumed_count": 1},
                    },
                )
                # Auto-delete the on-disk file if completely consumed,
                # but PRESERVE the DB entry as a depleted record so the
                # user can still see "I uploaded X rows, all consumed".
                if remaining == 0:
                    try:
                        Path(fp).unlink(missing_ok=True)
                    except Exception:
                        pass
                    await user_db["uploaded_resources"].update_one(
                        {"id": upload_data_file_id, "user_id": engine_user_id},
                        {"$set": {
                            "depleted": True,
                            "depleted_at": datetime.now(timezone.utc).isoformat(),
                            "file_path": None,
                            "item_count": 0,
                        }},
                    )
            except Exception as e:
                logger.debug(f"_live_remove_data_row failed: {e}")

    def pick_next_proxy() -> Optional[Dict[str, Any]]:
        """Round-robin pick a proxy, respecting no_repeated_proxy."""
        if no_repeated_proxy:
            for _ in range(len(parsed_proxies)):
                idx = state["proxy_idx"] % len(parsed_proxies)
                state["proxy_idx"] += 1
                raw = parsed_proxies[idx]["raw"]
                if raw not in used_proxy_set:
                    used_proxy_set.add(raw)
                    return parsed_proxies[idx]
            return None
        idx = state["proxy_idx"] % len(parsed_proxies)
        state["proxy_idx"] += 1
        return parsed_proxies[idx]

    def pick_next_ua() -> str:
        idx = state["ua_idx"] % len(uas_ok)
        state["ua_idx"] += 1
        return uas_ok[idx]

    def pick_next_row() -> Optional[Tuple[int, Dict[str, Any]]]:
        """Return (row_index, row_data) — skips rows already consumed (OK-submitted)
        AND rows flagged as invalid_data. If ALL rows are exhausted, returns None
        (caller should stop retrying)."""
        if not rows:
            return None
        total = len(rows)
        # Find a fresh row (not consumed, not invalid)
        for _ in range(total):
            idx = state["row_idx"] % total
            state["row_idx"] += 1
            if idx in consumed_row_indices or idx in invalid_row_indices:
                continue
            return (idx, rows[idx])
        return None  # all rows either used or invalid — nothing left

    def pick_next_row_for_state(state_code: str) -> Optional[Tuple[int, Dict[str, Any]]]:
        """State-matched row picker — returns a fresh row whose state == state_code.
        Round-robin within the candidate set for that state. Skips consumed/invalid.
        Returns None when no eligible row exists for this state."""
        if not state_code or state_code not in state_index:
            return None
        candidates = state_index[state_code]
        if not candidates:
            return None
        total_c = len(candidates)
        start_ptr = state_rr.get(state_code, 0)
        for _ in range(total_c):
            ptr = start_ptr % total_c
            start_ptr += 1
            idx = candidates[ptr]
            if idx in consumed_row_indices or idx in invalid_row_indices:
                continue
            state_rr[state_code] = start_ptr
            return (idx, rows[idx])
        # advance pointer even on failure so next call tries past it
        state_rr[state_code] = start_ptr
        return None

    # ── Resilient shared-browser holder ─────────────────────────────
    # Chromium occasionally crashes mid-job under heavy concurrency or
    # when a misbehaving proxy forces it to tear down. When that happens
    # every subsequent `browser.new_context(...)` throws
    # `TargetClosedError: Target page, context or browser has been closed`
    # and the whole job "fails" with 0 conversions. This holder wraps the
    # shared browser so workers can lazily relaunch it behind an async
    # lock if it ever drops offline.
    _browser_holder: Dict[str, Any] = {"b": None, "pw": None}
    _browser_lock = asyncio.Lock()

    async def _get_live_browser() -> Browser:
        """Return a live Playwright Browser — relaunches on the fly if the
        shared instance crashed / got disconnected. Safe to call from many
        concurrent workers (serialised via _browser_lock)."""
        b = _browser_holder.get("b")
        if b is not None:
            try:
                if b.is_connected():
                    return b
            except Exception:
                pass
        # Browser missing or disconnected — relaunch under lock.
        async with _browser_lock:
            b = _browser_holder.get("b")
            if b is not None:
                try:
                    if b.is_connected():
                        return b
                except Exception:
                    pass
            pw = _browser_holder.get("pw")
            if pw is None:
                # No Playwright runtime yet — create one.
                pw_cm_local = async_playwright()
                pw = await pw_cm_local.__aenter__()
                _browser_holder["pw"] = pw
                _browser_holder["pw_cm"] = pw_cm_local
            logger.warning(
                f"RUT job {job_id}: shared Chromium unavailable — relaunching…"
            )
            try:
                push_live_step(job_id, 0, "engine", "info",
                               "Chromium crashed — relaunching…")
            except Exception:
                pass
            new_b = await _launch_anti_detect_browser(pw)
            _browser_holder["b"] = new_b
            try:
                push_live_step(job_id, 0, "engine", "ok",
                               "Chromium relaunched — resuming visits")
            except Exception:
                pass
            return new_b

    async def process_one(i: int, shared_browser: Browser):
        # Short-circuit if user pressed Stop — don't waste a proxy/UA on it
        if cancel_event.is_set():
            return

        entry = {
            "visit_index": i + 1,
            "status": "pending",
            "proxy": "",
            "exit_ip": "",
            "country": "",
            "city": "",
            "timezone": "",
            "locale": "",
            "os": "",
            "ua": "",
            "viewport": "",
            "device_name": "",
            "http_status": "",
            "final_url": "",
            "landing_url": "",
            "conversion_page_reached": False,
            "trusted_form": "",
            "lead_id": "",
            "error": "",
            "screenshot": "",
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }

        # ── 2026-01 ROW-FIRST on-demand sequence ──────────────────────
        # When ProxyJet on-demand mode is on we INVERT the order:
        #   1) pick a fresh row,
        #   2) read its state,
        #   3) generate one state-matched ProxyJet IP,
        #   4) probe geo,
        #   5) if exit-IP is a duplicate of this user's prior clicks →
        #      throw it away and retry (up to `proxyjet_unique_retry_cap`),
        #   6) use the first unique IP for the visit.
        # Everything below this block (the legacy proxy/UA/geo pipeline)
        # is bypassed via the `_pj_on_demand_done` short-circuit so the
        # existing flow stays bit-identical for non-ProxyJet jobs.
        on_demand_row_pick: Optional[Tuple[int, Dict[str, Any]]] = None
        on_demand_proxy: Optional[Dict[str, Any]] = None
        on_demand_geo: Optional[Dict[str, Any]] = None
        if proxyjet_on_demand:
            # Step 1: pick row first — try sequential, then state-rotate
            # so we don't deadlock on a single state running out of leads.
            on_demand_row_pick = pick_next_row()
            if not on_demand_row_pick:
                entry["status"] = "failed"
                entry["error"] = "No remaining unconsumed rows in data file"
                push_live_step(job_id, i + 1, "row", "failed", "No rows left")
                await _record(job_id, entry, report, report_lock, db)
                return
            _row_idx, _row = on_demand_row_pick
            row_state_code = _normalize_state(_row.get(state_col)) if state_col else ""
            if not row_state_code:
                # Should be impossible — pre-flight rejected such rows.
                entry["status"] = "failed"
                entry["error"] = f"Row {_row_idx + 1} has an invalid state value (post-preflight)."
                push_live_step(job_id, i + 1, "row", "failed", entry["error"])
                await _record(job_id, entry, report, report_lock, db)
                return
            entry["row_index"] = _row_idx + 1
            entry["lead_state"] = row_state_code
            push_live_step(
                job_id, i + 1, "row", "ok",
                f"Row {_row_idx + 1} picked first · state={row_state_code} · fetching unique IP…",
            )

            # Step 2-5: loop until we get a non-duplicate exit IP.
            try:
                from proxyjet_module import generate_unique_proxies as _pj_gen  # noqa: WPS433
            except Exception as _imp_e:
                entry["status"] = "failed"
                entry["error"] = f"proxyjet_module unavailable: {_imp_e}"
                push_live_step(job_id, i + 1, "proxy", "failed", entry["error"])
                await _record(job_id, entry, report, report_lock, db)
                return

            attempt = 0
            cap = max(1, int(proxyjet_unique_retry_cap or 50))
            chosen_proxy: Optional[Dict[str, Any]] = None
            chosen_geo: Optional[Dict[str, Any]] = None
            last_reason = ""
            while attempt < cap:
                attempt += 1
                if cancel_event.is_set():
                    return
                try:
                    pj_lines = await _pj_gen(
                        db,
                        engine_user_id or "",
                        count=1,
                        country=proxyjet_country or "US",
                        state=row_state_code,
                        job_id=job_id,
                    )
                except Exception as e:  # noqa: BLE001
                    last_reason = f"ProxyJet generate failed: {type(e).__name__}: {e}"
                    push_live_step(job_id, i + 1, "proxy", "failed",
                                   f"Attempt {attempt}/{cap}: {last_reason}")
                    await asyncio.sleep(0.5)
                    continue
                if not pj_lines:
                    last_reason = "ProxyJet returned zero proxies"
                    push_live_step(job_id, i + 1, "proxy", "failed",
                                   f"Attempt {attempt}/{cap}: {last_reason}")
                    await asyncio.sleep(0.3)
                    continue
                parsed = _parse_proxy_line(pj_lines[0])
                if not parsed:
                    last_reason = "ProxyJet line failed to parse"
                    continue
                # Probe geo to learn the actual exit-IP — the only way to
                # check uniqueness against the duplicate_ip_set.
                _probe_ua = pick_next_ua()
                # Don't consume the UA pointer for probes — rewind
                state["ua_idx"] = max(0, state["ua_idx"] - 1)
                _geo = await _probe_proxy_geo(parsed, _probe_ua)
                if not _geo["ok"] or not _geo.get("exit_ip"):
                    last_reason = "exit-IP probe failed"
                    push_live_step(job_id, i + 1, "proxy", "failed",
                                   f"Attempt {attempt}/{cap}: probe failed, retrying")
                    continue
                exit_ip = _geo["exit_ip"]
                if skip_duplicate_ip and duplicate_ip_set and exit_ip in duplicate_ip_set:
                    last_reason = f"duplicate IP {exit_ip}"
                    # Don't log every single duplicate — too chatty. Log
                    # every 5th attempt so the user sees progress.
                    if attempt % 5 == 0 or attempt == 1:
                        push_live_step(job_id, i + 1, "proxy", "info",
                                       f"Attempt {attempt}/{cap}: duplicate {exit_ip}, retrying for {row_state_code}")
                    continue
                # Unique! Reserve it immediately so two parallel visits
                # don't both pick the same IP from concurrent probes.
                if duplicate_ip_set is not None:
                    duplicate_ip_set.add(exit_ip)
                # ── 2026-01 fix: CROSS-JOB persistence ────────────────
                # Without this, the in-memory `duplicate_ip_set` is lost
                # when the job ends, and the next job can pick the same
                # exit IP again from ProxyJet's pool (because clicks-DB
                # writes only happen on a successful click — a visit that
                # never reached a click would leak its IP back into the
                # pool for the next run). Persisting every picked exit
                # IP to `rut_burnt_ips` means EVERY future job loads it
                # into its `duplicate_ip_set` at startup → ProxyJet pool
                # IPs we've handed out are NEVER picked twice across runs.
                # Fire-and-forget so the request stays fast.
                try:
                    _spawn_live(_persist_burnt_ip(
                        db,
                        ip=exit_ip,
                        reason="proxyjet_picked",
                        user_id=engine_user_id or "",
                        offer_url=target_url or "",
                        state=row_state_code or "",
                        job_id=job_id or "",
                    ))
                except Exception:
                    # Never let persistence failure block the visit —
                    # in-memory set still protects within this job.
                    pass
                chosen_proxy = parsed
                chosen_geo = _geo
                break

            if not chosen_proxy or not chosen_geo:
                entry["status"] = "skipped_no_unique_ip"
                entry["error"] = (
                    f"No unique ProxyJet IP found for state {row_state_code} "
                    f"after {cap} attempts (last: {last_reason or 'unknown'})."
                )
                push_live_step(job_id, i + 1, "proxy", "failed", entry["error"])
                # Don't mark the row as consumed — give it back to the
                # picker so the next visit can try again with a fresh IP.
                try:
                    consumed_row_indices.discard(_row_idx)
                except Exception:
                    pass
                await _record(job_id, entry, report, report_lock, db)
                return

            on_demand_proxy = chosen_proxy
            on_demand_geo = chosen_geo
            push_live_step(
                job_id, i + 1, "proxy", "ok",
                f"Unique IP found on attempt {attempt}: {chosen_geo.get('exit_ip')} "
                f"· state {row_state_code}",
            )

        # ── Legacy proxy picker (skipped in on-demand mode) ──────────
        if proxyjet_on_demand:
            proxy = on_demand_proxy
        else:
            proxy = pick_next_proxy()
        if not proxy:
            entry["status"] = "failed"
            entry["error"] = "No more proxies available (no_repeated_proxy = on)"
            push_live_step(job_id, i + 1, "setup", "failed", "No proxies available")
            await _record(job_id, entry, report, report_lock, db)
            return
        # Mark this proxy raw as USED (attempted in a visit). Used by the
        # post-job upload-consume hook so only proxies actually consumed
        # in this run get removed from the saved batch — unused proxies
        # remain in the user's "Uploaded Things" library.
        try:
            raw_line = proxy.get("raw") or ""
            if raw_line:
                used_proxy_set.add(raw_line)
                # IMMEDIATE per-use deletion from the saved upload batch
                # (fire-and-forget — don't block the visit).
                _spawn_live(_live_remove_proxy(raw_line))
        except Exception:
            pass

        ua = pick_next_ua()
        # Track UA strings used so the upload-consume hook removes only
        # those that were actually attempted, not the entire UA batch.
        try:
            if ua:
                used_ua_set.add(ua)
                _spawn_live(_live_remove_ua(ua))
        except Exception:
            pass
        fp = _fingerprint_from_ua(ua)

        entry["proxy"] = proxy.get("server", "")
        entry["os"] = fp["os"]
        entry["ua"] = ua
        entry["viewport"] = f"{fp['viewport']['width']}x{fp['viewport']['height']}"
        entry["device_name"] = _device_name_from_ua(ua)
        entry["webgl_renderer"] = fp.get("webgl_renderer", "")
        entry["canvas_seed"] = fp.get("canvas_seed", 0)
        entry["hardware_concurrency"] = fp.get("hardware_concurrency", 0)
        entry["device_memory"] = fp.get("device_memory", 0)
        entry["device_scale_factor"] = fp.get("device_scale_factor", 0)
        push_live_step(job_id, i + 1, "setup", "info",
                       f"Proxy {entry['proxy']} · {entry['device_name']} · {entry['viewport']}")

        # Probe geo (also gives VPN flag) — REUSE the probe already
        # done during the on-demand row-first loop so we don't pay the
        # ip-api round-trip twice per visit (huge speed win).
        if proxyjet_on_demand and on_demand_geo is not None:
            geo = on_demand_geo
        else:
            geo = await _probe_proxy_geo(proxy, ua)
        entry["exit_ip"] = geo["exit_ip"] or ""
        entry["country"] = geo["country_name"]
        entry["city"] = geo["city"]
        entry["timezone"] = geo["timezone"]
        entry["locale"] = geo["locale"]
        push_live_step(job_id, i + 1, "geo", "ok" if geo["ok"] else "failed",
                       f"Exit {entry['exit_ip'] or '?'} · {entry['country'] or '?'}, {entry['city'] or '?'}")

        if not geo["ok"]:
            entry["status"] = "failed"
            entry["error"] = "Proxy unreachable (ip-api probe failed)"
            return await _record(job_id, entry, report, report_lock, db)

        # Pre-filter: country
        if allowed_countries_lc and geo["country_name"].lower() not in allowed_countries_lc:
            entry["status"] = "skipped_country"
            entry["error"] = f"{geo['country_name']} not in allowed list"
            push_live_step(job_id, i + 1, "filter", "skipped", f"Country not allowed: {geo['country_name']}")
            return await _record(job_id, entry, report, report_lock, db)

        # Pre-filter: VPN
        if skip_vpn and geo["is_vpn"]:
            entry["status"] = "skipped_vpn"
            entry["error"] = "Exit IP is flagged as VPN/hosting"
            push_live_step(job_id, i + 1, "filter", "skipped", "Exit IP flagged as VPN/hosting")
            return await _record(job_id, entry, report, report_lock, db)

        # Pre-filter: duplicate IP — already enforced inside the
        # on-demand loop above, so skip the redundant check.
        if not proxyjet_on_demand and skip_duplicate_ip and duplicate_ip_set and geo["exit_ip"] and geo["exit_ip"] in duplicate_ip_set:
            entry["status"] = "skipped_duplicate_ip"
            entry["error"] = "Exit IP already clicked this link before"
            push_live_step(job_id, i + 1, "filter", "skipped", f"Duplicate IP {geo['exit_ip']}")
            return await _record(job_id, entry, report, report_lock, db)

        # ── 2026-01: Target-URL reachability pre-check ──────────────
        # The geo probe (ipwho.is / ip-api.com) only confirms the
        # proxy can reach the internet, NOT that it can reach the
        # specific target URL. Residential pools regularly serve
        # exit nodes that pass geo but get blackholed / SSL-blocked /
        # 403'd by the actual offer's CDN. Catching those here saves
        # a full Chromium spawn + UA/lead consumption per dead proxy.
        #
        # 2026-01 (revised): If `target_url` is one of the user's own
        # tracker domains (api.krexion.com etc.), checking reachability
        # against the tracker itself is misleading — the tracker is
        # ALWAYS reachable server-side via the bypass path, but the
        # tracker just 302-redirects to the *real* offer URL
        # (e.g. trksy.org/aff_c?...), and THAT URL is where the
        # proxy actually has to land. So for tracker URLs we
        # pre-resolve the redirect server-side first (using the same
        # _resolve_tracker_via_localhost helper the bypass uses
        # after a failed goto) and then run the reachability check
        # on the resolved offer URL.
        #
        # 2026-01 (further revised): When server-side resolution
        # FAILS (e.g. Emergent preview pod can't resolve the user's
        # private tracker subdomain like api.krexion.com), instead
        # of silently skipping the precheck we now also probe the
        # target URL itself via the proxy. This catches "proxy
        # provider refuses this domain" 502 errors, "proxy can't
        # reach tracker" timeouts, and SSL/TLS handshake errors —
        # all of which would otherwise burn a UA + lead at the
        # browser-launch stage. On the user's actual VPS this branch
        # is rarely hit (local resolution works there), but it gives
        # the preview environment a useful filter too.
        try:
            from urllib.parse import urlparse as _t_up
            _t_host = (_t_up(target_url).hostname or "").lower()
        except Exception:
            _t_host = ""
        _is_tracker_target = _t_host in _bypass_hosts()
        _url_to_probe = target_url
        if _is_tracker_target and geo.get("exit_ip"):
            # Resolve the tracker server-side to find where it would
            # send a real click from this exit IP.
            try:
                _resolved_offer = await _resolve_tracker_via_localhost(
                    target_url, geo["exit_ip"], ua, timeout=8.0,
                )
            except Exception:
                _resolved_offer = None
            if _resolved_offer:
                _url_to_probe = _resolved_offer
                push_live_step(
                    job_id, i + 1, "filter", "info",
                    f"Tracker resolves to offer → probing {_url_to_probe[:80]}",
                )
            else:
                # Local resolution failed — fall back to probing the
                # tracker URL itself via the proxy.
                push_live_step(
                    job_id, i + 1, "filter", "info",
                    f"Tracker not resolvable server-side — probing target {_t_host} via proxy",
                )
        # Always run the reachability probe (target_url or resolved
        # offer URL), even for tracker targets. Failure → skip.
        _reach_ok, _reach_diag = await _probe_proxy_target_reachable(
            proxy, _url_to_probe, ua, timeout_s=12.0,
        )
        if not _reach_ok:
            entry["status"] = "skipped_dead_proxy"
            entry["error"] = (
                f"Proxy can't reach offer ({_reach_diag}) — skipped before browser launch"
            )
            push_live_step(
                job_id, i + 1, "filter", "skipped",
                f"Dead proxy: {_reach_diag}",
            )
            return await _record(job_id, entry, report, report_lock, db)
        push_live_step(
            job_id, i + 1, "filter", "ok",
            f"Offer reachable via proxy ({_reach_diag})",
        )

        # ── 2026-05 — PRE-BROWSER duplicate / VPN block probe ──────
        # The reachability check above only confirms transport-level
        # access (any HTTP status counts as "reachable"). It does NOT
        # inspect the actual page content, so an exit-IP that the
        # offer's server-side dedup list rejects with a "Duplicate IP"
        # or "VPN/proxy detected" page would still slip through and
        # we'd burn 10-30s of Chromium overhead per rejected IP.
        #
        # In ProxyJet on-demand mode we have an unlimited supply of
        # fresh exit-IPs, so it's MUCH cheaper to do one extra GET
        # via httpx (no browser, no JS, ~1-3s) and scan the response
        # body for the same block-page phrases that the post-load
        # detector would catch. If we see a match here we burn the
        # IP and raise _OfferBlockRetryNeeded — the worker() wrapper
        # retries with a fresh IP WITHOUT ever launching a browser
        # or producing a "Duplicate IP" thumbnail in Live Activity.
        #
        # Skipped in legacy proxy-list mode (`proxyjet_on_demand=False`)
        # because there's no IP-source loop to retry against — the
        # existing post-load detector still catches those cases.
        if proxyjet_on_demand:
            try:
                _pre_blk, _pre_reason, _pre_snip = await _probe_offer_duplicate_via_proxy(
                    proxy, _url_to_probe, ua, timeout_s=12.0,
                )
            except Exception:
                _pre_blk, _pre_reason, _pre_snip = (False, "", "")
            if _pre_blk:
                _burned_ip_pre = (geo.get("exit_ip") or entry.get("exit_ip") or "").strip()
                # In-job set update — affects next probe in this job
                if _burned_ip_pre and duplicate_ip_set is not None:
                    try:
                        duplicate_ip_set.add(_burned_ip_pre)
                    except Exception:
                        pass
                # Cross-job persistence — future jobs skip this IP
                if _burned_ip_pre:
                    _spawn_live(_persist_burnt_ip(
                        db,
                        ip=_burned_ip_pre,
                        reason=_pre_reason,
                        user_id=engine_user_id or "",
                        offer_url=target_url or "",
                        state=(entry.get("lead_state") or "").upper(),
                        job_id=job_id or "",
                    ))
                _pre_msg = (
                    f"Pre-browser {_pre_reason.replace('_', ' ')} block · burning "
                    f"{_burned_ip_pre or '?'} · persisted to rut_burnt_ips · "
                    f"retrying with fresh IP (no browser launch)"
                )
                push_live_step(
                    job_id, i + 1, "filter", "skipped", _pre_msg,
                )
                raise _OfferBlockRetryNeeded(
                    reason=_pre_reason, burnt_ip=_burned_ip_pre or "",
                )

        # Pick form-fill row — state-matched OR sequential
        # ── 2026-01 ROW-FIRST mode: row was already picked at the
        # top of process_one (we needed its state to fetch a matching
        # ProxyJet IP). Reuse that pick here.
        row_pick = None
        if proxyjet_on_demand and on_demand_row_pick is not None:
            row_pick = on_demand_row_pick
        elif form_fill_enabled:
            if state_match_enabled and state_col:
                # Match lead state to this proxy's exit-IP state.
                proxy_state_code = _normalize_state(geo.get("region")) or _normalize_state(geo.get("region_name"))
                entry["proxy_state"] = proxy_state_code or ""
                if proxy_state_code:
                    row_pick = pick_next_row_for_state(proxy_state_code)
                if not row_pick:
                    # No lead available for this proxy's state → skip this visit
                    entry["status"] = "skipped_state_mismatch"
                    entry["error"] = (
                        f"No unused lead for state {proxy_state_code or '?'}"
                        if proxy_state_code
                        else "Proxy state unknown (ip-api region missing)"
                    )
                    push_live_step(
                        job_id, i + 1, "filter", "skipped",
                        f"State mismatch: no lead for {proxy_state_code or '?'}",
                    )
                    return await _record(job_id, entry, report, report_lock, db)
            else:
                row_pick = pick_next_row()
        row_index, row = (row_pick if row_pick else (None, None))
        if row is not None:
            entry["row_index"] = (row_index or 0) + 1
            if state_col:
                entry["lead_state"] = _normalize_state(row.get(state_col)) or ""

        browser: Optional[Browser] = None
        context: Optional[BrowserContext] = None
        try:
            # Use the SHARED browser launched once at job start. Per-visit
            # isolation comes from a fresh BrowserContext with its own proxy,
            # cookies, storage, fingerprint, locale, timezone, viewport — which
            # is functionally identical to a fresh browser launch from any
            # detection script's perspective (canvas / WebGL / navigator are
            # all overridden per-context via init script). This drops RAM
            # usage 5-10x vs. per-visit Chromium launches and lets us safely
            # run 15+ concurrent visits without OOM.
            # NOTE: `_get_live_browser()` transparently relaunches Chromium
            # if it has crashed since the last visit — prevents the entire
            # job from failing with `TargetClosedError` the moment one
            # bad proxy or a Chromium bug kills the shared instance.
            browser = await _get_live_browser()
            try:
                # Auto-detect Referer from the user-agent — TikTok UAs get a
                # tiktok.com referer, FB in-app UAs get facebook.com, etc.
                # Plain browser UAs return "" and Referer header is omitted.
                _ua_referer = _get_referer_from_ua(ua)
                _ctx_headers = {"Accept-Language": geo["accept_language"]}
                if _ua_referer:
                    _ctx_headers["Referer"] = _ua_referer
                # 2026-01: Sec-CH-UA client hint headers matching UA's
                # Chrome version + OS. MaxMind / IPQS / Anura cross-check
                # these against the UA — a mismatch is a HARD bot signal.
                try:
                    _ctx_headers.update(_build_client_hint_headers(fp, ua))
                except Exception:
                    pass
                context = await browser.new_context(
                    proxy={
                        "server": proxy["server"],
                        **({"username": proxy["username"]} if proxy.get("username") else {}),
                        **({"password": proxy["password"]} if proxy.get("password") else {}),
                    },
                    user_agent=ua,
                    viewport=fp["viewport"],
                    device_scale_factor=fp["device_scale_factor"],
                    is_mobile=fp["is_mobile"],
                    has_touch=fp["has_touch"],
                    locale=geo["locale"],
                    timezone_id=geo["timezone"],
                    geolocation={"latitude": geo["lat"], "longitude": geo["lon"]},
                    permissions=["geolocation"],
                    extra_http_headers=_ctx_headers,
                )
            except Exception as _nce:
                # new_context can still race with a crash that happened
                # between is_connected() and the call. Give the holder one
                # last chance to relaunch, then retry the context.
                msg = str(_nce)
                if ("closed" in msg.lower()) or ("TargetClosed" in type(_nce).__name__):
                    browser = await _get_live_browser()
                    # Reuse same UA-derived Referer for the retry context.
                    _ua_referer_retry = _get_referer_from_ua(ua)
                    _ctx_headers_retry = {"Accept-Language": geo["accept_language"]}
                    if _ua_referer_retry:
                        _ctx_headers_retry["Referer"] = _ua_referer_retry
                    # 2026-01: Sec-CH-UA client hints on retry context too
                    try:
                        _ctx_headers_retry.update(_build_client_hint_headers(fp, ua))
                    except Exception:
                        pass
                    context = await browser.new_context(
                        proxy={
                            "server": proxy["server"],
                            **({"username": proxy["username"]} if proxy.get("username") else {}),
                            **({"password": proxy["password"]} if proxy.get("password") else {}),
                        },
                        user_agent=ua,
                        viewport=fp["viewport"],
                        device_scale_factor=fp["device_scale_factor"],
                        is_mobile=fp["is_mobile"],
                        has_touch=fp["has_touch"],
                        locale=geo["locale"],
                        timezone_id=geo["timezone"],
                        geolocation={"latitude": geo["lat"], "longitude": geo["lon"]},
                        permissions=["geolocation"],
                        extra_http_headers=_ctx_headers_retry,
                    )
                else:
                    raise
            await context.add_init_script(_build_stealth_script(fp, geo))

            # Defensive guard: block navigations to URLs that contain
            # unfilled tracker macros like `{{ccpa}}`, `{{sub}}`, etc.
            # Some affiliate offer pages render CCPA / opt-out anchors with
            # a `href="https://.../{{ccpa}}"` placeholder which the tracker
            # is supposed to replace at runtime. When the macro is missing
            # from the inbound tracker URL, the literal `{{...}}` leaks
            # through to the anchor and any aggressive "force-nav"
            # JS-evaluate step in the user's automation script ends up
            # clicking it, dead-ending the visit on a 404 path. Aborting
            # those requests at the network layer keeps the SPA on the
            # legitimate flow page so the form-fill steps can run.
            await context.route(
                "**/*",
                _make_macro_guard(job_id, i + 1),
            )

            if True:

                page = await context.new_page()

                # ── 2026-05 — TRACKER-SIDE duplicate-IP pre-flight check ─
                # Earlier attempts (ipwho.is via the browser) returned the
                # exit-IP from a DIFFERENT origin's connection pool — and
                # ProxyJet residential pools assign per-origin exits, so
                # the IP we measured did NOT match the IP the tracker
                # actually saw.
                #
                # Fix: probe a NEW endpoint on the SAME origin as the
                # tracker (`/api/_rut_ipcheck` on krexion.com / api.krexion.com)
                # before the tracker URL itself. Chromium pools connections
                # per origin, so both requests share the SAME upstream TCP
                # tunnel through ProxyJet → guaranteed SAME exit IP.
                #
                # 2026-05 — Probe BOTH origins (target host AND its parent
                # apex / sibling). Many setups use `api.krexion.com` →
                # 302 → `krexion.com` for the actual tracker check.
                # Probing only the first host misses the IP that the
                # tracker actually sees AND leaves the cross-origin
                # TCP tunnel cold (→ chrome-error://chromewebdata on the
                # second-leg CONNECT). Probing both warms up both
                # tunnels AND measures both IPs.
                if proxyjet_on_demand:
                    # Build the probe URLs from target_url's origin.
                    # target_url is e.g. https://krexion.com/api/t/abc or
                    # https://api.krexion.com/api/t/abc — keep scheme+host.
                    try:
                        from urllib.parse import urlparse as _urlparse
                        _u = _urlparse(target_url or "")
                        _scheme = _u.scheme or "https"
                        _host = (_u.netloc or "").lower().strip()
                    except Exception:
                        _scheme, _host = "https", ""

                    _probe_origins: List[str] = []
                    if _host:
                        _probe_origins.append(f"{_scheme}://{_host}")
                        # Add the apex / sibling host if the target is
                        # on a known multi-subdomain setup:
                        #   api.krexion.com → also probe krexion.com
                        #   krexion.com     → also probe api.krexion.com
                        #   www.example.com → also probe example.com
                        _parts = _host.split(".")
                        if len(_parts) >= 3 and _parts[0] in ("api", "www"):
                            _apex = ".".join(_parts[1:])
                            _probe_origins.append(f"{_scheme}://{_apex}")
                        elif len(_parts) == 2:
                            # Bare apex — also try `api.` subdomain
                            _probe_origins.append(f"{_scheme}://api.{_host}")

                    _tracker_ips_set: set = set()
                    _guard_ok = False
                    _data: Dict[str, Any] = {}
                    _BAD = {"unknown", "Unknown", "", "None", "none"}

                    for _probe_origin in _probe_origins:
                        _check_url = _probe_origin + "/api/_rut_ipcheck"
                        try:
                            _gr = await page.goto(
                                _check_url,
                                wait_until="domcontentloaded",
                                timeout=12000,
                            )
                            if _gr is None or not (200 <= (_gr.status or 0) < 400):
                                continue
                            try:
                                _body = await page.evaluate(
                                    "() => document.body ? document.body.innerText : ''"
                                )
                            except Exception:
                                _body = ""
                            try:
                                import json as _ipjson
                                _data_one = _ipjson.loads(_body or "{}")
                            except Exception:
                                _data_one = {}
                            if not _data_one.get("ok"):
                                continue
                            # Keep the FIRST successful response as
                            # `_data` (used to set the primary IP on the
                            # entry record). Both responses go into
                            # `_tracker_ips_set`.
                            if not _data:
                                _data = _data_one
                            for _k in ("primary", "ipv4"):
                                _v = (_data_one.get(_k) or "").strip()
                                if _v and ":" not in _v and _v not in _BAD:
                                    _tracker_ips_set.add(_v)
                            for _k in ("all", "proxy_ips"):
                                for _v in (_data_one.get(_k) or []):
                                    _v = (str(_v) or "").strip()
                                    if _v and ":" not in _v and _v not in _BAD:
                                        _tracker_ips_set.add(_v)
                        except Exception:
                            # Probe origin unreachable (chrome-error,
                            # timeout, etc.) — fall through to the next
                            # origin or to the unguarded visit path.
                            continue

                    _tracker_ips = list(_tracker_ips_set)
                    _guard_ok = bool(_tracker_ips)

                    if _guard_ok and _tracker_ips:
                        # Set entry["exit_ip"] to the primary tracker IP —
                        # this is what the report will display and what
                        # gets recorded in clicks DB on success.
                        _primary_ip = (
                            (_data.get("primary") or _data.get("ipv4") or _tracker_ips[0])
                            .strip()
                        )
                        if _primary_ip and _primary_ip not in _BAD:
                            entry["exit_ip"] = _primary_ip

                        # Check ALL tracker-visible IPs against duplicate_ip_set
                        _dup_hit = None
                        if skip_duplicate_ip and duplicate_ip_set is not None:
                            for _ip in _tracker_ips:
                                if _ip in duplicate_ip_set:
                                    _dup_hit = _ip
                                    break

                        if _dup_hit:
                            push_live_step(
                                job_id, i + 1, "filter", "skipped",
                                f"Tracker-visible duplicate IP {_dup_hit} "
                                f"(of {len(_tracker_ips)} tracker-visible IPs across "
                                f"{len(_probe_origins)} origin(s)) · "
                                f"closing profile + retrying with fresh ProxyJet IP "
                                f"(no offer load attempted)",
                            )
                            # Persist the duplicate to rut_burnt_ips so
                            # future probes skip it across jobs.
                            _spawn_live(_persist_burnt_ip(
                                db,
                                ip=_dup_hit,
                                reason="tracker_duplicate_ip",
                                user_id=engine_user_id or "",
                                offer_url=target_url or "",
                                state=(entry.get("lead_state") or "").upper(),
                                job_id=job_id or "",
                            ))
                            try:
                                await context.close()
                            except Exception:
                                pass
                            raise _OfferBlockRetryNeeded(
                                reason="tracker_duplicate_ip",
                                burnt_ip=_dup_hit,
                            )

                        # All tracker-visible IPs are unique. Reserve them
                        # in duplicate_ip_set so concurrent visits don't
                        # collide on the same exit.
                        if duplicate_ip_set is not None:
                            for _ip in _tracker_ips:
                                try:
                                    duplicate_ip_set.add(_ip)
                                except Exception:
                                    pass
                        push_live_step(
                            job_id, i + 1, "filter", "ok",
                            f"Tracker-side IP check passed · {_primary_ip or '?'} "
                            f"({len(_tracker_ips)} IPs across {len(_probe_origins)} "
                            f"origin(s)) unique · proceeding to tracker URL",
                        )
                    # else: probe endpoint unreachable on every origin —
                    # don't block the visit on a transient verification
                    # glitch; the existing post-load detector remains the
                    # safety net.

                push_live_step(job_id, i + 1, "browser", "info", f"Opening {target_url}")

                # Tunnel-error retry: ProxyJet sticky sessions sometimes
                # have a dead egress for a specific target host. Detect
                # ERR_TUNNEL_CONNECTION_FAILED / proxy-CONNECT errors and
                # transparently rotate to a fresh proxy up to 2 times so
                # the visit doesn't false-fail on a one-off bad tunnel.
                # 2026-01 — Tokens / patterns that indicate a proxy-side
                # transport problem (vs an offer-side problem). When we see
                # one of these AND we have proxies left in the pool, the
                # next "while" iteration transparently rotates to a fresh
                # proxy and re-tries the navigation — up to 2 times — so
                # the visit doesn't false-fail on a one-off bad tunnel /
                # slow-as-molasses residential exit.
                #
                # 2026-01 (revised): added timeout-style errors. Many
                # residential pools have exits that ACCEPT the TCP
                # tunnel cleanly (so the strict tunnel errors below
                # never fire) but then drop / throttle the actual
                # request, causing the goto to hit its 90s ceiling
                # without progressing. Rotating off these exits saves
                # the visit too.
                _TUNNEL_ERR_TOKENS = (
                    "ERR_TUNNEL_CONNECTION_FAILED",
                    "ERR_PROXY_CONNECTION_FAILED",
                    "ERR_HTTP_RESPONSE_CODE_FAILURE",
                    "ERR_CONNECTION_RESET",
                    "ERR_CONNECTION_CLOSED",
                    "ERR_CONNECTION_REFUSED",
                    "ERR_EMPTY_RESPONSE",
                    "ERR_SOCKET_NOT_CONNECTED",
                    "ERR_TIMED_OUT",
                    "Timeout 35000ms exceeded",
                    "Timeout 90000ms exceeded",
                    "net::ERR_NAME_NOT_RESOLVED",
                    "ERR_ADDRESS_UNREACHABLE",
                    "TLSV1_ALERT_INTERNAL_ERROR",
                    "WRONG_VERSION_NUMBER",
                    "ERR_SSL_PROTOCOL_ERROR",
                    # 2026-05 — chrome-error://chromewebdata happens when
                    # Chromium can't load a URL (typically after a
                    # cross-origin redirect through ProxyJet where the
                    # second-leg CONNECT to the new origin fails). Treat
                    # as a transient proxy issue, not a real visit
                    # failure — the dispatcher excludes these from the
                    # max_attempts budget so the run keeps going until
                    # real conversions land.
                    "chrome-error://chromewebdata",
                    "chrome-error://chromeweb",
                    "interrupted by another navigation",
                )
                MAX_TUNNEL_RETRIES = 3
                # ── 2026-05: same-proxy retry budget ──
                # When no fresh proxy is available (common when
                # no_repeated_proxy=True and all proxies are already
                # allocated to other visits), give the SAME proxy a few
                # more chances with a short backoff + a lighter
                # wait_until="commit" fallback. Many tunnel errors are
                # transient (TCP reset, brief network glitch, residential
                # exit ISP hiccup) and succeed on a second try seconds
                # later. This dramatically reduces the "after 1 attempts"
                # failure when proxy pool is exhausted.
                MAX_SAME_PROXY_RETRIES = 2
                same_proxy_retry = 0
                tunnel_attempt = 0
                resp = None
                goto_exc = None
                while True:
                    try:
                        # 2026-01 — lowered from 90s → 35s. Residential
                        # proxies that haven't started serving traffic
                        # within 35s are effectively dead; waiting the
                        # full 90s just wastes wall-clock + lead/UA
                        # budget per visit. The tunnel-retry loop above
                        # will rotate to a fresh proxy if this fails.
                        # 2026-05: on same-proxy retry, use "commit" so
                        # we just wait for the navigation to start (HTTP
                        # response received) instead of full DOM — many
                        # transient tunnel hiccups resolve mid-flight.
                        _wait_until = "commit" if same_proxy_retry > 0 else "domcontentloaded"
                        resp = await page.goto(target_url, timeout=35000, wait_until=_wait_until)
                        goto_exc = None
                        break
                    except Exception as _ge:
                        goto_exc = _ge
                        err_str = str(_ge)
                        is_tunnel = any(tok in err_str for tok in _TUNNEL_ERR_TOKENS)
                        if not is_tunnel or tunnel_attempt >= MAX_TUNNEL_RETRIES:
                            break
                        # Pick a fresh proxy and rebuild context+page
                        new_proxy = pick_next_proxy()
                        if not new_proxy:
                            # ── 2026-05 same-proxy retry fallback ──
                            # No spare proxy available — give the current
                            # proxy ANOTHER chance after a short backoff.
                            # This handles transient tunnel errors when
                            # the proxy pool is otherwise exhausted.
                            if same_proxy_retry < MAX_SAME_PROXY_RETRIES:
                                same_proxy_retry += 1
                                backoff_s = 1.5 * same_proxy_retry  # 1.5s, 3.0s
                                push_live_step(
                                    job_id, i + 1, "browser", "info",
                                    f"No fresh proxy left · retrying same proxy ({same_proxy_retry}/{MAX_SAME_PROXY_RETRIES}) in {backoff_s:.1f}s",
                                )
                                await asyncio.sleep(backoff_s)
                                continue
                            break
                        tunnel_attempt += 1
                        same_proxy_retry = 0  # fresh proxy → reset same-proxy counter
                        push_live_step(
                            job_id, i + 1, "browser", "info",
                            f"Tunnel failed · rotating proxy ({tunnel_attempt}/{MAX_TUNNEL_RETRIES}): {new_proxy.get('server','')}",
                        )
                        try:
                            raw_line = new_proxy.get("raw") or ""
                            if raw_line:
                                used_proxy_set.add(raw_line)
                                _spawn_live(_live_remove_proxy(raw_line))
                        except Exception:
                            pass
                        # close old context
                        try:
                            await context.close()
                        except Exception:
                            pass
                        try:
                            proxy = new_proxy
                            entry["proxy"] = proxy.get("server", "")
                            context = await browser.new_context(
                                proxy={
                                    "server": proxy["server"],
                                    **({"username": proxy["username"]} if proxy.get("username") else {}),
                                    **({"password": proxy["password"]} if proxy.get("password") else {}),
                                },
                                user_agent=ua,
                                viewport=fp["viewport"],
                                device_scale_factor=fp["device_scale_factor"],
                                is_mobile=fp["is_mobile"],
                                has_touch=fp["has_touch"],
                                locale=geo["locale"],
                                timezone_id=geo["timezone"],
                                geolocation={"latitude": geo["lat"], "longitude": geo["lon"]},
                                permissions=["geolocation"],
                                extra_http_headers=_ctx_headers,
                            )
                            await context.add_init_script(_build_stealth_script(fp, geo))
                            await context.route(
                                "**/*",
                                _make_macro_guard(job_id, i + 1),
                            )
                            page = await context.new_page()
                        except Exception as _rebuild_e:
                            goto_exc = _rebuild_e
                            break

                try:
                    if goto_exc is not None:
                        # ── 2026-01: Localhost-bypass fallback ─────────────
                        # If we just exhausted MAX_TUNNEL_RETRIES on a URL
                        # whose host is in RUT_LOCALHOST_BYPASS_HOSTS
                        # (e.g. our own krexion.com tracker), give it one
                        # last chance: pre-resolve the tracker server-side
                        # over 127.0.0.1 with the proxy's exit IP forged
                        # as X-Forwarded-For, then point the browser at
                        # the resulting offer URL through the proxy.
                        # The click is still recorded as the residential
                        # exit IP, the browser still navigates through
                        # the proxy for the offer page (which the proxy
                        # IS allowed to reach), and the rest of the visit
                        # — form-fill, conversion, screenshots — runs
                        # unmodified.
                        _err_text_bp = str(goto_exc)
                        _is_tunnel_bp = any(t in _err_text_bp for t in _TUNNEL_ERR_TOKENS)
                        _is_502_bp = (
                            "502" in _err_text_bp
                            or "bad gateway" in _err_text_bp.lower()
                            or "ERR_HTTP_RESPONSE_CODE_FAILURE" in _err_text_bp
                        )
                        # ── 2026-05 (revised): TRACKER-DOMAIN BYPASS ─────
                        # This bypass ONLY fires when the target URL is one
                        # of the user's OWN tracker domains (default
                        # `krexion.com, api.krexion.com, localhost,
                        # 127.0.0.1`). In that case the click registration
                        # is a SERVER→SERVER hit to the user's own backend
                        # (no external party sees the customer IP — it's
                        # internal infrastructure), and the proxy's exit
                        # IP is recorded as the click IP via the
                        # X-Forwarded-For header. The browser still
                        # navigates to the offer URL THROUGH the proxy.
                        #
                        # External offer URLs (e.g. anyunclaimedassets.com)
                        # NEVER take this path — they always go through
                        # the residential proxy or fail.
                        #
                        # Default = enabled, because residential proxies
                        # (ProxyJet, BrightData, etc.) routinely block
                        # custom tracker domains and without this bypass
                        # NO visit can ever complete. Set
                        # `RUT_STRICT_PROXY_ONLY=true` in backend/.env to
                        # disable even the tracker-domain bypass and force
                        # ALL hits — even your own tracker — through the
                        # proxy.
                        _strict_proxy = (
                            os.environ.get("RUT_STRICT_PROXY_ONLY", "false")
                            .strip().lower() in ("1", "true", "yes", "on")
                        )
                        if (
                            (_is_tunnel_bp or _is_502_bp)
                            and _url_host_matches_bypass(target_url)
                            and not _strict_proxy
                        ):
                            push_live_step(
                                job_id, i + 1, "bypass", "info",
                                "Proxy can't reach your own tracker — internal server→server hit "
                                "(click will still be recorded as the PROXY exit IP)…",
                            )
                            _exit_ip_for_bypass = await _get_exit_ip_via_proxy(proxy)
                            if _exit_ip_for_bypass:
                                _bypass_offer_url = await _resolve_tracker_via_localhost(
                                    target_url, _exit_ip_for_bypass, ua
                                )
                                if _bypass_offer_url:
                                    push_live_step(
                                        job_id, i + 1, "bypass", "ok",
                                        f"✓ Click registered as PROXY IP {_exit_ip_for_bypass} "
                                        f"(your machine's IP is NOT exposed — tracker is on your "
                                        f"own server). Browser → {_bypass_offer_url[:90]}",
                                    )
                                    try:
                                        # NOTE: We do NOT rebind the closure
                                        # variable `target_url` here. Doing
                                        # so would make Python treat
                                        # `target_url` as a LOCAL throughout
                                        # process_one (because any binding
                                        # in a function body switches the
                                        # whole-scope storage class), which
                                        # would raise UnboundLocalError on
                                        # every earlier read of target_url
                                        # (the "Opening …" log, the initial
                                        # page.goto, the bypass-host check)
                                        # before this branch ever runs.
                                        # Instead we tag the entry so
                                        # downstream auditing knows the
                                        # bypass kicked in, and we navigate
                                        # the browser to _bypass_offer_url
                                        # directly — the rest of the visit
                                        # operates on page.url() / page DOM
                                        # state, which already reflects the
                                        # offer URL after this navigation.
                                        entry["bypass_used"] = True
                                        entry["bypass_exit_ip"] = _exit_ip_for_bypass
                                        entry["bypass_offer_url"] = _bypass_offer_url
                                        resp = await page.goto(
                                            _bypass_offer_url,
                                            timeout=90000,
                                            wait_until="domcontentloaded",
                                        )
                                        goto_exc = None  # clear → success path
                                    except Exception as _bp_e:
                                        # Bypass-page itself failed too;
                                        # fall through to the original
                                        # failure-recording code.
                                        goto_exc = _bp_e
                                else:
                                    push_live_step(
                                        job_id, i + 1, "bypass", "failed",
                                        "Direct bypass returned no redirect — falling back.",
                                    )
                            else:
                                push_live_step(
                                    job_id, i + 1, "bypass", "failed",
                                    "Could not detect proxy exit IP — bypass skipped.",
                                )
                        if goto_exc is not None:
                            # ── 2026-05: Strict proxy mode notice ───────────
                            # When strict mode blocked the tracker-bypass,
                            # surface a clear reason so the user knows to
                            # either disable strict mode or use a proxy
                            # provider that allows their tracker domain.
                            if (
                                _strict_proxy
                                and (_is_tunnel_bp or _is_502_bp)
                                and _url_host_matches_bypass(target_url)
                            ):
                                push_live_step(
                                    job_id, i + 1, "bypass", "failed",
                                    "Strict proxy mode ON — even your own tracker domain "
                                    "is forced through the proxy. The proxy provider rejected "
                                    "the tunnel. Disable RUT_STRICT_PROXY_ONLY or use a proxy "
                                    "provider that allows your tracker domain.",
                                )
                            raise goto_exc
                    entry["http_status"] = str(resp.status) if resp else ""
                    # Detect chrome-error pages — happens when the residential
                    # proxy's egress tunnel breaks mid-navigation or DNS
                    # fails. Marking these as failures (instead of "ok")
                    # prevents false-positive success counts and triggers
                    # the upstream retry/reporting path correctly.
                    try:
                        cur_url = (page.url or "")
                    except Exception:
                        cur_url = ""
                    if cur_url.startswith("chrome-error://") or cur_url.startswith("chrome://network-error"):
                        entry["status"] = "failed"
                        entry["error"] = f"Browser navigation error (proxy tunnel broken): {cur_url}"
                        push_live_step(job_id, i + 1, "browser", "failed",
                                       f"Navigation error: {cur_url[:80]}")
                        await context.close()
                        return await _record(job_id, entry, report, report_lock, db)
                    # Grab a lightweight landing thumbnail so the Live Activity
                    # modal can prove the browser really loaded the page.
                    # ── 2026-05 (improved) ──
                    # User complaint: live activity ke thumbnails blank/teal
                    # aate the kyunki screenshot page fully render hone se
                    # PEHLE liya jaata tha. Ab proper sequence:
                    #   1. wait for "load" (DOM + main resources)
                    #   2. wait for "networkidle" (no in-flight requests)
                    #   3. wait for visible body content (height > 0)
                    #   4. small settle delay (250ms) for paint
                    #   5. take the screenshot
                    # Each step is best-effort with its own timeout —
                    # transient SPAs that never reach networkidle still
                    # produce a useful screenshot via the fallback paths.
                    try:
                        try:
                            await page.wait_for_load_state("load", timeout=8000)
                        except Exception:
                            pass
                        try:
                            await page.wait_for_load_state("networkidle", timeout=8000)
                        except Exception:
                            pass
                        # Ensure something has actually painted — wait until
                        # body has non-zero height (covers SPAs that finish
                        # network early but still render JS shortly after).
                        try:
                            await page.wait_for_function(
                                "() => document.body && document.body.scrollHeight > 100 && (document.body.innerText||'').trim().length > 0",
                                timeout=4000,
                            )
                        except Exception:
                            pass
                        # Final paint settle so first-contentful-paint frame
                        # isn't captured mid-animation.
                        await asyncio.sleep(0.25)
                        landing_shot = shots_dir / f"visit_{i+1:05d}_landing.png"
                        await page.screenshot(path=str(landing_shot), full_page=False, timeout=10000)
                        push_live_step(job_id, i + 1, "landing", "ok",
                                       f"📷 1/4 URL fully loaded (HTTP {entry['http_status'] or '?'})",
                                       screenshot=landing_shot.name)
                        entry["landing_screenshot"] = landing_shot.name
                    except Exception:
                        push_live_step(job_id, i + 1, "browser", "ok",
                                       f"Page loaded (HTTP {entry['http_status'] or '?'})")

                    # ── 2026-05: EARLY click logging ─────────────────────
                    # Insert this visit's click row into user_db.clicks RIGHT
                    # AWAY (with visit_status='pending') so the exit-IP is
                    # detectable as a duplicate by EVERY subsequent visit
                    # (this job AND any other job/tab/worker) within
                    # seconds — not minutes after the full automation
                    # finishes. The end-of-visit _log_click_for_link call
                    # in _record() will UPDATE this same row with final
                    # fields (visit_status, final_url, conversion). No
                    # double-counting, no double-insert.
                    try:
                        _j_for_log = RUT_JOBS.get(job_id, {})
                        await _log_click_for_link(entry, _j_for_log, db, early=True)
                    except Exception as _ele:
                        # Best-effort — duplicate-detection without this
                        # still works via in-memory duplicate_ip_set and
                        # the tracker's own click recording, so we just
                        # log and move on.
                        try:
                            logger.warning(f"RUT early click log failed: {_ele}")
                        except Exception:
                            pass
                except Exception as e:
                    entry["status"] = "failed"
                    err_text = str(e)
                    # Friendlier message for tunnel/proxy failures so the
                    # customer immediately understands it's a proxy-pool
                    # issue, not their setup.
                    friendly = err_text[:180]
                    is_tunnel_fail = any(tok in err_text for tok in _TUNNEL_ERR_TOKENS)
                    # Specific detection: 502 Bad Gateway from the proxy
                    # gateway means the provider is REFUSING the target
                    # domain (common with residential proxy ToS filters).
                    # Surface that distinctly so users don't keep blaming
                    # their own setup.
                    is_proxy_block_502 = (
                        "502" in err_text
                        or "bad gateway" in err_text.lower()
                        or "ERR_HTTP_RESPONSE_CODE_FAILURE" in err_text
                    )
                    if is_proxy_block_502:
                        friendly = (
                            f"Proxy provider returned 502 Bad Gateway after {tunnel_attempt + 1} attempt(s) — "
                            "your proxy provider is REFUSING this target domain. "
                            "Ask your proxy support to whitelist the domain, or use a different proxy provider."
                        )
                    elif is_tunnel_fail:
                        # ── 2026-05: combined attempts message ──
                        # Total goto attempts = proxy rotations + same-proxy retries.
                        # Surfaces both numbers so the user sees the system
                        # didn't just give up after 1 try.
                        total_attempts = (tunnel_attempt + 1) + same_proxy_retry
                        friendly = (
                            f"Proxy tunnel failed after {total_attempts} attempt(s) "
                            f"({tunnel_attempt} proxy rotations + {same_proxy_retry} same-proxy retries) — "
                            "your proxy provider couldn't reach the target. Try a different "
                            "US state, smaller batch, slower pacing, or reload proxies."
                        )
                    entry["error"] = f"goto failed: {friendly}"
                    # Tag tunnel/proxy-block failures so the dispatcher can
                    # optionally exclude them from the max_attempts budget
                    # (so transient proxy-provider hiccups don't
                    # prematurely end the run before real visits happen).
                    if is_tunnel_fail or is_proxy_block_502:
                        entry["tunnel_failed"] = True
                        try:
                            RUT_JOBS[job_id]["tunnel_fail_count"] = int(
                                RUT_JOBS[job_id].get("tunnel_fail_count", 0) or 0
                            ) + 1
                        except Exception:
                            pass
                    push_live_step(job_id, i + 1, "browser", "failed", f"goto failed: {friendly[:100]}")
                    await context.close()
                    return await _record(job_id, entry, report, report_lock, db)

                await page.wait_for_timeout(600 + random.randint(0, 500))

                # Wait for chained redirects (tracker-302 → parent domain →
                # JS-redirect → landing) to fully settle before we inspect
                # the DOM. Without this, on slower offer sites we inspect
                # an intermediate empty page and miss the CTA.
                # NOTE (2026-02): networkidle already happened inside the
                # landing-screenshot block above, so this is a short
                # belt-and-braces wait — reduced from 20s to 6s.
                try:
                    await page.wait_for_load_state("networkidle", timeout=6000)
                except Exception:
                    pass

                # Capture landing URL (post-tracker redirect settle, PRE form fill).
                # Used later to detect "conversion page reached" = host changed
                # after the submit compared to this landing host.
                try:
                    entry["landing_url"] = page.url
                except Exception:
                    pass

                # ── 2026-01: Human-like warm-up behaviour ─────────────
                # Real users read, scroll, hover before interacting.
                # Anura / IPQS / ArkoseLabs measure mouse-movement
                # entropy + dwell time — a bot that fills + submits in
                # 800ms is an instant fraud flag. _human_warmup adds
                # 3-8 seconds of realistic mouse/scroll/pause behaviour.
                # Errors swallowed inside the helper so a quirky page
                # can NEVER abort a visit.
                await _human_warmup(page, fp)

                if follow_redirect:
                    # Give the page a bit more time to do any JS redirect
                    try:
                        await page.wait_for_load_state("networkidle", timeout=4000)
                    except Exception:
                        pass

                # ── 2026-05: Offer-site Duplicate-IP / VPN hard-block detection ──
                # Some offers serve a "Duplicate IP / Access denied" or
                # "Please turn off your VPN" page before the form ever
                # renders. Burn this exit-IP from the in-job dup-set so
                # the on-demand ProxyJet probe will skip it on every
                # subsequent visit in this job. Also persist the IP to
                # the `rut_burnt_ips` MongoDB collection so EVERY future
                # job (even days later) loads this IP into its initial
                # `duplicate_ip_set` and skips it from the start.
                #
                # We check duplicate FIRST (more specific), then VPN.
                # Same handling for both — only the status reason text
                # and live-step message differ.
                _block_reason = ""          # "duplicate_ip" or "vpn"
                _block_snippet = ""
                try:
                    _is_dup_block, _dup_snippet = await _detect_offer_duplicate_ip_block(page)
                except Exception:
                    _is_dup_block, _dup_snippet = (False, "")
                if _is_dup_block:
                    _block_reason = "duplicate_ip"
                    _block_snippet = _dup_snippet
                else:
                    try:
                        _is_vpn_block, _vpn_snippet = await _detect_offer_vpn_block(page)
                    except Exception:
                        _is_vpn_block, _vpn_snippet = (False, "")
                    if _is_vpn_block:
                        _block_reason = "vpn"
                        _block_snippet = _vpn_snippet
                if _block_reason:
                    _burned_ip = (entry.get("exit_ip") or "").strip()
                    # In-job set update (instant — affects next probe in this job)
                    if _burned_ip and duplicate_ip_set is not None:
                        try:
                            duplicate_ip_set.add(_burned_ip)
                        except Exception:
                            pass
                    # Cross-job persistence (so future jobs load this IP)
                    if _burned_ip:
                        _spawn_live(_persist_burnt_ip(
                            db,
                            ip=_burned_ip,
                            reason=_block_reason,
                            user_id=engine_user_id or "",
                            offer_url=target_url or "",
                            state=(entry.get("lead_state") or "").upper(),
                            job_id=job_id or "",
                        ))
                    # Screenshot for evidence (different name per reason so
                    # the ZIP report makes the reason obvious at a glance).
                    try:
                        _shot_suffix = "dup_ip" if _block_reason == "duplicate_ip" else "vpn_block"
                        shot_path = shots_dir / f"visit_{i+1:05d}_{_shot_suffix}.png"
                        await page.screenshot(path=str(shot_path), full_page=False)
                        entry["screenshot"] = shot_path.name
                    except Exception:
                        pass
                    try:
                        entry["final_url"] = page.url
                    except Exception:
                        pass
                    if _block_reason == "duplicate_ip":
                        entry["status"] = "skipped_duplicate_ip"
                        entry["error"] = (
                            f"Offer-site rejected exit-IP {_burned_ip or '?'} as duplicate "
                            f"({(_block_snippet or 'duplicate IP block page')[:120]})"
                        )
                        _live_msg = (
                            f"Offer-side duplicate IP block · burning {_burned_ip or '?'} "
                            f"· persisted to rut_burnt_ips · retrying with fresh IP next visit"
                        )
                    else:
                        # Reuse the existing skipped_vpn status so the
                        # dashboard counters and downstream reports
                        # categorise this correctly without needing a
                        # new bucket.
                        entry["status"] = "skipped_vpn"
                        entry["error"] = (
                            f"Offer-site flagged exit-IP {_burned_ip or '?'} as VPN/proxy "
                            f"({(_block_snippet or 'VPN block page')[:120]})"
                        )
                        _live_msg = (
                            f"Offer-side VPN/proxy detection · burning {_burned_ip or '?'} "
                            f"· persisted to rut_burnt_ips · retrying with fresh IP next visit"
                        )
                    push_live_step(
                        job_id, i + 1, "filter", "skipped",
                        _live_msg,
                        screenshot=entry.get("screenshot", ""),
                    )
                    try:
                        await context.close()
                    except Exception:
                        pass
                    # ── 2026-05 — transparent retry in ProxyJet on-demand mode ─
                    # When the offer rejects the IP as duplicate or VPN,
                    # we DO NOT want to record a wasted visit entry. The
                    # worker() wrapper catches this exception and retries
                    # the SAME visit slot with a fresh ProxyJet IP — the
                    # burnt IP is already persisted above so future probes
                    # (within this job AND across jobs) will skip it.
                    # Legacy mode (proxyjet_on_demand=False) still records
                    # the entry as before since there's no IP-source loop
                    # to retry against.
                    if proxyjet_on_demand:
                        raise _OfferBlockRetryNeeded(
                            reason=_block_reason,
                            burnt_ip=_burned_ip or "",
                        )
                    return await _record(job_id, entry, report, report_lock, db)

                # No form fill → plain real click, just screenshot
                if not form_fill_enabled:
                    try:
                        entry["final_url"] = page.url
                    except Exception:
                        pass
                    shot_path = shots_dir / f"visit_{i+1:05d}.png"
                    try:
                        await page.screenshot(path=str(shot_path), full_page=False)
                        entry["screenshot"] = shot_path.name
                    except Exception:
                        pass
                    entry["status"] = "ok"
                    push_live_step(job_id, i + 1, "done", "ok",
                                   f"Visit complete → {entry.get('final_url', '')[:120]}",
                                   screenshot=entry.get("screenshot", ""))
                    await context.close()
                    return await _record(job_id, entry, report, report_lock, db)

                # Form-fill path — with validation-error detection + same-session
                # retry using the next available lead row (max 3 invalid retries).
                MAX_INVALID_RETRIES = 3
                retry_attempt = 0
                tried_row_ids: List[int] = []
                if row_index is not None:
                    tried_row_ids.append(row_index)

                while True:
                    if skip_captcha and await _page_has_captcha(page):
                        entry["status"] = "skipped_captcha"
                        entry["error"] = "Captcha detected on landing"
                        push_live_step(job_id, i + 1, "form", "skipped", "Captcha detected — skipping")
                        break

                    push_live_step(
                        job_id, i + 1, "form", "info",
                        (f"Filling form with lead #{entry.get('row_index', '?')}"
                         + (f" (retry {retry_attempt}/{MAX_INVALID_RETRIES})" if retry_attempt else ""))
                        if row else "Filling form",
                    )
                    # If user provided a custom Automation JSON, run that.
                    # Otherwise fall through to the smart auto-fill heuristic.
                    if automation_steps:
                        # Hand the runner a screenshot-callback so any
                        # {"action":"screenshot",...} step the user added
                        # via the Visual Recorder's Capture tool surfaces
                        # in the Live Activity panel (and is also saved
                        # alongside the visit's other shots for later
                        # inspection).
                        async def _on_user_capture(step_idx: int, name: str, png_bytes: bytes):
                            try:
                                safe_name = (name or f"step_{step_idx}").replace("/", "_")[:40]
                                shot_path = shots_dir / f"visit_{i+1:05d}_capture{step_idx:02d}_{safe_name}.png"
                                await asyncio.to_thread(shot_path.write_bytes, png_bytes)
                                entry.setdefault("capture_screenshots", []).append({
                                    "name": name,
                                    "step_index": step_idx,
                                    "path": str(shot_path),
                                })
                                # 2026-01 — Pass the saved screenshot filename
                                # along with the live step so the Live Activity
                                # panel can render the thumbnail at the exact
                                # point in the automation where the user placed
                                # the capture marker. Without this the panel
                                # only showed the text "📷 <name> (step N)" —
                                # which is why capture screenshots weren't
                                # surfacing in the live view.
                                push_live_step(
                                    job_id, i + 1, "capture", "ok",
                                    f"📷 {name} (step {step_idx})",
                                    screenshot=shot_path.name,
                                )
                            except Exception:
                                # Surface the failure but never crash
                                # the visit because of a capture.
                                pass

                        # 2026-01 (additive): per-visit live activity feed.
                        # Mirrors the Visual Recorder's live test feature so the
                        # RUT job UI can show a real-time grid of all in-flight
                        # visits (one tile per concurrency), each with the latest
                        # browser frame + current step. Keeps a SINGLE-frame
                        # state per visit (not a buffer) — UI just renders the
                        # most recent state. Failures are swallowed.
                        async def _visit_progress_cb(event: Dict[str, Any]) -> None:
                            try:
                                j_state = RUT_JOBS.get(job_id)
                                if j_state is None:
                                    return
                                lv = j_state.setdefault("live_visits", {})
                                vkey = str(i + 1)
                                v = lv.setdefault(vkey, {
                                    "visit_idx": i + 1,
                                    "started_at": time.time(),
                                    "events_count": 0,
                                    "latest_event": None,
                                    "latest_frame_b64": "",
                                    "page_url": "",
                                    "status": "running",
                                })
                                # Slim event copy (no heavy screenshot bytes)
                                v["latest_event"] = {
                                    k: ev_val for k, ev_val in event.items()
                                    if k != "screenshot_b64"
                                }
                                if event.get("screenshot_b64"):
                                    v["latest_frame_b64"] = event["screenshot_b64"]
                                if event.get("page_url"):
                                    v["page_url"] = event["page_url"]
                                # 2026-05 — Step-marker accumulator.
                                # Each successful element-targeted step
                                # arrives with `target_box` + `doc_size`;
                                # we accumulate a history so the
                                # frontend can overlay coloured dots on
                                # the latest screenshot showing where
                                # every step landed. Bounded at 50 markers
                                # so the polling payload doesn't bloat
                                # for very long recordings.
                                if event.get("target_box") and event.get("status") == "ok":
                                    markers = v.setdefault("step_markers", [])
                                    markers.append({
                                        "idx": event.get("idx"),
                                        "action": event.get("action"),
                                        "selector": (event.get("selector") or "")[:120],
                                        "box": event["target_box"],
                                        "status": "ok",
                                        "ts": event.get("timestamp_ms"),
                                    })
                                    if len(markers) > 50:
                                        del markers[: len(markers) - 50]
                                if event.get("doc_size"):
                                    v["doc_size"] = event["doc_size"]
                                v["events_count"] = int(v.get("events_count", 0)) + 1
                                if event.get("status") == "failed":
                                    v["status"] = "failed"
                                v["last_update"] = time.time()
                            except Exception:
                                pass

                        # Start a watchdog task that records a "stuck"
                        # event if the page URL hasn't changed for >25s
                        # while the automation script is running. The
                        # watchdog also captures a screenshot + body
                        # text snapshot for offline debugging, then
                        # cancels the steps task so the job moves on
                        # to the next visit instead of wasting the full
                        # automation budget on a dead page.
                        _steps_task = asyncio.create_task(
                            _execute_automation_steps(
                                page, row or {}, automation_steps, skip_captcha=skip_captcha,
                                self_heal=self_heal,
                                on_screenshot=_on_user_capture,
                                user_id=link_owner_id,   # enable self-healing aliases (2026-01)
                                on_step_progress=_visit_progress_cb,  # 2026-01: real-time per-visit feed
                            )
                        )

                        def _trigger_abort_steps():
                            if not _steps_task.done():
                                _steps_task.cancel()

                        # 2026-05: shared state dict so the watchdog can
                        # report whether the page ever progressed before
                        # going idle. If it did, the visit is treated as
                        # a successful submit (status="ok") rather than
                        # a dead-page abort (status="stuck").
                        _wd_state: Dict[str, Any] = {"progressed": False, "progression_count": 0}
                        _watchdog = asyncio.create_task(
                            _stuck_watchdog(
                                page, job_id, i + 1,
                                threshold_s=float(stuck_watchdog_seconds or 240.0),
                                shots_dir=shots_dir,
                                on_stuck=_trigger_abort_steps,
                                state=_wd_state,
                            )
                        )
                        try:
                            try:
                                step_res = await _steps_task
                            except asyncio.CancelledError:
                                # Watchdog aborted us. Two outcomes:
                                #   A) Page never progressed → genuinely
                                #      dead (proxy died, captcha wall,
                                #      blank screen). Status = "stuck".
                                #   B) Page progressed at least once then
                                #      went idle (submit succeeded, just
                                #      no more user-driven survey steps
                                #      in the JSON). Status = "ok" — the
                                #      submit DID happen, the visit
                                #      counts as successful.
                                _stuck_url_for_err = ""
                                try:
                                    _stuck_url_for_err = page.url or ""
                                except Exception:
                                    pass
                                _progressed = bool(_wd_state.get("progressed"))
                                _pcount = int(_wd_state.get("progression_count") or 0)
                                if _progressed:
                                    # Treat as successful submit — page
                                    # is alive, just idle on a post-
                                    # submit/survey screen the JSON
                                    # didn't continue past.
                                    step_res = {
                                        "status": "ok",
                                        "error": (
                                            f"Visit submitted; page progressed {_pcount}× then idled "
                                            f"on {_stuck_url_for_err[:160]} (auto-marked ok by watchdog — "
                                            f"not a dead page)"
                                        ),
                                        "executed_steps": 0,
                                    }
                                    try:
                                        push_live_step(
                                            job_id, i + 1, "wd_ok_idle", "ok",
                                            f"Watchdog: page progressed {_pcount}× — marking visit as OK",
                                        )
                                    except Exception:
                                        pass
                                else:
                                    step_res = {
                                        "status": "stuck",
                                        "error": f"Visit aborted by watchdog (page stuck >{int(stuck_watchdog_seconds or 240)}s on {_stuck_url_for_err[:200]})",
                                        "executed_steps": 0,
                                    }
                        finally:
                            _watchdog.cancel()
                            try:
                                await _watchdog
                            except (asyncio.CancelledError, Exception):
                                pass

                        # ── 2026-01: Multi-step SPA auto-continue fallback ─
                        # If the user's custom JSON completes (status=ok)
                        # OR is aborted by the watchdog (status=stuck),
                        # AND a target screenshot was provided (signalling
                        # the customer expects a multi-step survey →
                        # conversion-page flow), automatically attempt
                        # to drive the page forward through any remaining
                        # survey questions / Continue buttons / deal
                        # cards until either:
                        #     • the conversion page is visually reached,
                        #     • the page stops changing, or
                        #     • the fallback budget (~25s) is exhausted.
                        #
                        # Rationale: offer flows like
                        # anyunclaimedassets.com/indexform.php are
                        # JS-driven SPAs — Profile → Cash & Assets Survey
                        # → Benefits → Thank-you all on the same URL.
                        # A user JSON that only handles the Profile step
                        # would otherwise leave every visit stranded on
                        # the first survey page. This fallback uses the
                        # same survey_click_v2 + complete_random_deals +
                        # _dismiss_popups helpers the smart auto-fill
                        # already relies on, so behaviour is consistent
                        # with the non-custom-JSON path.
                        _fb_status = (step_res or {}).get("status") or ""
                        if (
                            target_screenshot_phash
                            and _fb_status in ("ok", "stuck")
                        ):
                            try:
                                from rut_flash_helpers import (
                                    survey_click_v2 as _fb_survey,
                                    complete_random_deals as _fb_deals,
                                )
                            except Exception:
                                _fb_survey = None
                                _fb_deals = None
                            try:
                                from form_filler import _dismiss_popups as _fb_dismiss
                            except Exception:
                                _fb_dismiss = None
                            _fb_continue_js = (
                                "(function(){"
                                " var KW=['continue','next','submit','i agree','agree','confirm','claim','yes','accept','proceed','get started','unlock','finish','complete','done','start'];"
                                " function vis(e){try{var s=window.getComputedStyle(e);return s&&s.display!=='none'&&s.visibility!=='hidden'&&e.offsetWidth>0&&e.offsetHeight>0;}catch(_){return false;}}"
                                " var nodes=Array.from(document.querySelectorAll('button,a,input[type=button],input[type=submit],[role=button]'));"
                                " for(var i=0;i<nodes.length;i++){var el=nodes[i];if(!vis(el))continue;"
                                "  var t=((el.innerText||el.textContent||el.value||'')+'').toLowerCase().replace(/\\s+/g,' ').trim();"
                                "  if(!t)continue;"
                                "  for(var k=0;k<KW.length;k++){if(t===KW[k]||t.indexOf(KW[k])>=0){"
                                "    try{el.scrollIntoView({block:'center'});}catch(_){};"
                                "    if(el.tagName==='A'&&el.href&&!el.target){window.location.assign(el.href);return true;}"
                                "    el.click();"
                                "    var f=el.form||(el.closest&&el.closest('form'));"
                                "    if(f&&(el.type==='submit'||(el.getAttribute&&el.getAttribute('type')==='submit'))){setTimeout(function(){try{if(!f._krx_fbs){f._krx_fbs=true;f.submit();}}catch(_){}},150);}"
                                "    return true;"
                                "  }}"
                                " }"
                                " return false;"
                                "})()"
                            )
                            _fb_deadline = __import__("time").monotonic() + 25.0
                            _fb_max_iter = 12
                            _fb_last_url = ""
                            try:
                                _fb_last_url = page.url or ""
                            except Exception:
                                pass
                            _fb_no_change = 0
                            _fb_iter_count = 0
                            push_live_step(
                                job_id, i + 1, "auto_continue", "info",
                                "Auto-continue fallback running — driving any remaining survey/continue pages…",
                            )
                            while (
                                _fb_iter_count < _fb_max_iter
                                and __import__("time").monotonic() < _fb_deadline
                            ):
                                _fb_iter_count += 1
                                _fb_progress = False
                                if _fb_dismiss is not None:
                                    try:
                                        await _fb_dismiss(page)
                                    except Exception:
                                        pass
                                if _fb_survey is not None:
                                    try:
                                        _sr = await _fb_survey(page, max_iterations=2, picker=None)
                                        if isinstance(_sr, dict) and int(_sr.get("clicks", 0)) > 0:
                                            _fb_progress = True
                                    except Exception:
                                        pass
                                if _fb_deals is not None:
                                    try:
                                        _dn = await _fb_deals(page, count_min=1, count_max=2)
                                        if int(_dn or 0) > 0:
                                            _fb_progress = True
                                    except Exception:
                                        pass
                                try:
                                    if bool(await page.evaluate(_fb_continue_js)):
                                        _fb_progress = True
                                except Exception:
                                    pass
                                try:
                                    await page.wait_for_timeout(1200)
                                except Exception:
                                    pass
                                try:
                                    await page.wait_for_load_state(
                                        "domcontentloaded", timeout=4000,
                                    )
                                except Exception:
                                    pass
                                try:
                                    _fb_cur = page.url or ""
                                except Exception:
                                    _fb_cur = _fb_last_url
                                if _fb_cur == _fb_last_url and not _fb_progress:
                                    _fb_no_change += 1
                                    if _fb_no_change >= 2:
                                        break
                                else:
                                    _fb_no_change = 0
                                    _fb_last_url = _fb_cur
                            # If fallback unstuck the visit (URL changed
                            # OR DOM-progress happened), upgrade status
                            # from "stuck" → "ok" so the downstream
                            # post_submit + conversion check don't treat
                            # this as a failure.
                            if _fb_status == "stuck" and _fb_iter_count > 0:
                                step_res["status"] = "ok"
                                if "error" in step_res:
                                    step_res["error"] = (
                                        (step_res.get("error") or "")
                                        + " | Recovered by auto-continue fallback"
                                    )[:300]
                                push_live_step(
                                    job_id, i + 1, "auto_continue", "ok",
                                    f"Auto-continue ran {_fb_iter_count} iterations — visit recovered",
                                )
                    else:
                        # Click through any CTA ("UNLOCK NOW", "Get Started", etc.)
                        # — up to 3 tries because some offers have a 2-step warm-up.
                        await _ensure_form_visible(page, max_tries=3)
                        if skip_captcha and await _page_has_captcha(page):
                            step_res = {"status": "skipped_captcha", "error": "Captcha after CTA click"}
                        else:
                            # Capture a screenshot right before every submit
                            # click inside the multi-stage form fill, so the
                            # user can inspect what data was in the form at
                            # the moment the bot clicked Continue.  Screenshots
                            # are named visit_<N>_pre_submit_<stage>.png and
                            # end up in the results ZIP + Live Activity log.
                            _pre_submit_shots: List[str] = []

                            async def _pre_submit_cb(label: str) -> None:
                                try:
                                    shot_name = f"visit_{i+1:05d}_pre_submit_{label}.png"
                                    shot_path_ps = shots_dir / shot_name
                                    await page.screenshot(
                                        path=str(shot_path_ps), full_page=True,
                                    )
                                    _pre_submit_shots.append(shot_name)
                                    push_live_step(
                                        job_id, i + 1, "form_filled", "info",
                                        f"📷 2/4 Form filled — about to submit ({label})",
                                        screenshot=shot_name,
                                    )
                                except Exception as e:  # noqa: BLE001
                                    logger.debug(f"pre-submit screenshot failed: {e}")

                            step_res = await _multi_step_fill(
                                page, row or {}, picker=ai_picker,
                                pre_submit_cb=_pre_submit_cb,
                            )
                            if _pre_submit_shots:
                                entry["pre_submit_screenshots"] = _pre_submit_shots

                    entry["status"] = step_res["status"]
                    if step_res.get("error"):
                        entry["error"] = step_res["error"]
                    # ── 2026-05: best-effort `screenshot` steps on failure ──
                    # If automation failed mid-way, the user's later
                    # {"action":"screenshot"} steps (added via Visual
                    # Recorder's Capture tool) would normally be lost.
                    # Run any remaining screenshot steps now as
                    # best-effort so the customer can see WHAT the page
                    # looked like when the submit broke (often reveals
                    # validation errors, captcha, blank pages, etc.).
                    #
                    # 2026-01-29 FIX: These captures fire AFTER a step
                    # failure, NOT after a real form submit — so they
                    # must NOT carry the user's original capture name
                    # verbatim (which earlier caused "FORM SUBMIT" to
                    # appear in Live Activity even when the form was
                    # never actually submitted). We now prefix the
                    # name with "[POST-FAILURE DEBUG]" so the operator
                    # can clearly distinguish a real submit capture
                    # from a debug snapshot taken AFTER a mid-form
                    # failure.
                    if (
                        step_res.get("status") == "failed"
                        and step_res.get("remaining_steps")
                        and automation_steps
                    ):
                        for _rem in step_res["remaining_steps"]:
                            try:
                                if (
                                    isinstance(_rem, dict)
                                    and (_rem.get("action") or "").strip().lower() == "screenshot"
                                ):
                                    _png = await page.screenshot(
                                        type="png", timeout=8000, full_page=True
                                    )
                                    _orig_nm = str(_rem.get("name") or "post_failure").strip() or "post_failure"
                                    # Clear post-failure tag so the user
                                    # knows the capture fired in DEBUG
                                    # mode (their preceding step failed),
                                    # not on a successful submit.
                                    _nm = f"[POST-FAILURE DEBUG] {_orig_nm}"
                                    await _on_user_capture(0, _nm, _png)
                            except Exception:  # noqa: BLE001
                                pass
                    # FlashRewards-style: track survey answers + deals
                    # completed on this attempt (if helper ran). These
                    # surface in the UI report and are required for
                    # at-least-2-deals success criterion.
                    if "deals_completed" in step_res:
                        entry["deals_completed"] = int(step_res["deals_completed"] or 0)
                    if "survey_answers" in step_res:
                        entry["survey_answers"] = int(step_res["survey_answers"] or 0)
                    # Stash survey picks (q_sig, answer_text) so we can record
                    # outcomes once thank_you_reached is determined further down.
                    if "survey_picks" in step_res and isinstance(step_res["survey_picks"], list):
                        entry["_survey_picks"] = step_res["survey_picks"]

                    # Detect server-side / inline validation errors. Only run
                    # when submit actually went through (ok / submitted-but-no-
                    # redirect); skipped_captcha / no_fields_matched handled above.
                    # GATED by user toggle — default OFF because many landing
                    # pages have consent banners / alerts that trigger false
                    # positives on the form page itself (before submit).
                    is_invalid_data = False
                    val_error = ""
                    if invalid_detection_enabled and entry["status"] in ("submitted_but_no_redirect", "ok"):
                        try:
                            is_invalid_data, val_error = await _detect_validation_errors(page)
                        except Exception:
                            is_invalid_data = False

                    if is_invalid_data and row_index is not None:
                        # Mark the CURRENT row as invalid (drops from pending_leads)
                        async with report_lock:
                            invalid_row_indices.add(row_index)
                        # Per-use immediate deletion from saved data file
                        _spawn_live(_live_remove_data_row(row_index))

                        push_live_step(
                            job_id, i + 1, "submit", "failed",
                            f"Invalid data on lead #{row_index + 1}: {val_error[:110]}",
                        )

                        if retry_attempt < MAX_INVALID_RETRIES:
                            # Pick next available row and reload the form page
                            # to clear previous error state. State-match aware.
                            if state_match_enabled and state_col:
                                proxy_state_code = _normalize_state(geo.get("region")) or _normalize_state(geo.get("region_name"))
                                next_pick = pick_next_row_for_state(proxy_state_code) if proxy_state_code else None
                            else:
                                next_pick = pick_next_row()
                            if not next_pick:
                                entry["status"] = "invalid_data"
                                entry["error"] = f"Invalid: {val_error[:160]} (no more leads to retry)"
                                push_live_step(job_id, i + 1, "form", "failed",
                                               "No more leads available to retry")
                                break

                            row_index, row = next_pick
                            entry["row_index"] = row_index + 1
                            tried_row_ids.append(row_index)
                            retry_attempt += 1

                            # Reload form page so invalid state is cleared
                            try:
                                await page.goto(target_url, timeout=90000,
                                                wait_until="domcontentloaded")
                                await page.wait_for_timeout(700 + random.randint(0, 500))
                                try:
                                    await page.wait_for_load_state("networkidle", timeout=15000)
                                except Exception:
                                    pass
                            except Exception as e:
                                entry["status"] = "invalid_data"
                                entry["error"] = f"Page reload after invalid failed: {str(e)[:120]}"
                                break

                            push_live_step(
                                job_id, i + 1, "form", "info",
                                f"Retry on same form with next lead #{row_index + 1}",
                            )
                            continue  # retry the while loop
                        else:
                            # Max retries reached — finalize as invalid_data
                            entry["status"] = "invalid_data"
                            entry["error"] = f"Invalid (max {MAX_INVALID_RETRIES} retries): {val_error[:140]}"
                            push_live_step(
                                job_id, i + 1, "submit", "failed",
                                f"Max {MAX_INVALID_RETRIES} invalid retries reached",
                            )
                            break

                    # Normal outcome (ok, submitted_but_no_redirect without validation
                    # error, no_fields_matched, skipped_*, etc.) — exit retry loop.
                    break

                # Stash retry metadata
                entry["retry_attempts"] = retry_attempt
                entry["tried_row_indices"] = [r + 1 for r in tried_row_ids]

                push_live_step(
                    job_id, i + 1, "submit",
                    "ok" if entry["status"] == "ok" else ("skipped" if "skipped" in str(entry["status"]) else "failed"),
                    f"{entry['status']}{' — ' + entry['error'] if entry.get('error') else ''}"[:180],
                )

                # Mark the FINAL lead row as CONSUMED if submit succeeded — it
                # will be excluded from the pending_leads.xlsx so the user
                # never reuses it.
                if entry["status"] == "ok" and row_index is not None:
                    async with report_lock:
                        consumed_row_indices.add(row_index)
                    # IMMEDIATE per-use deletion from saved data file
                    _spawn_live(_live_remove_data_row(row_index))

                # Grab TrustedForm / LeadID proofs (if the landing page uses them)
                try:
                    lead = await page.evaluate("""() => {
                        const grab = sel => {
                            const el = document.querySelector(sel);
                            return el ? (el.value || el.getAttribute('value') || '') : '';
                        };
                        return {
                            trusted_form: grab('[name="xxTrustedFormCertUrl"]') || grab('[name="xxTrustedFormToken"]'),
                            lead_id: grab('#leadid_token') || grab('[name="universal_leadid"]') || grab('[name="LeadiD"]'),
                        };
                    }""")
                    entry["trusted_form"] = lead.get("trusted_form", "")
                    entry["lead_id"] = lead.get("lead_id", "")
                except Exception:
                    pass

                try:
                    entry["final_url"] = page.url
                except Exception:
                    pass
                try:
                    await page.wait_for_load_state("networkidle", timeout=5000)
                except Exception:
                    pass

                # Legacy conversion-page signal (host change only) kept in
                # the entry for debugging comparison with the new strict
                # thank-you detection. The strict check runs below after
                # the post-submit wait and overrides conversion_page_reached.
                try:
                    entry["host_changed_after_submit"] = _did_reach_conversion(
                        entry.get("landing_url") or "",
                        entry.get("final_url") or "",
                    )
                except Exception:
                    entry["host_changed_after_submit"] = False

                # Post-submit wait so the "thank-you" / offers-flow page
                # fully renders BEFORE we screenshot it. Default 6s,
                # user-configurable 3–15s.
                if entry["status"] == "ok":
                    wait_ms = max(3000, min(int(post_submit_wait * 1000), 600000))
                    await page.wait_for_timeout(wait_ms)
                else:
                    await page.wait_for_timeout(900)

                # Re-read final_url after the post-submit wait — redirect chains
                # to the actual thank-you page often happen during this sleep.
                try:
                    entry["final_url"] = page.url
                except Exception:
                    pass

                # ALWAYS capture a post-submit screenshot — regardless of
                # whether the visit is ultimately counted as a conversion.
                # User explicitly asked for this: "jab submit hone k 7 second
                # bad jo page ay os k ss ay". File lands in the results ZIP
                # and is also pushed into the Live Activity panel so the
                # user can click to enlarge.
                try:
                    post_shot_name = f"visit_{i+1:05d}_post_submit.png"
                    post_shot_path = shots_dir / post_shot_name
                    await page.screenshot(path=str(post_shot_path), full_page=True)
                    entry["post_submit_screenshot"] = post_shot_name
                    push_live_step(
                        job_id, i + 1, "post_submit", "info",
                        f"📷 3/4 Post-submit page loaded ({post_submit_wait}s after submit) → {entry.get('final_url', '')[:90]}",
                        screenshot=post_shot_name,
                    )
                except Exception as e:  # noqa: BLE001
                    logger.debug(f"post-submit screenshot failed: {e}")

                # Detect mid-navigation proxy tunnel breaks — if after the
                # post-submit wait the page is sitting on a Chromium internal
                # error URL, the visit did NOT actually reach any real page.
                # Mark as failed so conversion stats and downstream retries
                # behave correctly.
                _fu = entry.get("final_url", "") or ""
                if _fu.startswith("chrome-error://") or _fu.startswith("chrome://network-error"):
                    entry["status"] = "failed"
                    if not entry.get("error"):
                        entry["error"] = f"Browser navigation error (proxy tunnel broken mid-flow): {_fu}"
                    push_live_step(job_id, i + 1, "nav", "failed",
                                   f"Navigation error after submit: {_fu[:80]}")

                # STRICT thank-you page detection: needs host change + URL
                # keyword + page text keyword. Only TRUE thank-you pages
                # count as conversions and get screenshotted.
                page_title_str = ""
                page_text_str = ""
                try:
                    page_title_str = await page.title()
                except Exception:
                    pass
                try:
                    page_text_str = await page.evaluate(
                        "() => (document.body ? document.body.innerText : '').slice(0, 4000)"
                    )
                except Exception:
                    pass

                try:
                    entry["thank_you_reached"] = _is_thank_you_page(
                        entry.get("landing_url") or "",
                        entry.get("final_url") or "",
                        page_text_str,
                        page_title_str,
                    )
                except Exception:
                    entry["thank_you_reached"] = False

                # ── 2026-01 status upgrade ─────────────────────────────
                # If the visit reached the thank-you / conversion page
                # but was later aborted by the stuck-watchdog (the page
                # was already on the success URL, nothing more to do, so
                # the URL stopped changing and the watchdog killed the
                # tab), this is actually a SUCCESS — the postback /
                # conversion fired before the abort. Upgrade the status
                # so the counters show 'succeeded' instead of 'stuck'
                # for these wins. We only upgrade `stuck` (watchdog
                # outcome) — other failures (network, captcha, invalid
                # data, etc.) keep their original status even if the
                # offer's tracking pixel happened to fire.
                if entry.get("thank_you_reached") and entry.get("status") == "stuck":
                    entry["status"] = "ok"
                    entry["error"] = ""
                    entry["_upgraded_from_stuck"] = True

                entry["page_title"] = (page_title_str or "")[:200]
                # Conversion count is now driven STRICTLY by thank-you-reached
                # (not just any host-change). Matches the user's explicit ask:
                # "jo form thanks page tak complete ho wahi conversion count ho".
                entry["conversion_page_reached"] = bool(entry.get("thank_you_reached"))

                # AI answer-learning: record outcomes for survey picks now
                # that we know whether this visit converted. Fire-and-forget
                # so it never blocks the visit lifecycle.
                # ── 2026-05: skip when pure_json_mode is ON ──
                try:
                    picks = entry.get("_survey_picks") or []
                    if picks and db is not None and not pure_json_mode:
                        from rut_answer_learning import record_outcomes
                        asyncio.create_task(
                            record_outcomes(
                                db, target_url, picks,
                                bool(entry.get("thank_you_reached")),
                            )
                        )
                except Exception as e:  # noqa: BLE001
                    logger.debug(f"record_outcomes failed: {e}")

                # Screenshot logic: ONLY capture the final screenshot when
                # the thank-you page was reached. Drops disk usage and matches
                # the user's ask ("jo form thanks page tak complete ho os ka
                # screenshot ho"). We keep the small landing thumbnail taken
                # earlier for debugging / proof that the browser opened at all.
                if entry.get("thank_you_reached"):
                    shot_path = shots_dir / f"visit_{i+1:05d}_thankyou.png"
                    try:
                        # Wait until the final page has rendered something
                        # visible — without this we sometimes capture a
                        # blank intermediate frame between SPA redirects.
                        # Reduced timeout from 12s to 4s for speed optimization
                        try:
                            await page.wait_for_load_state(
                                "networkidle", timeout=4000,
                            )
                        except Exception:
                            pass
                        # Poll for visible body content (>100 chars or
                        # img/svg present) — stop blank captures.
                        # Reduced attempts from 8 to 4 and sleep from 1s to 0.5s
                        for _shot_attempt in range(4):
                            try:
                                ok = await page.evaluate(
                                    """() => {
                                        const t = (document.body ? document.body.innerText : '').trim();
                                        const hasImg = !!document.querySelector('img,svg,canvas');
                                        return t.length > 100 || hasImg;
                                    }"""
                                )
                                if ok:
                                    break
                            except Exception:
                                pass
                            await asyncio.sleep(0.5)
                        # Reduced settle time from 1.2s to 0.3s for speed
                        await asyncio.sleep(0.3)
                        await page.screenshot(path=str(shot_path), full_page=True)
                        entry["screenshot"] = shot_path.name
                        # Push its own live-activity step with the explicit
                        # 4/4 label so users see all 4 checkpoints clearly
                        # in the "View Backend Activity" modal.
                        push_live_step(
                            job_id, i + 1, "final", "ok",
                            f"📷 4/4 Final conversion page → {entry.get('final_url', '')[:90]}",
                            screenshot=shot_path.name,
                        )

                        # Target Screenshot Verification — pHash compare
                        if target_screenshot_phash:
                            try:
                                from screenshot_verifier import (
                                    compute_phash,
                                    compare_phashes,
                                    similarity_pct,
                                    is_match,
                                )
                                with open(shot_path, "rb") as f:
                                    live_bytes = f.read()
                                live_phash = compute_phash(live_bytes)
                                if live_phash:
                                    dist = compare_phashes(
                                        target_screenshot_phash, live_phash,
                                    )
                                    sim = similarity_pct(dist or 64)
                                    matched = is_match(dist or 64, target_screenshot_threshold)
                                    entry["screenshot_match_distance"] = dist
                                    entry["screenshot_match_similarity"] = sim
                                    entry["screenshot_match"] = bool(matched)

                                    # Check whether the final URL lands on
                                    # a known deal-page host. If yes, the
                                    # URL is authoritative — pHash is only
                                    # advisory, and we keep the conversion
                                    # as TRUE even when mobile/desktop
                                    # viewport or theme differences cause
                                    # a high Hamming distance.
                                    final_url_for_conv = entry.get("final_url") or ""
                                    try:
                                        from urllib.parse import urlparse as _up
                                        _fh = (_up(final_url_for_conv).netloc or "").lower().lstrip("www.")
                                    except Exception:
                                        _fh = ""
                                    url_says_converted = _matches_high_confidence_host(_fh)

                                    if url_says_converted:
                                        entry["thank_you_reached"] = True
                                        entry["conversion_page_reached"] = True
                                        push_live_step(
                                            job_id, i + 1, "screenshot_verify",
                                            "ok",
                                            f"target match: {sim}% similar (distance={dist}, threshold={target_screenshot_threshold}) "
                                            f"— URL host '{_fh}' is a known deal page → VERIFIED conversion (URL-authoritative)",
                                        )
                                    else:
                                        # Normal pHash-override behaviour
                                        entry["thank_you_reached"] = bool(matched)
                                        entry["conversion_page_reached"] = bool(matched)
                                        push_live_step(
                                            job_id, i + 1, "screenshot_verify",
                                            "ok" if matched else "skipped",
                                            f"target match: {sim}% similar (distance={dist}, threshold={target_screenshot_threshold}) → "
                                            f"{'VERIFIED conversion' if matched else 'NOT matching target — counting as miss'}",
                                        )
                            except Exception as e:  # noqa: BLE001
                                logger.debug(f"target screenshot compare failed: {e}")
                    except Exception as e:
                        logger.debug(f"thank-you screenshot failed: {e}")
                else:
                    # ── 2026-05: Failure-debug screenshot ───────────────────
                    # Form/visit didn't reach the thank-you page. Capture
                    # the page as-is so the customer can see WHY in the UI
                    # (e.g. validation error, captcha, blank, redirect to
                    # error page, etc.). Saved as `_final.png` to avoid
                    # confusion with successful conversion shots.
                    try:
                        fail_shot = shots_dir / f"visit_{i+1:05d}_final.png"
                        # short settle so animations / errors render
                        await asyncio.sleep(0.5)
                        await page.screenshot(path=str(fail_shot), full_page=True, timeout=6000)
                        entry["final_screenshot"] = fail_shot.name
                        # If no other screenshot field is set, use this for
                        # the visit row preview so the user always sees
                        # SOMETHING.
                        if not entry.get("screenshot"):
                            entry["screenshot"] = fail_shot.name
                        push_live_step(
                            job_id, i + 1, "final", "info",
                            f"📷 4/4 Final page (no conversion) → {entry.get('final_url', '')[:90]}",
                            screenshot=fail_shot.name,
                        )
                    except Exception as e:  # noqa: BLE001
                        logger.debug(f"failure screenshot failed: {e}")

                # Final live step — always push, include screenshot if we have one
                push_live_step(
                    job_id, i + 1, "done",
                    "ok" if entry["status"] == "ok" else ("skipped" if "skipped" in str(entry["status"]) else "failed"),
                    f"Visit {entry['status']}{' — ✓ converted' if entry.get('thank_you_reached') else ''} → {entry.get('final_url', '')[:120]}",
                    screenshot=entry.get("screenshot", ""),
                )

                # If we are in target-conversions mode and hit the target,
                # flip the DRAIN flag (not the hard-cancel flag) so the
                # dispatcher stops spawning NEW visits but every in-flight
                # visit still runs to completion. This keeps already-picked
                # leads / proxies / UAs from being wasted.
                try:
                    tgt = int(RUT_JOBS[job_id].get("target_conversions") or 0)
                    if tgt > 0 and entry.get("thank_you_reached"):
                        cur = int(RUT_JOBS[job_id].get("conversions") or 0)
                        # +1 because _record hasn't incremented yet for this entry
                        if (cur + 1) >= tgt and not target_drain_event.is_set():
                            target_drain_event.set()
                            RUT_JOBS[job_id]["target_reached"] = True
                            RUT_JOBS[job_id]["target_reached_at"] = datetime.now(timezone.utc).isoformat()
                            push_live_step(
                                job_id, 0, "done", "ok",
                                f"🎯 Target {tgt} conversions reached — no new visits will start. Waiting for in-flight visits to complete (so their leads/proxies aren't wasted)…",
                            )
                except Exception:
                    pass

                await context.close()
                # Force garbage collection after each visit to prevent memory buildup
                import gc
                gc.collect()
        except _OfferBlockRetryNeeded:
            # ── 2026-05 — propagate to worker() retry loop ─────────────
            # Offer rejected this exit-IP. The IP is already burned
            # (in-job set + persisted to rut_burnt_ips). The worker()
            # wrapper catches this and re-runs process_one() with a
            # fresh ProxyJet IP, up to proxyjet_unique_retry_cap times.
            # We MUST NOT call _record below or this wasted attempt
            # would clutter the report with a "skipped_duplicate_ip" /
            # "skipped_vpn" row that the user doesn't want to see.
            raise
        except Exception as e:
            entry["status"] = "failed"
            entry["error"] = f"{type(e).__name__}: {str(e)[:180]}"
        finally:
            # Browser is shared across visits — we only close the per-visit
            # context here. The shared browser is closed by the parent job
            # once ALL workers have finished.
            # Aggressive cleanup to prevent memory leaks
            if context is not None:
                try:
                    await context.close()
                except Exception:
                    pass
            # Small delay to let cleanup complete
            await asyncio.sleep(0.1)

        await _record(job_id, entry, report, report_lock, db)

    # Launch with concurrency + optional pacing
    # Extreme speed mode: Max 50 concurrent workers for ultra-fast processing
    # Can handle 100 conversions in 15-20 minutes with proper resources
    # — but on low-RAM boxes (8 GB laptops) the operator can hard-cap the
    # ceiling via RUT_MAX_CONCURRENCY env var so a careless slider value
    # of 10 doesn't OOM the host. Defaults to 50 (the old hard ceiling).
    _rut_hard_cap = max(1, int(os.environ.get("RUT_MAX_CONCURRENCY", "50")))
    semaphore = asyncio.Semaphore(max(1, min(int(concurrency or 1), _rut_hard_cap)))
    conc = max(1, min(int(concurrency or 1), _rut_hard_cap))

    logger.info(f"RUT Speed Mode: {conc} concurrent workers enabled (env cap={_rut_hard_cap})")

    async def worker(i: int, shared_browser: Browser):
        # Per-visit pacing: target time for this visit = i * delay_between
        if delay_between > 0:
            target_t = state["start_time"] + i * delay_between
            # sleep in small chunks so cancel is responsive
            while time.time() < target_t:
                # Either hard-cancel OR target-drain stops a NOT-YET-STARTED
                # visit — neither has any in-flight resource yet so abandoning
                # is free.
                if cancel_event.is_set() or target_drain_event.is_set():
                    return
                await asyncio.sleep(min(0.5, target_t - time.time()))
        if cancel_event.is_set() or target_drain_event.is_set():
            return
        async with semaphore:
            # Skip both on hard-cancel AND on target-drain. We haven't
            # picked any proxy / UA / lead row yet (those happen at the
            # top of process_one), so abandoning here is FREE — no
            # resource is wasted. Once we're inside process_one we'll
            # run to completion regardless of target_drain_event.
            if cancel_event.is_set() or target_drain_event.is_set():
                return
            # ── 2026-05 — Offer-block retry loop ─────────────────────
            # When ProxyJet on-demand mode is on, process_one() may
            # raise _OfferBlockRetryNeeded if the offer rejects the
            # exit-IP as "Duplicate IP" or "VPN/proxy". The IP is
            # already burned & persisted before the raise, so the
            # retry below will pick a FRESH ProxyJet IP that the
            # offer hasn't seen yet. Capped at
            # `proxyjet_unique_retry_cap` (default 50) attempts per
            # visit slot — beyond that we record a final
            # `skipped_no_unique_ip` so the job doesn't loop forever
            # when the offer's filter is unrealistically strict.
            max_offer_retries = (
                max(1, int(proxyjet_unique_retry_cap or 50))
                if proxyjet_on_demand else 1
            )
            retry_attempts = 0
            while True:
                try:
                    await process_one(i, shared_browser)
                    return
                except _OfferBlockRetryNeeded as _ob:
                    retry_attempts += 1
                    if retry_attempts >= max_offer_retries:
                        # Cap exhausted — record a single skipped entry so
                        # the user can see this visit slot was abandoned.
                        push_live_step(
                            job_id, i + 1, "filter", "skipped",
                            f"Offer-side block · {retry_attempts}/{max_offer_retries} fresh IPs all rejected — giving up this visit",
                        )
                        _final_entry = {
                            "visit_index": i + 1,
                            "status": "skipped_no_unique_ip",
                            "proxy": "",
                            "exit_ip": "",
                            "country": "",
                            "city": "",
                            "timezone": "",
                            "locale": "",
                            "os": "",
                            "ua": "",
                            "viewport": "",
                            "device_name": "",
                            "http_status": "",
                            "final_url": "",
                            "landing_url": "",
                            "conversion_page_reached": False,
                            "trusted_form": "",
                            "lead_id": "",
                            "error": (
                                f"Offer-site rejected {retry_attempts} fresh ProxyJet IPs "
                                f"as duplicate/VPN (last: {_ob.burnt_ip or '?'})"
                            ),
                            "screenshot": "",
                            "timestamp": datetime.now(timezone.utc).isoformat(),
                        }
                        await _record(job_id, _final_entry, report, report_lock, db)
                        return
                    push_live_step(
                        job_id, i + 1, "filter", "info",
                        f"Offer-side block · burnt {_ob.burnt_ip or '?'} · retrying with fresh IP "
                        f"(attempt {retry_attempts + 1}/{max_offer_retries})",
                    )
                    # Short breather before next attempt — gives the burnt-IP
                    # persistence task a moment to flush so future probes
                    # see it in rut_burnt_ips.
                    await asyncio.sleep(0.5)
                    if cancel_event.is_set() or target_drain_event.is_set():
                        return
                    continue

    # ── Launch ONE shared Chromium browser for the WHOLE job ─────────
    # All visits create their own isolated BrowserContext from this single
    # browser. This is the standard anti-detection pattern (used by
    # Multilogin/GoLogin/AdsPower under the hood) — every context has its
    # own cookies, storage, proxy, fingerprint and is undetectable from
    # the website's side. RAM cost drops 5-10x vs. per-visit launches,
    # which lets concurrency=15 run safely in the pod.
    # NOTE: The browser + Playwright handle are stashed in `_browser_holder`
    # so `_get_live_browser()` can transparently relaunch Chromium if it
    # ever crashes mid-job (prevents the TargetClosedError death-spiral).
    pw_cm = async_playwright()
    pw = await pw_cm.__aenter__()
    _browser_holder["pw"] = pw
    _browser_holder["pw_cm"] = pw_cm
    shared_browser: Optional[Browser] = None
    try:
        shared_browser = await _launch_anti_detect_browser(pw)
        _browser_holder["b"] = shared_browser
    except Exception as e:
        try:
            await pw_cm.__aexit__(type(e), e, None)
        except Exception:
            pass
        await _finalise_and_persist(db, job_id, "failed",
                  f"Playwright browser launch failed: {type(e).__name__}: {str(e)[:160]}")
        return
    push_live_step(job_id, 0, "preflight", "ok",
                   f"Shared Chromium ready · concurrency={conc}")

    # ── Dispatcher ──────────────────────────────────────────────────
    # Two modes:
    #   clicks:       run exactly `total` visits (legacy behaviour)
    #   conversions:  keep spawning visits until `target_conversions` hit OR
    #                 `max_attempts` reached. Respects `concurrency`.
    if (
        target_mode == "conversions"
        and int(target_conversions or 0) > 0
    ):
        target_conv = int(target_conversions)
        max_att = int(max_attempts or 0)
        if max_att <= 0:
            max_att = max(target_conv * 20, target_conv + 50)  # safety default
        RUT_JOBS[job_id]["max_attempts"] = max_att
        RUT_JOBS[job_id]["total"] = max_att  # UI progress bar denominator
        # Ensure tunnel_fail_count is initialised so reads never KeyError.
        RUT_JOBS[job_id].setdefault("tunnel_fail_count", 0)

        # ── 2026-01: Tunnel-fail-aware budget ────────────────────────────
        # Tunnel/proxy-block failures (ERR_TUNNEL_*, 502 Bad Gateway from
        # proxy gateway, etc.) are typically caused by the proxy provider
        # — not by the visit logic — and they don't represent a real
        # visit attempt against the target site. Counting them in the
        # max_attempts budget means a flaky proxy pool can prematurely
        # end the run with 0 real visits.
        #
        # New default behaviour: tunnel failures are FREE — they don't
        # decrement the budget, so the dispatcher keeps trying until
        # either a real visit hits the budget cap, the proxy pool truly
        # exhausts ("No more proxies available"), the conversion target
        # is hit, or the absolute safety cap (HARD_CAP) below is reached.
        #
        # Set RUT_TUNNEL_FAIL_COUNTS_IN_BUDGET=true to revert to the
        # legacy "every attempt counts" behaviour.
        _tunnel_counts_in_budget = os.environ.get(
            "RUT_TUNNEL_FAIL_COUNTS_IN_BUDGET", "false"
        ).lower() == "true"
        # Absolute safety cap to prevent runaway jobs if every single
        # spawn hits a tunnel failure (e.g. proxy provider hard-blocks
        # the entire domain). Configurable via RUT_HARD_CAP_MULTIPLIER
        # (default 10× max_att, floor 1000).
        _hard_cap_mult = max(int(os.environ.get("RUT_HARD_CAP_MULTIPLIER", "10")), 1)
        HARD_CAP = max(max_att * _hard_cap_mult, 1000)

        attempt_counter = 0
        in_flight: set = set()
        try:
            while True:
                # HARD cancel (user pressed Stop) → break immediately;
                # `await asyncio.gather(*in_flight)` below still waits for
                # in-flight visits to either short-circuit or complete.
                if cancel_event.is_set():
                    break
                cur_conv = int(RUT_JOBS[job_id].get("conversions") or 0)
                if cur_conv >= target_conv:
                    # SOFT drain: stop spawning new visits but let any
                    # in-flight visit run to completion. The proxies/UAs/
                    # leads they already picked up won't be wasted.
                    if not target_drain_event.is_set():
                        target_drain_event.set()
                        RUT_JOBS[job_id]["target_reached"] = True
                        RUT_JOBS[job_id]["target_reached_at"] = datetime.now(timezone.utc).isoformat()
                        push_live_step(
                            job_id, 0, "done", "ok",
                            f"🎯 Target {target_conv} conversions reached — no new visits will start. Draining {len(in_flight)} in-flight visit(s) so their leads aren't wasted…",
                        )
                    break
                # Compute effective attempts — tunnel failures are excluded
                # from the budget by default (see _tunnel_counts_in_budget).
                # 2026-05: also exclude silent-skipped burnt-IP visits so
                # ProxyJet Auto users don't burn through max_attempts on
                # offer/tracker duplicate-IP blocks.
                _tunnel_fails_so_far = int(RUT_JOBS[job_id].get("tunnel_fail_count", 0) or 0)
                _silent_skips_so_far = int(RUT_JOBS[job_id].get("silent_skip_count", 0) or 0)
                _effective_attempts = (
                    attempt_counter
                    if _tunnel_counts_in_budget
                    else max(attempt_counter - _tunnel_fails_so_far - _silent_skips_so_far, 0)
                )
                if _effective_attempts >= max_att:
                    push_live_step(
                        job_id, 0, "done", "info",
                        f"Max {max_att} attempts exhausted — stopping (conversions: {cur_conv}/{target_conv})",
                    )
                    break
                # Absolute runaway-protection cap.
                if attempt_counter >= HARD_CAP:
                    push_live_step(
                        job_id, 0, "done", "info",
                        f"Hard cap {HARD_CAP} attempts reached — stopping "
                        f"(conversions: {cur_conv}/{target_conv}, tunnel_fails: {_tunnel_fails_so_far}). "
                        "Most attempts failed at the proxy tunnel — check your proxy provider.",
                    )
                    break

                # Fill the pool up to `concurrency` in-flight visits
                while (
                    len(in_flight) < conc
                    and _effective_attempts < max_att
                    and attempt_counter < HARD_CAP
                    and not cancel_event.is_set()
                    and not target_drain_event.is_set()
                ):
                    # 2026-05 — Route through worker() instead of
                    # process_one() so the _OfferBlockRetryNeeded retry
                    # loop (added in commit ad037b2) actually fires in
                    # conversions mode too. Without this wrapper the
                    # retry exception was leaking out of the task and
                    # quietly burning the max_attempts budget.
                    t = asyncio.create_task(worker(attempt_counter, shared_browser))
                    in_flight.add(t)
                    t.add_done_callback(in_flight.discard)
                    _register_visit_task(attempt_counter, t)  # 2026-05 manual-kill registry
                    attempt_counter += 1
                    # Re-compute effective attempts inside the spawn loop
                    # too — otherwise we'd over-spawn when a burst of
                    # tunnel-fail entries lands between iterations.
                    _tunnel_fails_so_far = int(RUT_JOBS[job_id].get("tunnel_fail_count", 0) or 0)
                    _silent_skips_so_far = int(RUT_JOBS[job_id].get("silent_skip_count", 0) or 0)
                    _effective_attempts = (
                        attempt_counter
                        if _tunnel_counts_in_budget
                        else max(attempt_counter - _tunnel_fails_so_far - _silent_skips_so_far, 0)
                    )

                if not in_flight:
                    await asyncio.sleep(0.2)
                    continue

                # Wait for any visit to finish before re-evaluating target
                await asyncio.wait(
                    in_flight,
                    return_when=asyncio.FIRST_COMPLETED,
                    timeout=1.5,
                )
        except Exception as e:
            logger.warning(f"RUT conversions-mode dispatcher error: {e}")

        # Graceful finish: let in-flight visits complete naturally. We do
        # NOT cancel them here — only `cancel_event` (hard stop) can cause
        # early termination, and that has already been handled inside
        # process_one. If only `target_drain_event` is set, every in-flight
        # visit runs to its natural conclusion below.
        if in_flight:
            await asyncio.gather(*in_flight, return_exceptions=True)

        # Update total to reflect actual attempts launched
        RUT_JOBS[job_id]["total"] = attempt_counter
    else:
        # ── 2026-05: Budget-based clicks-mode dispatcher ──────────────
        # Legacy behaviour was a flat `gather([process_one(i) for i in
        # range(total)])` — exactly `total` spawns, no retries. That
        # broke when ProxyJet Auto kept handing us IPs that the
        # offer/tracker rejected as duplicate or VPN: the user asked
        # for 5 clean clicks and got 5 attempts (44 silent skips +
        # 5 visible).
        #
        # New behaviour (only when `silent_skip_burnt_ip` is on, which
        # is auto-set by ProxyJet Auto Mode):
        # spawn visits in a while-loop, evaluate after each completion
        # whether VISIBLE `processed` (which now EXCLUDES silent skips)
        # has reached `total`, and keep going up to a HARD_CAP. This
        # guarantees the user gets exactly `total` UI-visible visits,
        # however many burnt IPs ProxyJet returns under the hood.
        _silent_mode = bool(RUT_JOBS[job_id].get("silent_skip_burnt_ip"))
        if not _silent_mode:
            # Legacy fixed-size gather — unchanged behaviour for jobs
            # that don't use ProxyJet Auto (uploaded-proxy mode etc.).
            tasks = [asyncio.create_task(worker(i, shared_browser)) for i in range(total)]
            # 2026-05 — manual-kill registry (legacy fixed-size gather path)
            for _i, _t in enumerate(tasks):
                _register_visit_task(_i, _t)
            try:
                await asyncio.gather(*tasks, return_exceptions=True)
            except Exception as e:
                logger.warning(f"RUT gather error: {e}")
        else:
            # Budget-based dispatcher: keep spawning until VISIBLE
            # `processed` >= `total` (or HARD_CAP / cancel hit).
            # HARD_CAP intentionally GENEROUS — when ProxyJet returns
            # 90 %+ dirty IPs for the offer, we still need hundreds of
            # attempts to surface `total` clean visits. The earlier
            # 20× cap (1000 floor) terminated jobs at ~9 visible
            # visits, which is what the user reported. New cap is
            # 100× user's target with a 5000 floor, so a 5-click
            # request can absorb up to 5000 silent skips and a
            # 50-click request up to 5000. Configurable via
            # RUT_HARD_CAP_MULTIPLIER for users who want tighter or
            # looser bounds.
            _hard_cap_mult = max(int(os.environ.get("RUT_HARD_CAP_MULTIPLIER", "100")), 1)
            HARD_CAP = max(total * _hard_cap_mult, 5000)
            RUT_JOBS[job_id]["hard_cap"] = HARD_CAP
            # Secondary early-exit: if we've spawned ≥150 attempts
            # AND the visible-success ratio is <2 % (i.e. the
            # ProxyJet pool is genuinely exhausted for this offer,
            # not just unlucky), give up gracefully so the user
            # doesn't burn through their full ProxyJet credit on
            # a hopeless run. 150 is enough to be statistically
            # confident — 0/150 = exhausted, 3/150 = 2 % = keep going.
            _MIN_ATTEMPTS_FOR_RATIO_CHECK = 150
            _MIN_PROGRESS_RATIO = 0.02
            attempt_counter = 0
            in_flight: set = set()
            try:
                while True:
                    if cancel_event.is_set() or target_drain_event.is_set():
                        break
                    visible_processed = int(RUT_JOBS[job_id].get("processed") or 0)
                    if visible_processed >= total:
                        # Target reached — let in-flight finish via the
                        # gather below.
                        break
                    if attempt_counter >= HARD_CAP:
                        _silent = int(RUT_JOBS[job_id].get("silent_skip_count") or 0)
                        push_live_step(
                            job_id, 0, "done", "info",
                            f"Hard cap {HARD_CAP} attempts reached — stopping "
                            f"(visible: {visible_processed}/{total}, silently skipped: {_silent}). "
                            "Most ProxyJet IPs were rejected by the tracker — try a different state or refresh the ProxyJet pool.",
                        )
                        break
                    # ProxyJet-pool-exhausted early exit
                    if (
                        attempt_counter >= _MIN_ATTEMPTS_FOR_RATIO_CHECK
                        and visible_processed < int(attempt_counter * _MIN_PROGRESS_RATIO)
                    ):
                        _silent = int(RUT_JOBS[job_id].get("silent_skip_count") or 0)
                        push_live_step(
                            job_id, 0, "done", "info",
                            f"ProxyJet pool exhausted for this offer — stopping at attempt {attempt_counter} "
                            f"(visible: {visible_processed}/{total}, silently skipped: {_silent}). "
                            "Less than 2 % of attempts produced a visible visit — try a different US state, "
                            "refresh your ProxyJet pool, or try again later when more IPs rotate in.",
                        )
                        break
                    # Spawn up to concurrency
                    while (
                        len(in_flight) < conc
                        and attempt_counter < HARD_CAP
                        and not cancel_event.is_set()
                        and not target_drain_event.is_set()
                    ):
                        # 2026-05 — Route through worker() so the
                        # _OfferBlockRetryNeeded retry loop fires here
                        # too (silent clicks-mode dispatcher). Without
                        # the wrapper, _OfferBlockRetryNeeded exceptions
                        # leak out of process_one and corrupt the
                        # silent-skip accounting.
                        t = asyncio.create_task(worker(attempt_counter, shared_browser))
                        in_flight.add(t)
                        t.add_done_callback(in_flight.discard)
                        _register_visit_task(attempt_counter, t)  # 2026-05 manual-kill registry
                        attempt_counter += 1
                    if not in_flight:
                        await asyncio.sleep(0.2)
                        continue
                    # Wait for any visit to finish before re-evaluating
                    await asyncio.wait(
                        in_flight,
                        return_when=asyncio.FIRST_COMPLETED,
                        timeout=1.5,
                    )
            except Exception as e:
                logger.warning(f"RUT clicks-mode silent dispatcher error: {e}")
            # Graceful drain — let in-flight visits finish naturally.
            if in_flight:
                await asyncio.gather(*in_flight, return_exceptions=True)
            # Reflect actual spawn count for diagnostics (does NOT
            # change the UI's TOTAL field, which stays at `total`).
            RUT_JOBS[job_id]["attempts_spawned"] = attempt_counter

    # ── Close the shared browser & playwright runtime ────────────────
    # Prefer the holder's current browser (may have been relaunched mid-job)
    # over the original `shared_browser` variable.
    final_browser = _browser_holder.get("b") or shared_browser
    try:
        if final_browser is not None:
            await final_browser.close()
            logger.info(f"Job {job_id}: Shared browser closed successfully")
    except Exception as e:
        logger.debug(f"shared browser close failed: {e}")
    
    # Force cleanup
    import gc
    gc.collect()
    
    # Use holder's pw_cm if present (in case _get_live_browser created a
    # new one after the original pw_cm was discarded).
    final_pw_cm = _browser_holder.get("pw_cm") or pw_cm
    try:
        await final_pw_cm.__aexit__(None, None, None)
    except Exception as e:
        logger.debug(f"playwright runtime exit failed: {e}")
    _browser_holder["b"] = None
    _browser_holder["pw"] = None
    _browser_holder["pw_cm"] = None

    # Remember whether Stop was pressed — used later to set final status
    was_cancelled = cancel_event.is_set()

    # Build Excel report + leftover leads + ZIP
    try:
        _write_excel_report(job_dir / "report.xlsx", report)
    except Exception as e:
        logger.warning(f"Excel report failed: {e}")

    # Write leads_with_status.xlsx — ORIGINAL schema + one extra "status" column
    # marking each row as "used" (submit OK), "invalid" (server-side validation
    # rejected), or "not_used". Color-coded:
    #   green  = used        (remove from pending)
    #   red    = invalid     (remove from pending)
    #   orange = not_used    (keeps — goes into pending_leads.xlsx)
    status_path = None
    pending_path = None
    if rows:
        # Build union of column keys across ALL rows (not just rows[0]) so
        # sparse Excel uploads don't silently lose columns.
        seen = set()
        orig_cols: List[str] = []
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    seen.add(k)
                    orig_cols.append(k)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter

            status_path = job_dir / "leads_with_status.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            headers = orig_cols + ["status"]
            ws.append(headers)
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="374151")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
            # Fills
            used_fill = PatternFill("solid", fgColor="C6EFCE")      # light green
            invalid_fill = PatternFill("solid", fgColor="FFC7CE")   # light red
            unused_fill = PatternFill("solid", fgColor="FFE699")    # light orange
            for idx, r in enumerate(rows):
                if idx in consumed_row_indices:
                    status_val, fill = "used", used_fill
                elif idx in invalid_row_indices:
                    status_val, fill = "invalid", invalid_fill
                else:
                    status_val, fill = "not_used", unused_fill
                row_vals = [r.get(c, "") for c in orig_cols] + [status_val]
                ws.append(row_vals)
                excel_row = idx + 2  # +1 header, +1 one-indexed
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=excel_row, column=col_idx).fill = fill
            # Auto-size columns (approximate)
            for col_idx, col_name in enumerate(headers, start=1):
                max_len = max(
                    [len(str(col_name))]
                    + [len(str(r.get(col_name, ""))) for r in rows[:50]]
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
            wb.save(status_path)
        except Exception as e:
            logger.warning(f"leads_with_status.xlsx write failed: {e}")
            status_path = None

        # ── pending_leads.xlsx — ONLY the unused rows, ready for next run ──
        # Identical schema to the uploaded lead file (no extra columns), so the
        # user can re-upload this file directly as the next run's data source.
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            pending_path = job_dir / "pending_leads.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Pending Leads"
            ws.append(orig_cols)
            # Header styling
            try:
                from openpyxl.styles import PatternFill, Font
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="374151")
                for col_idx in range(1, len(orig_cols) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
            except Exception:
                pass

            pending_count = 0
            for idx, r in enumerate(rows):
                if idx in consumed_row_indices or idx in invalid_row_indices:
                    continue
                ws.append([r.get(c, "") for c in orig_cols])
                pending_count += 1

            for col_idx, col_name in enumerate(orig_cols, start=1):
                max_len = max(
                    [len(str(col_name))]
                    + [len(str(r.get(col_name, ""))) for r in rows[:50]]
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

            wb.save(pending_path)
            RUT_JOBS[job_id]["pending_leads_count"] = pending_count
            RUT_JOBS[job_id]["pending_leads_path"] = str(pending_path)
        except Exception as e:
            logger.warning(f"pending_leads.xlsx write failed: {e}")
            pending_path = None

    zip_path = job_dir / "results.zip"
    try:
        # Categorise every completed visit so the operator can audit each
        # bucket separately inside the downloaded ZIP:
        #   Processed/   → every visit (raw, unfiltered)
        #   Succeeded/   → status == "ok" (form submitted cleanly)
        #   Conversions/ → conversion_page_reached == True (hit final/thank-you page)
        #   Leads_Left/  → rows NOT consumed (pending_leads.xlsx + full leads_with_status.xlsx)
        # Each bucket gets its own `report.xlsx` + matching `screenshots/`
        # subfolder (only the screenshots that belong to the visits in
        # that bucket), so the user doesn't have to manually sift through
        # one giant flat list.
        def _bucket_visits(predicate) -> List[Dict[str, Any]]:
            return [e for e in report if predicate(e)]

        succeeded_visits = _bucket_visits(lambda e: str(e.get("status", "")) == "ok")
        conversion_visits = _bucket_visits(lambda e: bool(e.get("conversion_page_reached")))

        def _shots_for_visits(visits: List[Dict[str, Any]]) -> List[Path]:
            idxs = {int(v.get("visit_index") or 0) for v in visits if v.get("visit_index") is not None}
            if not idxs:
                return []
            out: List[Path] = []
            for p in shots_dir.glob("*.png"):
                # filenames look like visit_00001.png / visit_00001_thankyou.png /
                # visit_00001_capture02_xyz.png — extract the leading numeric id.
                try:
                    name = p.name
                    if not name.startswith("visit_"):
                        continue
                    num = int(name[6:11])  # 5-digit zero-padded
                    if num in idxs:
                        out.append(p)
                except Exception:
                    continue
            return out

        # Build per-bucket Excel reports (filtered subset of `report`).
        processed_report_path = job_dir / "_bucket_processed_report.xlsx"
        succeeded_report_path = job_dir / "_bucket_succeeded_report.xlsx"
        conversions_report_path = job_dir / "_bucket_conversions_report.xlsx"
        try:
            _write_excel_report(processed_report_path, report)
        except Exception as _be:
            logger.debug(f"processed bucket report failed: {_be}")
            processed_report_path = None
        try:
            _write_excel_report(succeeded_report_path, succeeded_visits)
        except Exception as _be:
            logger.debug(f"succeeded bucket report failed: {_be}")
            succeeded_report_path = None
        try:
            _write_excel_report(conversions_report_path, conversion_visits)
        except Exception as _be:
            logger.debug(f"conversions bucket report failed: {_be}")
            conversions_report_path = None

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # ── 1. Top-level legacy artefacts (kept for backward compat) ──
            for p in shots_dir.glob("*.png"):
                zf.write(p, arcname=f"screenshots/{p.name}")
            if (job_dir / "report.xlsx").exists():
                zf.write(job_dir / "report.xlsx", arcname="report.xlsx")
            if status_path and status_path.exists():
                zf.write(status_path, arcname="leads_with_status.xlsx")
            if pending_path and pending_path.exists():
                zf.write(pending_path, arcname="pending_leads.xlsx")

            # ── 2. Processed/ (every visit) ──
            if processed_report_path and processed_report_path.exists():
                zf.write(processed_report_path, arcname="Processed/report.xlsx")
            for p in shots_dir.glob("*.png"):
                zf.write(p, arcname=f"Processed/screenshots/{p.name}")

            # ── 3. Succeeded/ (status == ok) ──
            if succeeded_report_path and succeeded_report_path.exists():
                zf.write(succeeded_report_path, arcname="Succeeded/report.xlsx")
            for p in _shots_for_visits(succeeded_visits):
                zf.write(p, arcname=f"Succeeded/screenshots/{p.name}")

            # ── 4. Conversions/ (conversion_page_reached == True) ──
            if conversions_report_path and conversions_report_path.exists():
                zf.write(conversions_report_path, arcname="Conversions/report.xlsx")
            for p in _shots_for_visits(conversion_visits):
                zf.write(p, arcname=f"Conversions/screenshots/{p.name}")

            # ── 5. Leads_Left/ (rows still pending + full status xlsx) ──
            if pending_path and pending_path.exists():
                zf.write(pending_path, arcname="Leads_Left/pending_leads.xlsx")
            if status_path and status_path.exists():
                zf.write(status_path, arcname="Leads_Left/leads_with_status.xlsx")
    except Exception as e:
        logger.warning(f"zip build failed: {e}")
    finally:
        # Clean up the temporary per-bucket Excel files (they're already
        # inside the zip; no need to leave them in the job dir).
        for _tmp in (
            job_dir / "_bucket_processed_report.xlsx",
            job_dir / "_bucket_succeeded_report.xlsx",
            job_dir / "_bucket_conversions_report.xlsx",
        ):
            try:
                if _tmp.exists():
                    _tmp.unlink()
            except Exception:
                pass

    # ── Await all pending live-remove tasks BEFORE finalising ────────
    # The fire-and-forget _spawn_live() calls scheduled per-visit $pull /
    # XLSX-rewrite tasks. If we finalise the job before they complete the
    # LAST visit's deletion is sometimes lost (testing agent caught this:
    # consumed_count ended at N-1 instead of N). Drain the queue here so
    # uploaded_resources reflects the FULL set of consumed items by the
    # time the job is marked completed.
    logger.info(
        f"RUT job {job_id}: draining {len(_live_pending_tasks)} live-remove "
        f"tasks (proxy={len(_live_proxy_pulled)} ua={len(_live_ua_pulled)})"
    )
    if _live_pending_tasks:
        try:
            results = await asyncio.wait_for(
                asyncio.gather(*_live_pending_tasks, return_exceptions=True),
                timeout=30.0,
            )
            errs = [r for r in results if isinstance(r, Exception)]
            if errs:
                logger.warning(f"RUT job {job_id}: {len(errs)} live-remove tasks raised: {errs[:3]}")
        except asyncio.TimeoutError:
            logger.warning(
                f"RUT job {job_id}: {len(_live_pending_tasks)} live-remove tasks "
                f"didn't finish in 30s — proceeding with finalise anyway"
            )
        except Exception as e:
            logger.debug(f"live-remove drain error: {e}")

    # Engine self-diagnosis: if a job ends with 0 visits, the user has
    # NO useful information from the existing fields. Capture the actual
    # reason here so the frontend's View modal + Past Jobs list can
    # surface a one-line explanation. This is the difference between
    # "stopped 0/100 ¯\_(ツ)_/¯" and "stopped: cancelled by user 3s into
    # the run — no proxies were tried yet" (or whatever actually happened).
    final_status = "stopped" if was_cancelled else "completed"
    self_diagnosis = ""
    processed = int(RUT_JOBS.get(job_id, {}).get("processed") or 0)
    if processed == 0:
        if was_cancelled:
            elapsed = 0
            try:
                started_at_iso = RUT_JOBS[job_id].get("started_at") or RUT_JOBS[job_id].get("created_at")
                if started_at_iso:
                    elapsed = max(0, int((datetime.now(timezone.utc) - datetime.fromisoformat(started_at_iso.replace("Z","+00:00"))).total_seconds()))
            except Exception:
                pass
            cancel_at = RUT_JOBS[job_id].get("cancel_requested_at") or "(unknown)"
            self_diagnosis = (
                f"Engine cancelled before any visit completed (elapsed ~{elapsed}s, "
                f"cancel signal received at {cancel_at}). Possible causes: "
                f"(1) you clicked Stop too soon after Submit; "
                f"(2) backend container restarted mid-run — check `docker logs krexion-backend` "
                f"for OOMKilled / SIGTERM; "
                f"(3) Playwright chromium failed to launch — check the same log for 'browser' / "
                f"'chromium' errors; "
                f"(4) Cloudflare tunnel dropped, your browser fired a stop on the recovery path."
            )
        else:
            # Engine ran to completion but did 0 visits (very unusual —
            # would mean total_clicks loop or target_conversions/max_attempts
            # math hit 0). Record an explicit diagnosis.
            self_diagnosis = (
                f"Engine completed normally but processed 0 visits. "
                f"target_mode={target_mode}, total_clicks={total_clicks}, "
                f"max_attempts={max_attempts}, target_conversions={target_conversions}. "
                f"Likely an empty proxy/UA list slipped through validation, "
                f"or the inner loop bailed early. Check `docker logs krexion-backend` "
                f"for the full engine trace."
            )

    RUT_JOBS[job_id].update({
        "status": final_status,
        "finished_at": datetime.now(timezone.utc).isoformat(),
        "report": report[-200:],  # keep last 200 in memory
        "zip_path": str(zip_path),
        "leftover_leads_count": (len(rows) - len(consumed_row_indices) - len(invalid_row_indices)) if rows else 0,
        "consumed_leads_count": len(consumed_row_indices),
        "invalid_leads_count": len(invalid_row_indices),
        # Self-diagnosis — empty when the run produced visits.
        "diagnosis": self_diagnosis,
        # Tracked so the post-finish upload-consume hook can selectively
        # prune ONLY the proxies / UAs actually used during this run from
        # the saved upload batches (not the entire batch).
        "used_proxy_raws": list(used_proxy_set),
        "used_ua_strings": list(used_ua_set),
    })
    # Remove the non-serializable asyncio.Event before any DB persist
    RUT_JOBS[job_id].pop("_cancel_event", None)
    RUT_JOBS[job_id].pop("_target_drain_event", None)

    # ── Auto-consume any "Uploaded Things" batches BEFORE persisting ─
    # The live-remove tasks above pull each used proxy / UA / row in
    # real-time. This batched consume is now a SAFETY-NET that mops up
    # anything the live path missed (e.g. a $pull that raced with
    # auto-delete of an empty batch). We run it BEFORE _persist so that
    # by the time the API reports status=completed, the upload doc is
    # already at its final shape — frontend / tests / users will not see
    # a stale snapshot during the brief window between persist and consume.
    consume_upload_ids: List[str] = []
    if db is not None:
        try:
            jr = await db.real_user_traffic_jobs.find_one(
                {"job_id": job_id},
                {"_id": 0, "consume_upload_ids": 1, "user_id": 1, "pending_leads_path": 1},
            )
            if jr:
                consume_upload_ids = jr.get("consume_upload_ids") or []
                uid = jr.get("user_id")
                if consume_upload_ids and uid:
                    try:
                        from server import _consume_uploads
                        await _consume_uploads(
                            uid,
                            consume_upload_ids,
                            used_proxy_raws=list(used_proxy_set),
                            used_ua_strings=list(used_ua_set),
                            pending_leads_path=jr.get("pending_leads_path") or "",
                        )
                        logger.info(
                            f"RUT job {job_id}: pruned {len(consume_upload_ids)} uploaded batch(es) — "
                            f"removed {len(used_proxy_set)} used proxies, "
                            f"{len(used_ua_set)} used UAs (live + safety-net pass)"
                        )
                    except Exception as e:
                        logger.warning(f"RUT job {job_id}: upload consume failed: {e}")
        except Exception as e:
            logger.warning(f"RUT job {job_id}: pre-persist consume hook failed: {e}")

    if db is not None:
        # Mark consume IDs as cleared so we don't double-process later.
        if consume_upload_ids:
            RUT_JOBS[job_id]["consume_upload_ids"] = []
            RUT_JOBS[job_id]["consumed_upload_ids_final"] = consume_upload_ids
        await _persist(db, job_id)


# ──────────────────────────────────────────────────────────────────
# Custom Automation JSON executor
# ──────────────────────────────────────────────────────────────────
# Supported actions: goto, click, fill, select, check, uncheck, press, wait,
# wait_for_selector, wait_for_navigation, scroll, screenshot, evaluate.
# Every step can take: selector, value, ms, timeout, optional, wait_nav.
# Placeholders in `value`:
#   {{row.FIELD}}  or  {{FIELD}}   → Excel row value (case-insensitive)
#   {{random.N}}                    → N-digit random number
#   {{randomletters.N}}             → N random letters
# ──────────────────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────────────────
# 2026-01: Column synonym map for {{placeholder}} substitution
# ──────────────────────────────────────────────────────────────────
# Visual Recorder users sometimes bind a field to a column name that
# doesn't EXACTLY match the data file column (e.g. recorder used
# `{{cellphone}}` but Excel has `phone`). Previously this left the
# form field empty → form validation failed → CONTINUE didn't advance
# → conversion lost. We now do a 2-stage lookup:
#   1. Exact case-insensitive match (existing behaviour, unchanged)
#   2. Synonym fallback — common field name aliases
# Pure additive: if exact match succeeds we return it; synonym lookup
# only fires when exact lookup misses. Production behaviour for
# matching column names is therefore identical.
_PLACEHOLDER_SYNONYMS: Dict[str, List[str]] = {
    # phone
    "phone":      ["cellphone", "cell_phone", "cell", "mobile", "mobilephone", "mobile_phone", "telephone", "tel", "phonenumber", "phone_number"],
    "cellphone":  ["phone", "cell", "mobile", "mobilephone", "telephone", "tel", "phonenumber", "phone_number", "cell_phone"],
    "cell":       ["phone", "cellphone", "mobile", "tel", "phonenumber", "phone_number"],
    "mobile":     ["phone", "cellphone", "cell", "tel", "phonenumber", "phone_number"],
    "telephone":  ["phone", "cellphone", "cell", "mobile", "tel"],
    "tel":        ["phone", "cellphone", "telephone", "cell", "mobile"],
    "phonenumber": ["phone", "cellphone", "cell", "mobile", "tel"],
    "phone_number": ["phone", "cellphone", "cell", "mobile", "tel"],
    # name
    "first":      ["firstname", "first_name", "fname", "givenname", "given_name"],
    "firstname":  ["first", "first_name", "fname", "givenname"],
    "first_name": ["first", "firstname", "fname", "givenname"],
    "fname":      ["first", "firstname", "first_name", "givenname"],
    "last":       ["lastname", "last_name", "lname", "surname", "familyname", "family_name"],
    "lastname":   ["last", "last_name", "lname", "surname", "familyname"],
    "last_name":  ["last", "lastname", "lname", "surname"],
    "lname":      ["last", "lastname", "last_name", "surname"],
    "surname":    ["last", "lastname", "last_name"],
    "fullname":   ["full_name", "name", "fname", "firstname"],
    # address
    "address":    ["street", "streetaddress", "street_address", "address1", "addr", "address_line_1", "addressline1"],
    "street":     ["address", "streetaddress", "street_address", "address1"],
    "streetaddress": ["address", "street", "street_address", "address1"],
    "address1":   ["address", "street", "streetaddress", "address_line_1"],
    "address2":   ["apt", "unit", "suite", "address_line_2", "addressline2"],
    "apt":        ["address2", "unit", "suite"],
    # city / state / zip
    "city":       ["town", "locality"],
    "state":      ["region", "province", "state_code", "statecode"],
    "zip":        ["zipcode", "zip_code", "postal", "postalcode", "postal_code", "postcode"],
    "zipcode":    ["zip", "zip_code", "postal", "postalcode", "postal_code", "postcode"],
    "zip_code":   ["zip", "zipcode", "postal", "postalcode", "postal_code", "postcode"],
    "postal":     ["zip", "zipcode", "postal_code", "postalcode", "postcode"],
    "postalcode": ["zip", "zipcode", "postal", "postal_code", "postcode"],
    "postcode":   ["zip", "zipcode", "postal", "postalcode", "postal_code"],
    # email
    "email":      ["emailaddress", "email_address", "mail", "e_mail", "e-mail"],
    "emailaddress": ["email", "email_address", "mail"],
    "email_address": ["email", "emailaddress", "mail"],
    # DOB
    "day":        ["birth_day", "birthday_day", "dob_day", "dday", "bday", "birth_d"],
    "birth_day":  ["day", "birthday_day", "dob_day", "dday", "bday"],
    "month":      ["birth_month", "birthmonth", "dob_month", "dmonth", "bmonth", "birth_m"],
    "birth_month": ["month", "birthmonth", "dob_month", "dmonth"],
    "year":       ["birth_year", "birthyear", "dob_year", "dyear", "byear", "birth_y"],
    "birth_year": ["year", "birthyear", "dob_year", "dyear"],
    "dob":        ["birthdate", "birth_date", "date_of_birth"],
    "birthdate":  ["dob", "birth_date", "date_of_birth"],
    # gender
    "gender":     ["sex"],
    "sex":        ["gender"],
}


# ── 2026-06 — Legacy evaluate-script upgrade helpers ──────────────────
# The Visual Recorder emits "click by text" / "random pick by text"
# steps as `action: evaluate` containing a synthetic JS that does
# `el.click()`. Under SPA frameworks (React, Vue) and iframe-based
# offer walls (stacks.app, uplevelrewards, etc.) synthetic clicks
# frequently fail silently — page doesn't navigate, subsequent
# optional steps skip past, and the visit "completes" without
# actually doing anything.
#
# These helpers parse the script's `labels=[...]` (random-pick) or
# `t='...'` (single text click) arrays so the RUT engine can run
# the click via Playwright's NATIVE locator API — which fires real
# pointer/mouse events, walks all frames, and waits for actionable
# state. Backwards-compatible: if extraction fails OR native click
# fails, the engine falls back to running the original JS.

def _extract_random_pick_labels(script: Any) -> Optional[List[str]]:
    """Parse `var labels=['a','b','c']` from a legacy random-pick
    evaluate script. Returns the list of labels or None if pattern
    doesn't match. Handles JS-escaped quotes / backslashes.
    """
    import re as _re
    if not isinstance(script, str):
        return None
    m = _re.search(r"var\s+labels\s*=\s*\[([^\]]*)\]", script)
    if not m:
        return None
    body = m.group(1)
    items = _re.findall(r"'((?:[^'\\]|\\.)*)'", body)
    out: List[str] = []
    for it in items:
        t = it.replace("\\'", "'").replace("\\\\", "\\").strip()
        if t:
            out.append(t)
    return out or None


def _extract_text_click_label(script: Any) -> Optional[str]:
    """Parse `var t='Continue'` from a legacy text-click evaluate
    script. Returns the label or None. Only matches the simple
    single-text-click builder, NOT the random-pick variant (which
    uses `var labels=[...]`).
    """
    import re as _re
    if not isinstance(script, str):
        return None
    # Bail out if this looks like a random-pick script.
    if _re.search(r"var\s+labels\s*=\s*\[", script):
        return None
    m = _re.search(r"var\s+t\s*=\s*'((?:[^'\\]|\\.)*)'", script)
    if not m:
        return None
    raw = m.group(1)
    t = raw.replace("\\'", "'").replace("\\\\", "\\").strip()
    return t or None


async def _native_click_by_text(page: Any, text: str, timeout_ms: int = 8000) -> Tuple[bool, str, str]:
    """Click an element by visible text using Playwright's native
    locator API. Searches the main frame AND every sub-frame
    (iframe / same-origin). Tries role-based locators first
    (button / link → most reliable on SPA pages) then falls back
    to plain `get_by_text`.

    Returns:
        (clicked: bool, frame_url: str, error: str)
    """
    if not isinstance(text, str):
        return False, "", "non-string text"
    text = text.strip()
    if not text:
        return False, "", "empty text"

    last_err = ""
    try:
        frames = list(page.frames)
    except Exception:
        frames = []
    if not frames:
        try:
            frames = [page.main_frame]
        except Exception:
            frames = []

    for frame in frames:
        try:
            frame_url = getattr(frame, "url", "") or ""
        except Exception:
            frame_url = ""

        # Strategy 1: role-based (button / link) — most reliable on SPA
        for role in ("button", "link"):
            try:
                loc = frame.get_by_role(role, name=text).first
                if await loc.count() > 0:
                    try:
                        await loc.scroll_into_view_if_needed(timeout=2000)
                    except Exception:
                        pass
                    try:
                        await loc.click(timeout=timeout_ms)
                        return True, frame_url, ""
                    except Exception as e:  # noqa: BLE001
                        last_err = f"role={role}: {type(e).__name__}: {str(e)[:90]}"
            except Exception as e:  # noqa: BLE001
                last_err = f"role={role}: {type(e).__name__}: {str(e)[:90]}"

        # Strategy 2: plain text locator (fuzzy)
        try:
            loc = frame.get_by_text(text, exact=False).first
            if await loc.count() > 0:
                try:
                    await loc.scroll_into_view_if_needed(timeout=2000)
                except Exception:
                    pass
                try:
                    await loc.click(timeout=timeout_ms)
                    return True, frame_url, ""
                except Exception as e:  # noqa: BLE001
                    last_err = f"text: {type(e).__name__}: {str(e)[:90]}"
        except Exception as e:  # noqa: BLE001
            last_err = f"text: {type(e).__name__}: {str(e)[:90]}"

    return False, "", last_err or "no match in any frame"




def _substitute(template: str, row: Dict[str, Any]) -> str:
    if not isinstance(template, str):
        return template
    import re
    # Build a normalised lookup dict ONCE per call so synonym lookups
    # are O(1). Lowercase + strip + replace common separators.
    row_norm: Dict[str, Any] = {}
    for k, v in row.items():
        if k is None:
            continue
        nk = str(k).strip().lower().replace("-", "_").replace(" ", "_")
        if v is not None and str(v).strip() != "":
            row_norm[nk] = v

    def repl(m):
        raw_inside = m.group(1).strip()
        # 2026-01: formatter pipeline — split `{{key|fmt1|fmt2:arg}}` into
        # key + pipeline. Pipeline is applied AFTER key resolution. When
        # pipeline is empty (legacy `{{key}}`), behaviour unchanged.
        if _EXT_LOADED and "|" in raw_inside:
            key, pipeline = _ext_split_pipeline(raw_inside)
        else:
            key, pipeline = raw_inside, ""

        def _resolve(key: str) -> str:
            if key.lower().startswith("row."):
                key = key[4:]
            if key.lower().startswith("random."):
                try:
                    n = int(key.split(".", 1)[1])
                    return "".join(random.choice("0123456789") for _ in range(max(1, n)))
                except Exception:
                    return ""
            if key.lower().startswith("randomletters."):
                try:
                    n = int(key.split(".", 1)[1])
                    import string
                    return "".join(random.choice(string.ascii_lowercase) for _ in range(max(1, n)))
                except Exception:
                    return ""
            # 1. Exact case-insensitive row lookup (legacy behaviour preserved)
            for k, v in row.items():
                if str(k).strip().lower() == key.lower():
                    return "" if v is None else str(v)
            # 2. 2026-01 synonym fallback — common field aliases so
            # `{{cellphone}}` resolves when the data file has `phone`, etc.
            norm_key = key.strip().lower().replace("-", "_").replace(" ", "_")
            # 2a. Direct normalised match (handles hyphen/space differences)
            if norm_key in row_norm:
                v = row_norm[norm_key]
                return "" if v is None else str(v)
            # 2b. Stripped-underscore variant (e.g. "first_name" -> "firstname")
            collapsed = norm_key.replace("_", "")
            for cand_key in row_norm:
                if cand_key.replace("_", "") == collapsed:
                    v = row_norm[cand_key]
                    return "" if v is None else str(v)
            # 2c. Synonym list lookup
            for syn in _PLACEHOLDER_SYNONYMS.get(norm_key, []):
                if syn in row_norm:
                    v = row_norm[syn]
                    return "" if v is None else str(v)
                # Try collapsed variant of each synonym too
                syn_col = syn.replace("_", "")
                for cand_key in row_norm:
                    if cand_key.replace("_", "") == syn_col:
                        v = row_norm[cand_key]
                        return "" if v is None else str(v)
            # 2d. Substring match — last resort (e.g. "homephone" → "phone")
            if len(norm_key) >= 4:
                for cand_key in row_norm:
                    if norm_key in cand_key or cand_key in norm_key:
                        v = row_norm[cand_key]
                        return "" if v is None else str(v)
            return ""

        resolved = _resolve(key)
        # Apply formatter pipeline if present (e.g. `|upper|first:5`)
        if pipeline and _EXT_LOADED:
            try:
                resolved = _ext_apply_formatters(resolved, pipeline)
            except Exception:
                pass
        return resolved
    return re.sub(r"\{\{\s*([^}]+?)\s*\}\}", repl, template)


# ── 2026-05: Field-type-aware selector aliases for form inputs ──
# Why this exists:
#   `_smart_wait_for_selector` already derives generic id/name/token
#   fallbacks (e.g. `#phone` → `[name="phone"]` → `[name*="phone" i]`).
#   That handles RENAMED ids/names — but doesn't help when the page
#   uses a completely different mechanism for the field. Real example
#   from a customer's offer page: the recorded selector `#phone` no
#   longer matches because the page now ONLY exposes the phone input as
#   `<input type="tel" autocomplete="tel">` (no id, no name="phone").
#   Result in RUT replay: phone fill silently skipped (step had
#   `optional: true` set by the recorder) → next step (`select #birth_month`)
#   hits a form whose phone validation hasn't been satisfied → month
#   dropdown never enables → visit fails at step #16.
#   Visual Recorder live-test never hit this because at recording time
#   the user clicked the actual element and the recorder captured its
#   live state. Replay-time DOM divergence is exactly what this map fixes.
#
# Returns a list of CSS selector candidates given a recorded selector's
# "key" (the field name/id substring). Each candidate is a robust
# attribute-based match that doesn't depend on the page's id/name
# choices — they cover `type`, `autocomplete`, `name*=`, `id*=`,
# `placeholder*=`, `aria-label*=` so even fully-renamed pages match.
def _step_fallbacks(step: Any) -> List[str]:
    """2026-05 — Read the `fallbacks` dict embedded in a Visual-Recorder
    step (see visual_recorder._build_fallbacks) and produce a list of
    Playwright-compatible selector alternatives, ordered most-specific
    first.

    Returns [] for steps without a `fallbacks` dict so callers can
    unconditionally concatenate the result with the legacy alts —
    this is what keeps OLD recordings (no fallbacks key) working
    exactly as before. Pure additive, never replaces existing alts.

    Strategy ordering (each entry is a Playwright selector string):
      1. xpath_stable        — survives id/name renames if ANY stable
                               attr was captured on the element or its
                               ancestors.
      2. xpath_abs           — survives selector renames when DOM tree
                               shape matches recording.
      3. attribute combos    — id, name, data-testid, placeholder,
                               aria-label, role — exact + case-insens
                               variants, tag-scoped + tag-free.
      4. text-based scoped   — `button:has-text("Continue")` etc. for
                               clickable/labelled tags.
      5. text= engine match  — Playwright's built-in text engine
                               (case-insensitive substring).

    All entries returned here will be fed into `_smart_wait_for_selector`'s
    `extra_alts` parameter and tried with `state="attached"` (since
    custom-UI dropdowns hide the real <select>; "visible" would fail).
    """
    if not isinstance(step, dict):
        return []
    fb = step.get("fallbacks")
    if not isinstance(fb, dict) or not fb:
        return []
    out: List[str] = []

    # 1. xpath_stable
    xs = (fb.get("xpath") or "").strip()
    if xs:
        out.append(f"xpath={xs}")

    # 2. xpath_abs (only if it differs from xpath_stable)
    xa = (fb.get("xpath_abs") or "").strip()
    if xa and xa != xs:
        out.append(f"xpath={xa}")

    # 3. attribute combos
    attrs = fb.get("attrs") or {}
    tag = (fb.get("tag") or "").lower()
    if isinstance(attrs, dict):
        # Priority attribute list — data-testid/data-cy first because
        # those are explicitly meant to survive refactors.
        attr_priority = (
            "data-testid", "data-test", "data-cy", "data-qa", "data-id",
            "id", "name", "aria-label", "placeholder", "for",
            "role", "type", "autocomplete", "title", "alt",
        )
        for k in attr_priority:
            v = attrs.get(k)
            if not isinstance(v, str) or not v:
                continue
            # Escape any embedded double-quote
            v_esc = v.replace('"', '\\"')
            # Exact match (tag-free + tag-scoped)
            out.append(f'[{k}="{v_esc}"]')
            if tag:
                out.append(f'{tag}[{k}="{v_esc}"]')
            # Case-insensitive substring (rescues partial renames)
            if k in ("id", "name", "placeholder", "aria-label", "title"):
                out.append(f'[{k}*="{v_esc}" i]')
                if tag:
                    out.append(f'{tag}[{k}*="{v_esc}" i]')

    # 4 + 5. Text-based fallbacks (only for clickable / labelled tags;
    # text on a wrapper <div> tends to be too broad and matches the
    # wrong element).
    txt = (fb.get("text") or "").strip()
    if txt and 3 <= len(txt) <= 80:
        # Escape backslashes + double-quotes for Playwright string
        txt_esc = txt.replace('\\', '\\\\').replace('"', '\\"')
        if tag in ("button", "a", "label", "input", "span"):
            out.append(f'{tag}:has-text("{txt_esc}")')
        # Generic text= engine (case-insensitive substring match)
        out.append(f'text="{txt_esc}"')

    # Dedup while preserving order
    seen: set = set()
    uniq: List[str] = []
    for s in out:
        if s and s not in seen:
            seen.add(s)
            uniq.append(s)
    return uniq


def _field_type_alts_for(selector: str) -> List[str]:
    """Return field-type-aware fallback selectors for a recorded form
    selector (e.g. `#phone` → tel/mobile aliases, `#email` → email
    type/autocomplete aliases, etc.). Empty list if the selector
    doesn't look like a common form field."""
    import re as _re_local
    if not selector:
        return []
    sel = selector.strip()
    key = ""
    m = _re_local.match(r"^#([\w\-]+)$", sel)
    if m:
        key = m.group(1)
    else:
        m = _re_local.search(r"\[(?:name|id)\s*=\s*[\"']([^\"']+)[\"']\]", sel)
        if m:
            key = m.group(1)
        else:
            m = _re_local.search(r"\[name\s*=\s*([^\]\s]+)\]", sel)
            if m:
                key = m.group(1).strip("\"'")
    if not key:
        return []
    kl = key.lower()
    # Group-specific aliases. Each group covers all common HTML5 type
    # attributes, autocomplete values, AND wildcard name/id/placeholder
    # matches so we never miss a renamed-but-semantically-identical field.
    if any(p in kl for p in ("phonenumber", "phone_number", "phone", "mobile", "tel", "cell", "contactnumber", "contact_number", "contactphone")):
        return [
            'input[type="tel"]',
            'input[autocomplete="tel"]',
            'input[autocomplete*="phone" i]',
            'input[autocomplete*="tel" i]',
            'input[name*="phone" i]',
            'input[name*="tel" i]',
            'input[name*="mobile" i]',
            'input[name*="cell" i]',
            'input[id*="phone" i]',
            'input[id*="tel" i]',
            'input[id*="mobile" i]',
            'input[placeholder*="phone" i]',
            'input[placeholder*="mobile" i]',
            'input[aria-label*="phone" i]',
            'input[aria-label*="mobile" i]',
        ]
    if any(p in kl for p in ("emailaddress", "email_address", "email", "e_mail", "mail")):
        return [
            'input[type="email"]',
            'input[autocomplete="email"]',
            'input[autocomplete*="email" i]',
            'input[name*="email" i]',
            'input[name*="mail" i]',
            'input[id*="email" i]',
            'input[id*="mail" i]',
            'input[placeholder*="email" i]',
            'input[aria-label*="email" i]',
        ]
    if any(p in kl for p in ("firstname", "first_name", "fname", "first", "givenname", "given")):
        return [
            'input[autocomplete="given-name"]',
            'input[name*="first" i]',
            'input[name*="fname" i]',
            'input[name*="given" i]',
            'input[id*="first" i]',
            'input[id*="fname" i]',
            'input[placeholder*="first" i]',
        ]
    if any(p in kl for p in ("lastname", "last_name", "lname", "last", "surname", "familyname", "family")):
        return [
            'input[autocomplete="family-name"]',
            'input[name*="last" i]',
            'input[name*="lname" i]',
            'input[name*="surname" i]',
            'input[name*="family" i]',
            'input[id*="last" i]',
            'input[id*="lname" i]',
            'input[placeholder*="last" i]',
        ]
    if any(p in kl for p in ("zipcode", "zip_code", "postal", "postcode", "zip", "post_code")):
        return [
            'input[autocomplete="postal-code"]',
            'input[name*="zip" i]',
            'input[name*="postal" i]',
            'input[name*="postcode" i]',
            'input[id*="zip" i]',
            'input[id*="postal" i]',
            'input[placeholder*="zip" i]',
            'input[placeholder*="postal" i]',
        ]
    if any(p in kl for p in ("city", "town", "locality")):
        return [
            'input[autocomplete="address-level2"]',
            'input[name*="city" i]',
            'input[name*="town" i]',
            'input[id*="city" i]',
        ]
    if any(p in kl for p in ("state", "region", "province", "addresslevel1")):
        return [
            'input[autocomplete="address-level1"]',
            'select[autocomplete="address-level1"]',
            'input[name*="state" i]',
            'select[name*="state" i]',
            'input[name*="region" i]',
        ]
    if any(p in kl for p in ("streetaddress", "street_address", "address", "street", "addr", "addressline")):
        return [
            'input[autocomplete="street-address"]',
            'input[autocomplete="address-line1"]',
            'input[name*="address" i]',
            'input[name*="street" i]',
            'input[name*="addr" i]',
            'input[id*="address" i]',
            'input[id*="street" i]',
            'input[placeholder*="address" i]',
            'input[placeholder*="street" i]',
        ]
    if any(p in kl for p in ("dob", "birth", "birthday", "birthdate", "dateofbirth")):
        # Birth-date container/composite; specific month/day/year handled below
        return [
            'input[autocomplete*="bday" i]',
            'select[autocomplete*="bday" i]',
            'input[name*="dob" i]',
            'input[name*="birth" i]',
            'select[name*="birth" i]',
            'input[id*="birth" i]',
            'select[id*="birth" i]',
        ]
    if any(p in kl for p in ("ssn", "social", "socialsecurity")):
        return [
            'input[name*="ssn" i]',
            'input[name*="social" i]',
            'input[id*="ssn" i]',
            'input[placeholder*="ssn" i]',
        ]
    return []


# ── 2026-05: Robust wait_for_selector with state + selector fallbacks ──
# Why this exists:
#   Visual Recorder dropdown-binding emits a `wait_for_selector` step
#   right before the `select` step, with `state="visible"` (Playwright
#   default). On modern lead-gen pages the actual `<select id="birth_month">`
#   is often:
#     1. Hidden via CSS (`opacity:0` / `position:absolute; left:-9999px`)
#        behind a custom dropdown UI (React-Select, Bootstrap-Select,
#        Chosen, Select2 etc.) — element exists but is NOT "visible" in
#        Playwright's strict sense → 25s timeout → visit fails as
#        "Automation crashed: Page.wait_for_selector: Timeout 25000ms
#        exceeded. Call log: - waiting for locator(\"#birth_month\") to
#        be visible".
#     2. Rendered after a network round-trip so it shows up later than
#        the surrounding fields.
#     3. The page changed the element's id (`#birth_month` → `#dob_month`)
#        between recording and replay.
#
#   The downstream `select` action ALREADY handles all three via
#   _smart_select_with_fallback (it has a JS-driven PHASE 1 that sets
#   hidden <select> values directly). So the wait_for_selector step is
#   actually redundant — we just need it to NOT fail when the element
#   exists but isn't visible. This helper:
#     • Tries the requested state first (default "visible") with a
#       reduced 6s budget — fast path stays fast for normal forms.
#     • Falls back to state="attached" (exists in DOM) — handles hidden
#       <select> behind custom UIs.
#     • Tries selector fallbacks derived from the original
#       (`#birth_month` → `[name="birth_month"]` → `select[name*="birth" i][name*="month" i]`)
#       — handles renamed ids.
#     • If everything fails AND step is optional, soft-skips with a
#       warning (the subsequent select step still has its own fallbacks).
#     • If everything fails AND step is required, raises with a clear
#       "tried N selectors / states" message.
async def _smart_wait_for_selector(page, selector: str, *,
                                    state: str = "visible",
                                    timeout: int = 25000,
                                    extra_alts: Optional[List[str]] = None) -> str:
    """Robust wait_for_selector with state + selector fallback.
    Returns the selector that actually matched (useful for downstream
    debug logging). Raises Playwright TimeoutError on total failure.

    Total budget is `timeout`. Divided across phases:
      Phase 1: original selector with requested state (≤ 40% of budget)
      Phase 2: original selector with state="attached" (≤ 20%)
      Phase 3: fallback selectors with state="attached" (rest)

    `extra_alts` (2026-01) — alias selectors loaded from the user's
    permanent Selector Aliases store (self-healing replay). Inserted
    AHEAD of the token-derived fallbacks so a known-good rename wins
    before we burn time on guesses.
    """
    import re as _re_local
    sel = (selector or "").strip()
    if not sel:
        raise ValueError("wait_for_selector: empty selector")

    requested_state = (state or "visible").strip().lower()
    if requested_state not in ("visible", "attached", "hidden", "detached"):
        requested_state = "visible"

    # ── Build selector fallback chain (same logic as smart_select) ──
    selectors: List[str] = [sel]
    # User-aliased selectors come BEFORE token-derived guesses — they
    # were already proven correct by the user via the Edit modal.
    if extra_alts:
        for a in extra_alts:
            a = (a or "").strip()
            if a and a != sel:
                selectors.append(a)
    key = ""
    m = _re_local.match(r"^#([\w\-]+)$", sel)
    if m:
        key = m.group(1)
    else:
        m = _re_local.search(r"\[(?:name|id)\s*=\s*[\"']([^\"']+)[\"']\]", sel)
        if m:
            key = m.group(1)
        else:
            m = _re_local.search(r"\[name\s*=\s*([^\]\s]+)\]", sel)
            if m:
                key = m.group(1).strip("\"'")
    if key:
        selectors.extend([
            f'[id="{key}"]',
            f'[name="{key}"]',
            f'select#{key}',
            f'select[name="{key}"]',
            f'input[name="{key}"]',
            f'[id*="{key}" i]',
            f'[name*="{key}" i]',
        ])
        tokens = [t for t in _re_local.split(r"[_\-\s]+", key) if t and len(t) >= 2]
        if len(tokens) >= 2:
            selectors.append("".join(f'[name*="{t}" i]' for t in tokens))
            selectors.append("".join(f'[id*="{t}" i]' for t in tokens))

    # Dedup while preserving order
    seen: set = set()
    uniq: List[str] = []
    for s in selectors:
        if s and s not in seen:
            seen.add(s)
            uniq.append(s)

    total = max(2000, int(timeout))
    # Tight budgets: we don't want to burn the full 25s budget waiting
    # for an element that's hidden behind a custom UI to suddenly become
    # "visible". Fail the visibility check fast and fall back to
    # state="attached" which succeeds instantly for elements that
    # already exist in DOM.
    phase1_t = max(1500, min(int(total * 0.2), 5000))   # ≤5s on visibility
    phase2_t = max(1500, min(int(total * 0.15), 3000))  # ≤3s on attached for original
    remaining = max(2000, total - phase1_t - phase2_t)
    # Per-fallback budget — at least 800ms each, up to ~2.5s
    per_fallback = max(800, min(remaining // max(1, len(uniq)), 2500))

    last_err: Optional[Exception] = None

    # Phase 1 — original selector, requested state
    try:
        await page.wait_for_selector(sel, timeout=phase1_t, state=requested_state)
        return sel
    except Exception as e:
        last_err = e

    # Phase 2 — original selector, attached (handles hidden-behind-custom-UI)
    if requested_state != "attached":
        try:
            await page.wait_for_selector(sel, timeout=phase2_t, state="attached")
            return sel
        except Exception as e:
            last_err = e

    # Phase 3 — fallback selectors with attached state
    for alt in uniq[1:]:  # skip the original we already tried
        try:
            await page.wait_for_selector(alt, timeout=per_fallback, state="attached")
            return alt
        except Exception as e:
            last_err = e
            continue

    # ── 2026-05: Phase 4 — Lazy-load trigger + final retry ──
    # If we exhausted all variants WITHOUT a match, the element might
    # not yet exist in the DOM because the page is lazy-rendering it
    # below the fold (common on long offer / survey pages). We force-
    # scroll through the page to trigger any IntersectionObserver /
    # lazy-import handlers, then give the ORIGINAL selector one more
    # short attempt. User asked for: "job k doran pora page he scan ho
    # ta k agr page pr koi step nazar na b a raha ho to step skip na ho
    # — proper har step follow ho".
    try:
        await page.evaluate(
            """async () => {
                // Scroll to bottom in steps so IntersectionObservers fire,
                // then back to top so layout settles before the retry.
                const h = document.documentElement.scrollHeight;
                const stepPx = Math.max(400, Math.floor(window.innerHeight * 0.8));
                for (let y = 0; y < h; y += stepPx) {
                    window.scrollTo(0, y);
                    await new Promise(r => setTimeout(r, 60));
                }
                window.scrollTo(0, 0);
                await new Promise(r => setTimeout(r, 120));
            }"""
        )
    except Exception:
        pass
    # Retry original + each fallback once more with a short budget
    # (≤ 1200ms each) so we don't blow the overall timeout.
    _retry_budget = 1200
    for cand in uniq:
        try:
            await page.wait_for_selector(cand, timeout=_retry_budget, state="attached")
            return cand
        except Exception as e:
            last_err = e
            continue

    # Total failure — re-raise the last error with a richer message so
    # the failed-visit row in the UI says WHY (not just "timeout 25s").
    tried_n = len(uniq)
    raise type(last_err)(
        f"wait_for_selector exhausted {tried_n} selector variants "
        f"(+ lazy-load scroll retry) — original selector "
        f"{sel!r}. Last error: {last_err}"
    )


# ── 2026-05: Robust select-option with selector + match-strategy fallbacks ──
# Why this exists:
#   The original automation steps tried `page.select_option(selector, value=X)`
#   once with the EXACT selector the user recorded. Real-world breakage:
#     1. Page changes the element's `id` (e.g. `#birth_month` becomes `#dob_month`)
#        — selector times out at 25s.
#     2. The customer's data file has human-readable labels ("May") but the
#        underlying <option value="5"> — value-only match fails.
#     3. The dropdown is rendered later than the surrounding fields, so 25s
#        isn't enough for the locator.
#   This helper makes select robust to ALL three cases without changing the
#   user's recorded JSON:
#     • Tries the user's exact selector first (fast path — no penalty when it works).
#     • If that fails, derives 4-6 fallback selectors from the original
#       (e.g. `#birth_month` → `[name="birth_month"]` → `select[name*="birth" i][name*="month" i]`).
#     • For each selector, tries label → value → index strategies (covers
#       Excel-style "May" data AND raw "5" data).
#   Total budget is the original `timeout` — divided across attempts so a
#   bad selector doesn't burn 25s while the right one is one fallback away.
async def _smart_select_with_fallback(page, selector: str, value: Any,
                                       match_by: str = "label",
                                       timeout: int = 25000) -> None:
    """Robust select_option: tries multiple match strategies AND fallback
    selectors. Raises the last underlying exception on total failure."""
    import re as _re_local
    val = "" if value is None else str(value)
    match_by = (match_by or "label").lower().strip()

    # ── 2026-05 (RUT vs Visual-Recorder parity fix) ─────────────────────
    # Build value-variants the same way the Visual Recorder live-test
    # path does (visual_recorder._smart_select_option). Symptom this
    # fixes: a JSON recorded with `#birth_month` value="6" works during
    # the recorder's own live test (because the recorder tries "6" →
    # "06" → "June" → "Jun"), but RUT replay used to try ONLY "6" and
    # the offer page's dropdown only has `<option>June</option>` labels
    # (no numeric `value=` attribute) — so step #16 (select #birth_month)
    # times out and the whole visit fails. By mirroring the recorder's
    # candidate list here, RUT now succeeds on the SAME JSON the user
    # tested in the recorder.
    _MONTHS_FULL = (
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    )
    val_variants: List[str] = []
    _seen_v: set = set()
    def _add_v(x: str) -> None:
        if x and x not in _seen_v:
            _seen_v.add(x)
            val_variants.append(x)
    _add_v(val)
    _add_v(val.strip())
    if val.strip().isdigit():
        _n = int(val.strip())
        _add_v(f"{_n:02d}")     # zero-padded ("6"  → "06")
        _add_v(str(_n))          # un-padded   ("06" → "6")
        if 1 <= _n <= 12:        # month-number → month-name
            _full = _MONTHS_FULL[_n - 1]
            _abbrev = _full[:3]
            for _v in (_full, _abbrev, _full.lower(), _abbrev.lower(),
                       _full.upper(), _abbrev.upper()):
                _add_v(_v)
    else:
        _lv = val.strip().lower()
        for _i, _m in enumerate(_MONTHS_FULL, start=1):
            if _lv == _m.lower() or _lv == _m.lower()[:3]:
                _add_v(str(_i))
                _add_v(f"{_i:02d}")
                break

    # ── Build match-strategy attempts (label / value / index) ──
    # We iterate every value-variant × (label, value) so date / month /
    # number dropdowns with EITHER raw "6" data OR human "June" data
    # connect to options that store EITHER value="6" / value="06" OR
    # only label="June". Order honours `match_by` user preference.
    strategies: List[Tuple[str, Any]] = []
    if match_by == "value":
        _modes = ("value", "label")
    else:
        _modes = ("label", "value")
    for _cand in val_variants:
        for _m in _modes:
            strategies.append((_m, _cand))
    # Also try numeric index as last resort (e.g. value "0", "1", "2")
    if val.strip().isdigit():
        strategies.append(("index", int(val.strip())))

    # ── Build selector candidates ──
    selectors_to_try: List[str] = [selector] if selector else []
    key = ""
    if selector:
        # CSS ID selector: #birth_month → birth_month
        m = _re_local.match(r"^#([\w\-]+)$", selector.strip())
        if m:
            key = m.group(1)
        else:
            # [name="X"] or [id="X"] (single or double quotes)
            m = _re_local.search(
                r"\[(?:name|id)\s*=\s*[\"']([^\"']+)[\"']\]", selector
            )
            if m:
                key = m.group(1)
            else:
                # Bare [name=X]
                m = _re_local.search(r"\[name\s*=\s*([^\]\s]+)\]", selector)
                if m:
                    key = m.group(1).strip("\"'")
    if key:
        # Direct attribute matches
        selectors_to_try.extend([
            f'select#{key}',
            f'select[name="{key}"]',
            f'[name="{key}"]',
            f'select[id="{key}"]',
            f'select[name*="{key}" i]',
            f'select[id*="{key}" i]',
        ])
        # Token-based match: birth_month → both "birth" AND "month" required
        tokens = [t for t in _re_local.split(r"[_\-\s]+", key) if t and len(t) >= 2]
        if len(tokens) >= 2:
            selectors_to_try.append(
                "select" + "".join(f'[name*="{t}" i]' for t in tokens)
            )
            selectors_to_try.append(
                "select" + "".join(f'[id*="{t}" i]' for t in tokens)
            )

    # Deduplicate while preserving order
    _seen: set = set()
    uniq_selectors: List[str] = []
    for s in selectors_to_try:
        if s and s not in _seen:
            _seen.add(s)
            uniq_selectors.append(s)
    if not uniq_selectors:
        raise ValueError(f"select: empty selector (value={val!r})")

    last_err: Optional[Exception] = None

    # ── PHASE 1 (instant, ~10 ms): JS-driven set ─────────────────────
    # Why we try this FIRST instead of last:
    #   1. It's a single page.evaluate round-trip — no actionability wait,
    #      no visibility wait, no 25s-timeout per selector.
    #   2. It handles Bootstrap-Select (`class="selectpicker"`), Select2,
    #      Chosen, and ANY plugin that hides the native <select> and
    #      replaces it with a custom UI. The native element still drives
    #      form submission, so setting .value + dispatching change events
    #      is the correct fix — and Playwright's select_option would
    #      otherwise wait forever for the hidden element to become visible.
    #   3. It tries label → value → numeric-index → label-substring
    #      matching ALL inside a single JS pass.
    #   4. It dispatches native `input`+`change` events AND jQuery
    #      `.trigger('change')` + selectpicker.refresh / chosen:updated,
    #      so the visible plugin UI and any form validators stay in sync.
    # If JS fails (e.g. element in cross-origin iframe), we fall through
    # to Playwright's select_option in PHASE 2.
    _js_set_select = """
    (function(args) {
        var selectors = args.selectors;
        var rawList = args.rawList && args.rawList.length ? args.rawList : [args.raw];
        var byLabel = args.byLabel;
        function findOpt(el, raw, byLabel) {
            const want = String(raw);
            const wantTrim = want.trim().toLowerCase();
            for (let i = 0; i < el.options.length; i++) {
                const o = el.options[i];
                if (byLabel) {
                    const t = (o.text || '').trim().toLowerCase();
                    const l = (o.label || '').trim().toLowerCase();
                    if (t === wantTrim || l === wantTrim) return o;
                } else {
                    if (String(o.value) === want) return o;
                }
            }
            // Try the OTHER strategy if the primary didn't match
            for (let i = 0; i < el.options.length; i++) {
                const o = el.options[i];
                if (byLabel) {
                    if (String(o.value) === want) return o;
                } else {
                    const t = (o.text || '').trim().toLowerCase();
                    if (t === wantTrim) return o;
                }
            }
            // Substring on label (handles "May 2024" vs "May")
            for (let i = 0; i < el.options.length; i++) {
                const o = el.options[i];
                const t = (o.text || '').trim().toLowerCase();
                if (t && wantTrim && (t === wantTrim || t.indexOf(wantTrim) === 0)) return o;
            }
            return null;
        }
        for (let s = 0; s < selectors.length; s++) {
            let el = null;
            try { el = document.querySelector(selectors[s]); } catch (e) { continue; }
            if (!el || el.tagName !== 'SELECT') continue;
            // Try EACH value-variant in turn against this element.
            // The first hit wins. We DON'T fall back to index-by-number
            // across variants because a numeric variant like "6" could
            // bind to options[6] (placeholder + 5 months off-by-one) —
            // recorder parity is the goal, so only exact / substring /
            // label / value matches are accepted here. (RUT's PHASE 2
            // Playwright fallback still includes the index strategy.)
            let opt = null;
            for (let r = 0; r < rawList.length && !opt; r++) {
                opt = findOpt(el, rawList[r], byLabel);
            }
            if (!opt) {
                // Per-element index fallback (last resort, single value only)
                const first = String(rawList[0] || '');
                if (/^\\d+$/.test(first)) {
                    const idx = parseInt(first, 10);
                    if (idx >= 0 && idx < el.options.length) opt = el.options[idx];
                }
            }
            if (!opt) continue;
            el.value = opt.value;
            try { el.dispatchEvent(new Event('input',  {bubbles: true})); } catch (e) {}
            try { el.dispatchEvent(new Event('change', {bubbles: true})); } catch (e) {}
            try {
                if (window.jQuery && window.jQuery(el).length) {
                    window.jQuery(el).val(opt.value).trigger('change');
                    try { window.jQuery(el).selectpicker('refresh'); } catch (e) {}
                    try { window.jQuery(el).trigger('chosen:updated'); } catch (e) {}
                }
            } catch (e) {}
            return {ok: true, selector: selectors[s], value: opt.value, label: opt.text};
        }
        return {ok: false};
    })
    """
    try:
        js_result = await asyncio.wait_for(
            page.evaluate(
                _js_set_select,
                {
                    "selectors": uniq_selectors,
                    "raw": val,
                    "rawList": val_variants,
                    "byLabel": (match_by == "label"),
                },
            ),
            timeout=4.0,
        )
        if isinstance(js_result, dict) and js_result.get("ok"):
            return  # SUCCESS via JS — works for native AND hidden plugin selects
    except Exception as e:
        last_err = e

    # ── PHASE 2 (slower, ~3 s per selector): Playwright select_option ──
    # Fallback for cases JS couldn't reach (cross-origin iframes) or
    # forms that ONLY respond to fully-simulated user events. Per-selector
    # timeout is the original `timeout` divided across remaining selectors
    # so a bad pattern can't eat the entire 25s budget.
    per_sel_timeout = (
        timeout if len(uniq_selectors) == 1
        else max(3000, int(timeout / max(2, len(uniq_selectors))))
    )

    for sel in uniq_selectors:
        for strategy, payload in strategies:
            try:
                if strategy == "label":
                    await page.select_option(sel, label=str(payload), timeout=per_sel_timeout)
                elif strategy == "value":
                    await page.select_option(sel, value=str(payload), timeout=per_sel_timeout)
                elif strategy == "index":
                    await page.select_option(sel, index=int(payload), timeout=per_sel_timeout)
                return  # SUCCESS
            except Exception as e:
                last_err = e
                continue

    if last_err is not None:
        raise last_err


# ── 2026-05: Robust check/uncheck with proxy-element + JS fallbacks ─────
# Same family of problem as `_smart_select_with_fallback`:
#   • Modern offer forms use CSS-styled checkboxes — the real <input
#     type="checkbox"> is `display:none` and a wrapping <label> + visible
#     <span>/<div> acts as the click proxy. Playwright's `page.check()`
#     waits for the input to be visible → times out at 25s.
#   • Form validation (jQuery / vanilla) only fires when the native
#     `change` event bubbles up from a real click → just setting
#     `.checked = true` via JS often gets reset by the page's own logic.
# This helper tries 4 strategies in order:
#   1. Native page.check/uncheck (works for plain visible checkboxes).
#   2. Click the wrapping <label> (correct user-flow for CSS-styled
#      checkboxes — triggers default label→input toggle).
#   3. Click the visible sibling span/div inside the label (proxy click
#      surface).
#   4. JS set + dispatch `change`+`input`+jQuery('.trigger') as last
#      resort.
# Returns silently on success; raises last exception on total failure.
async def _smart_check_with_fallback(page, selector: str, want_checked: bool = True,
                                     timeout: int = 25000) -> None:
    """Robust check/uncheck handling CSS-styled (hidden) checkboxes."""
    last_err: Optional[Exception] = None

    # ── Strategy 1: native Playwright check/uncheck (short timeout) ──
    # Fast-path for plain visible checkboxes. We give it ~25% of the budget
    # so a hidden checkbox doesn't burn the whole 25s before we fall back.
    short_timeout = max(2000, int(timeout * 0.25))
    try:
        if want_checked:
            await page.check(selector, timeout=short_timeout)
        else:
            await page.uncheck(selector, timeout=short_timeout)
        return
    except Exception as e:
        last_err = e

    # ── Strategy 2-4: JS-driven smart click (instant, handles hidden) ──
    js_smart_check = """
    (function(args) {
        const selector = args.selector;
        const wantChecked = args.wantChecked;
        const cb = document.querySelector(selector);
        if (!cb) return {ok: false, reason: 'not_found'};
        if (cb.checked === wantChecked) return {ok: true, already: true};

        // 2a: Click the wrapping <label> (proper user-flow for CSS checkboxes)
        const wrapLabel = cb.closest('label');
        if (wrapLabel) {
            try { wrapLabel.click(); } catch (e) {}
            if (cb.checked === wantChecked) return {ok: true, strategy: 'label_click'};
        }
        // 2b: Click a label[for="<id>"] outside the wrap
        if (cb.id) {
            const extLabel = document.querySelector('label[for="' + cb.id + '"]');
            if (extLabel) {
                try { extLabel.click(); } catch (e) {}
                if (cb.checked === wantChecked) return {ok: true, strategy: 'for_label_click'};
            }
        }
        // 2c: Click a visible sibling element inside the wrap label
        //     (CSS pseudo-checkbox like <span class="custom-check">)
        if (wrapLabel) {
            const sibs = wrapLabel.querySelectorAll('span, div, i, em');
            for (const s of sibs) {
                if (s.offsetWidth > 0 && s.offsetHeight > 0) {
                    try { s.click(); } catch (e) {}
                    if (cb.checked === wantChecked) return {ok: true, strategy: 'sibling_click'};
                }
            }
        }
        // 2d: Click the checkbox itself — works even when display:none
        //     because programmatic .click() doesn't require visibility.
        try { cb.click(); } catch (e) {}
        if (cb.checked === wantChecked) return {ok: true, strategy: 'cb_click'};
        // 2e: LAST RESORT — set property + dispatch full event suite.
        //     Only useful when the form's validators read .checked directly.
        cb.checked = wantChecked;
        try { cb.dispatchEvent(new Event('input', {bubbles: true})); } catch (e) {}
        try { cb.dispatchEvent(new Event('change', {bubbles: true})); } catch (e) {}
        try { cb.dispatchEvent(new Event('click', {bubbles: true})); } catch (e) {}
        if (window.jQuery) {
            try { window.jQuery(cb).prop('checked', wantChecked).trigger('change').trigger('click'); } catch (e) {}
        }
        return {ok: cb.checked === wantChecked, strategy: 'set_dispatch'};
    })
    """
    try:
        r = await asyncio.wait_for(
            page.evaluate(js_smart_check,
                          {"selector": selector, "wantChecked": bool(want_checked)}),
            timeout=4.0,
        )
        if isinstance(r, dict) and r.get("ok"):
            return  # SUCCESS
        if isinstance(r, dict) and r.get("reason") == "not_found":
            last_err = ValueError(f"check: element not found ({selector!r})")
    except Exception as e:
        if last_err is None:
            last_err = e

    if last_err is not None:
        raise last_err


async def _execute_automation_steps(
    page: Page,
    row: Dict[str, Any],
    steps: List[Dict[str, Any]],
    skip_captcha: bool = True,
    self_heal: bool = True,
    on_screenshot: Optional[Callable[[int, str, bytes], Awaitable[None]]] = None,
    collect_timings: bool = False,
    user_id: Optional[str] = None,
    on_step_progress: Optional[Callable[[Dict[str, Any]], Awaitable[None]]] = None,
) -> Dict[str, Any]:
    """Execute a user-provided automation script step-by-step. Returns
    {status, error?, executed_steps, step_results?}.  Each step format:
        {"action": "click", "selector": "a.btn-primary", "wait_nav": true}
        {"action": "fill",  "selector": "input[name=first]", "value": "{{first}}"}
        {"action": "select","selector": "select[name=dobmonth]", "value": "{{month}}"}
        {"action": "wait",  "ms": 2000}
        {"action": "screenshot", "name": "after_submit"}

    When `self_heal=True`, if a non-optional step fails we take a screenshot
    of the current page and ask Gemini 2.5 Pro for a single recovery action
    (dismiss popup, click Continue, etc.). We try the main step one more time
    after applying the recovery.

    When an `on_screenshot` callback is provided, every `screenshot` action
    captures a real PNG of the current page and hands it to the callback
    (along with the step index and the user-supplied name) so the live
    activity / per-visit storage layer can surface it for the operator.

    2026-05 — When `collect_timings=True` (Visual-Recorder live-test mode),
    we wrap each step in a perf-counter envelope and capture per-step
    {idx, action, selector, ok, error, ms} into the result. This powers
    the "Run Live Test" panel in the recorder so the user can see which
    step is slow/failing BEFORE committing the automation to a RUT job.
    Production RUT visits leave this OFF (default) so there's zero
    overhead in the hot loop.
    """
    import time as _time_mod
    executed = 0
    heal_used = 0
    MAX_HEAL = 3  # total AI recovery attempts per row
    step_results: List[Dict[str, Any]] = []
    _t_total_start = _time_mod.perf_counter() if collect_timings else 0.0

    # ── 2026-01: Selector Aliases (self-healing replay) ──────────────
    # If user_id is provided, pre-load all alias mappings for the
    # current page's domain. When a step's selector fails, the smart
    # wait/select/click helpers will additionally try these aliases
    # AHEAD of the token-derived guesses. This rescues recordings
    # whose target website renamed a form field after the recording
    # was made (the classic "#birth_month → #dob_month" case).
    aliases_for_domain: Dict[str, List[str]] = {}
    alias_user_id: Optional[str] = user_id
    alias_domain: str = ""
    if user_id:
        try:
            import selector_aliases as _sa
            try:
                alias_domain = _sa.extract_domain(page.url or "")
            except Exception:
                alias_domain = ""
            if alias_domain:
                aliases_for_domain = await _sa.get_aliases_for_domain(user_id, alias_domain)
                if aliases_for_domain:
                    logger.info(f"[selector_aliases] loaded {len(aliases_for_domain)} alias mappings for domain={alias_domain}")
        except Exception as _ae:  # noqa: BLE001
            logger.debug(f"[selector_aliases] preload skipped: {_ae}")

    def _alias_alts_for(sel: str) -> List[str]:
        """Return alias selectors for `sel` (or empty list)."""
        if not sel or not aliases_for_domain:
            return []
        return list(aliases_for_domain.get(sel.strip(), []))

    # 2026-01 (additive): One-shot cookie/GDPR banner auto-dismiss at
    # the start of the automation. Many lead-gen pages show OneTrust /
    # Cookiebot / Quantcast banners that block clicks until the user
    # accepts — bots get stuck on the first click. We dismiss BEFORE
    # step 0 so the rest of the recorded automation runs on a clean
    # page. Safe — failures swallowed.
    if _EXT_LOADED:
        try:
            await _ext_dismiss_cookies(page, log_label="pre-automation")
        except Exception:
            pass

    # ── 2026-05: Hard per-step timeout ceilings ─────────────────────
    # User report: "JSON perfect banaya pr phr b fail/stuck ho raha hai".
    # Root cause investigation showed that a single malformed step
    # (e.g. `wait: ms=600000` from a paused-during-record session, or a
    # `fill` whose selector never matches on production HTML) would
    # block the visit's concurrency slot for 60-180s while waiting on
    # internal Playwright timeouts. The job-wide 240s stuck-watchdog
    # would eventually rescue it, but by then 3-5 minutes of proxy
    # budget had been wasted on a dead visit. These ceilings GUARANTEE
    # no single step can hold a slot for more than the cap below,
    # regardless of what the JSON says — bad recordings can no longer
    # poison the whole concurrency pool.
    _STEP_TIMEOUT_CEILINGS_MS: Dict[str, int] = {
        "wait": 30_000,                  # explicit sleeps capped at 30s
        "wait_for_load": 30_000,
        "wait_for_networkidle": 30_000,
        "wait_for_navigation": 30_000,
        "wait_for_selector": 30_000,
        "wait_for_text": 30_000,
        "screenshot": 20_000,
        "evaluate": 15_000,              # custom JS shouldn't run long
        "goto": 45_000,                  # navigation can be slow
        # All element-targeted actions (fill/click/select/check/uncheck
        # /type/press/hover/scroll_into_view) share the default 45s cap.
        "_default": 45_000,
    }

    def _capped_timeout(_action: str, requested_ms: int) -> int:
        cap = _STEP_TIMEOUT_CEILINGS_MS.get(_action, _STEP_TIMEOUT_CEILINGS_MS["_default"])
        if requested_ms is None or requested_ms <= 0:
            return cap
        return min(int(requested_ms), cap)

    # ── 2026-05: Per-step heartbeat task ────────────────────────────
    # While a step is awaiting Playwright (e.g. smart-pre-wait for a
    # selector for up to 30s), the Live Visual Grid would otherwise
    # show the SAME "step #N running" event with no update for 30
    # seconds — which user (correctly) interpreted as "stuck". We now
    # spawn a tiny task that re-emits the running event every 6
    # seconds with an elapsed-time field, so the UI shows live
    # progress: "step #13 wait — running 18s". Cancels itself when
    # the step's main awaitable returns.
    async def _start_step_heartbeat(_idx: int, _action: str, _selector: str, _val_preview: str) -> Optional[asyncio.Task]:
        if on_step_progress is None:
            return None
        _start_ts = _time_mod.time()

        async def _beat():
            try:
                while True:
                    await asyncio.sleep(6.0)
                    elapsed_s = int(_time_mod.time() - _start_ts)
                    try:
                        await on_step_progress({
                            "idx": _idx,
                            "action": _action,
                            "selector": (_selector or "")[:200],
                            "value_preview": (_val_preview or "")[:80],
                            "total_steps": len(steps or []),
                            "status": "running",
                            "elapsed_s": elapsed_s,   # NEW: UI shows "(18s)"
                            "timestamp_ms": int(_time_mod.time() * 1000),
                            "heartbeat": True,         # mark so UI can dedupe
                        })
                    except Exception:
                        pass
            except asyncio.CancelledError:
                return

        return asyncio.create_task(_beat())

    try:
        for idx, step in enumerate(steps or []):
            if not isinstance(step, dict):
                continue
            action = (step.get("action") or "").strip().lower()
            _t_step_start = _time_mod.perf_counter() if collect_timings else 0.0
            _step_ok = True
            _step_err: Optional[str] = None
            selector = step.get("selector") or ""
            value = _substitute(step.get("value", ""), row)
            # ── 2026-05: longer default per-step timeout ──
            # Bumped from 10s → 25s. Visual Recorder doesn't emit a
            # per-step `timeout` so the default IS what every fill /
            # click / select uses. Real residential-proxy pages can
            # easily take 12-18s to render the next form field after
            # a click; the old 10s default caused steps to time out
            # silently (with optional=true), the bot did nothing,
            # the page stayed frozen, and the 60s watchdog killed
            # the visit. 25s aligns with realistic human patience
            # while still bailing out on truly dead pages — and the
            # actionable-pre-wait helper added below makes most
            # steps return as SOON as the selector is ready, so
            # this larger ceiling only kicks in on slow pages.
            timeout = _capped_timeout(action, int(step.get("timeout") or 25000))
            optional = bool(step.get("optional") or False)
            wait_nav = bool(step.get("wait_nav") or False)

            # 2026-01 (additive): real-time step progress callback. When
            # the caller provides `on_step_progress`, we emit a
            # "running" event BEFORE the action executes so the UI can
            # show "step #N (action) — running…" live. Followed by an
            # "ok" / "failed" event in the respective code paths below.
            # Failures here are swallowed — the callback is purely
            # informational; never blocks the automation.
            if on_step_progress is not None:
                try:
                    await on_step_progress({
                        "idx": idx,
                        "action": action,
                        "selector": (selector or "")[:200],
                        "value_preview": (str(value) if value else "")[:80],
                        "total_steps": len(steps or []),
                        "status": "running",
                        "timestamp_ms": int(_time_mod.time() * 1000),
                    })
                except Exception:
                    pass

            # 2026-05 — start the heartbeat so the UI sees live elapsed
            # time during long steps (the Visual Grid otherwise looks
            # frozen for ~30s on slow `fill`/`wait`/`wait_for_load`
            # steps and users assume the visit is stuck).
            _hb_task = await _start_step_heartbeat(idx, action, selector, str(value) if value else "")

            if skip_captcha and action not in ("wait", "screenshot", "evaluate"):
                if await _page_has_captcha(page):
                    return {"status": "skipped_captcha", "error": f"Captcha at step {idx+1}", "executed_steps": executed}

            # ── 2026-05: smart pre-wait for actionable elements ──
            # For every action that needs a real DOM element (fill,
            # click, select, check/uncheck, type, press, hover,
            # scroll-into-view), ACTIVELY wait up to the step's
            # timeout for that selector to appear & be visible BEFORE
            # we even try to interact. This single addition solves
            # the long-standing "watchdog kills visit at 60s because
            # the bot did nothing" failure mode: previously every
            # `fill` with optional=true would silently skip when the
            # field wasn't in DOM yet, the bot would do nothing on a
            # half-loaded form, and 60s later watchdog reaped the
            # visit. Now we wait for the field to actually render
            # before filling — exactly as a human would.
            #
            # Wait is skipped for:
            #   • selectorless actions (wait, screenshot, evaluate,
            #     wait_for_load, scroll without selector, goto)
            #   • wait_for_selector itself (handled by Playwright)
            #
            # If the wait times out and the step is `optional`, the
            # outer try/except still swallows it as before.
            _PRE_WAIT_ACTIONS = {
                "fill", "click", "type", "select",
                "check", "uncheck", "press", "hover",
            }
            # ── 2026-01 (additive): if_exists pre-check ──
            # When a step has `if_exists: true` (and it has a selector),
            # we quickly probe the DOM. If the selector is NOT present
            # within `if_exists_timeout` ms (default 2000), we SOFT-SKIP
            # the step without raising. Useful for cookie/upsell popups
            # that only appear randomly.
            if step.get("if_exists") and selector and action in _PRE_WAIT_ACTIONS:
                _ie_timeout = int(step.get("if_exists_timeout") or 2000)
                try:
                    await page.wait_for_selector(selector, timeout=_ie_timeout, state="attached")
                except Exception:
                    # Element absent — skip this step quietly.
                    executed += 1
                    if collect_timings:
                        step_results.append({
                            "idx": idx, "action": action,
                            "selector": (selector or "")[:200],
                            "ok": True,
                            "error": "(skipped — if_exists check failed)",
                            "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                            "optional": True,
                            "if_exists_skipped": True,
                        })
                    continue

            # 2026-05 — Pre-wait WAS in its own try block but that meant
            # selector-not-found errors propagated to the outer "Automation
            # crashed" handler instead of the per-step except below
            # (skipping the optional/retry/self-heal logic). The pre-wait
            # call now lives INSIDE the action try block, so any
            # selector-resolution failure is handled identically to a
            # plain Playwright timeout from page.fill/click — keeping
            # the per-step ceiling intact (no double-wait of pre-wait
            # 45s + action 45s = 90s) while preserving optional skip.
            try:
                if action in _PRE_WAIT_ACTIONS and selector:
                    if action in ("select", "check", "uncheck"):
                        resolved_sel = await _smart_wait_for_selector(
                            page, selector, state="attached", timeout=timeout,
                            extra_alts=(_step_fallbacks(step) + _alias_alts_for(selector) + _field_type_alts_for(selector)),
                        )
                    else:
                        resolved_sel = await _smart_wait_for_selector(
                            page, selector, state="visible", timeout=timeout,
                            extra_alts=(_step_fallbacks(step) + _alias_alts_for(selector) + _field_type_alts_for(selector)),
                        )
                    if resolved_sel and resolved_sel != selector:
                        logger.info(
                            f"[selector-fallback] step #{idx+1} ({action}) "
                            f"'{selector}' → '{resolved_sel}' (rescued)"
                        )
                        selector = resolved_sel

                if action == "goto":
                    await page.goto(value or selector, timeout=timeout, wait_until="domcontentloaded")
                elif action == "click":
                    if wait_nav:
                        # Expect navigation to fire as a result of the click.
                        # Many modern lead-gen pages attach JS handlers to the
                        # submit button that fire analytics/tracking but DO NOT
                        # actually submit the form — so a bare `page.click` +
                        # wait_for_load_state("networkidle") misses the fact
                        # that the form never POSTed. We use expect_navigation
                        # to detect this explicitly, and fall back to calling
                        # form.submit() on the button's parent form if nothing
                        # navigated within the timeout.
                        nav_timeout = min(timeout, 30000)
                        navigated = False
                        try:
                            async with page.expect_navigation(timeout=nav_timeout, wait_until="load"):
                                await page.click(selector, timeout=timeout)
                            navigated = True
                        except Exception:
                            navigated = False
                        if not navigated:
                            # Click might have succeeded but no navigation fired.
                            # Give any pending JS (LeadId / TrustedForm token
                            # collectors that attach onsubmit handlers and only
                            # populate hidden fields on the first click) a brief
                            # window to finish, then call plain form.submit()
                            # which BYPASSES onsubmit handlers and forces the
                            # POST through.
                            try:
                                await page.wait_for_timeout(2500)
                            except Exception:
                                pass
                            try:
                                async with page.expect_navigation(timeout=nav_timeout, wait_until="load"):
                                    await page.evaluate(
                                        "(sel) => {"
                                        "  var el = document.querySelector(sel);"
                                        "  var f = el && (el.form || el.closest('form'));"
                                        "  if (f) { try { f.submit(); } catch(e) {} }"
                                        "}",
                                        selector,
                                    )
                                navigated = True
                            except Exception:
                                pass
                        # Best-effort wait for the post-navigation page to
                        # settle (non-fatal if already idle).
                        try:
                            await page.wait_for_load_state("networkidle", timeout=15000)
                        except Exception:
                            pass
                    else:
                        await page.click(selector, timeout=timeout)
                elif action == "fill":
                    # 2026-01 Anti-detect: humanise the fill (see notes
                    # at the second fill handler ~4675 for the rationale).
                    # NEW (2026-01) — per-step opt-out: if user set
                    # `humanize: false` via the Edit-step UI (Visual
                    # Recorder), skip slow per-char typing and use
                    # page.fill() directly. Useful for live-test
                    # debugging or internal forms where stealth doesn't
                    # matter. Defaults to True (humanised) so existing
                    # recorded steps keep their anti-detect behaviour.
                    if step.get("humanize") is False:
                        await page.fill(selector, str(value), timeout=timeout)
                    else:
                        try:
                            from form_filler import _human_type_field as _htf, _human_tab_or_pause as _htp
                            el_h = await page.query_selector(selector)
                            if el_h is not None:
                                ok_h = await _htf(page, el_h, str(value))
                                if ok_h:
                                    await _htp(page)
                                else:
                                    await page.fill(selector, str(value), timeout=timeout)
                            else:
                                await page.fill(selector, str(value), timeout=timeout)
                        except Exception:
                            await page.fill(selector, str(value), timeout=timeout)
                elif action == "type":
                    # Slower per-char typing (more human) — now with
                    # variable delay + thinking pauses via the helper.
                    # Same per-step humanize opt-out as `fill` above.
                    if step.get("humanize") is False:
                        await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
                    else:
                        try:
                            from form_filler import _human_type_field as _htf, _human_tab_or_pause as _htp
                            el_h = await page.query_selector(selector)
                            if el_h is not None:
                                ok_h = await _htf(page, el_h, str(value))
                                if ok_h:
                                    await _htp(page)
                                else:
                                    await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
                            else:
                                await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
                        except Exception:
                            await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
                elif action == "select":
                    # 2026-05: robust select with match-strategy + selector
                    # fallbacks (see _smart_select_with_fallback docstring).
                    await _smart_select_with_fallback(
                        page, selector, value,
                        match_by=str(step.get("match_by") or "label"),
                        timeout=timeout,
                    )
                elif action == "check":
                    # 2026-05: handles CSS-styled hidden checkboxes via
                    # label-click / sibling-click / JS-set fallbacks.
                    await _smart_check_with_fallback(
                        page, selector, want_checked=True, timeout=timeout,
                    )
                elif action == "uncheck":
                    await _smart_check_with_fallback(
                        page, selector, want_checked=False, timeout=timeout,
                    )
                elif action == "press":
                    await page.press(selector or "body", value or "Enter", timeout=timeout)
                elif action == "wait":
                    # 2026-05 — apply hard ceiling. Even if the JSON says
                    # ms=600000 (10 minutes — usually a recording bug),
                    # we never sleep more than the cap. Prevents one bad
                    # step from blocking a concurrency slot forever.
                    _wait_ms_req = int(step.get("ms") or 1000)
                    _wait_ms = min(_wait_ms_req, _STEP_TIMEOUT_CEILINGS_MS["wait"])
                    if _wait_ms < _wait_ms_req:
                        logger.info(
                            f"[step-cap] step #{idx+1} wait: requested {_wait_ms_req}ms "
                            f"capped to {_wait_ms}ms (engine ceiling)"
                        )
                    await page.wait_for_timeout(_wait_ms)
                elif action == "wait_for_selector":
                    # 2026-05: Use smart wait helper that handles hidden
                    # <select> behind custom dropdown UIs (state=attached
                    # fallback) + selector renames (#birth_month →
                    # [name="birth_month"] etc). Fixes "Page.wait_for_selector:
                    # Timeout 25000ms exceeded — waiting for locator(
                    # '#birth_month') to be visible" on dropdown-binding
                    # steps where the native <select> is styled hidden.
                    # 2026-01: also fed user-saved alias selectors so
                    # known renames win instantly.
                    await _smart_wait_for_selector(
                        page, selector,
                        state=step.get("state") or "visible",
                        timeout=timeout,
                        extra_alts=(_step_fallbacks(step) + _alias_alts_for(selector)),
                    )
                elif action in ("wait_for_navigation", "wait_for_load", "wait_for_networkidle"):
                    # 2026-01 — Smart load wait for SPA-friendly pages.
                    # The plain `wait_for_load_state("networkidle", timeout=60000)`
                    # used to block for the full 60s whenever the offer
                    # was a single-page app (no real <load> event after a
                    # JS-driven route transition) — and worse, during
                    # those 60s the URL also didn't change, so the
                    # stuck-watchdog killed the visit while the SPA was
                    # actually rendering the next step. Now we race
                    # several "page progressed" signals and return as
                    # soon as ANY of them fires, capped at the
                    # caller's timeout (or 15s if none was supplied).
                    _wfl_timeout = int(timeout or 15000)
                    try:
                        _wfl_start_url = page.url or ""
                    except Exception:
                        _wfl_start_url = ""
                    _wfl_progress_js = (
                        "()=>{try{var b=document.body;"
                        "var t=(b&&b.innerText)?b.innerText.length:0;"
                        "var c=(b&&b.querySelectorAll)?b.querySelectorAll('*').length:0;"
                        "return [t,c,document.title||''];"
                        "}catch(e){return [0,0,''];}}"
                    )
                    try:
                        _wfl_start_progress = await asyncio.wait_for(
                            page.evaluate(_wfl_progress_js), timeout=2.0,
                        )
                    except Exception:
                        _wfl_start_progress = [0, 0, ""]

                    async def _wait_state_dom():
                        try:
                            await page.wait_for_load_state(
                                "domcontentloaded", timeout=_wfl_timeout,
                            )
                        except Exception:
                            pass

                    async def _wait_state_net():
                        try:
                            await page.wait_for_load_state(
                                "networkidle", timeout=_wfl_timeout,
                            )
                        except Exception:
                            pass

                    async def _wait_url_or_dom_change():
                        # Poll for either a URL change or a meaningful
                        # DOM-fingerprint change every 250 ms. SPA route
                        # transitions don't emit load events but DO
                        # change the DOM noticeably (different number of
                        # nodes / different page title / different text
                        # length), so this poll catches them within
                        # half a second of the transition completing.
                        _deadline = (
                            __import__("time").monotonic() + (_wfl_timeout / 1000.0)
                        )
                        while __import__("time").monotonic() < _deadline:
                            try:
                                await asyncio.sleep(0.25)
                            except asyncio.CancelledError:
                                return
                            try:
                                _cu = page.url or ""
                            except Exception:
                                _cu = _wfl_start_url
                            if _cu != _wfl_start_url:
                                return
                            try:
                                _cp = await asyncio.wait_for(
                                    page.evaluate(_wfl_progress_js), timeout=1.0,
                                )
                            except Exception:
                                continue
                            try:
                                _t_d = abs(int(_cp[0]) - int(_wfl_start_progress[0]))
                                _c_d = abs(int(_cp[1]) - int(_wfl_start_progress[1]))
                                _title_d = _cp[2] != _wfl_start_progress[2]
                            except Exception:
                                _t_d = _c_d = 0
                                _title_d = False
                            if _title_d or _t_d >= 20 or _c_d >= 5:
                                return
                    try:
                        # Whichever signal fires first wins.
                        _wfl_tasks = [
                            asyncio.create_task(_wait_state_dom()),
                            asyncio.create_task(_wait_state_net()),
                            asyncio.create_task(_wait_url_or_dom_change()),
                        ]
                        try:
                            _done, _pending = await asyncio.wait(
                                _wfl_tasks,
                                return_when=asyncio.FIRST_COMPLETED,
                                timeout=(_wfl_timeout / 1000.0) + 1.0,
                            )
                        finally:
                            for _pt in _wfl_tasks:
                                if not _pt.done():
                                    _pt.cancel()
                    except Exception:
                        pass
                elif action == "scroll":
                    try:
                        if selector:
                            el = await page.query_selector(selector)
                            if el:
                                await el.scroll_into_view_if_needed()
                        else:
                            await page.evaluate(f"window.scrollBy(0,{int(step.get('y') or 500)})")
                    except Exception:
                        pass
                elif action == "evaluate":
                    js = _substitute(step.get("script") or step.get("js") or "", row)
                    # ── 2026-06 native-click upgrade ──────────────────
                    # The Visual Recorder emits "click by text" and
                    # "random pick by text" steps as `action: evaluate`
                    # containing a synthetic JS that does `el.click()`.
                    # Under SPA frameworks (React, Vue) and iframe-based
                    # offer walls (stacks.app, uplevelrewards, etc.)
                    # synthetic clicks frequently fail silently — page
                    # doesn't navigate, optional follow-up steps skip
                    # past, visit "completes" without any work done.
                    #
                    # We pre-scan the script for known legacy patterns:
                    #   • `var labels=[...]` → random-pick — pick one
                    #     label here in Python so the choice is logged
                    #     and the same label survives a retry loop.
                    #   • `var t='...'`      → single text-click.
                    # If matched, route through Playwright's native
                    # locator (real mouse events, walks all frames,
                    # waits for actionable state). On native success we
                    # SKIP the original JS to avoid double-clicking.
                    # On native failure we fall back to JS execution so
                    # nothing previously-working regresses.
                    _native_handled = False
                    _native_picked: Optional[str] = None
                    try:
                        _rp_labels = _extract_random_pick_labels(js)
                        if _rp_labels:
                            import random as _rnd_pick
                            _native_picked = _rnd_pick.choice(_rp_labels)
                            _ok_n, _frame_url_n, _err_n = await _native_click_by_text(
                                page, _native_picked, timeout_ms=8000
                            )
                            if _ok_n:
                                logger.info(
                                    f"[evaluate→native_click] random-pick='{_native_picked}' "
                                    f"frame='{(_frame_url_n or '')[:60]}'"
                                )
                                _native_handled = True
                            else:
                                logger.warning(
                                    f"[evaluate→native_click] random-pick='{_native_picked}' "
                                    f"failed: {_err_n[:100]} — falling back to JS"
                                )
                        else:
                            _tc_label = _extract_text_click_label(js)
                            if _tc_label:
                                _native_picked = _tc_label
                                _ok_n, _frame_url_n, _err_n = await _native_click_by_text(
                                    page, _tc_label, timeout_ms=8000
                                )
                                if _ok_n:
                                    logger.info(
                                        f"[evaluate→native_click] text-click='{_tc_label}' "
                                        f"frame='{(_frame_url_n or '')[:60]}'"
                                    )
                                    _native_handled = True
                                else:
                                    logger.warning(
                                        f"[evaluate→native_click] text-click='{_tc_label}' "
                                        f"failed: {_err_n[:100]} — falling back to JS"
                                    )
                    except Exception as _pre_e:  # noqa: BLE001
                        logger.warning(f"[evaluate→native_click] pre-processor error: {_pre_e}")
                    # ── 2026-05 fix ───────────────────────────────────────
                    # Visual Recorder emits `evaluate` steps that may
                    # trigger navigation (e.g. `el.click()` on a button
                    # whose onclick sets `location.href`, or
                    # `window.location.assign(el.href)` from the new
                    # anchor-navigation fix). Plain `page.evaluate()`
                    # does not await the resulting navigation — the
                    # next `fill` step then tries to target a selector
                    # on the OLD page and times out.
                    #
                    # We snapshot the URL before, run the JS, and if
                    # the URL has changed (or a load is in-flight)
                    # within a short window, wait for the new page
                    # to reach a stable state. This is safe for
                    # non-navigating scripts because the polling
                    # window is bounded and best-effort.
                    _url_before = ""
                    try:
                        _url_before = page.url
                    except Exception:
                        pass
                    # ── 2026-01 fix ──────────────────────────────────
                    # When the JS inside `evaluate` performs a click
                    # that triggers immediate navigation, Playwright's
                    # JS execution context gets DESTROYED before the
                    # promise returned by `evaluate` resolves, raising
                    # "Execution context was destroyed, most likely
                    # because of a navigation". This is NOT a failure
                    # — it actually means the click worked exactly as
                    # intended. We catch the error, verify the URL
                    # changed (proof of navigation), and continue
                    # normally to the next step. The downstream
                    # wait_for_load_state below handles the new page.
                    if not _native_handled:
                        try:
                            await page.evaluate(js)
                        except Exception as _ev_err:
                            _ev_msg = str(_ev_err).lower()
                            _navigated = False
                            try:
                                _navigated = (page.url != _url_before)
                            except Exception:
                                _navigated = False
                            # If the destroyed-context happened because of a
                            # navigation (URL changed OR error explicitly
                            # mentions navigation), treat as success.
                            if (
                                "execution context was destroyed" in _ev_msg
                                or "execution context" in _ev_msg
                                or "navigation" in _ev_msg
                                or _navigated
                            ):
                                logger.info(
                                    f"[evaluate] context destroyed by navigation — treating as success "
                                    f"(was: {_url_before}, now: {page.url if _navigated else '?'})"
                                )
                            else:
                                # Real script error (syntax, ReferenceError, etc.)
                                raise
                    # Short polling window to detect navigation kicked
                    # off by the JS (location.assign / location.href /
                    # form.submit / el.click() on anchor or submit btn).
                    # Form submits (POST → server → redirect) can take
                    # 1-2s on slow proxies, so we poll up to ~2.5s.
                    try:
                        for _ in range(25):  # up to ~2.5s
                            await page.wait_for_timeout(100)
                            try:
                                if page.url != _url_before:
                                    # URL change detected — wait for new
                                    # page to finish loading before the
                                    # next automation step runs.
                                    try:
                                        await page.wait_for_load_state(
                                            "domcontentloaded", timeout=20000
                                        )
                                    except Exception:
                                        pass
                                    try:
                                        await page.wait_for_load_state(
                                            "networkidle", timeout=15000
                                        )
                                    except Exception:
                                        pass
                                    break
                            except Exception:
                                pass
                    except Exception:
                        pass
                elif action == "wait_for_text":
                    # 2026-01 (additive): wait until visible body text
                    # contains a substring. Useful for conversion
                    # detection ("Thank you", "Order confirmed").
                    if _EXT_LOADED:
                        _txt = _substitute(step.get("text") or "", row)
                        _ci = bool(step.get("case_insensitive", True))
                        _ok = await _ext_wait_text(page, _txt, timeout_ms=timeout, case_insensitive=_ci)
                        if not _ok and not optional:
                            raise Exception(f"wait_for_text: '{_txt[:60]}' not found in {timeout}ms")
                elif action == "wait_for_url":
                    # 2026-01 (additive): wait until URL matches a
                    # predicate (contains / equals / regex pattern).
                    if _EXT_LOADED:
                        _ok = await _ext_wait_url(
                            page,
                            contains=_substitute(step.get("contains") or "", row) or None,
                            equals=_substitute(step.get("equals") or "", row) or None,
                            pattern=step.get("pattern") or None,
                            timeout_ms=timeout,
                        )
                        if not _ok and not optional:
                            raise Exception(
                                f"wait_for_url: predicate not met in {timeout}ms "
                                f"(contains={step.get('contains')}, equals={step.get('equals')}, pattern={step.get('pattern')})"
                            )
                elif action == "extract":
                    # 2026-01 (additive): extract text from a selector
                    # into row[store_key] so subsequent steps can use it
                    # via `{{store_key}}` substitution.
                    if _EXT_LOADED:
                        store_key = str(step.get("store_key") or step.get("var") or "").strip()
                        if not store_key:
                            if not optional:
                                raise Exception("extract: store_key (variable name) required")
                        else:
                            _ok, _val = await _ext_extract(
                                page, selector, store_key, row,
                                attribute=step.get("attribute"),
                                regex=step.get("regex"),
                                timeout_ms=timeout,
                            )
                            if not _ok and not optional:
                                raise Exception(f"extract: {_val}")
                            logger.info(f"[extract] {store_key} = '{(_val or '')[:80]}'")
                elif action == "dismiss_popups":
                    # 2026-01 (additive): explicit cookie/popup banner
                    # dismissal step. Safe — failures swallowed.
                    if _EXT_LOADED:
                        try:
                            await _ext_dismiss_cookies(page, log_label=f"step-{idx+1}")
                        except Exception:
                            pass
                        try:
                            await _dismiss_popups(page)
                        except Exception:
                            pass
                elif action == "screenshot":
                    # User-defined intermediate capture — take a real
                    # PNG and forward it to the on_screenshot callback
                    # so the live-activity panel can surface it as
                    # visible visit progress.
                    # ── 2026-05 (mandatory capture) ──
                    # User requirement: "jahan capture lagaya hai json
                    # mein wo lazmi ss ay" — screenshot steps MUST
                    # execute and MUST appear in Live Activity.
                    # Sequence:
                    #   1. wait for "load" + "networkidle" so the
                    #      capture reflects the truly-loaded state
                    #   2. wait for body to have visible content
                    #   3. small paint settle delay (250ms)
                    #   4. take screenshot with generous timeout (15s)
                    #   5. forward to live-activity callback if present
                    # Failures are NEVER fatal to the visit but they
                    # ARE surfaced as a warning in live activity so the
                    # operator knows the step didn't capture.
                    shot_name = str(
                        step.get("name") or f"Step {idx + 1}"
                    ).strip() or f"Step {idx + 1}"
                    try:
                        try:
                            await page.wait_for_load_state("load", timeout=8000)
                        except Exception:
                            pass
                        try:
                            await page.wait_for_load_state("networkidle", timeout=8000)
                        except Exception:
                            pass
                        try:
                            await page.wait_for_function(
                                "() => document.body && document.body.scrollHeight > 100 && (document.body.innerText||'').trim().length > 0",
                                timeout=4000,
                            )
                        except Exception:
                            pass
                        await asyncio.sleep(0.25)
                        png_bytes = await page.screenshot(
                            type="png", timeout=15000, full_page=False
                        )
                        if on_screenshot is not None:
                            try:
                                await on_screenshot(idx + 1, shot_name, png_bytes)
                            except Exception as cb_err:  # noqa: BLE001
                                logger.debug(f"on_screenshot callback failed for '{shot_name}': {cb_err}")
                    except Exception as ss_err:  # noqa: BLE001
                        logger.warning(f"screenshot step '{shot_name}' failed: {ss_err}")
                        # Surface failure to live activity so the operator
                        # can see WHICH capture didn't fire.
                        if on_screenshot is not None:
                            try:
                                # Best-effort retry without the load-state
                                # waits in case the page is mid-navigation.
                                png_bytes = await page.screenshot(
                                    type="png", timeout=10000, full_page=False
                                )
                                await on_screenshot(idx + 1, shot_name, png_bytes)
                            except Exception:
                                pass
                elif action in ("close", "close_browser", "browser_close"):
                    # ── 2026-05: User-requested explicit browser-close step ──
                    # The user noticed that visits keep the browser around
                    # for a moment after the LAST recorded step finishes
                    # (the per-visit context is closed in the parent
                    # finally-block, but that runs AFTER the offer page's
                    # post-submit redirect / 3rd-party pixel chain settles).
                    # On medium-RAM VPSes this stacks up: a finished visit
                    # still holds ~150-300 MB until the parent loop reaches
                    # the finally. Recording an explicit `close` step at the
                    # end of the JSON (or anywhere mid-flow if the operator
                    # has already captured the conversion) lets the operator
                    # free that browser IMMEDIATELY so the next worker can
                    # start. Safe because:
                    #   • page.close() / context.close() are idempotent —
                    #     the outer `finally` in process_one() still calls
                    #     context.close() but the second call is a no-op.
                    #   • We return status="ok" with a flag so the caller
                    #     knows the step deliberately terminated the visit
                    #     early (NOT a failure — `executed` is incremented).
                    #   • Any steps AFTER `close` in the JSON are ignored
                    #     by design — that's the whole point of this step.
                    try:
                        # Close the page first (releases DOM / JS heap).
                        await page.close()
                    except Exception as _pe:
                        logger.debug(f"close step: page.close() failed (continuing): {_pe}")
                    # Try to close the parent context too, if it's still
                    # accessible — this is what actually frees the proxy
                    # connection + cookies + storage.
                    try:
                        _ctx = page.context
                        if _ctx is not None:
                            await _ctx.close()
                    except Exception as _ce:
                        logger.debug(f"close step: context.close() failed (continuing): {_ce}")
                    executed += 1
                    if collect_timings:
                        step_results.append({
                            "idx": idx, "action": action,
                            "selector": "", "ok": True, "error": None,
                            "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                            "optional": optional,
                            "closed_early": True,
                        })
                    # Emit a final progress event so Live Activity grid
                    # shows "✓ closed" instead of the visit getting stuck
                    # at the prior step indefinitely.
                    if on_step_progress is not None:
                        try:
                            await on_step_progress({
                                "idx": idx + 1,
                                "action": action,
                                "selector": "",
                                "total_steps": len(steps or []),
                                "status": "ok",
                                "note": "browser closed by JSON step",
                                "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                                "timestamp_ms": int(_time_mod.time() * 1000),
                                "screenshot_b64": "",
                                "page_url": "",
                            })
                        except Exception:
                            pass
                    return {
                        "status": "ok",
                        "executed_steps": executed,
                        "closed_early": True,
                        **({"step_results": step_results,
                            "total_ms": int((_time_mod.perf_counter() - _t_total_start) * 1000)}
                           if collect_timings else {}),
                    }
                elif action in ("auto_continue", "auto_continue_survey"):
                    # 2026-01 — Smart multi-page survey continuation.
                    # After the user's custom JSON has filled the initial
                    # form & submitted, append this step at the end of
                    # the automation to let the engine drive the bot
                    # through any post-submit survey pages (Yes/No
                    # questions, "Continue", "Claim", deal cards, etc.)
                    # until either the conversion page is reached, the
                    # page stops changing, or max_iterations is hit.
                    # Re-uses the same helpers that the auto-fill heuristic
                    # already relies on (survey_click_v2 + complete_random_deals
                    # + _dismiss_popups + a generic "click any continue-like
                    # button" loop), so behaviour is consistent with the
                    # non-custom-JSON path.
                    try:
                        from rut_flash_helpers import (
                            survey_click_v2 as _ac_survey,
                            complete_random_deals as _ac_deals,
                        )
                    except Exception:
                        _ac_survey = None
                        _ac_deals = None
                    try:
                        from form_filler import _dismiss_popups as _ac_dismiss
                    except Exception:
                        _ac_dismiss = None
                    max_iter = int(step.get("max_iterations") or 15)
                    per_iter_wait_ms = int(step.get("iteration_wait_ms") or 1500)
                    stop_on_host = (step.get("stop_on_host") or "").strip().lower()
                    _ac_continue_js = (
                        "(function(){"
                        " var KW=['continue','next','submit','i agree','agree','confirm','claim','yes','accept','proceed','get started','unlock','finish','complete'];"
                        " function vis(e){try{var s=window.getComputedStyle(e);return s&&s.display!=='none'&&s.visibility!=='hidden'&&e.offsetWidth>0&&e.offsetHeight>0;}catch(_){return false;}}"
                        " var nodes=Array.from(document.querySelectorAll('button,a,input[type=button],input[type=submit],[role=button]'));"
                        " for(var i=0;i<nodes.length;i++){var el=nodes[i];if(!vis(el))continue;"
                        "  var t=((el.innerText||el.textContent||el.value||'')+'').toLowerCase().replace(/\\s+/g,' ').trim();"
                        "  if(!t)continue;"
                        "  for(var k=0;k<KW.length;k++){if(t===KW[k]||t.indexOf(KW[k])>=0){"
                        "    try{el.scrollIntoView({block:'center'});}catch(_){};"
                        "    if(el.tagName==='A'&&el.href&&!el.target){window.location.assign(el.href);return true;}"
                        "    el.click();"
                        "    var f=el.form||(el.closest&&el.closest('form'));"
                        "    if(f&&(el.type==='submit'||(el.getAttribute&&el.getAttribute('type')==='submit'))){setTimeout(function(){try{if(!f._krx_acs){f._krx_acs=true;f.submit();}}catch(_){}},150);}"
                        "    return true;"
                        "  }}"
                        " }"
                        " return false;"
                        "})()"
                    )
                    _last_url = ""
                    _no_change_count = 0
                    try:
                        _last_url = page.url or ""
                    except Exception:
                        pass
                    for _ai in range(max_iter):
                        progress = False
                        # Dismiss popups first so they don't block clicks
                        if _ac_dismiss is not None:
                            try:
                                await _ac_dismiss(page)
                            except Exception:
                                pass
                        # Stage 1: survey answer clicks (Yes/No chips, etc)
                        if _ac_survey is not None:
                            try:
                                _sres = await _ac_survey(page, max_iterations=3, picker=None)
                                if isinstance(_sres, dict) and int(_sres.get("clicks", 0)) > 0:
                                    progress = True
                            except Exception:
                                pass
                        # Stage 2: deal cards (FlashRewards/RetailProductsUSA)
                        if _ac_deals is not None:
                            try:
                                _dn = await _ac_deals(page, count_min=1, count_max=2)
                                if int(_dn or 0) > 0:
                                    progress = True
                            except Exception:
                                pass
                        # Stage 3: generic continue-button click
                        try:
                            _clicked = bool(await page.evaluate(_ac_continue_js))
                            if _clicked:
                                progress = True
                        except Exception:
                            pass
                        # Settle wait — let any nav / async render finish
                        try:
                            await page.wait_for_timeout(per_iter_wait_ms)
                        except Exception:
                            pass
                        try:
                            await page.wait_for_load_state("domcontentloaded", timeout=8000)
                        except Exception:
                            pass
                        # URL-change detection — stop early if nothing's
                        # changing for two consecutive iterations AND nothing
                        # was clicked this round.
                        try:
                            _cur = page.url or ""
                        except Exception:
                            _cur = _last_url
                        # Optional explicit stop signal: caller can pass
                        # stop_on_host="thank-you.example.com" to break the
                        # loop as soon as a conversion-style host is reached.
                        if stop_on_host:
                            try:
                                from urllib.parse import urlparse as _up
                                if stop_on_host in (_up(_cur).hostname or "").lower():
                                    break
                            except Exception:
                                pass
                        if _cur == _last_url and not progress:
                            _no_change_count += 1
                            if _no_change_count >= 2:
                                break
                        else:
                            _no_change_count = 0
                            _last_url = _cur
                else:
                    if not optional:
                        if collect_timings:
                            _step_ok = False
                            _step_err = f"Unknown action '{action}'"
                            step_results.append({
                                "idx": idx, "action": action,
                                "selector": (selector or "")[:200],
                                "ok": False, "error": _step_err,
                                "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                                "optional": optional,
                            })
                        return {"status": "failed", "error": f"Unknown action '{action}' at step {idx+1}", "executed_steps": executed,
                                **({"step_results": step_results} if collect_timings else {})}
                executed += 1
                if collect_timings:
                    step_results.append({
                        "idx": idx, "action": action,
                        "selector": (selector or "")[:200],
                        "ok": True, "error": None,
                        "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                        "optional": optional,
                    })
                # 2026-01: emit "ok" progress event + live screenshot
                if on_step_progress is not None:
                    # 2026-05 — stop heartbeat BEFORE we emit the final
                    # ok event so the UI receives "running heartbeat"
                    # last (a quick test would otherwise occasionally
                    # show a heartbeat tick AFTER ok).
                    if _hb_task is not None and not _hb_task.done():
                        _hb_task.cancel()
                        try:
                            await _hb_task
                        except (asyncio.CancelledError, Exception):
                            pass
                    # Capture a small JPEG of the current viewport so
                    # the frontend Live Activity panel can show a real
                    # "what is the page doing right now" image. Best-
                    # effort — failures don't break the automation.
                    # 2026-05 — switched to full_page=True so the
                    # operator can see the WHOLE offer page (scroll
                    # below the fold inside the expanded tile).
                    # Quality dropped to 35 because full-page JPEGs
                    # can be 5-10× bigger than viewport — at 35
                    # quality a typical 3000px-tall offer page is
                    # ~80-150KB which is fine for 800ms polling.
                    _live_b64 = ""
                    try:
                        import base64 as _b64
                        _shot = await page.screenshot(
                            type="jpeg", quality=35, full_page=True, timeout=4000,
                        )
                        # Hard cap on b64 payload size — if a page is
                        # unusually tall (50k px lazy-load mega-pages),
                        # fall back to viewport-only so the tile still
                        # updates instead of the polling endpoint
                        # choking on a 5 MB JSON blob.
                        if len(_shot) > 800_000:
                            _shot = await page.screenshot(
                                type="jpeg", quality=45, full_page=False, timeout=2500,
                            )
                        _live_b64 = "data:image/jpeg;base64," + _b64.b64encode(_shot).decode("ascii")
                    except Exception:
                        # Retry once with viewport-only — for cases
                        # where full_page fails (e.g. page mid-nav).
                        try:
                            import base64 as _b64
                            _shot = await page.screenshot(
                                type="jpeg", quality=50, full_page=False, timeout=2000,
                            )
                            _live_b64 = "data:image/jpeg;base64," + _b64.b64encode(_shot).decode("ascii")
                        except Exception:
                            pass

                    # ── 2026-05: Step-marker bounding box capture ──
                    # User ask: "Show Step Markers — har step ki target
                    # location screenshot pe colored dots overlay ho".
                    # We grab the resolved element's bounding box in
                    # FULL-PAGE coordinates (post-scroll-offset) right
                    # after a successful step so the frontend SVG
                    # overlay can position a dot at the exact target.
                    # Also includes `doc_size` so the frontend can
                    # scale the dots to the rendered image dimensions.
                    # Best-effort — failure leaves the keys absent so
                    # the marker for this step just isn't drawn.
                    _target_box = None
                    _doc_size = None
                    if selector and action in (
                        "click", "fill", "type", "select", "check",
                        "uncheck", "hover", "press",
                    ):
                        try:
                            _box_data = await page.evaluate(
                                """(sel) => {
                                    let el = null;
                                    if (sel.startsWith('xpath=')) {
                                        const xp = sel.slice(6);
                                        const r = document.evaluate(xp, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
                                        el = r.singleNodeValue;
                                    } else if (sel.startsWith('text=')) {
                                        const t = sel.slice(5).replace(/^"|"$/g, '');
                                        const all = document.querySelectorAll('button, a, input, label, span, div');
                                        for (const e of all) {
                                            if ((e.innerText || e.textContent || '').trim().includes(t)) {
                                                el = e; break;
                                            }
                                        }
                                    } else {
                                        try { el = document.querySelector(sel); } catch(_) {}
                                    }
                                    const docW = Math.max(
                                        document.documentElement.scrollWidth,
                                        document.body ? document.body.scrollWidth : 0,
                                    );
                                    const docH = Math.max(
                                        document.documentElement.scrollHeight,
                                        document.body ? document.body.scrollHeight : 0,
                                    );
                                    if (!el || !el.getBoundingClientRect) return {doc: {w: docW, h: docH}};
                                    const r = el.getBoundingClientRect();
                                    return {
                                        box: {
                                            x: Math.round(r.left + window.scrollX),
                                            y: Math.round(r.top + window.scrollY),
                                            w: Math.round(r.width),
                                            h: Math.round(r.height),
                                        },
                                        doc: {w: docW, h: docH},
                                    };
                                }""",
                                selector,
                            )
                            if isinstance(_box_data, dict):
                                if isinstance(_box_data.get("box"), dict):
                                    _target_box = _box_data["box"]
                                if isinstance(_box_data.get("doc"), dict):
                                    _doc_size = _box_data["doc"]
                        except Exception:
                            pass

                    try:
                        await on_step_progress({
                            "idx": idx,
                            "action": action,
                            "selector": (selector or "")[:200],
                            "total_steps": len(steps or []),
                            "status": "ok",
                            "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                            "timestamp_ms": int(_time_mod.time() * 1000),
                            "screenshot_b64": _live_b64,
                            "page_url": (page.url or "")[:300] if hasattr(page, "url") else "",
                            # 2026-05 Step-marker overlay payload
                            **({"target_box": _target_box} if _target_box else {}),
                            **({"doc_size": _doc_size} if _doc_size else {}),
                        })
                    except Exception:
                        pass
            except Exception as e:
                # 2026-05 — stop the heartbeat the moment a step errors,
                # otherwise the UI would keep showing "running 30s" while
                # the retry/self-heal code paths do their work below.
                if _hb_task is not None and not _hb_task.done():
                    _hb_task.cancel()
                    try:
                        await _hb_task
                    except (asyncio.CancelledError, Exception):
                        pass
                # 2026-01 (additive): per-step retry. When the step
                # carries `retry: N`, attempt the SAME step up to N more
                # times with `retry_delay` ms between attempts BEFORE
                # falling through to optional-skip / self-heal. Defaults
                # to 0 so existing recordings are unchanged.
                _retry_count = int(step.get("retry") or 0)
                _retry_delay_ms = int(step.get("retry_delay") or 1000)
                if _retry_count > 0:
                    _retry_step = dict(step)
                    _retry_step["retry"] = 0  # prevent infinite recursion
                    _retry_step["optional"] = False
                    _retry_last_err = e
                    _retry_recovered = False
                    for _rn in range(_retry_count):
                        try:
                            await asyncio.sleep(_retry_delay_ms / 1000.0)
                        except Exception:
                            pass
                        try:
                            _sub = await _execute_automation_steps(
                                page, row, [_retry_step],
                                skip_captcha=skip_captcha,
                                self_heal=False,
                                user_id=alias_user_id,
                            )
                            if _sub.get("status") == "ok":
                                _retry_recovered = True
                                logger.info(
                                    f"[retry] step #{idx+1} ({action}) recovered on attempt {_rn+2}/{_retry_count+1}"
                                )
                                break
                            _retry_last_err = Exception(_sub.get("error") or "retry failed")
                        except Exception as _re:
                            _retry_last_err = _re
                    if _retry_recovered:
                        executed += 1
                        if collect_timings:
                            step_results.append({
                                "idx": idx, "action": action,
                                "selector": (selector or "")[:200],
                                "ok": True,
                                "error": f"(retry-recovered) {str(e)[:120]}",
                                "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                                "optional": False,
                                "retry_recovered": True,
                            })
                        continue
                    # All retries exhausted — fall through with the last
                    # error so optional / self-heal still get a chance.
                    e = _retry_last_err
                if optional:
                    executed += 1
                    if collect_timings:
                        step_results.append({
                            "idx": idx, "action": action,
                            "selector": (selector or "")[:200],
                            "ok": True, "error": f"(skipped — optional) {str(e)[:140]}",
                            "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                            "optional": True,
                        })
                    continue
                # ── Self-heal: ask AI to propose a recovery action ──────
                if self_heal and heal_used < MAX_HEAL:
                    heal_used += 1
                    try:
                        heal_action = await _try_self_heal(page, step, str(e))
                    except Exception:
                        heal_action = None
                    if heal_action:
                        try:
                            await _execute_automation_steps(
                                page, row, [heal_action],
                                skip_captcha=skip_captcha, self_heal=False,
                                user_id=alias_user_id,
                            )
                        except Exception:
                            pass
                        # Retry the original step ONCE after recovery
                        try:
                            await _dispatch_single_action(
                                page, action, selector, value, step, timeout,
                                wait_nav, row,
                            )
                            executed += 1
                            if collect_timings:
                                step_results.append({
                                    "idx": idx, "action": action,
                                    "selector": (selector or "")[:200],
                                    "ok": True,
                                    "error": f"(self-healed) {str(e)[:120]}",
                                    "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                                    "optional": False,
                                    "self_healed": True,
                                })
                            continue
                        except Exception as e2:
                            if collect_timings:
                                step_results.append({
                                    "idx": idx, "action": action,
                                    "selector": (selector or "")[:200],
                                    "ok": False, "error": str(e2)[:200],
                                    "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                                    "optional": False,
                                })
                            return {"status": "failed",
                                    "error": f"Step {idx+1} ({action}) failed after self-heal: {str(e2)[:200]}",
                                    "executed_steps": executed,
                                    **({"step_results": step_results} if collect_timings else {})}
                if collect_timings:
                    step_results.append({
                        "idx": idx, "action": action,
                        "selector": (selector or "")[:200],
                        "ok": False, "error": str(e)[:200],
                        "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                        "optional": False,
                        # 2026-01: friendly Roman-Urdu/English hint
                        "friendly_hint": (_ext_friendly_error(str(e)) if _EXT_LOADED else ""),
                    })
                _hint = _ext_friendly_error(str(e)) if _EXT_LOADED else ""
                _err_msg = f"Step {idx+1} ({action}) failed: {str(e)[:200]}"
                if _hint:
                    _err_msg = f"{_err_msg} | Hint: {_hint}"
                # 2026-01: emit "failed" progress event + final screenshot
                if on_step_progress is not None:
                    _live_b64 = ""
                    try:
                        import base64 as _b64
                        # 2026-05 — full_page=True so operator can see
                        # the WHOLE page state at moment of failure
                        # (often the failed selector IS below the fold,
                        # which the old viewport-only capture missed).
                        _shot = await page.screenshot(
                            type="jpeg", quality=35, full_page=True, timeout=4000,
                        )
                        if len(_shot) > 800_000:
                            _shot = await page.screenshot(
                                type="jpeg", quality=45, full_page=False, timeout=2500,
                            )
                        _live_b64 = "data:image/jpeg;base64," + _b64.b64encode(_shot).decode("ascii")
                    except Exception:
                        try:
                            import base64 as _b64
                            _shot = await page.screenshot(
                                type="jpeg", quality=50, full_page=False, timeout=2000,
                            )
                            _live_b64 = "data:image/jpeg;base64," + _b64.b64encode(_shot).decode("ascii")
                        except Exception:
                            pass
                    try:
                        await on_step_progress({
                            "idx": idx,
                            "action": action,
                            "selector": (selector or "")[:200],
                            "total_steps": len(steps or []),
                            "status": "failed",
                            "error": str(e)[:200],
                            "friendly_hint": _hint,
                            "ms": int((_time_mod.perf_counter() - _t_step_start) * 1000),
                            "timestamp_ms": int(_time_mod.time() * 1000),
                            "screenshot_b64": _live_b64,
                            "page_url": (page.url or "")[:300] if hasattr(page, "url") else "",
                        })
                    except Exception:
                        pass
                return {"status": "failed", "error": _err_msg, "executed_steps": executed,
                        "failed_at_idx": idx, "remaining_steps": list(steps[idx+1:]),
                        "friendly_hint": _hint,
                        **({"step_results": step_results} if collect_timings else {})}
            finally:
                # 2026-05 — Belt-and-suspenders heartbeat cleanup.
                # Some action branches inside the inner try block
                # (close/auto_continue/etc.) `return` from the function
                # before reaching the success-path cancellation above.
                # This finally runs after EVERY step iteration (including
                # those early returns) so no orphan heartbeat task is
                # left writing into a closed Playwright page.
                if _hb_task is not None and not _hb_task.done():
                    _hb_task.cancel()
                    try:
                        await _hb_task
                    except (asyncio.CancelledError, Exception):
                        pass
        return {"status": "ok", "executed_steps": executed,
                **({"step_results": step_results, "total_ms": int((_time_mod.perf_counter() - _t_total_start) * 1000)} if collect_timings else {})}
    except Exception as e:
        return {"status": "failed", "error": f"Automation crashed: {str(e)[:200]}", "executed_steps": executed,
                **({"step_results": step_results} if collect_timings else {})}


async def _wait_for_actionable_selector(
    page: Page,
    selector: str,
    timeout: int,
    visible: bool = True,
    extra_alts: Optional[List[str]] = None,
) -> bool:
    """
    ── 2026-05: smart pre-wait for the actionable selector ──

    Before any fill / click / select / type, ACTIVELY wait for the
    element to appear in the DOM (and optionally become visible). This
    is the customer's most-requested fix: "agr first name input kr le
    to next agr last name hai to last name put hone tak auto wait phr
    next step hone tak auto wait".

    2026-01: `extra_alts` (user-saved selector aliases for self-healing
    replay). If the original selector fails AND we have aliases, try
    each in turn against the remaining budget so a known-good rename
    rescues the step transparently.

    Returns True if the selector became actionable; raises the
    underlying Playwright exception otherwise. (Caller's
    try/except + `optional` flag decides whether to swallow.)
    """
    if not selector:
        return True
    state = "visible" if visible else "attached"
    try:
        await page.wait_for_selector(selector, timeout=timeout, state=state)
        return True
    except Exception as primary_err:
        # 2026-01 (additive): iframe-aware fallback. Many forms today
        # live inside iframes (Stripe Checkout, embedded widgets,
        # third-party lead-gen forms). If the selector did not match
        # on the main page, scan all frames for it. Detection only —
        # the actual interaction will still happen on the original
        # page (Playwright's strict selector mode requires it). For
        # iframe-internal interactions users should record again
        # AFTER switching to the iframe — but at minimum detecting
        # presence here lets the actionable wait succeed instead of
        # raising a misleading "selector not found" error which
        # would abandon an otherwise-recoverable visit.
        if _EXT_LOADED:
            try:
                frame_hit = await _ext_find_frame(page, selector, timeout_ms=min(3000, timeout))
                if frame_hit is not None:
                    # Probe the frame with the requested visibility
                    # state. If it passes, we treat the selector as
                    # actionable — and the downstream click/fill will
                    # naturally route to the frame via Playwright's
                    # locator-by-frame resolver.
                    try:
                        await frame_hit.wait_for_selector(selector, timeout=min(3000, timeout), state=state)
                        logger.info(f"[iframe-fallback] '{selector[:60]}' located inside iframe — treating as actionable.")
                        return True
                    except Exception:
                        pass
            except Exception:
                pass
        # If user has saved aliases for this selector, try them — they
        # were proven correct by the user via the Edit modal, so they
        # deserve a real attempt before we give up.
        if not extra_alts:
            raise
        per_alt = max(800, timeout // max(1, len(extra_alts) + 1))
        last_err = primary_err
        for alt in extra_alts:
            alt = (alt or "").strip()
            if not alt or alt == selector:
                continue
            try:
                await page.wait_for_selector(alt, timeout=per_alt, state=state)
                logger.info(f"[selector_aliases] HIT — '{selector}' rescued by alias '{alt}'")
                return True
            except Exception as e:
                last_err = e
                continue
        raise last_err


async def _dispatch_single_action(page: Page, action: str, selector: str,
                                  value: Any, step: Dict[str, Any],
                                  timeout: int, wait_nav: bool,
                                  row: Dict[str, Any]) -> None:
    """Run one action (no retry, no self-heal). Raises on failure."""
    if action == "goto":
        await page.goto(value or selector, timeout=timeout, wait_until="domcontentloaded")
    elif action == "click":
        await page.click(selector, timeout=timeout)
        if wait_nav:
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
    elif action == "fill":
        # 2026-01 Anti-detect: route the recorded "fill" through the
        # human-typing helper too. Real users never paste an entire
        # value instantly into a field, even if the original recording
        # used Playwright's .fill(). Fall back to .fill() only if the
        # element can't be queried (e.g. selector refers to a hidden
        # form sync target).
        # NEW (2026-01) — per-step opt-out: `humanize: false` skips
        # human-typing for this step. See parallel block in
        # `_execute_automation_steps` for full rationale.
        if step.get("humanize") is False:
            await page.fill(selector, str(value), timeout=timeout)
        else:
            try:
                from form_filler import _human_type_field as _htf, _human_tab_or_pause as _htp
                el_h = await page.query_selector(selector)
                if el_h is not None:
                    ok_h = await _htf(page, el_h, str(value))
                    if ok_h:
                        await _htp(page)
                    else:
                        await page.fill(selector, str(value), timeout=timeout)
                else:
                    await page.fill(selector, str(value), timeout=timeout)
            except Exception:
                await page.fill(selector, str(value), timeout=timeout)
    elif action == "type":
        # 2026-01 Anti-detect: humanise per-char typing — variable
        # delay + occasional pause replaces the flat delay=50 which
        # detectors histogram as a bot signature.
        # Same per-step humanize opt-out as `fill`.
        if step.get("humanize") is False:
            await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
        else:
            try:
                from form_filler import _human_type_field as _htf, _human_tab_or_pause as _htp
                el_h = await page.query_selector(selector)
                if el_h is not None:
                    ok_h = await _htf(page, el_h, str(value))
                    if ok_h:
                        await _htp(page)
                    else:
                        await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
                else:
                    await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
            except Exception:
                await page.type(selector, str(value), delay=int(step.get("delay") or 50), timeout=timeout)
    elif action == "select":
        # 2026-05: routed through _smart_select_with_fallback so this
        # path benefits from selector fallbacks too (e.g. #birth_month
        # → [name="birth_month"] → select[name*="birth" i][name*="month" i]).
        # Match-strategy fallback (label → value → index) preserved from
        # the original Visual Recorder behaviour.
        await _smart_select_with_fallback(
            page, selector, value,
            match_by=str(step.get("match_by") or "label"),
            timeout=timeout,
        )
    elif action == "check":
        # 2026-05: route through smart helper — handles CSS-styled hidden
        # checkboxes (label-click / sibling-click / JS-set fallbacks).
        await _smart_check_with_fallback(
            page, selector, want_checked=True, timeout=timeout,
        )
    elif action == "uncheck":
        await _smart_check_with_fallback(
            page, selector, want_checked=False, timeout=timeout,
        )
    elif action == "press":
        await page.press(selector or "body", value or "Enter", timeout=timeout)
    elif action == "wait":
        await page.wait_for_timeout(int(step.get("ms") or 1000))
    elif action == "wait_for_selector":
        await page.wait_for_selector(selector, timeout=timeout, state=step.get("state") or "visible")
    elif action in ("wait_for_navigation", "wait_for_load", "wait_for_networkidle"):
        await page.wait_for_load_state("networkidle", timeout=timeout)
    elif action == "scroll":
        if selector:
            el = await page.query_selector(selector)
            if el:
                await el.scroll_into_view_if_needed()
        else:
            await page.evaluate(f"window.scrollBy(0,{int(step.get('y') or 500)})")
    elif action == "evaluate":
        js = _substitute(step.get("script") or step.get("js") or "", row)
        # Mirror the navigation-aware behaviour from
        # _execute_automation_steps so single-step self-heal and other
        # callers also benefit from waiting on JS-triggered navigation.
        _url_before = ""
        try:
            _url_before = page.url
        except Exception:
            pass
        # 2026-01: swallow "Execution context destroyed" if it was
        # caused by a navigation (URL changed). See parallel handler in
        # _execute_automation_steps for full rationale.
        try:
            await page.evaluate(js)
        except Exception as _ev_err:
            _ev_msg = str(_ev_err).lower()
            _navigated = False
            try:
                _navigated = (page.url != _url_before)
            except Exception:
                _navigated = False
            if (
                "execution context" in _ev_msg
                or "navigation" in _ev_msg
                or _navigated
            ):
                logger.info(f"[evaluate/dispatch] context destroyed by navigation — treating as success")
            else:
                raise
        try:
            for _ in range(25):  # up to ~2.5s
                await page.wait_for_timeout(100)
                try:
                    if page.url != _url_before:
                        try:
                            await page.wait_for_load_state("domcontentloaded", timeout=20000)
                        except Exception:
                            pass
                        try:
                            await page.wait_for_load_state("networkidle", timeout=15000)
                        except Exception:
                            pass
                        break
                except Exception:
                    pass
        except Exception:
            pass


async def _try_self_heal(page: Page, failed_step: Dict[str, Any],
                         error_msg: str) -> Optional[Dict[str, Any]]:
    """Take a screenshot + ask Gemini for a recovery action. Returns a step
    dict or None. Keeps the call short so it doesn't stall the job."""
    try:
        from ai_automation_generator import suggest_self_heal_action
    except Exception as e:
        logger.warning(f"self-heal disabled (import failed): {e}")
        return None

    try:
        import tempfile, os as _os
        tmpdir = tempfile.gettempdir()
        path = _os.path.join(tmpdir, f"rut_heal_{uuid.uuid4().hex[:8]}.png")
        try:
            await page.screenshot(path=path, full_page=False, timeout=5000)
        except Exception:
            return None
        title = ""
        url = ""
        try:
            title = await page.title()
            url = page.url
        except Exception:
            pass
        action = await suggest_self_heal_action(
            screenshot_path=path,
            page_title=title,
            page_url=url,
            failed_step={**failed_step, "_error": error_msg[:200]},
        )
        try:
            _os.remove(path)
        except Exception:
            pass
        return action
    except Exception as e:
        logger.warning(f"self-heal call failed: {e}")
        return None


async def _multi_step_fill(
    page: Page,
    row: Dict[str, Any],
    picker: Any = None,
    pre_submit_cb: Optional[Callable[[str], Awaitable[None]]] = None,
) -> Dict[str, Any]:
    """
    The pre_submit_cb (if provided) is invoked RIGHT BEFORE each attempt
    to click the page's main submit/continue button, with a short label
    like "stage_1" / "stage_2_retry". Callers typically use it to capture
    an intermediate screenshot so the user can inspect exactly what
    the form looked like at the moment the bot was about to submit it
    (fields filled, dropdowns selected, gender clicked, etc.). Failures
    inside the callback are swallowed — it MUST NOT break the visit.
    """
    # Auto-fill blank required fields (gender, DOB day/month/year, zip) so
    # FlashRewards-style forms don't fail for users whose Excel only has
    # name + address + email + phone.
    try:
        from rut_flash_helpers import (
            enrich_row_random,
            survey_click_random_answers,
            survey_click_v2,
            complete_random_deals,
        )
        enrich_row_random(row)
    except Exception:  # noqa: BLE001
        survey_click_random_answers = None  # type: ignore
        survey_click_v2 = None  # type: ignore
        complete_random_deals = None  # type: ignore

    max_steps = 10  # Increased for FlashRewards 6-stage flow (A→B→C→D→E→F)
    deals_completed_count = 0
    survey_clicks_count = 0
    survey_picks: List[Any] = []  # list of (q_sig, answer_text)

    for step in range(max_steps):
        await page.wait_for_timeout(500 + random.randint(0, 400))
        await _dismiss_popups(page)

        # FlashRewards-style multi-question survey: answers are <a>/<button>,
        # not <input>. _fill_form skips them. We click answers FIRST so the
        # survey advances to the registration form. AI-biased picker is used
        # when provided, else uniform random.
        # V2 handles BOTH Stage A (Yes/No pre-pop) AND Stage D (long survey
        # wall with multi-select + sponsored-ad bypass). V1 is fallback.
        survey_fn = survey_click_v2 if survey_click_v2 is not None else survey_click_random_answers
        if survey_fn is not None:
            try:
                if survey_fn is survey_click_v2:
                    sres = await survey_fn(
                        page, max_iterations=30, picker=picker,
                    )
                else:
                    sres = await survey_fn(
                        page, max_questions=8, picker=picker,
                    )
                if isinstance(sres, dict):
                    survey_clicks_count += int(sres.get("clicks", 0))
                    new_picks = sres.get("picks") or []
                    if isinstance(new_picks, list):
                        survey_picks.extend(new_picks)
                if survey_clicks_count:
                    await page.wait_for_timeout(800)
                    await _dismiss_popups(page)
            except Exception as e:  # noqa: BLE001
                logger.debug(f"survey click err: {e}")

        fill_info = await _fill_form(page, row)
        step_filled = len(fill_info.get("filled") or [])
        if step == 0 and step_filled == 0:
            # If we're on a deals page already (no form fields, just deal cards),
            # skip "no_fields_matched" and jump to deal completion.
            if complete_random_deals is not None:
                try:
                    deals_completed_count = await complete_random_deals(
                        page, count_min=2, count_max=3
                    )
                    if deals_completed_count >= 2:
                        return {
                            "status": "ok",
                            "deals_completed": deals_completed_count,
                            "survey_answers": survey_clicks_count,
                            "survey_picks": survey_picks,
                        }
                except Exception as e:  # noqa: BLE001
                    logger.debug(f"deals completion err: {e}")

            # Survey clicks happened or page may be still loading? Don't
            # give up on iteration 0 — wait extra 4s and retry the loop.
            # FlashRewards/RetailProductsUSA SPA can take 5-10s to render
            # answer chips on the very first page load behind a proxy.
            if survey_clicks_count > 0:
                # Survey is progressing — let next iteration try _fill_form
                # again on the new page after the click.
                continue
            try:
                await page.wait_for_timeout(4500)
                await page.wait_for_load_state("networkidle", timeout=5000)
            except Exception:
                pass
            # If on next iteration we STILL find nothing, the catch below
            # at step==max_steps-1 will return ok with survey count.
            continue
        if step > 0 and step_filled == 0:
            # All forms filled. Try deal completion before returning ok.
            if complete_random_deals is not None:
                try:
                    deals_completed_count = await complete_random_deals(
                        page, count_min=2, count_max=3
                    )
                except Exception as e:  # noqa: BLE001
                    logger.debug(f"deals completion err: {e}")
            return {
                "status": "ok",
                "deals_completed": deals_completed_count,
                "survey_answers": survey_clicks_count,
                "survey_picks": survey_picks,
            }
        await _tick_consent_checkboxes(page)
        try:
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight/2)")
        except Exception:
            pass
        start_url = page.url
        # Pre-submit checkpoint — callback captures screenshot so user can
        # inspect exactly what fields the bot filled before submitting.
        if pre_submit_cb is not None:
            try:
                await pre_submit_cb(f"stage_{step+1}")
            except Exception:
                pass
        await _click_submit(page)
        for _ in range(2):
            try:
                await page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass
            await page.wait_for_timeout(900)
            if page.url != start_url:
                break
            if await _dismiss_review_modal(page):
                await _tick_consent_checkboxes(page)
                await page.wait_for_timeout(400)
                if pre_submit_cb is not None:
                    try:
                        await pre_submit_cb(f"stage_{step+1}_retry")
                    except Exception:
                        pass
                await _click_submit(page)
                continue
            break
        if page.url == start_url:
            return {"status": "submitted_but_no_redirect"}

    # Final loop end: try deals one last time on whatever page we landed on.
    # Wait extra long so post-submit redirects to the deals page have time
    # to render — FlashRewards often chains 2-3 redirects after the email
    # signup before landing on the first deal selection screen.
    if complete_random_deals is not None and deals_completed_count < 2:
        try:
            for _wait_round in range(6):
                await page.wait_for_timeout(2500)
                try:
                    await page.wait_for_load_state("networkidle", timeout=8000)
                except Exception:
                    pass
                # Dismiss any popups appearing on the deals page
                try:
                    await _dismiss_popups(page)
                except Exception:
                    pass
                extra = await complete_random_deals(page, count_min=2, count_max=3)
                if extra:
                    deals_completed_count += extra
                    break
        except Exception as e:  # noqa: BLE001
            logger.debug(f"deals retry err: {e}")

    return {
        "status": "ok",
        "deals_completed": deals_completed_count,
        "survey_answers": survey_clicks_count,
        "survey_picks": survey_picks,
    }


async def _log_click_for_link(
    entry: Dict[str, Any],
    job_info: Dict[str, Any],
    main_db,
    early: bool = False,
) -> None:
    """Mirror the /api/t/ tracker: create a click document in the link
    owner's per-user DB and bump the link's click counter.

    2026-05: now supports an EARLY logging mode. When `early=True` this
    is called RIGHT AFTER `page.goto(offer_url)` succeeds (before any
    automation runs) so the exit-IP enters the user's clicks collection
    immediately. The click_doc is stamped with `entry['_early_click_id']`
    and a `visit_status='pending'` so the final call (with `early=False`,
    from `_record()`) updates the SAME document with the visit's final
    state (final_url, conversion_page_reached, real status) rather than
    inserting a second row. Result: 1 click row per visit (no double
    counting) AND the IP is duplicate-checkable seconds after goto
    instead of seconds after the entire 30-90s visit completes."""
    import uuid as _uuid
    link_id = job_info.get("link_id")
    owner_id = job_info.get("link_owner_id")
    short_code = job_info.get("link_short_code") or ""
    if not link_id or not owner_id or main_db is None:
        return
    try:
        # Access the per-user DB on the same client. IMPORTANT: Must match
        # server.py::get_user_db() exactly — that helper uses a 20-char
        # truncated, underscore-normalised key:
        #     f"krexion_user_{user_id.replace('-', '_')[:20]}"
        # If we use the raw owner_id (with hyphens) here, the click docs go
        # into a SEPARATE database and the dashboard / Clicks page reads
        # from the truncated DB and sees ZERO clicks — exactly the bug
        # users have reported ("tracker link use kia pr click count nahi hoa").
        client = main_db.client
        db_name = f"krexion_user_{owner_id.replace('-', '_')[:20]}"
        user_db = client[db_name]

        exit_ip = (entry.get("exit_ip") or "").strip()
        is_vpn = bool(entry.get("status") == "skipped_vpn" or entry.get("is_vpn"))
        ua = entry.get("ua") or ""
        device_display = entry.get("device_name") or entry.get("os") or "Unknown"

        # ── Branch A: end-of-visit UPDATE for early-logged clicks ─────
        early_id = entry.get("_early_click_id")
        if (not early) and early_id:
            try:
                await user_db.clicks.update_one(
                    {"id": early_id},
                    {"$set": {
                        "visit_status": entry.get("status") or "",
                        "final_url": entry.get("final_url") or "",
                        "conversion_page_reached": bool(entry.get("conversion_page_reached")),
                        "is_vpn": is_vpn,
                        # If we somehow learned a better/canonical exit
                        # IP between early-insert and now, refresh those
                        # fields too. exit_ip is usually unchanged.
                        **({"ip_address": exit_ip} if exit_ip else {}),
                        **({"ipv4": exit_ip} if exit_ip and ":" not in exit_ip else {}),
                    }},
                )
            except Exception as _ue:
                logger.warning(f"RUT click UPDATE failed: {_ue}")
            return

        # ── Branch B: regular INSERT (early=True OR no early row) ─────
        new_id = str(_uuid.uuid4())
        click_doc = {
            "id": new_id,
            "click_id": str(_uuid.uuid4()),
            "link_id": link_id,
            "user_id": owner_id,
            "short_code": short_code,
            "ip_address": exit_ip or "unknown",
            "ipv4": exit_ip if exit_ip and ":" not in exit_ip else "",
            "all_ips": [exit_ip] if exit_ip else [],
            "country": entry.get("country") or "Unknown",
            "city": entry.get("city") or "",
            "timezone": entry.get("timezone") or "",
            "is_vpn": is_vpn,
            "is_proxy": bool(entry.get("proxy")),
            "vpn_score": 0,
            "user_agent": ua,
            "device": (entry.get("os") or "desktop").lower(),
            "device_type": (entry.get("os") or "desktop").lower(),
            "device_display": device_display,
            "device_brand": "",
            "device_model": "",
            "os_name": entry.get("os") or "",
            "os_version": "",
            "browser": "Chrome",
            "browser_version": "",
            "referrer": "",
            "referrer_source": "rut",
            "referrer_source_name": "Real User Traffic",
            "source": "real_user_traffic",
            "visit_status": ("pending" if early else (entry.get("status") or "")),
            "final_url": entry.get("final_url") or "",
            "conversion_page_reached": bool(entry.get("conversion_page_reached")),
            "created_at": entry.get("timestamp") or datetime.now(timezone.utc).isoformat(),
        }
        await user_db.clicks.insert_one(click_doc)
        if early:
            # Stash the id so the end-of-visit call updates this row
            # instead of inserting a duplicate.
            entry["_early_click_id"] = new_id
        # Bump link-level click counter on the main DB (only once per
        # visit — on the EARLY insert if that path runs, else on the
        # late insert. The branch-A update path returns above without
        # bumping, so this is safe).
        await main_db.links.update_one({"id": link_id}, {"$inc": {"clicks": 1}})
    except Exception as e:
        # Best-effort — never crash the visit because click logging failed
        logger.warning(f"RUT click log failed (job_id-unknown, early={early}): {e}")


async def _record(
    job_id: str,
    entry: Dict[str, Any],
    report: List[Dict[str, Any]],
    lock: asyncio.Lock,
    db,
):
    async with lock:
        j = RUT_JOBS.setdefault(job_id, {})
        s = entry.get("status", "failed")

        # ── 2026-05: Silent-skip for burnt-IP / dead-proxy visits ──────
        # When `silent_skip_burnt_ip` is true (auto-enabled by ProxyJet
        # Auto Mode), visits where the ATTEMPT NEVER REACHED the offer
        # form because of an offer/tracker-side block or a proxy-side
        # failure are TREATED AS IF THEY NEVER HAPPENED from the UI's
        # perspective:
        #   • skipped_duplicate_ip — offer/tracker said "Duplicate IP"
        #   • skipped_vpn          — offer said "Please turn off VPN"
        #   • skipped_dead_proxy   — ProxyJet returned a proxy that
        #                            couldn't even open a TCP / SSL
        #                            connection to the offer (Read
        #                            timeout, 502, WRONG_VERSION_NUMBER…)
        #   • skipped_no_unique_ip — ProxyJet's pool exhausted state
        #                            without giving us a unique IP
        # All four categories share one trait: the user's lead row,
        # UA, and click budget were NEVER consumed against the offer
        # — they're pure "infrastructure noise". Hiding them keeps the
        # UI to a clean "what actually happened against the offer"
        # view. The dispatcher's HARD_CAP-bounded while-loop keeps
        # spawning more visits until `total_clicks` VISIBLE visits
        # have been recorded.
        _SILENT_SKIP_STATUSES = (
            "skipped_duplicate_ip",
            "skipped_vpn",
            "skipped_dead_proxy",
            "skipped_no_unique_ip",
        )
        silent_skip = (
            s in _SILENT_SKIP_STATUSES
            and bool(j.get("silent_skip_burnt_ip"))
        )
        if silent_skip:
            # Tear down the early-logged click_doc + counter bump
            # (only burnt-IP visits get an early click_doc — the
            # dead-proxy / no-unique-IP paths bail out before
            # page.goto so they don't have one. Safe to attempt
            # delete either way — delete_one of a non-existent id
            # is a no-op.)
            _early_id = entry.get("_early_click_id")
            _link_id = j.get("link_id")
            _owner_id = j.get("link_owner_id")
            if _early_id and _link_id and _owner_id and db is not None:
                try:
                    client = db.client
                    db_name = f"krexion_user_{_owner_id.replace('-', '_')[:20]}"
                    user_db = client[db_name]
                    await user_db.clicks.delete_one({"id": _early_id})
                except Exception as _de:
                    logger.debug(f"silent-skip click delete failed: {_de}")
                try:
                    await db.links.update_one(
                        {"id": _link_id}, {"$inc": {"clicks": -1}}
                    )
                except Exception as _de2:
                    logger.debug(f"silent-skip link counter decrement failed: {_de2}")
            # Bump diagnostics counter only.
            j["silent_skip_count"] = int(j.get("silent_skip_count") or 0) + 1
            # Also keep a breakdown (helpful in Diagnostics panel).
            _bd = j.setdefault("silent_skip_breakdown", {})
            _bd[s] = int(_bd.get(s) or 0) + 1
            # NOTE: We intentionally skip processed++, counter_key++,
            # report.append, events.append. The visit is invisible
            # from the UI but its IP burn IS persistent.
            return

        # ── Visible visit accounting (original logic, unchanged) ─────────
        j["processed"] = int(j.get("processed") or 0) + 1
        key_map = {
            "ok": "succeeded",
            "skipped_captcha": "skipped_captcha",
            "skipped_country": "skipped_country",
            "skipped_os": "skipped_os",
            "skipped_duplicate_ip": "skipped_duplicate_ip",
            "skipped_vpn": "skipped_vpn",
            "skipped_state_mismatch": "skipped_state_mismatch",
            "skipped_no_unique_ip": "skipped_no_unique_ip",
            "skipped_dead_proxy": "skipped_dead_proxy",
            "invalid_data": "invalid_data",
        }
        counter_key = key_map.get(s, None)
        # 2026-01 — Catch-all for any future "skipped_*" status so a
        # proxy/quality skip never gets miscounted as a hard failure.
        if counter_key is None and isinstance(s, str) and s.startswith("skipped"):
            counter_key = s
            j.setdefault(s, 0)
        if counter_key is None:
            counter_key = "failed"
        j[counter_key] = int(j.get(counter_key) or 0) + 1
        # Conversion counter: visits where final URL redirected OFF the form page
        if entry.get("conversion_page_reached"):
            j["conversions"] = int(j.get("conversions") or 0) + 1
        report.append(entry)

        # ── Log this visit as a click against the link (so dashboard's
        #    Clicks page + duplicate-IP detection both see RUT traffic) ──
        try:
            await _log_click_for_link(entry, j, db)
        except Exception:
            pass

        events = j.setdefault("events", [])
        events.append({
            "row": entry["visit_index"],
            "status": entry["status"],
            "proxy": entry["proxy"],
            "exit_ip": entry["exit_ip"],
            "country": entry["country"],
            "city": entry["city"],
            "device": f"{entry.get('device_name') or entry['os']} · {entry['viewport']}",
            "final_url": entry["final_url"],
            "conversion_page_reached": bool(entry.get("conversion_page_reached")),
            # FlashRewards-style metrics (visible to UI live-tracking)
            "survey_answers": int(entry.get("survey_answers") or 0),
            "deals_completed": int(entry.get("deals_completed") or 0),
            "thank_you_reached": bool(entry.get("thank_you_reached")),
            "row_index": int(entry.get("row_index") or 0),
            # Target Screenshot Verification metrics — null when feature unused
            "screenshot_match": entry.get("screenshot_match"),
            "screenshot_match_similarity": entry.get("screenshot_match_similarity"),
            "screenshot_match_distance": entry.get("screenshot_match_distance"),
            "error": (entry["error"] or "")[:140],
            "ts": entry["timestamp"],
        })
        if len(events) > 80:
            del events[:-80]
        if db is not None and j["processed"] % 5 == 0:
            try:
                await _persist(db, job_id)
            except Exception:
                pass


def _did_reach_conversion(landing_url: str, final_url: str) -> bool:
    """True if after form submit the user ended up on a DIFFERENT host than
    the original landing (e.g. thnkspg.com after stimulusassistforall.com),
    OR on the same host but on a different page stem than the form page
    (indexform / index-form / index). Indicates the offer accepted the lead
    and redirected the user forward."""
    try:
        from urllib.parse import urlparse
        lu = urlparse(landing_url or "")
        fu = urlparse(final_url or "")
        lh = (lu.netloc or "").lower().lstrip("www.")
        fh = (fu.netloc or "").lower().lstrip("www.")
        if not fh:
            return False
        if lh and fh and lh != fh:
            # Different second-level domain → classic conversion redirect
            lh_root = ".".join(lh.split(".")[-2:])
            fh_root = ".".join(fh.split(".")[-2:])
            if lh_root != fh_root:
                return True
        # Same host: compare path stems. Landing is usually index-form.php
        # or indexform.php; a conversion moves to offers-flow.php etc.
        form_stems = ("index-form", "indexform", "/index.php", "/index-form.php")
        lp = (lu.path or "").lower()
        fp = (fu.path or "").lower()
        if fp and lp != fp:
            if not any(s in fp for s in form_stems):
                return True
        return False
    except Exception:
        return False


# ─── Strict thank-you page detection ──────────────────────────────
# Per user request: only count a visit as a CONVERSION (and only take the
# final screenshot) when we are confident the browser reached the offer's
# thank-you / confirmation page. We combine three signals:
#   1. Host root changed vs. the landing page (e.g. stimulusassistforall.com
#      → thnkspg.com)
#   2. URL path / query contains a thank-you / success / claim / offer keyword
#   3. Page title or body text contains strong thank-you text (e.g.
#      "Claim Your $750", "Thank You", "Ways to Earn", "Congratulations")
# At least TWO of the three positive signals must match → avoids counting
# captcha-redirects, error pages, or same-host follow-ups as conversions.
_THANKYOU_URL_KEYWORDS = [
    "thank", "thanks", "thnk", "/ty", "ty.php", "thnks", "thnkspg",
    "success", "confirm", "confirmation", "completed",
    "claim", "offer", "offers-flow", "offer-flow", "offerwall", "offer-wall",
    "reward", "congrat", "congrats", "received", "submitted",
]
_THANKYOU_TEXT_KEYWORDS = [
    "thank you", "thank-you", "thankyou", "thanks for",
    "congratulations", "congrats",
    "claim your", "claim $", "claim 1 deal", "claim 1 prize",
    "your reward", "your prize", "your claim",
    "successfully submitted", "submission received", "submission successful",
    "we received", "we've received", "order confirmation",
    "complete paid offers", "complete the offers", "ways to earn",
    "pending offers", "pending offer", "offer wall",
    "you qualify", "you're qualified", "you have been matched",
    # FlashRewards / Reward4Spot deal-page markers — exact text seen on
    # eward4spot.com after a successful pre-pop → survey → deals flow.
    "best match for you", "level 1 deals", "complete 1 deal",
    "complete 1 deal on this level", "my deals", "my reward", "how to earn",
    "$750 target", "$750 reward", "target reward", "next step:",
]
# Page texts that STRONGLY indicate we're still on the form page (not converted)
_FORM_PAGE_TEXT_NEGATIVES = [
    "enter your first name", "fill out the form", "complete the form below",
    "please fill", "submit below", "please correct the errors",
]

# ─────────────────────────────────────────────────────────────────────
# HIGH-CONFIDENCE CONVERSION HOSTS — any visit that ends on one of these
# domains is ALWAYS counted as a conversion, regardless of pHash match,
# host-change, or text-keyword checks. These are the actual deals/offer-
# wall pages of the known partner networks — landing here literally means
# the user completed the funnel. Added for RetailProductsUSA / FlashRewards
# / Reward4Spot / DisplayOpt / PrizeGrab flow.
#
# Match is a SUFFIX match on the lower-cased netloc (after stripping
# "www."), so sub-domains are included (e.g. `offers.eward4spot.com`
# also counts). Do NOT add the user's OWN domain (krexion.com) or
# the tracker host here — that would always false-positive.
# ─────────────────────────────────────────────────────────────────────
_HIGH_CONFIDENCE_CONVERSION_HOSTS = [
    "eward4spot.com",          # Reward4Spot deals page (FlashRewards final)
    "reward4spot.com",         # alt spelling — some campaigns use this
    "flashrewards.club",       # FlashRewards legacy offer wall
    "flashrewards.online",     # FlashRewards new offer wall
    "prizegrab.com",           # PrizeGrab deal page
    "offers.displayoptoffers.com",  # DisplayOpt final offer wall
    "swagify.io",              # Swagify reward wall (sometimes seen in FR)
]


def _matches_high_confidence_host(netloc: str) -> bool:
    """True if the given lower-cased netloc ends with any of the known
    deal-page hosts. Strips a leading 'www.' prefix first."""
    if not netloc:
        return False
    h = netloc.lower().lstrip()
    if h.startswith("www."):
        h = h[4:]
    return any(h == dom or h.endswith("." + dom) for dom in _HIGH_CONFIDENCE_CONVERSION_HOSTS)


def _is_thank_you_page(landing_url: str, final_url: str,
                       page_text: str = "", page_title: str = "") -> bool:
    """Strict thank-you / conversion page check. Returns True only when at
    least TWO of {host-change, URL-keyword, page-text-keyword} match."""
    try:
        from urllib.parse import urlparse
        lu = urlparse(landing_url or "")
        fu = urlparse(final_url or "")
    except Exception:
        return False

    if not fu.netloc:
        return False

    lh = (lu.netloc or "").lower().lstrip("www.")
    fh = (fu.netloc or "").lower().lstrip("www.")

    # 0. HIGH-CONFIDENCE HOST SHORT-CIRCUIT — if the final URL lands on one
    # of the known deal-page / offer-wall hosts (e.g. eward4spot.com), we
    # treat it as a conversion unconditionally. This is the single biggest
    # signal and overrides viewport/pHash mismatches that otherwise waste
    # real conversions. Host is matched as a suffix so sub-domains count.
    if _matches_high_confidence_host(fh):
        return True

    # 1. host root change
    host_changed = False
    if lh and fh and lh != fh:
        lh_root = ".".join(lh.split(".")[-2:])
        fh_root = ".".join(fh.split(".")[-2:])
        host_changed = lh_root != fh_root

    # 2. URL keyword match
    full_url = ((fu.geturl() or "") + " " + (fu.path or "") + " " + (fu.query or "")).lower()
    url_keyword_hit = any(k in full_url for k in _THANKYOU_URL_KEYWORDS)

    # 3. page text / title keyword match
    text_combined = ((page_title or "") + " " + (page_text or "")).lower()[:6000]
    text_keyword_hit = any(k in text_combined for k in _THANKYOU_TEXT_KEYWORDS)

    # Strong negative: if we see clear form-page text, require EXTRA evidence
    is_still_on_form = any(k in text_combined for k in _FORM_PAGE_TEXT_NEGATIVES)

    positives = sum([host_changed, url_keyword_hit, text_keyword_hit])
    if is_still_on_form:
        # Need all three signals to overrule strong form-page evidence
        return positives >= 3
    return positives >= 2


def _write_excel_report(out_path: Path, report: List[Dict[str, Any]]):
    if not report:
        df = pd.DataFrame([{"info": "no visits completed"}])
    else:
        df = pd.DataFrame(report)
        # nice column order
        preferred = [
            "visit_index", "status", "conversion_page_reached", "proxy", "exit_ip",
            "country", "city",
            "timezone", "locale", "os", "viewport", "device_scale_factor",
            "hardware_concurrency", "device_memory", "webgl_renderer", "canvas_seed",
            "ua", "http_status", "landing_url", "final_url", "trusted_form", "lead_id",
            "screenshot", "error", "timestamp",
        ]
        cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
        df = df[cols]
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name="Visits", index=False)
        if report:
            # Summary sheet
            status_counts = df["status"].value_counts().reset_index()
            status_counts.columns = ["status", "count"]
            status_counts.to_excel(xl, sheet_name="Summary", index=False)
            # Conversion summary
            if "conversion_page_reached" in df.columns:
                conv_true = int(df["conversion_page_reached"].fillna(False).astype(bool).sum())
                conv_total = int(len(df))
                conv_ok = int(((df["status"] == "ok") & (df["conversion_page_reached"].fillna(False).astype(bool))).sum())
                conv_df = pd.DataFrame([
                    {"metric": "total_visits", "count": conv_total},
                    {"metric": "status_ok", "count": int((df["status"] == "ok").sum())},
                    {"metric": "conversion_page_reached", "count": conv_true},
                    {"metric": "conversion_and_ok", "count": conv_ok},
                    {"metric": "conversion_rate_pct",
                     "count": round(100.0 * conv_true / conv_total, 2) if conv_total else 0.0},
                ])
                conv_df.to_excel(xl, sheet_name="Conversions", index=False)


def _finalise(job_id: str, status: str, error: str = ""):
    j = RUT_JOBS.setdefault(job_id, {})
    j["status"] = status
    if error:
        j["error"] = error
    j["finished_at"] = datetime.now(timezone.utc).isoformat()


async def _finalise_and_persist(db, job_id: str, status: str, error: str = ""):
    """Same as _finalise but ALSO writes the failed/stopped state to MongoDB
    so the Past Jobs row + REST endpoint reflect the error message instead
    of leaving the job stuck on 'queued' forever."""
    _finalise(job_id, status, error)
    if db is not None:
        try:
            await _persist(db, job_id)
        except Exception as e:
            logger.debug(f"_finalise_and_persist persist failed: {e}")


async def _persist(db, job_id: str):
    j = RUT_JOBS.get(job_id, {})
    if not j:
        return
    # Filter out non-serializable entries (e.g. asyncio.Event)
    safe = {k: v for k, v in j.items() if not k.startswith("_")}
    try:
        await db.real_user_traffic_jobs.update_one(
            {"job_id": job_id},
            {"$set": {**safe, "job_id": job_id}},
            upsert=True,
        )
    except Exception as e:
        logger.debug(f"persist real_user_traffic_jobs failed: {e}")
    # ── Flush macro-leak + stuck-event diagnostics into MongoDB ──
    # We accumulate these in-memory during the run (cheap, non-blocking
    # for Playwright handlers) and flush at job-finalise so the admin
    # "Diagnostics" view can show exactly which URLs leaked unfilled
    # macros and where each visit got stuck. Bounded to 200 records per
    # bucket per job to keep MongoDB writes small.
    try:
        leaks = _MACRO_LEAK_BUFFER.pop(job_id, []) or []
        if leaks:
            try:
                await db.rut_diagnostics.insert_many([
                    {**ev, "kind": "macro_leak"} for ev in leaks
                ])
            except Exception as _ie:
                logger.debug(f"macro-leak flush failed: {_ie}")
    except Exception:
        pass
    try:
        stucks = _STUCK_EVENT_BUFFER.pop(job_id, []) or []
        if stucks:
            try:
                await db.rut_diagnostics.insert_many([
                    {**ev, "kind": "stuck"} for ev in stucks
                ])
            except Exception as _ie:
                logger.debug(f"stuck-event flush failed: {_ie}")
    except Exception:
        pass


def request_job_cancel(job_id: str) -> bool:
    """Flip the in-memory cancel flag on a running job. Returns True if the
    job was found and signalled, False otherwise (job finished / unknown)."""
    j = RUT_JOBS.get(job_id)
    if not j:
        return False
    ev = j.get("_cancel_event")
    if ev is None:
        return False
    try:
        ev.set()
        j["cancel_requested_at"] = datetime.now(timezone.utc).isoformat()
        return True
    except Exception:
        return False


def create_rut_job(
    job_id: str,
    user_id: str,
    target_url: str,
    total: int,
    form_fill_enabled: bool,
) -> Dict[str, Any]:
    RUT_JOBS[job_id] = {
        "job_id": job_id,
        "user_id": user_id,
        "target_url": target_url,
        "total": total,
        "form_fill_enabled": form_fill_enabled,
        "status": "queued",
        "processed": 0,
        "succeeded": 0,
        "skipped_captcha": 0,
        "skipped_country": 0,
        "skipped_os": 0,
        "skipped_duplicate_ip": 0,
        "skipped_vpn": 0,
        "skipped_state_mismatch": 0,
        "skipped_no_unique_ip": 0,
        "skipped_dead_proxy": 0,
        "failed": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    return RUT_JOBS[job_id]


def cleanup_rut_job(job_id: str):
    import shutil
    d = RESULTS_ROOT / job_id
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)
    RUT_JOBS.pop(job_id, None)


# ── Live step-log (for the "what's happening now" modal) ────────────
# Each job keeps a bounded ring-buffer of recent steps so the UI can stream
# them without any backend cost until the modal is actually opened.
_MAX_LIVE_STEPS = 300

def push_live_step(job_id: str, visit: int, stage: str, status: str, detail: str = "",
                   screenshot: str = ""):
    j = RUT_JOBS.get(job_id)
    if j is None:
        return
    buf = j.setdefault("live_steps", [])
    buf.append({
        "idx": len(buf) + 1,
        "visit": visit,
        "stage": stage,              # "setup" | "geo" | "filter" | "browser" | "form" | "submit" | "done"
        "status": status,            # "info" | "ok" | "skipped" | "failed"
        "detail": (detail or "")[:200],
        "screenshot": screenshot,    # filename only; served by /jobs/{id}/screenshot/{file}
        "ts": datetime.now(timezone.utc).isoformat(),
    })
    # Cap buffer
    if len(buf) > _MAX_LIVE_STEPS:
        del buf[:-_MAX_LIVE_STEPS]
    # 2026-01 (additive): also mirror the most recent stage into
    # `live_visits` so the Visual Live Grid shows ALL active visits —
    # including ones still in proxy/browser setup BEFORE the automation
    # steps callback starts firing. Without this, the grid stays empty
    # for the first ~5-15s of each visit (proxy pick + browser launch),
    # giving the false impression that nothing is happening.
    # Skip visit==0 (job-level events that aren't tied to a single
    # visit slot, e.g. "preflight" / "engine_check").
    if visit and isinstance(visit, int) and visit > 0:
        lv = j.setdefault("live_visits", {})
        vkey = str(visit)
        v = lv.setdefault(vkey, {
            "visit_idx": visit,
            "started_at": time.time(),
            "events_count": 0,
            "latest_event": None,
            "latest_frame_b64": "",
            "page_url": "",
            "status": "running",
        })
        # Only overwrite latest_event when the new event is NOT a
        # screenshot-bearing _execute_automation_steps callback (those
        # carry a richer payload via _visit_progress_cb and we want
        # them to take precedence over generic stage messages).
        existing = v.get("latest_event") or {}
        if not (existing.get("action") and existing.get("status") == "running"):
            v["latest_event"] = {
                "stage": stage,
                "status": status,
                "detail": (detail or "")[:200],
                "timestamp_ms": int(time.time() * 1000),
            }
        v["events_count"] = int(v.get("events_count", 0)) + 1
        if status == "failed":
            v["status"] = "failed"
        v["last_update"] = time.time()


def get_live_steps(job_id: str, since: int = 0) -> Dict[str, Any]:
    """Return steps with idx > since (used by the frontend modal)."""
    j = RUT_JOBS.get(job_id)
    if j is None:
        return {"steps": [], "cursor": since, "running": False}
    buf = j.get("live_steps") or []
    new = [s for s in buf if s.get("idx", 0) > since]
    return {
        "steps": new,
        "cursor": buf[-1]["idx"] if buf else since,
        "running": j.get("status") in ("running", "queued", "preparing"),
        "status": j.get("status", "unknown"),
        "processed": j.get("processed", 0),
        "total": j.get("total", 0),
    }


def cancel_visit(job_id: str, visit_index: int) -> Dict[str, Any]:
    """2026-05 — Manually abort ONE in-flight visit without stopping
    the whole job.

    Called by the per-tile "kill" button in the Live Visual Grid UI.
    The dispatcher's spawn loop will automatically replenish the freed
    concurrency slot with the next pending visit.

    Returns a small status dict — never raises.
    """
    j = RUT_JOBS.get(job_id)
    if j is None:
        return {"ok": False, "reason": "job_not_found"}
    if j.get("status") not in ("running", "queued", "preparing"):
        return {"ok": False, "reason": "job_not_running"}

    vid = str(int(visit_index))
    visit_tasks = j.get("_visit_tasks") or {}
    t = visit_tasks.get(vid)
    if t is None:
        return {"ok": False, "reason": "visit_not_found_or_already_done"}
    if t.done():
        return {"ok": False, "reason": "already_done"}

    try:
        t.cancel()
    except Exception as _e:  # noqa: BLE001
        return {"ok": False, "reason": f"cancel_failed: {_e}"}

    # Mirror into live_visits so the UI sees the tile flip to "cancelled"
    # on the next 800 ms poll, even if process_one's finally-block has
    # not yet finished cleaning up the Playwright context.
    #
    # Order matters here: push_live_step writes to the same dict and
    # would overwrite v["status"] = "cancelled" with "failed" because
    # of its built-in `if status == "failed": v["status"] = "failed"`
    # branch. We therefore call push_live_step FIRST and stamp the
    # "cancelled" status LAST so it wins.
    try:
        push_live_step(
            job_id, int(visit_index), "manual_cancel", "failed",
            "Visit cancelled by user — slot freed for next visit",
        )
    except Exception:
        pass

    try:
        lv = j.setdefault("live_visits", {})
        v = lv.setdefault(vid, {})
        v["status"] = "cancelled"
        v["latest_event"] = {
            "stage": "manual_cancel",
            "status": "failed",
            "detail": "Cancelled by user from Live Visual Grid",
            "timestamp_ms": int(time.time() * 1000),
        }
        v["last_update"] = time.time()
    except Exception:
        pass

    return {"ok": True, "visit_index": int(visit_index)}
