"""
Demo Real-User-Traffic test:
- Loads the user-provided JSON automation script
- Loads test 01.xlsx (181 rows of lead data)
- Launches headed-style headless Chromium (no proxy — preview-pod IP)
- Walks through the same _execute_automation_steps() path the prod
  job uses, taking screenshots at every screenshot/key step.
- Prints the final URL + heuristic conversion check.

NOTE: this is a DEMO run only. Without a residential proxy the offer
will most likely flag the IP as bot/hosting and may short-circuit the
flow before the final "deal" page — but the path through the form-fill
+ JSON steps is what we want to validate here. If the form is filled
correctly and submission triggers a navigation, we know the script +
data combination works; the production VPS run with proxies should
then succeed.
"""
import asyncio
import json
import os
import sys
import time
from pathlib import Path

import openpyxl

# Make the backend package importable
sys.path.insert(0, "/app/backend")
from real_user_traffic import _execute_automation_steps, _block_unfilled_macro_request  # noqa: E402

OFFER_URL = os.environ.get(
    "DEMO_OFFER_URL",
    "https://giftclick.org/aff_c?offer_id=250&aff_id=179687",
)

DEMO_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/121.0.0.0 Safari/537.36"
)

SCRIPT_PATH = Path("/tmp/rut_script.txt")
XLSX_PATH = Path("/tmp/test01.xlsx")
OUT_DIR = Path("/app/backend/tests/demo_results")
OUT_DIR.mkdir(parents=True, exist_ok=True)


async def _progress_logger(page, started_ts):
    while True:
        await asyncio.sleep(5)
        try:
            print(f"  [{time.time()-started_ts:6.1f}s] url={page.url}", flush=True)
        except Exception:
            pass


def load_row(idx: int = 0) -> dict:
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1] if c.value]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    raw = rows[idx]
    row = {h: (str(raw[i]) if raw[i] is not None else "") for i, h in enumerate(headers)}
    # Normalise into the keys the JSON template expects
    row["first"] = row.get("first", "")
    row["last"] = row.get("last", "")
    row["address"] = row.get("address", "")
    row["city"] = row.get("city", "")
    row["state"] = row.get("state", "")
    row["zip"] = row.get("zip", "")
    row["cellphone"] = row.get("cellphone", "")
    row["email"] = row.get("email", "")
    row["month"] = row.get("month", "")
    row["day"] = row.get("day", "")
    row["year"] = row.get("year", "")
    return row


def load_steps() -> list:
    with open(SCRIPT_PATH) as f:
        return json.load(f)


async def main():
    from playwright.async_api import async_playwright

    row = load_row(0)
    steps = load_steps()
    print("Row :", {k: row[k] for k in ("first", "last", "state", "zip", "cellphone", "month", "day", "year")})
    print("Steps:", len(steps))

    started = time.time()
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--headless=new",
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-features=IsolateOrigins,site-per-process",
                "--disable-dev-shm-usage",
            ],
        )
        ctx = await browser.new_context(
            user_agent=DEMO_UA,
            viewport={"width": 1366, "height": 768},
            locale="en-US",
        )
        # Apply the same defensive guard the prod RUT runner uses
        await ctx.route(
            "**/*",
            lambda route, request: asyncio.ensure_future(
                _block_unfilled_macro_request(route, request)
            ),
        )
        page = await ctx.new_page()

        url_log = []

        def _on_nav(frame):
            if frame == page.main_frame:
                try:
                    url_log.append((time.time() - started, frame.url))
                except Exception:
                    pass

        page.on("framenavigated", _on_nav)

        print(f"Goto: {OFFER_URL}")
        try:
            await page.goto(OFFER_URL, timeout=60000, wait_until="domcontentloaded")
        except Exception as e:
            print("Initial goto error:", e)

        shot_idx = [0]
        async def on_shot(idx, name, png):
            shot_idx[0] += 1
            out = OUT_DIR / f"{shot_idx[0]:02d}_{idx}_{name}.png"
            out.write_bytes(png)
            print(f"  screenshot → {out.name}  ({len(png)} bytes)")

        # Periodically log current page URL so we can see progress
        progress_task = asyncio.create_task(_progress_logger(page, started))
        try:
            result = await _execute_automation_steps(
                page, row, steps, skip_captcha=False, self_heal=False,
                on_screenshot=on_shot,
            )
        finally:
            progress_task.cancel()
        print("Result:", result, flush=True)

        # Take a final screenshot regardless
        try:
            final_png = await page.screenshot(full_page=False)
            (OUT_DIR / "ZZ_final.png").write_bytes(final_png)
        except Exception:
            pass

        print("Final URL:", page.url)
        print("Navigation timeline (last 15):")
        for ts, u in url_log[-15:]:
            print(f"  +{ts:6.2f}s  {u}")

        await browser.close()

    print("Done in", round(time.time() - started, 1), "s")


if __name__ == "__main__":
    asyncio.run(main())
